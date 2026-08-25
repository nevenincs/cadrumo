"""A registry revision must not span an AEAT record-design re-layout.

A revision carries exactly ONE export layout, so every filing year inside its
period selector is written at the SAME byte offsets. AEAT re-lays out its Diseños
de Registro whenever a block gains a rung, and every downstream offset shifts. A
revision whose span crosses such a boundary therefore encodes one byte layout
across two incompatible designs, and one of them is wrong.

The harm is not a wrong number in a box. Measured on Modelo 390: an export for a
filing year on the older side of a boundary succeeds and produces bytes laid out
for the newer design, so a reader using the correct design for that year finds
the declared total ABSENT where it expects it, a real value in a different box's
slot, and content past the end of the record.

WHY THIS IS KEYED ON DESIGN-TO-DESIGN AGREEMENT rather than on comparing each
layout field against its published box. The obvious framing needs a casilla ->
official box number mapping, and on several modelos that mapping barely exists:
casillas are semantic ids, the design is box-numbered, and the two vocabularies
do not intersect. A number-keyed check would report hundreds of false absences on
Modelo 390 alone. Design-to-design agreement needs no such mapping and states the
actual defect: one layout cannot cover years AEAT laid out differently.

The expected offsets are PARSED out of the bundled designs, never transcribed, so
the corpus is the authority and an author's misreading cannot enter the gate.

WHAT THIS DOES NOT CHECK, stated plainly so its silence is not over-read. It does
NOT verify that a revision's layout matches the design for its year -- only that
the revision does not CLAIM years whose designs disagree. A revision confined to
one design's years passes here even if every offset in its layout is wrong,
because that comparison needs the casilla-to-box mapping this module deliberately
avoids depending on. The two checks are complements: this one bounds the span,
and a per-modelo offset gate (where the box numbers exist) bounds the contents.

It also does NOT enforce the authoring policy that governs how a span is split.
The accepted posture is to split only at boundaries inside a modelo's reachable
filing window -- defined by prescripción, four years from the voluntary filing
deadline (LGT arts. 66-67), computed at implementation time -- and to refuse
export for years before the earliest split. This gate knows nothing of that. It
compares designs across whatever span a revision CLAIMS, so it catches a span
widened back over a boundary and objects to nothing if a revision is split where
the window no longer requires one. The asymmetry is deliberate: this instrument
guards byte-correctness, that policy guards authoring cost, and they answer
different questions. Its silence about a split is not approval of the split.

That window is itself dated and moves. Exercise 2021 prescribed on 2026-01-30,
so a boundary that required a split in December 2025 does not require one now,
and the next expiry shifts the answer again. Recompute it rather than reading a
boundary set off any record, including this one.

ANTI-VACUITY. A parser that cannot read a design returns the same answer as a
design with no divergence, so silence has to be loud: a design file this module
claims to read but extracts nothing from is a FAILURE, not a skip. Without that,
the gate goes green by not looking, which is this instrument's most likely rot
path.

No count is hardcoded. The number of designs, boundaries and shared boxes all
vary as the corpus grows; gating on any of them would encode today and detect
nothing tomorrow.

RUNTIME, stated here so the next reader meets it in the code rather than in CI.
Reading the design SOURCES rather than their markdown derivatives took this
module from roughly 28s to roughly 185s, a five-fold increase, because it now
parses spreadsheets and PDFs instead of pre-extracted text. Two things keep that
acceptable and both are worth knowing before anyone tries to "optimise" it. The
parsers are ``lru_cache``d, so the cost is paid ONCE PER SESSION rather than per
test -- the number that matters for CI is the one-off, not a multiple of it. And
it cannot be scoped away: restricting the parse to modelos that actually declare
an export layout removes only ~37% of the files and ~15% of the bytes, because
the largest designs belong to modelos that do export. The cost buys the offsets
and the field occupancy the derivatives do not carry, which is what the box and
retirement signals are made of.

SIX INDEPENDENT SIGNALS, ONE VERDICT -- and the occupancy one reports two directions.
This heading has now been wrong three times in the same direction: TWO for as long as
the occupancy signal existed, THREE for as long as the box-SET signal has, and FOUR
for as long as the description-flip signal has, which is worth stating rather than
quietly correcting: a module that miscounts its own instruments invites a reader to act
on the ones it names and miss the rest. It compounded, too -- box-SET membership and
description-flip were both labelled "FOURTH SIGNAL" at their own definitions, an
internal collision nobody using either docstring alone would notice.

The signals, renumbered here rather than left to drift again: (1) box-offset
displacement, (2) page-length / record-count, (3) reserved-space occupancy
retire/revive, (4) box-SET membership -- a box added or removed with nothing
displaced, which the movement check structurally cannot see because it iterates only
the boxes both designs share -- (5) unnumbered-slot description flip, including its
no-separable-leaf branch, and (6) box-FREE position-SET membership -- a field added or
removed at a fixed offset with no bracketed number to key on at all, which (4)
structurally cannot see because its key is the box number itself. Measured directly:
62 of 174 bundled designs carry zero bracketed box numbers anywhere, so (4) never runs
for them and (6) is not a refinement of it but the only membership signal that exists
there.

The box-offset diff sees which boxes moved
but needs bracketed box markers. The page-length diff sees only that a page
changed size, but reads designs the box table cannot -- several older PDF
extractions publish their page totals while yielding no box markers -- so it
measures years that would otherwise be blind. Neither subsumes the other: a
re-layout preserving every page length is caught by the first, a year only the
second can read is caught by the second. A year is reported UNMEASURED only when
BOTH are blind.

They report through ONE assertion rather than two, because reporting separately
was the instrument's own defect. The two see overlapping but DIFFERENT boundary
sets, so a fix owner acting on either list alone splits a revision at some of its
boundaries and leaves the rest standing -- a gate still red, reading as an
incomplete fix rather than a wrong one. Modelo 303 is the live case: two of its
six boundaries are visible only to the page-length signal. The failure text is
therefore the split specification, naming per revision every boundary, which
signal saw it, and how many revisions the span actually needs.

THIS MODULE IS LANDED RED, DELIBERATELY, AND THE FAILURES ARE THE FINDING RATHER
THAN A REGRESSION. It names two confirmed live defects: Modelo 390's single
revision spans five re-layouts, and Modelo 303's revisions span six more --
including a 2025-to-2026 shift affecting filings made today, where the box diff
shows 120 of 163 shared boxes moving and the page diff independently shows the
Liquidación page growing by five bytes. The Modelo 390 case was proved end to end
-- an export at an earlier filing year succeeds and writes bytes laid out for the
newest design. Weakening the assertions to land green would delete the evidence;
all of it goes green when the revisions are split at the boundaries the failure
text names, which is the fix.

The coverage guard is red for a different and much smaller reason: one year
inside a gated span has a design neither signal can read.

Mutation-proved from outside the repository, three directions. Narrowing Modelo
390's claimed span to the newest design removes exactly its own violations and
leaves every other modelo's standing. Widening every revision to claim all design
years implicates further modelos, so the gate detects a span that grows into a
boundary rather than only the spans that exist today. Breaking the box pattern
makes the coverage guard refuse instead of passing on an empty parse.
"""

from __future__ import annotations

import re
import unicodedata
from bisect import bisect_right
from datetime import date
from functools import cache, lru_cache
from itertools import combinations, pairwise
from pathlib import Path

import pytest

from .....core import PeriodKind, RegistryAuthorityGrade, registry_period_kind
from .....core.directory_scan import DirectoryEntryKind, scan_directory
from .....core.external_constants import PDF_EXTENSION as _PDF_EXTENSION
from .....core.external_constants import XLS_EXTENSION as _XLS_EXTENSION
from .....core.resources import bundled_path
from .....tests.registry_tree import bundled_registry_tree
from .. import ModeloDefinition, ModeloRevision, SourceReference
from .._authority import ValidatedRegistryAuthority
from .._record_design import (
    _clean_pdf_line,
    _extract_pdf_text_lines,
    extract_record_design_pdf,
    extract_record_design_workbook,
    extract_record_design_xls_workbook,
)
from .._record_design_coverage import _CASILLA_TAG_RE
from .._record_design_schema import RecordDesignSheet

pytestmark = [pytest.mark.unit, pytest.mark.hex_domain]


_DESIGN_ROOT_PARTS = ("corpus", "aeat_official", "disenos_registro")

#: The bracketed box number AEAT embeds in a design field description, IMPORTED from the
#: registry's canonical definition rather than re-declared here.
#:
#: This module carried its own copy capped at FOUR digits, and Modelo 200 numbers its
#: boxes with FIVE. Measured on its newest bundled design: 5538 of 5561 bracketed tokens
#: are five digits, so the private copy keyed 23 boxes -- under 0.4% of the modelo -- and
#: the box-offset and box-set signals were effectively switched off there while reporting
#: nothing wrong. The description-keyed population is defined as the slots carrying NO
#: box number, so the same cap also mis-classified those 5538 numbered fields as
#: unnumbered.
#:
#: The worst part was not the blindness but that it presented as agreement: an
#: independently derived boundary union and this gate's verdict both said Modelo 200 had
#: exactly one boundary and no gap, because BOTH used a four-digit marker. Agreement
#: between two instruments sharing one blind spot is worth nothing, and unlike a wrong
#: answer it offers nothing to notice.
#:
#: The canonical definition is bounded at five digits deliberately rather than left open:
#: its own rationale records that an unbounded ``\d+`` would admit amounts, NIF fragments
#: and position offsets that appear bracketed in the same columns. That reasoning is not
#: restated here, because restating it is how a third copy begins.
_BOX_MARKER = _CASILLA_TAG_RE
_DESIGN_YEAR = re.compile(r"ejercicios?-(\d{4})(?:-(y|a|hasta)-(\d{4}))?")
#: A design filename may name TWO explicit ejercicios ("ejercicio-2015-y-2016"),
#: and the corpus holds four such spans (2007-y-2008, 2008-y-2009, 2015-y-2016,
#: 2019-y-2020). Taking only the first match attributed the design to its opening
#: year and left the second attributed to NOTHING -- 2016 and 2020 each had a
#: bundled design and no year claiming it. "y-siguientes" is deliberately NOT
#: expanded: its span is open-ended, so enumerating it would invent years.
#: The orden year in a filename ("orden-hap-2373-2014-...-ejercicio-2018") is not
#: a coverage year and must never be read as one, which is why this anchors on
#: "ejercicio-" rather than scanning for any four-digit run.
#: Case-INSENSITIVE on purpose. AEAT writes both `Reservado para la AEAT` and
#: `RESERVADO PARA LA A.E.A.T.`, and a case-sensitive substring test silently
#: misses the uppercase form -- which reads a reserved block's own growth as a
#: moved field and invents a boundary that is not there.
_RESERVED_FIELD = re.compile(r"reservado", re.IGNORECASE)
#: Design SOURCE suffixes, in preference order for a year bundled more than once.
#:
#: ``.xlsm`` is accepted because the corpus bundles two Modelo 220 designs in it
#: (ejercicios 2022 and 2023) and openpyxl reads it exactly as it reads ``.xlsx``.
#: Omitting it dropped both files from every map here for no reason anyone chose:
#: a macro-enabled container is a packaging detail, not a different document.
_DESIGN_SUFFIXES = (".xlsx", ".xlsm", ".xls", ".pdf")
#: The synthetic single-sheet name the PDF backend falls back to when a document
#: declares no per-page sections. Its presence means the parse is FLATTENED, so
#: per-page signals must abstain rather than compare one synthetic page against
#: a spreadsheet's real ones.
_PDF_FLATTENED_SHEET = "PDF record design"

# Each page closes with a "TOTAL <n> POSICIONES" row, in the pipe-delimited
# spreadsheet extractions and the space-delimited PDF ones alike. The sequence of
# page lengths is a SECOND, box-number-free signal: if a page's byte length
# changed, something inside it moved. It reaches designs the box table cannot --
# the older PDF extractions carry these rows while yielding no bracketed boxes.
_PAGE_TOTAL = re.compile(r"TOTAL\s*\|?\s*\|?\s*(\d+|variable)\s*\|?\s*POSICIONES", re.IGNORECASE)

#: The ejercicio a design states in its OWN title text. Bounded to a short window
#: after the word deliberately: "Ejercicio 2023" is a coverage assertion, while
#: "rendimientos obtenidos en ejercicios iniciados antes de 1999" is prose about
#: the filer's data, and a greedy gap between the two admits the second.
_TITLE_EJERCICIO = re.compile(r"(?i)\bejercicio\b[^\n]{0,25}?\b(19[89]\d|20[0-4]\d)\b")
#: A pago-fraccionado designation: a period token immediately followed by its
#: ejercicio, optionally naming a second ("3p-2013-y-2014"). Anchored on the
#: period token so an orden number cannot supply the year -- ``hap-523-2015``
#: carries no period token and contributes nothing.
_PERIOD_QUALIFIED_YEAR = re.compile(r"(?<![0-9a-z])\d+[pt]-(\d{4})(?:-y-(\d{4}))?", re.IGNORECASE)
#: An ``Ejercicio`` slot whose declared value is a CONSTANT year, in either word
#: order AEAT writes it. The constant is what distinguishes the design fixing a
#: year from the design reserving a slot for the filer to write one in.
_CONSTANT_EJERCICIO = re.compile(r"(?i)\bejercicio\b[^|]{0,60}?\bconstante\b[^|]{0,20}?\b(19[89]\d|20[0-4]\d)\b")
_CONSTANT_EJERCICIO_REVERSED = re.compile(
    r"(?i)\bconstante\b[^|]{0,40}?\bejercicio\b[^|]{0,30}?\b(19[89]\d|20[0-4]\d)\b"
)


def _authority() -> ValidatedRegistryAuthority:
    return ValidatedRegistryAuthority.load(bundled_path("registry", "aeat"), source_root=bundled_path())


def _design_dir(modelo_id: str) -> Path:
    return bundled_path(*_DESIGN_ROOT_PARTS, f"modelo_{modelo_id}")


def _sources_by_year(modelo_id: str) -> tuple[tuple[int, Path], ...]:
    """``(design year, source path)`` pairs, first source per year winning."""
    seen: dict[int, Path] = {}
    for path in _design_sources(modelo_id):
        if not _design_sheets(path):
            continue
        for year in _design_coverage_years(path):
            seen.setdefault(year, path)
    return tuple(sorted(seen.items()))


def _design_years(name: str) -> tuple[int, ...]:
    """Every ejercicio a design filename claims, expanded and ordered.

    AEAT names these four ways and they do NOT mean the same thing:

    * ``ejercicio-2024``            -- one year.
    * ``ejercicios-2024``           -- one year; the plural is a naming
      habit, not a range, and it appears on 61 bundled files.
    * ``ejercicio-2015-y-2016``     -- TWO discrete years ("and").
    * ``ejercicios-2004-a-2009`` /
      ``ejercicios-2016-hasta-2018`` -- an INCLUSIVE RANGE ("through"),
      so 2004-a-2009 claims six years, not two endpoints.

    Reading ``a``/``hasta`` as two endpoints silently drops every year
    between them, and matching only the singular ``ejercicio-`` misses 40 of
    the 209 bundled design files across 15 modelos entirely -- a design that
    is invisible to enumeration reads exactly like one that does not exist.

    ``y-siguientes`` is still NOT expanded: its span is open-ended, so
    enumerating it would invent years.

    A PERIOD-QUALIFIED ejercicio counts too: AEAT writes a pago-fraccionado
    design's coverage as ``ejercicio-2p-y-3p-2016``, ``ejercicios-3p-2013-y-2014``
    or a bare ``1p-2016``, and the year-follows-``ejercicio-`` anchor misses all
    three because a period token sits between the word and the digits. This is NOT
    the orden-year anti-pattern relaxed: the year still follows an explicit period
    designation, which is a coverage statement, and the anchor moves from
    ``ejercicio-`` to the period token rather than being dropped.
    """
    matched = _DESIGN_YEAR.search(name)
    if matched is None:
        return _period_qualified_years(name)
    first, connector, second = matched.group(1), matched.group(2), matched.group(3)
    if second is None:
        return (int(first),)
    if connector == "y":
        return (int(first), int(second))
    return tuple(range(int(first), int(second) + 1))


def _period_qualified_years(name: str) -> tuple[int, ...]:
    """Years a period-qualified designation claims, e.g. ``3p-2013-y-2014``."""
    years: set[int] = set()
    for matched in _PERIOD_QUALIFIED_YEAR.finditer(name):
        years.add(int(matched.group(1)))
        if matched.group(2) is not None:
            years.add(int(matched.group(2)))
    return tuple(sorted(years))


@lru_cache(maxsize=512)
def _title_ejercicio_years(path: Path) -> tuple[int, ...]:
    """Every ejercicio the design's own TITLE TEXT states, or empty.

    AEAT prints ``Ejercicio NNNN`` as a title line on a design's first page. That
    is the document asserting which ejercicio it governs, and it is a different
    fact from anything in the filename: measured across the corpus, a filename
    records when AEAT last touched the published page while the title records the
    ejercicio the layout is for, and the two diverge by up to SEVEN years --
    ``03-180-orden-hap-1732-2014-de-24-de-septiembre.pdf`` states ``Ejercicio
    2021``, and ``01-165-diseno-de-registro-actualizado-en-2023.pdf`` states
    ``Ejercicio 2026``.

    WHY THIS IS NOT READ FROM THE PARSED FIELDS. Only eight designs in the whole
    corpus declare an ejercicio as a field constant, so a field-level scan finds
    almost nothing; the assertion lives in the document's heading, above the
    column header the extractor starts at.

    Deliberately reads a BOUNDED prefix -- a PDF's first sixty text lines, a
    workbook's first six rows on its first four sheets -- rather than the whole
    document. Beyond the heading the same phrase appears inside guidance prose
    ("para ejercicios anteriores a 1999"), which is a statement about the DATA a
    filer reports, not about the design's coverage, and admitting it would attribute
    Modelo 193's designs to 1999.
    """
    try:
        lines = _pdf_title_lines(path) if path.suffix.lower() == _PDF_EXTENSION else _workbook_title_lines(path)
    except Exception:
        # An unreadable container attributes nothing, exactly as the sibling parse
        # helper does: a file this cannot open must not be attributed to a year by
        # guesswork, and the coverage guard reports it as unattributable instead.
        return ()
    joined = "\n".join(lines)
    return tuple(sorted({int(matched.group(1)) for matched in _TITLE_EJERCICIO.finditer(joined)}))


def _pdf_title_lines(path: Path) -> list[str]:
    lines = _extract_pdf_text_lines(path.read_bytes(), source_label=str(path))
    return [cleaned for line in lines[:60] if (cleaned := _clean_pdf_line(line))]


def _workbook_title_lines(path: Path) -> list[str]:
    if path.suffix.lower() == _XLS_EXTENSION:
        import xlrd

        book = xlrd.open_workbook(str(path), on_demand=True)
        try:
            return [
                text
                for name in book.sheet_names()[:4]
                for row in range(min(6, book.sheet_by_name(name).nrows))
                if (
                    text := " ".join(
                        str(cell).strip() for cell in book.sheet_by_name(name).row_values(row) if str(cell).strip()
                    )
                )
            ]
        finally:
            book.release_resources()
    from openpyxl import load_workbook

    book = load_workbook(path, read_only=True, data_only=True)
    try:
        return [
            text
            for worksheet in book.worksheets[:4]
            for row in worksheet.iter_rows(max_row=6, values_only=True)
            if (text := " ".join(str(cell).strip() for cell in row if cell is not None and str(cell).strip()))
        ]
    finally:
        book.close()


@lru_cache(maxsize=512)
def _constant_ejercicio_years(path: Path) -> tuple[int, ...]:
    """Every ejercicio a design declares as a FIXED FIELD VALUE, or empty.

    The second place a design states its own coverage: a slot labelled ``Ejercicio``
    whose declared content is a CONSTANT rather than something the filer supplies.
    Modelo 714 writes ``Ejercicio | Constante 2025`` and Modelo 390 writes
    ``2. Devengo - Ejercicio. | Constante "2016"``.

    ``Constante`` is load-bearing, not decoration. Modelo 303 declares ``Ejercicio de
    devengo (EEEE)`` at a four-byte slot with no constant, because that is a value the
    FILER writes; reading it as coverage would attribute every M303 design to whatever
    year happened to appear beside it. Requiring the constant is what separates "the
    design fixes this to 2025" from "the design reserves four bytes for a year".

    Rarer than the title assertion -- eight designs corpus-wide -- but better
    corroborated per instance: seven of the eight also carry a trustworthy filename
    ejercicio and all seven match it.
    """
    years: set[int] = set()
    for sheet in _design_sheets(path):
        for parsed_field in sheet.fields:
            blob = " ".join(
                text for text in (parsed_field.description, parsed_field.content, parsed_field.validation) if text
            )
            for pattern in (_CONSTANT_EJERCICIO, _CONSTANT_EJERCICIO_REVERSED):
                matched = pattern.search(blob)
                if matched is not None:
                    years.add(int(matched.group(1)))
                    break
    return tuple(sorted(years))


def _content_ejercicio_years(path: Path) -> tuple[int, ...]:
    """Every ejercicio the DESIGN ITSELF states, from either content signal."""
    return tuple(sorted(set(_title_ejercicio_years(path)) | set(_constant_ejercicio_years(path))))


@cache
def _catalogue_ejercicio_span() -> dict[str, tuple[int, int]]:
    """Design filename -> the ejercicio span its SOURCE ENTRY declares, where it states one.

    A design that names no ejercicio in its filename or title enters no
    comparison, and that is a hole in boundary detection rather than only a gap
    in a ledger: Modelo 347's ``Orden EHA/3378/2011`` design is cited by its
    revision, carries an authored coverage span in the source catalogue, and is
    still never paired against its neighbour, so the re-layout between them goes
    unreported.

    ADMITTED ONLY FROM A CLOSED, EJERCICIO-ALIGNED SPAN, and both conditions are
    load-bearing. An OPEN span (``applies_to`` unset) cannot be enumerated for
    the same reason ``y-siguientes`` is not expanded -- it would invent every
    year to the end of time. A span whose ends do not fall on 1 January and 31
    December is not describing ejercicios at all: Modelo 180's design runs
    2000-11-28 to 2014-09-26, which are promulgation and repeal dates, and
    Modelo 210's run 2022-06-01 to 2025-12-31, which is a DEVENGO span on an
    axis that is not an ejercicio. Reading either as coverage would repeat the
    update-date-as-governed-period conflation already corrected twice.

    So the rule is checked against the dates themselves rather than against a
    list of modelos, and the designs it admits are exactly those whose catalogue
    entry states a whole number of ejercicios.
    """
    spans: dict[str, tuple[int, int]] = {}
    for source in _authority().catalogues.sources.values():
        if source.kind != "record_design":
            continue
        start, end_date = getattr(source, "applies_from", None), getattr(source, "applies_to", None)
        if start is None or end_date is None:
            continue
        if (start.month, start.day, end_date.month, end_date.day) != (1, 1, 12, 31):
            continue
        spans[str(source.corpus_path).rsplit("/", 1)[-1]] = (start.year, end_date.year)
    return spans


def _design_coverage_years(path: Path) -> tuple[int, ...]:
    """Every ejercicio a design covers, from its content and its filename TOGETHER.

    THE UNION, NOT A PRECEDENCE, and that correction is load-bearing. Ranking the
    content above the filename looks right -- the filename is a publication
    artefact and the title is the document's own assertion -- but the two do not
    describe the same shape of fact. **A design's title names a POINT and its
    filename can name a SPAN.** A design published for ejercicio 2009 and applying
    through 2014 heads its first page ``Ejercicio 2009`` and is filed as
    ``ejercicios-2009-a-2014``; preferring the title there discards 2010 to 2014.

    Measured before this was written that way round: four designs would have lost
    twelve design-years between them, Modelo 130 and Modelo 131 five each. A
    content signal that can only ADD is safe in a way one that can also REPLACE is
    not, and the asymmetry matters here because dropping a year is the QUIET
    failure -- an unattributed design forms no boundary, so the verdict gets
    shorter and reads as progress.

    The union also means the two signals cannot narrow each other by accident,
    which leaves genuine contradiction -- disjoint claims -- as the only thing
    :func:`test_a_design_title_never_contradicts_a_trustworthy_filename_year` has
    to catch, and it is the only thing that check should be asked to judge.

    Returns empty for a design that states no ejercicio anywhere. That is a real
    answer rather than a defect -- Modelo 036 is scoped by an in-force DATE and
    Modelo 210 by a devengo span, neither of which is an ejercicio -- and the
    coverage guard reports such designs as unattributable rather than guessing.
    """
    stated = set(_content_ejercicio_years(path)) | set(_design_years(path.name))
    if stated:
        return tuple(sorted(stated))
    # The document says nothing; fall back to a coverage span its source entry
    # states outright. See _catalogue_ejercicio_span for why only a closed,
    # ejercicio-aligned span qualifies.
    span = _catalogue_ejercicio_span().get(path.name)
    if span is None:
        return ()
    return tuple(range(span[0], span[1] + 1))


#: AEAT's declared period bound for a design that covers only part of an ejercicio.
#: ``hasta-periodo(s)-06`` covers the year UP TO that period, so its coverage STARTS at
#: 01; ``desde-periodo-07`` and ``a-partir-de-periodos-09`` cover FROM that period on, so
#: their coverage starts there. Ordering on the declared coverage start is what puts the
#: two halves of a mid-split ejercicio in publication order.
_COVERAGE_BOUND = re.compile(r"(hasta|a-partir-de|desde)-periodos?-(\d{2})", re.IGNORECASE)


def _coverage_start_period(name: str) -> int | None:
    """The first period a design's own AEAT designation claims, or ``None``.

    WHY THIS IS NOT READ FROM THE FILE BODY, stated because the honest answer is that
    it cannot be. The design workbooks carry NO assertion of the periods they govern:
    measured across Modelo 303's two 2024 halves, both declare the identical sheet set
    and the identical ``Ejercicio de devengo (EEEE)`` and ``Período. (PP)`` header
    fields, because those are slots the FILER fills in, not metadata about the design's
    own coverage. Zero coverage assertions exist inside either file. AEAT's published
    designation is therefore the only place the boundary is stated at all, which is why
    this reads the designation rather than the document.

    WHAT THIS IS NOT. It is not the numeric filename prefix. AEAT numbers its published
    listing NEWEST FIRST -- Modelo 303's ``01`` is the 2026 design and its ``06`` is
    2025 -- so a filename sort is not merely arbitrary, it is close to REVERSED, and
    within one year it silently pairs the LATE half before the EARLY one. Nothing here
    depends on that numbering, so a change to AEAT's listing convention cannot move it.

    Returns ``None`` when the designation declares no period bound, which is a real
    case rather than a defect: Modelo 303's two 2018 designs are ``ejercicio-2018`` and
    ``ejercicio-2018-salvo-ultimo-periodo-12m-4t``, and "except the last period" fixes
    no start while the unqualified sibling asserts nothing at all. Ordering those two
    would be inference by elimination dressed as a measurement, so the caller is told
    the year is unorderable instead.
    """
    matched = _COVERAGE_BOUND.search(name)
    if matched is None:
        return None
    kind, period = matched.group(1).lower(), int(matched.group(2))
    return 1 if kind == "hasta" else period


def _design_fingerprint(path: Path) -> tuple[object, ...]:
    """Format-independent identity of a design: what it DECLARES, not how it is packaged.

    A raw byte hash is the obvious identity and it is the wrong one. AEAT bundles the
    same design twice in two container formats, and an ``.xls`` and an ``.xlsx`` of one
    document are different bytes while being the same design. Measured: Modelo 200
    carries such a twin for every ejercicio from 2015 to 2025, and a byte-keyed
    deduplication reported all ELEVEN of those years as carrying two designs -- which
    reads as eleven mid-course splits on a modelo that has none. Fingerprinting the
    parsed declaration collapses the twins and leaves genuine splits standing: on
    Modelo 303 the 2018, 2021 and 2024 pairs survive, because those really are two
    different designs.

    Byte-keying also has to stay dead rather than merely unused, because it looks
    correct: it DOES collapse the corpus's other duplicate shape, the same file bundled
    twice under a truncated filename, so a reader checking it against that case
    concludes it works.
    """
    return tuple(
        (
            sheet.name,
            sheet.total_positions,
            tuple(
                (field.offset, field.length, field.type_code, " ".join(field.description.split()))
                for field in sheet.fields
            ),
        )
        for sheet in _design_sheets(path)
    )


def _designs_in_publication_order(modelo_id: str) -> tuple[tuple[Path, ...], tuple[int, ...]]:
    """``(designs oldest-first, years whose designs could not be ordered)``.

    Deduplicated by CONTENT, because the corpus bundles several designs twice under
    names differing only by a truncated extension, and a path-keyed pass reports the
    duplicate as a second design.

    A year holding two designs where at least one declares no coverage bound is
    reported in the second element and its members are left in a stable but
    UNASSERTED order, so a consumer can refuse rather than trust it.
    """
    by_year: dict[int, list[Path]] = {}
    seen: set[object] = set()
    for path in _design_sources(modelo_id):
        if not _design_sheets(path):
            continue
        years = _design_coverage_years(path)
        if not years:
            continue
        marker = _design_fingerprint(path)
        if marker in seen:
            continue
        seen.add(marker)
        by_year.setdefault(min(years), []).append(path)

    ordered: list[Path] = []
    unorderable: list[int] = []
    for year in sorted(by_year):
        members = by_year[year]
        starts = [_coverage_start_period(path.name) for path in members]
        if len(members) > 1 and any(start is None for start in starts):
            unorderable.append(year)
            ordered.extend(sorted(members, key=lambda path: path.name))
            continue
        ordered.extend(sorted(members, key=lambda path: (_coverage_start_period(path.name) or 1, path.name)))
    return tuple(ordered), tuple(unorderable)


def _designs_by_year(modelo_id: str) -> dict[int, tuple[Path, ...]]:
    """``{year: every readable design source claiming it}`` -- one to MANY.

    AEAT splits an ejercicio mid-course, so a year can have two designs that differ
    in byte layout. Modelo 303 does it three times: 2018, 2021 and 2024 each ship a
    "hasta periodo N" design and a "desde/a partir de periodo N" one. The year-keyed
    maps below keep the FIRST by filename sort and silently discard the other, so a
    caller asking what a year looked like gets an arbitrary half with no signal a
    second exists. This map keeps both, so a consumer deriving epoch boundaries can
    see the split rather than land a boundary on the wrong side of it.
    """
    grouped: dict[int, list[Path]] = {}
    for path in _design_sources(modelo_id):
        if not _design_sheets(path):
            continue
        for year in _design_coverage_years(path):
            grouped.setdefault(year, []).append(path)
    return {year: tuple(paths) for year, paths in sorted(grouped.items())}


def _design_sources(modelo_id: str) -> list[Path]:
    """Every bundled design SOURCE for one modelo, deterministically ordered.

    Walks the modelo directory RECURSIVELY rather than globbing ``files/*``. Almost
    every modelo keeps its designs in a ``files`` subdirectory, but Modelo 210
    keeps ``dr210_2011.pdf`` directly in the modelo directory, and the fixed-depth
    glob excluded it from every map in this module. Nothing decided that; it is a
    directory-shape assumption that happened to hold for all but one modelo, which
    is the shape of assumption that survives longest unnoticed.
    """
    return sorted(
        (
            path
            for path in scan_directory(_design_dir(modelo_id), recursive=True, select=DirectoryEntryKind.FILES)
            if path.suffix.lower() in _DESIGN_SUFFIXES
        ),
        key=lambda path: (path.name, path.suffix.lower()),
    )


def _design_sheets(path: Path) -> tuple[RecordDesignSheet, ...]:
    """Parse one design SOURCE, dispatching on its suffix.

    The sources are read directly rather than their ``.extracted.md``
    derivatives. The markdown extraction of a PDF drops the offset column and
    the page-total rows, so every PDF-sourced design read that way reports as
    unreadable while the shipped parser reads the same file correctly -- 21 of
    the 24 designs this module called unmeasured were readable all along, six
    of them consecutive Modelo 100 years. Reading the source removes the whole
    class rather than special-casing it.

    Returns an empty tuple for a genuinely unparseable file, so the caller's
    unreadable-reporting path still names it. A parse that fails must stay
    visible as unmeasured; only a parse that never ran should disappear.
    """
    parsers = {
        ".xlsx": extract_record_design_workbook,
        ".xlsm": extract_record_design_workbook,
        ".xls": extract_record_design_xls_workbook,
        ".pdf": extract_record_design_pdf,
    }
    parser = parsers.get(path.suffix.lower())
    if parser is None:
        return ()
    try:
        # ACCEPTS a partial read deliberately. This module compares designs against
        # each other, and a design read in part still carries real evidence about
        # the sheets it did read; refusing it here would replace a comparison that
        # sees most of a boundary with one that sees none of it. The completeness
        # of each read is reported by the coverage guard rather than resolved here.
        return parser(path).accept_partial()
    except Exception:
        return ()


def _unparseable_design_sources(modelo_id: str) -> tuple[Path, ...]:
    """Every bundled design file for one modelo whose sheets fail to parse AT ALL.

    A THIRD SHAPE, distinct from the two this instrument otherwise reports. A
    design that parses (even partially, per :func:`_design_sheets`'s own
    ``accept_partial`` posture) and states no ejercicio anywhere is a real,
    legitimate answer -- some modelos are scoped by something other than an
    ejercicio, and the coverage guard reports that honestly. A year with NO
    bundled file at all is a genuine acquisition gap -- fetching one from AEAT
    is the only fix. A file in THIS set is neither: it physically exists in the
    corpus, so acquisition does nothing, but the parser returns nothing usable
    for it AT ALL, so whatever ejercicio it might state -- filename or content
    -- never reaches :func:`_design_coverage_years`. That is an extraction-layer
    defect, and conflating it with either of the other two shapes sends a
    fix-owner to acquire a design that is already bundled, or to treat a
    legitimately non-ejercicio design as if something were missing from it.

    Measured live: Modelo 036 carries three such files
    (``...-ejercicio-2023-y-siguientes...``, two ``...-ejercicio-2021-y-siguientes...``),
    each with a real filename-derivable ejercicio year that never reaches any
    comparison because :func:`_designs_in_publication_order`'s own first filter
    drops an unparseable design before year-attribution is ever consulted --
    loud at the extraction layer (the parser genuinely fails), but that loudness
    does not reach this module, because the design is dropped before it can be
    reported.

    PROSPECTIVE, NOT SPECULATIVE -- stated because the distinction matters for
    how dormant code is read. As of this writing no revision's declared span
    currently overlaps Modelo 036's three unparseable years, so the UNPARSEABLE
    branch in the two gate failure messages that consult this function is
    proven by construction (bite-proofed synthetically, see the mutation tests
    in this module) but currently unexercised by the live corpus -- correct and
    dormant, not correct and moot. The three unparseable files still exist and
    a future span change (Modelo 036's own, or a currently-absent modelo
    gaining a declared revision over years an unparseable file already claims)
    brings this branch back into range without anyone touching it. A dormant
    branch is still a rot risk: if this signal's wording or matching logic ever
    drifts silently, nothing in today's corpus would catch it until a real case
    arrives. Re-run the synthetic bite proof, not just the live gate, when this
    function changes.
    """
    return tuple(path for path in _design_sources(modelo_id) if not _design_sheets(path))


def _parse_extracted(path: Path) -> dict[str, int]:
    """box number -> offset from a design's ``.extracted.md`` derivative.

    Retained as a FALLBACK, not as the primary read. Measured: switching to the
    sources alone closed 21 PDF-sourced designs but OPENED new blind spots on
    `.xls` files whose markdown parses while the binary does not, so the
    unreadable count moved 24 -> 21 rather than 24 -> 3. Source-first with a
    derivative fallback is strictly better than either alone -- the source
    yields offsets and occupancy the markdown cannot, and the markdown covers
    the binaries the parsers refuse.
    """
    derivative = path.with_name(path.name + ".extracted.md")
    if not derivative.is_file():
        return {}
    table: dict[str, int] = {}
    for line in derivative.read_text(encoding="utf-8", errors="replace").splitlines():
        if "|" not in line:
            continue
        columns = [column.strip() for column in line.split("|")]
        if len(columns) < 5:
            continue
        try:
            offset = int(columns[1])
        except (IndexError, ValueError):
            continue
        boxes = _BOX_MARKER.findall(line)
        if boxes:
            table.setdefault(boxes[-1], offset)
    return table


def _parse_design(path: Path) -> dict[str, int]:
    """box number -> record offset for one design, source first, derivative second."""
    table: dict[str, int] = {}
    for sheet in _design_sheets(path):
        for field in sheet.fields:
            offset = field.offset
            boxes = _BOX_MARKER.findall(field.description)
            if not boxes:
                continue
            # A field's OWN number is the LAST bracket on its row. Formula totals cite
            # their operands first -- "Total cuota devengada ([152]+[167]+...) [27]"
            # -- so first-match keying would file the total under an operand, and
            # requiring a single bracket would drop it entirely. Modelo 303 has
            # eleven such rows, including its three most load-bearing totals; Modelo
            # 390 has none, which is why this only surfaced on the second modelo.
            #
            # A box can also appear once per régimen segment at the same offset; keep
            # the first and never overwrite, so a later segment cannot mask a move.
            #
            # Modelo 390 proves this is not hypothetical: box [114] carries one
            # `Código CNAE` per prorrata row -- once in the 2016 design and FIVE
            # times in 2017. A last-wins map compares 2016's first row against
            # 2017's fifth and reports a +332 "move" that never happened.
            table.setdefault(boxes[-1], offset)
    return table or _parse_extracted(path)


def _page_lengths(path: Path) -> tuple[str, ...]:
    """The per-sheet declared record length, source first, derivative second.

    ABSTAINS on a flattened PDF parse. The PDF backend returns ONE synthetic
    sheet covering the whole document, so its "page lengths" are a one-element
    tuple describing no page at all. Compared against a spreadsheet's nine real
    per-page totals it differs unconditionally, which reports a re-layout
    between every PDF-sourced year and its neighbour -- a boundary manufactured
    by the parser shape rather than by AEAT.

    Measured: with the flattened tuple emitted, Modelo 390 gained a false
    2015->2016 boundary, while a box-keyed comparison of the same two designs
    finds 345 shared boxes and ZERO moved. Returning nothing lets the box signal
    carry those years and keeps this one honest about what it cannot see.
    """
    sheets = _design_sheets(path)
    if len(sheets) == 1 and sheets[0].name == _PDF_FLATTENED_SHEET:
        return ()
    lengths = tuple("variable" if sheet.total_positions is None else str(sheet.total_positions) for sheet in sheets)
    if lengths:
        return lengths
    derivative = path.with_name(path.name + ".extracted.md")
    if not derivative.is_file():
        return ()
    return tuple(str(total) for total in _PAGE_TOTAL.findall(derivative.read_text(encoding="utf-8", errors="replace")))


def _occupancy(path: Path) -> dict[tuple[str, int], bool]:
    """``(sheet, offset) -> is_reserved`` for every field in one design.

    The reserved flag is carried as a VALUE rather than applied as a filter, and
    that is the whole point. Excluding reserved rows from a comparison makes a
    field RETIRED INTO reserved space invisible: it is real on one side and
    reserved on the other, so the exclusion drops one side and the pair is never
    compared. Measured on Modelo 390, three `Reg. Simplificado - Reducción
    aplicable` slots were retired between the 2024 and 2025 designs at offsets
    223, 543 and 1205, and a reserved-excluding diff reported the two years
    identical.

    A filter that drops a category cannot see movement into or out of that
    category. Keying on the position and classifying afterwards makes the
    transition a first-class result.
    """
    return {
        (sheet.name, field.offset): _RESERVED_FIELD.search(field.description) is not None
        for sheet in _design_sheets(path)
        for field in sheet.fields
    }


def _page_lengths_for(modelo_id: str) -> dict[int, tuple[str, ...]]:
    """``{design year: page-length sequence}`` for every readable design."""
    lengths: dict[int, tuple[str, ...]] = {}
    for path in _design_sources(modelo_id):
        found = _page_lengths(path)
        if not found:
            continue
        for year in _design_coverage_years(path):
            lengths.setdefault(year, found)
    return lengths


def _designs_for(modelo_id: str) -> tuple[dict[int, dict[str, int]], dict[int, str]]:
    """Return ``{design year: {box: offset}}`` and ``{year: filename}`` for unreadable ones.

    A design whose year cannot be read off the filename is excluded from both
    maps: it can neither be compared nor attributed to a revision's span, so
    counting it either way would be a guess.
    """
    parsed: dict[int, dict[str, int]] = {}
    unreadable: dict[int, str] = {}
    for path in _design_sources(modelo_id):
        table = _parse_design(path)
        for year in _design_coverage_years(path):
            if table:
                parsed.setdefault(year, table)
            elif year not in parsed:
                unreadable.setdefault(year, path.name)
    return parsed, unreadable


def _span_years(revision: ModeloRevision) -> set[int]:
    """Every filing year the revision's period selector claims."""
    selector = revision.period_selector
    if selector.years:
        return set(selector.years)
    if selector.year_from is None:
        return set()
    upper = selector.year_to
    if upper is None:
        # Open-ended: the gate can only speak about years the corpus covers, so
        # the span is bounded by the newest bundled design rather than by a
        # literal ceiling that would go stale.
        return {selector.year_from}
    return set(range(selector.year_from, upper + 1))


def _filing_revisions() -> list[tuple[ModeloDefinition, str, ModeloRevision]]:
    """Every revision that claims filing support, including incomplete claims.

    Filing grade is the authority boundary. Filtering on populated layouts would
    hide the exact genuine gap this suite must expose: a filing-supported revision
    whose official export evidence or authored layout is still missing.
    """
    return [
        (modelo, revision_id, revision)
        for modelo in _authority().modelos
        for revision_id, revision in modelo.revisions.items()
        if revision.effective_authority_grade is RegistryAuthorityGrade.FILING
    ]


def _declared_revisions() -> list[tuple[ModeloDefinition, str, ModeloRevision]]:
    """Every revision, for detector anti-vacuity controls outside support policy."""
    return [
        (modelo, revision_id, revision)
        for modelo in _authority().modelos
        for revision_id, revision in modelo.revisions.items()
    ]


def _filing_supported_revisions() -> list[tuple[ModeloDefinition, str, ModeloRevision]]:
    """Every revision that explicitly claims filing support.

    Kept as the raw-loader counterpart to :func:`_filing_revisions`: operational
    signal gates use validated authority, while the coverage-completeness gate
    must still run and report the exact structural gap if unrelated validation is
    red elsewhere in the tree.

    Reads through the raw loader (:func:`load_registry_tree`), never
    :func:`_authority`: the coverage-completeness question is structural
    (declared revisions, period selectors, bundled corpus) and needs no
    business-rule or filing-grade validation to answer. Coupling it to
    :class:`ValidatedRegistryAuthority` would make this gate's own
    reliability hostage to an unrelated validation defect anywhere else in
    the tree -- the exact silent-failure shape this gate exists to
    remove, now reproduced as "the coverage gate didn't even run" instead of
    "the coverage gate reported clean".
    """
    modelos, _catalogues = bundled_registry_tree()
    return [
        (modelo, revision_id, revision)
        for modelo in modelos
        for revision_id, revision in modelo.revisions.items()
        if revision.effective_authority_grade is RegistryAuthorityGrade.FILING
    ]


def _claimed_years(revision: ModeloRevision, design_years: set[int]) -> set[int]:
    """Design years the revision claims, honouring an open-ended upper bound."""
    selector = revision.period_selector
    explicit = _span_years(revision)
    if selector.years or selector.year_from is None:
        return explicit & design_years
    if selector.year_to is None:
        return {year for year in design_years if year >= selector.year_from}
    return explicit & design_years


@cache
def _source_reference_by_id() -> dict[str, SourceReference]:
    """Loaded source catalogue, used as the revision's dependency receipts."""
    _modelos, catalogues = bundled_registry_tree()
    return {str(ref): source for ref, source in catalogues.sources.items()}


def _layout_authority_receipts(modelo_id: str, revision: ModeloRevision) -> tuple[SourceReference, ...]:
    """Layout-authority sources explicitly cited by a revision.

    Fixed-width modelos normally cite ``record_design`` sources.  Modelo 100
    is filed from AEAT's published dictionaries and XSD instead; those sources
    carry the same ``layout_authority`` tier and must not be mistaken for an
    absent design merely because their transport grammar is not a workbook.
    """
    sources = _source_reference_by_id()
    return tuple(
        source
        for ref in revision.source_refs
        if (source := sources.get(str(ref))) is not None
        and source.evidence_tier == "layout_authority"
        and f"/modelo_{modelo_id}/" in f"/{str(source.corpus_path).replace('\\', '/')}"
    )


def _receipt_covers_year(source: SourceReference, year: int) -> bool:
    selector = source.period_selector
    if selector is not None:
        return selector.includes_year(year)
    if source.applies_from is None or source.applies_from.year > year:
        return False
    return source.applies_to is None or source.applies_to.year >= year


def _source_epoch_proves_revision_span(modelo_id: str, revision: ModeloRevision) -> tuple[bool, str]:
    """Whether cited record-design receipts cover the revision's declared year span.

    An open revision needs an open receipt; a closed revision needs every claimed
    year covered. This consumes authored source authority and never manufactures
    extra annual designs merely to reach a count. Corpus-detected relayouts are
    checked first by the caller and therefore always override this positive proof.
    """
    receipts = _layout_authority_receipts(modelo_id, revision)
    if not receipts:
        return False, "revision cites no layout-authority source receipt"
    selector = revision.period_selector
    if selector.years:
        years = set(selector.years)
    elif selector.year_from is not None and selector.year_to is not None:
        years = set(range(selector.year_from, selector.year_to + 1))
    elif selector.year_from is not None:
        open_receipts = tuple(
            source
            for source in receipts
            if source.applies_to is None
            and (
                (
                    source.period_selector is not None
                    and source.period_selector.year_from is not None
                    and source.period_selector.year_from <= selector.year_from
                    and source.period_selector.year_to is None
                )
                or (
                    source.period_selector is None
                    and source.applies_from is not None
                    and source.applies_from.year <= selector.year_from
                )
            )
        )
        if open_receipts:
            return True, f"open source epoch(s) {[source.id for source in open_receipts]!r}"
        return False, "open revision has no cited open-ended layout-authority receipt"
    else:
        return False, "revision declares no filing-year span"
    missing = sorted(year for year in years if not any(_receipt_covers_year(source, year) for source in receipts))
    if missing:
        return False, f"cited layout-authority receipts do not cover filing year(s) {missing!r}"
    return True, f"bounded source epoch receipt(s) {[source.id for source in receipts]!r}"


def test_the_design_parser_reads_every_markdown_design_it_claims() -> None:
    """A design that parses to nothing must fail, never pass by silence.

    This is the gate's anti-vacuity guard and its most likely rot path. If the
    extraction shape changes, or a regex stops matching, every comparison below
    would run over an empty table and agree trivially. An unreadable design is
    reported as UNMEASURED here rather than quietly dropped, because a parser
    that cannot read a format returns the same answer as a format with no
    divergence.
    """
    blind_spots: list[str] = []
    measured = 0
    for modelo, revision_id, revision in _filing_revisions():
        designs, unreadable = _designs_for(modelo.id)
        measured += len(designs)
        if not unreadable:
            continue
        # Only a design INSIDE a gated revision's span can hide a divergence from
        # this gate. One outside every span is a corpus file nothing compares,
        # so failing on it would be noise about the extractor rather than about
        # coverage.
        # A year is only a blind spot when BOTH signals are blind. A design that
        # yields no bracketed boxes but does publish its per-page lengths is
        # still measured by the page-length check below, so reporting it here
        # would overstate the gap.
        page_lengths = _page_lengths_for(modelo.id)
        opaque = {year for year in unreadable if year not in page_lengths}
        in_span = _claimed_years(revision, opaque)
        blind_spots.extend(
            f"modelo {modelo.id} revision {revision_id!r} claims {year}, but its design "
            f"{unreadable[year]!r} yields neither box offsets nor page lengths"
            for year in sorted(in_span)
        )
    assert measured, "no bundled design parsed at all; the extractor or the corpus path has moved"
    assert not blind_spots, (
        "these years are UNMEASURED rather than clean -- the gate cannot see a re-layout there, "
        "so its silence about them means nothing:\n  " + "\n  ".join(sorted(set(blind_spots)))
    )


@cache
def _corpus_path_by_source_ref() -> dict[str, str]:
    """Every source ref mapped to the corpus path it records, unparsed."""
    _modelos, catalogues = bundled_registry_tree()
    return {
        ref: str(entry.corpus_path) for ref, entry in catalogues.sources.items() if getattr(entry, "corpus_path", None)
    }


@cache
def _design_fingerprint_for_ref(ref: str) -> tuple[object, ...] | None:
    """The fingerprint of the design a source ref names, or ``None`` if it names none.

    Resolved lazily, one ref at a time. Fingerprinting PARSES the design, so
    building this for every catalogue source up front costs a full corpus read
    on a question only a handful of refs per revision ever ask.

    Fingerprint rather than file name, and that is what makes the match work.
    The corpus bundles some designs twice under names differing only by a
    truncated extension; :func:`_designs_in_publication_order` collapses those
    twins and keeps whichever sorts first, which is not necessarily the one the
    catalogue cites. Modelo 303's late 2024 design is exactly that: the
    catalogue names ``...-381-kb-xls.xlsx`` while the walk keeps its
    byte-identical ``...-381-kb-x.xlsx``. Comparing names fails on two files
    that are the same design; comparing the identity this module already uses to
    collapse twins does not.
    """
    corpus_path = _corpus_path_by_source_ref().get(ref)
    if corpus_path is None:
        return None
    resolved = bundled_path() / corpus_path
    if not resolved.is_file() or not _design_sheets(resolved):
        return None
    return _design_fingerprint(resolved)


def _cited_design_fingerprints(revision: ModeloRevision) -> set[tuple[object, ...]]:
    """The designs this revision's own source refs name, by fingerprint."""
    found = (_design_fingerprint_for_ref(str(ref)) for ref in revision.source_refs)
    return {fingerprint for fingerprint in found if fingerprint is not None}


def _mid_year_span(revision: ModeloRevision) -> int | None:
    """The year a revision sits WHOLLY inside while covering less than all of it.

    ``None`` for a revision that covers a full year, several years, or is
    open-ended -- those claim their years outright and every design in them.
    """
    valid_from, valid_to = revision.valid_from, revision.valid_to
    if valid_from is None or valid_to is None or valid_from.year != valid_to.year:
        return None
    covers_whole_year = (valid_from.month, valid_from.day) == (1, 1) and (valid_to.month, valid_to.day) == (12, 31)
    return None if covers_whole_year else valid_from.year


def _designs_claimed_by(modelo_id: str, revision: ModeloRevision) -> tuple[Path, ...]:
    """The designs a revision's span claims, in publication order.

    KEYED ON THE DESIGN FILE, NOT ON THE PARSED YEAR, and that is the whole point. The
    per-signal inventories this replaced each kept ONE design per year through a
    ``setdefault`` over a filename sort, so where AEAT split an ejercicio mid-course the
    second half was silently discarded and the boundary INSIDE that year could not be
    seen by any signal. Modelo 303 does it three times, in 2018, 2021 and 2024, and the
    mid-2024 boundary is inside the reachable filing window.

    One walk feeds all three signals, rather than each rebuilding its own inventory.
    Three separate walks is how they came to disagree about which designs existed: a
    year readable by one signal and not another entered one map and not the others, so
    the same boundary was keyed differently per signal and the evidence for it split
    across keys that never met.
    """
    ordered, _unorderable = _designs_in_publication_order(modelo_id)
    every_year = {year for path in ordered for year in _design_coverage_years(path)}
    claimed = _claimed_years(revision, every_year)
    within = tuple(path for path in ordered if set(_design_coverage_years(path)) & claimed)

    # A revision covering only PART of one year claims only the design it cites
    # for that year. AEAT splits an ejercicio mid-course by publishing two
    # designs with the same coverage year, so a year-keyed claim hands both to
    # each half -- and the halves then report a (2024, 2024) boundary they do
    # not span. Modelo 303's 2024 halves and modelo 490's 2022 halves are the
    # cases: each declares its own months in its id AND names one design in its
    # source refs, and the design filenames say the same thing
    # ("hasta-periodos-08-y-2t" beside "a-partir-de-periodos-09-y-3t").
    #
    # Deliberately narrow. A revision covering a whole year, several years, or
    # an open-ended span is untouched, so the genuine cross-year spans this gate
    # exists to find -- modelo 184, 200, 322 and 347 -- keep reporting. And the
    # narrowing applies only where the revision actually cites a design, so a
    # revision citing none claims its years outright as before.
    mid_year = _mid_year_span(revision)
    if mid_year is None:
        return within
    cited = _cited_design_fingerprints(revision)
    if not cited:
        return within
    kept = tuple(
        path
        for path in within
        if _design_fingerprint(path) in cited or mid_year not in set(_design_coverage_years(path))
    )
    return kept or within


def _box_set_evidence(before_boxes: dict[str, int], after_boxes: dict[str, int]) -> str | None:
    """Evidence that the box SET changed, whether or not anything moved.

    A SEPARATE SIGNAL from displacement, not a refinement of it. The displacement check
    iterates the boxes two designs SHARE, so a box present in one and absent in the
    other falls outside its loop entirely. That is not a lesser event: a box the later
    design declares and the earlier one does not cannot be declared at all under the
    earlier layout, and a box the earlier one declares and the later one drops is a
    value written into space the later design puts to another use.

    Measured on Modelo 390, where the whole class was invisible: 2015 to 2016 adds six
    boxes and 2016 to 2017 removes twenty, both with ZERO movement, no readable
    page-length difference and no occupancy transition, so no signal in this module
    reported either boundary. Adding this one took that revision's verdict from six
    re-layouts to eight.

    The blindness survived because it is MASKED wherever membership changes alongside
    movement: the 2017 to 2018 boundary drops seventy-two boxes, and the displacement
    check reports that boundary anyway on its ninety-seven moved boxes, so a reader
    spot-checking the signal against that pair sees a membership change duly reported
    and concludes the set is compared.

    Extracted as a named helper rather than inlined so the signal has a seam a mutation
    can suppress on its own, leaving every other signal running. A mutation that has to
    break the whole comparison proves the module can fail, not that this signal works.
    """
    added = sorted(set(after_boxes) - set(before_boxes), key=int)
    removed = sorted(set(before_boxes) - set(after_boxes), key=int)
    if not added and not removed:
        return None
    parts = []
    if added:
        parts.append(f"{len(added)} added (e.g. {', '.join(f'[{box}]' for box in added[:3])})")
    if removed:
        parts.append(f"{len(removed)} removed (e.g. {', '.join(f'[{box}]' for box in removed[:3])})")
    return (
        f"box SET changed: {' and '.join(parts)} -- a box only one side declares cannot be carried "
        "by the other layout at all, which no displacement, length or digest check sees"
    )


def _position_content(path: Path) -> dict[tuple[str, int, int], str]:
    """``(sheet, offset, length) -> normalised description`` for EVERY field, boxed or not.

    Unlike :func:`_unnumbered_labels`, NOT filtered to box-token-free slots -- every
    field's declared content at its position. This is what lets
    :func:`_position_set_evidence` see a field added or removed where no box number
    exists to key membership on at all, closing a class of blindness measured
    directly: 62 of 174 bundled designs carry ZERO bracketed box numbers anywhere
    (29 modelos, including 111, 180, 232, 349, 360, 369 and 720 -- three of which
    can already file today), so :func:`_box_set_evidence`'s key never exists for
    them, not merely loses precision on them.

    ABSTAINS on a flattened PDF parse, for the same reason :func:`_unnumbered_labels`
    does: the synthetic single-sheet shape makes ``(sheet, offset, length)`` collide
    across unrelated pages.
    """
    sheets = _design_sheets(path)
    if len(sheets) == 1 and sheets[0].name == _PDF_FLATTENED_SHEET:
        return {}
    table: dict[tuple[str, int, int], str] = {}
    for sheet in sheets:
        for field in sheet.fields:
            table.setdefault((sheet.name, field.offset, field.length), " ".join(field.description.split()))
    return table


def _position_set_evidence(earlier: Path, later: Path) -> str | None:
    """SIXTH SIGNAL: a field added or removed at a position, independent of box number.

    Generalises :func:`_box_set_evidence`'s membership idea past its box-number key,
    the same relationship the box-SET signal has to the displacement check: a
    SEPARATE signal, not a refinement, because a field one side declares and the
    other does not falls outside a shared-key loop entirely. The box-SET signal is
    structurally blind wherever no bracketed box number exists to key on at all --
    a bracket-free field cannot be a member of a box-number SET, period, not merely
    an imprecisely-tracked one.

    KEYED ON ``(sheet, offset, length)``, NEVER ON DESCRIPTION TEXT, so a genuine
    content CHANGE at an UNMOVED position is deliberately NOT reported here -- that
    is :func:`_description_flip_evidence`'s job, including its no-separable-leaf
    branch. This signal reports only when a POSITION exists in one design and not
    the other: a field the later design declares that the earlier one has nowhere
    to put, or one the earlier design declares that the later one no longer
    reserves space for. The two signals are complementary rather than redundant --
    each answers a question the other structurally cannot.

    RESERVED POSITIONS ARE EXCLUDED, and this exclusion is measured, not assumed.
    Modelo 202's two 2019-era designs (the May 2020 and September 2019 updates)
    declare the identical reserved byte range 516-689 -- one as a single 173-byte
    ``Reservado para la Administracion`` field, the other split into a 1-byte field
    at 516 and a 172-byte field at 517. A first version of this signal, without the
    exclusion, reported that as one field added and two removed: a false positive
    from re-partitioning empty space, not a real layout change. A reserved slot
    changing SHAPE while staying reserved on both sides is not a field appearing or
    disappearing; a reserved slot changing OCCUPANCY (becoming real, or a real slot
    becoming reserved) is the occupancy signal's own job and stays there rather than
    being double-reported here.
    """
    before, after = _position_content(earlier), _position_content(later)
    added = sorted(key for key in set(after) - set(before) if not _RESERVED_FIELD.search(after[key]))
    removed = sorted(key for key in set(before) - set(after) if not _RESERVED_FIELD.search(before[key]))
    if not added and not removed:
        return None
    parts = []
    if added:
        parts.append(
            f"{len(added)} added (e.g. {', '.join(f'{sheet} offset {offset}' for sheet, offset, _l in added[:3])})"
        )
    if removed:
        parts.append(
            f"{len(removed)} removed (e.g. {', '.join(f'{sheet} offset {offset}' for sheet, offset, _l in removed[:3])})"
        )
    return (
        f"field SET changed at these positions: {' and '.join(parts)} -- independent of box number, "
        "so this catches a field added or removed where no bracket exists to key membership on"
    )


#: Marks a boundary whose ONLY evidence is the description-keyed pass, which is the
#: least precise signal here. Used by the verdict text and by the review assertion.
_DESCRIPTION_ONLY = "DESCRIPTION-KEYED PASS ONLY"

#: AEAT joins the containing block to a field's own name with this separator, so the
#: final segment is the slot's own label and everything before it is context.
_LABEL_SEPARATOR = " - "


def _unnumbered_labels(path: Path) -> dict[tuple[str, int, int], str]:
    """``(sheet, offset, length) -> description`` for slots carrying NO box number.

    ABSTAINS on a flattened PDF parse by returning nothing. The PDF backend collapses a
    document to one synthetic sheet, so ``(sheet, offset, length)`` stops identifying a
    slot and starts colliding across pages -- measured, a corpus-wide run without this
    abstention returned 15366 "changes" that were overwhelmingly unrelated fields
    compared against each other.
    """
    sheets = _design_sheets(path)
    if len(sheets) == 1 and sheets[0].name == _PDF_FLATTENED_SHEET:
        return {}
    table: dict[tuple[str, int, int], str] = {}
    for sheet in sheets:
        for field in sheet.fields:
            if _BOX_MARKER.findall(field.description):
                continue
            table.setdefault((sheet.name, field.offset, field.length), " ".join(field.description.split()))
    return table


def _description_flip_evidence(earlier: Path, later: Path) -> str | None:
    """FIFTH SIGNAL: an UNNUMBERED slot whose declared meaning changes at a fixed position.

    The box-number key is structurally blind to a slot carrying no bracketed number, and
    two such slots on Modelo 303 do change meaning between the 2024 halves -- a one-byte
    flag and the reference beside it go from declaring a complementaria and its prior
    receipt to declaring an autoliquidacion rectificativa and its identifying receipt.
    Byte-valid, length-valid, digest-valid, and declaring something else.

    THE DISCRIMINATION, which is what makes this shippable. A text diff cannot tell a
    changed meaning from a reworded label, and the accepted sub-year record says so.
    AEAT writes these descriptions hierarchically, so this compares the FINAL segment: a
    changed leaf is the slot's own meaning, while an unchanged leaf under a changed
    prefix is the containing block being relabelled and is NOT reported. Validated
    against three hand-judged cases -- Modelo 390's ``Lorca`` becoming
    ``Reducciones (nota 2)`` at a fixed 17-byte slot is reported, Modelo 131 dropping La
    Palma from a one-byte deduction flag is reported, and Modelo 111's
    ``Identificacion. Ejercicio`` becoming ``Devengo. Ejercicio`` is correctly NOT, since
    only the heading above the field moved.

    WHERE IT CANNOT SEPARATE A LEAF, IT STILL ASSERTS -- as a WEAKER, DISTINCTLY MARKED
    finding, never as silence. A design carrying NO bracketed box numbers at all (an
    informative-return form, never a hierarchical ``Block - Field`` label) fails the
    leaf-separation test on every one of its changed slots, which is what silently
    discarded real divergence before this signal was fixed: Modelo 347's two boundaries
    each change dozens of unnumbered slots -- 38 of 41 shared, then 32 of 40 -- with EVERY
    ONE unseparable, so ``flipped`` stayed empty and the whole comparison returned
    ``None``, a "no boundary" verdict for a revision spanning a seventeen-year re-layout.
    That is not evidence of identity; it is the instrument declaring a limit and reporting
    silence instead of the limit. A design with no box tokens and no hierarchical labels is
    not thereby unexaminable -- position-content divergence at an unchanged offset and
    width is still real signal, it simply cannot be NAMED the way a separable leaf can.

    PRECISION, stated so the verdict is read correctly, for BOTH assertion shapes. On
    individual verdicts the separable-leaf pass runs roughly one false positive in three,
    and a measured example survives in the corpus: Modelo 303's 2014/2015 pair reports a
    leaf going from ``regimen simplificado`` to ``Regimen Simplificado (RS)``, which is a
    rewording. That costs nothing THERE because three other signals already name that
    boundary -- a false positive on an already-named boundary adds noise to evidence, not
    a wrong split. The unseparable-only shape is coarser still -- it can say THAT content
    at a position differs, never WHAT changed, so it cannot even apply the leaf-rewording
    filter -- which is the accepted trade for not discarding real divergence outright. The
    case that matters, for either shape, is a boundary this pass names ALONE, which the
    verdict marks so a reader knows it rests on the weakest instrument.

    Reserved transitions are excluded: those belong to the occupancy signal, and counting
    them here would double-report one event under two headings.
    """
    before, after = _unnumbered_labels(earlier), _unnumbered_labels(later)
    flipped: list[tuple[tuple[str, int, int], str, str]] = []
    unseparable = 0
    for slot in sorted(set(before) & set(after)):
        was, now = before[slot], after[slot]
        if _normalised(was) == _normalised(now):
            continue
        if _RESERVED_FIELD.search(was) or _RESERVED_FIELD.search(now):
            continue
        if _LABEL_SEPARATOR in was and _LABEL_SEPARATOR in now:
            leaf_was = was.rsplit(_LABEL_SEPARATOR, 1)[1]
            leaf_now = now.rsplit(_LABEL_SEPARATOR, 1)[1]
            if _normalised(leaf_was) == _normalised(leaf_now):
                continue
            flipped.append((slot, leaf_was, leaf_now))
        else:
            unseparable += 1
    if flipped:
        shown = "; ".join(
            f"{sheet} offset {offset} len {length}: {was!r} -> {now!r}"
            for (sheet, offset, length), was, now in flipped[:3]
        )
        note = (
            f"{len(flipped)} unnumbered slot(s) re-described at an unchanged position and width "
            f"(e.g. {shown}) -- the box-number key cannot see these, and no offset, length or "
            "digest check detects a slot that keeps its place while declaring something else"
        )
        if unseparable:
            note += (
                f" [plus {unseparable} slot(s) whose text changed but carries no separable leaf, "
                "NOT individually named -- see the instrument-limit note if this is the only signal]"
            )
        return note
    if unseparable:
        # No separable leaf anywhere, so nothing can be individually named -- but the
        # position-content divergence is still real, measured, and must not silently
        # collapse to "no boundary". Deliberately reuses the "unnumbered slot(s)
        # re-described" phrase so the same DESCRIPTION-ONLY marking in the callers below
        # covers this shape without a second matcher.
        return (
            f"{unseparable} unnumbered slot(s) re-described at an unchanged position and width, "
            "with NO separable leaf to name what changed -- INSTRUMENT LIMIT: this is the "
            "weakest signal in this module, reporting THAT content differs at a fixed position "
            "without being able to say WHAT, most often because the design carries no "
            "hierarchical Block-Field labels at all (an informative-return form); real "
            "divergence still, never proof of identity"
        )
    return None


def _normalised(text: str) -> str:
    """Case- and diacritic-insensitive form, so an accent or casing fix is not a flip."""
    folded = unicodedata.normalize("NFKD", " ".join(text.split()).casefold())
    return "".join(char for char in folded if not unicodedata.combining(char))


def _boundary_label(earlier: Path, later: Path) -> tuple[int, int]:
    """``(left year, right year)``; the two are EQUAL for a mid-course split."""
    return max(_design_coverage_years(earlier)), min(_design_coverage_years(later))


def _boundaries_for(modelo_id: str, revision: ModeloRevision) -> dict[tuple[int, int], list[str]]:
    """Every re-layout boundary inside one revision's span, keyed year-pair to evidence.

    Both signals contribute to ONE verdict rather than reporting separately,
    because they see overlapping-but-different boundary sets and a reader
    unioning two lists by hand will miss the ones only the weaker signal saw.

    A key whose two years are EQUAL is a mid-course split, where AEAT re-laid out a
    form partway through one ejercicio.
    """
    boundaries: dict[tuple[int, int], list[str]] = {}
    claimed_designs = _designs_claimed_by(modelo_id, revision)

    for earlier, later in pairwise(claimed_designs):
        evidence = _compare_design_pair(earlier, later)
        if evidence:
            boundaries[_boundary_label(earlier, later)] = evidence

    return boundaries


def _compare_design_pair(earlier: Path, later: Path) -> list[str]:
    """Every signal's evidence that two designs diverge; empty when they agree.

    THE ONE INSTRUMENT this module compares designs with, extracted so it has
    exactly one caller-independent body. :func:`_boundaries_for` calls this once
    per ADJACENT pair inside a revision's own claimed span, to prove no
    re-layout crosses the span. The single-year neighbour check
    (:func:`_neighbour_divergence`) calls the SAME function once against the
    immediately adjacent revision's design, to prove a single-year split was
    warranted. Two questions, one comparator -- never a second, parallel diff.
    """
    evidence: list[str] = []
    before_lengths, after_lengths = _page_lengths(earlier), _page_lengths(later)

    def _record_count_delta(
        before: tuple[str, ...] = before_lengths, after: tuple[str, ...] = after_lengths
    ) -> str | None:
        """``'9 -> 10 records'`` when the design's record SET changed, else None."""
        if not before or not after or len(before) == len(after):
            return None
        return f"{len(before)} -> {len(after)} records"

    before_boxes, after_boxes = _parse_design(earlier), _parse_design(later)
    shared = set(before_boxes) & set(after_boxes)
    moved = sorted(box for box in shared if before_boxes[box] != after_boxes[box])
    if moved:
        sample = ", ".join(f"[{box}] {before_boxes[box]}->{after_boxes[box]}" for box in moved[:3])
        note = f"{len(moved)} of {len(shared)} shared boxes moved (e.g. {sample})"
        # A displacement count measured across a decomposition change is not a
        # clean in-record figure: a box that migrated into a NEW record counts
        # as "moved" alongside one that shifted within its own. Both are real
        # movement, but comparing the magnitude against a same-record
        # boundary's is comparing different quantities.
        if _record_count_delta():
            note += " -- NOT a clean in-record displacement: the record set also changed"
        evidence.append(note)

    # FOURTH SIGNAL: the box SET changed, whether or not anything moved.
    #
    # The comparison above reads only DISPLACEMENT -- it iterates the boxes the two
    # designs SHARE -- so a box present in one design and absent in the other is
    # outside its loop entirely. That is not a lesser event: a box the later design
    # declares and the earlier one does not cannot be declared at all under the
    # earlier layout, and a box the earlier one declares and the later one drops is
    # a value written into space the later design puts to another use.
    #
    # Measured on Modelo 390, where the whole class was invisible: 2015 to 2016 adds
    # six boxes and 2016 to 2017 removes twenty, both with ZERO movement, identical
    # or unreadable page lengths and no occupancy transition, so no signal in this
    # module reported either boundary.
    #
    # The blindness survived because it is MASKED wherever membership changes
    # alongside movement: the 2017 to 2018 boundary drops seventy-two boxes, and the
    # displacement check reports that boundary anyway on its ninety-seven moved
    # boxes, so a reader spot-checking the signal against that pair sees a
    # membership change duly reported and concludes the set is compared.
    membership = _box_set_evidence(before_boxes, after_boxes)
    if membership:
        evidence.append(membership)

    if before_lengths and after_lengths and before_lengths != after_lengths:
        delta = _record_count_delta()
        # Say what a page-length change MEANS before showing the raw tuples. A
        # record-count change is a different and larger event than a page growing,
        # and stated as bare tuples it was under-read for hours by everyone
        # looking at it, including its author.
        headline = (
            f"RECORD SET CHANGED ({delta}) -- the design's record decomposition differs, so this is not an offset shift"
            if delta
            else "page byte-lengths differ, so something moved inside a record"
        )
        evidence.append(f"{headline}: {before_lengths} vs {after_lengths}")

    evidence.extend(_occupancy_evidence(earlier, later))

    # SIXTH SIGNAL: a field added or removed at a position, independent of box number.
    # See _position_set_evidence's own docstring for the full rationale and the
    # measured false positive (Modelo 202's reserved-space repartition) its exclusion
    # closes.
    position_membership = _position_set_evidence(earlier, later)
    if position_membership:
        evidence.append(position_membership)

    description = _description_flip_evidence(earlier, later)
    if description:
        evidence.append(description)

    # SEVENTH SIGNAL: a pure displacement, which every signal above is blind to.
    straddle = _straddle_evidence(earlier, later)
    if straddle:
        evidence.append(straddle)

    return evidence


def _straddle_evidence(earlier: Path, later: Path) -> str | None:
    """SEVENTH SIGNAL: a field DISPLACED so it overlaps another without containing it.

    The signal every other one here is blind to, and the blindness is structural
    rather than incidental. The box signals key on a bracketed number, so a
    design that prints none -- Modelo 347's do not -- gives them nothing to
    watch. The membership signals key on a field being added or removed, so a
    pure WIDENING leaves them seeing the same set before and after. Modelo 347's
    2010 and 2011 declarante records have the SAME field count and the same
    descriptions; all that changed is that ``IMPORTE TOTAL ANUAL`` grew from 15
    bytes to 16 and pushed everything after it one position along.

    That is a re-layout by any useful definition -- a filing written at the
    wrong one is a byte out from position 145 to the end of the record -- and it
    went unreported.

    WHAT STRADDLING MEANS, AND WHY IT IS THE RIGHT TEST. Where a field of one
    design sits wholly INSIDE a field of the other, the narrower is a
    subdivision of the wider: AEAT split or merged a slot, and the bytes still
    correspond. Where two fields overlap with neither containing the other, each
    covers bytes the other does not, and no correspondence survives. So
    containment is tolerated and straddling is evidence, which is the same
    distinction the record-design contiguity rules already turn on.

    Reads BYTES, so it needs neither a box number nor a description change.
    """
    before = {sheet.name: sheet for sheet in _design_sheets(earlier)}
    after = {sheet.name: sheet for sheet in _design_sheets(later)}

    straddles: list[str] = []
    for name in sorted(set(before) & set(after)):
        later_fields = sorted(after[name].fields, key=lambda field: field.offset)
        starts = [field.offset for field in later_fields]
        for a in before[name].fields:
            a_start, a_end = a.offset, a.offset + a.length - 1
            # Only fields starting at or before a_end can overlap; walk back far
            # enough to catch one that starts earlier and reaches into a.
            index = bisect_right(starts, a_end)
            for b in reversed(later_fields[:index]):
                b_start, b_end = b.offset, b.offset + b.length - 1
                if b_end < a_start:
                    break
                if (a_start >= b_start and a_end <= b_end) or (b_start >= a_start and b_end <= a_end):
                    continue
                straddles.append(f"{name} @{a_start}-{a_end} vs @{b_start}-{b_end}")

    if not straddles:
        return None
    return f"{len(straddles)} field(s) displaced so they straddle the other design's boundaries (e.g. {straddles[0]})"


def _occupancy_evidence(earlier: Path, later: Path) -> list[str]:
    """THIRD SIGNAL: a slot moving into or out of reserved space.

    It moves no box and changes no page length -- the reserved block absorbs the freed
    bytes exactly -- so neither signal above can see it, and a digest cannot either. It
    is still a re-layout: a field present in one design and absent in the next means a
    filing written under the older layout puts declared values into space AEAT now marks
    reserved.

    Measured live on Modelo 390, where three ``Reg. Simplificado - Reducción aplicable``
    slots were retired between the 2024 and 2025 designs while both signals above
    reported the years identical.

    BOTH DIRECTIONS are asserted, and the reverse one was withheld on a claim that was
    never checked against the corpus it described. This module used to record that
    reserved -> real "measures zero across the whole bundled corpus, so an assertion for
    it would ship vacuous and pass silently forever." Measured through these very
    helpers, it is 32 transitions across four modelos and twelve boundaries -- twice the
    16 retirements the direction that WAS asserted finds. A rationale for withholding an
    assertion is itself a measurement, and this one was reasoned rather than run.

    Nor is it the lesser half. A slot revived OUT of reserved space is a field the later
    design declares and the earlier one does not, so a filing written under the earlier
    layout cannot declare that quantity at all while the later one can -- the same harm
    as a retirement with the two sides exchanged, and equally invisible to an offset
    check, a length check and a digest. On Modelo 303 it is the only signal in this
    module that names a boundary at 2017/2018.
    """
    before, after = _occupancy(earlier), _occupancy(later)
    shared = set(before) & set(after)
    evidence: list[str] = []
    for slots, headline in (
        (
            sorted(slot for slot in shared if not before[slot] and after[slot]),
            "RETIRED into reserved space",
        ),
        (
            sorted(slot for slot in shared if before[slot] and not after[slot]),
            "REVIVED out of reserved space",
        ),
    ):
        if not slots:
            continue
        sample = ", ".join(f"{sheet} offset {offset}" for sheet, offset in slots[:3])
        evidence.append(
            f"{len(slots)} slot(s) {headline} (e.g. {sample}) -- no box moved and no page "
            "length changed, so one side of this boundary declares a quantity at a position "
            "the other side marks reserved"
        )
    return evidence


def test_no_revision_spans_a_design_relayout() -> None:
    """One revision, one byte layout — so its span must not cross a re-layout.

    ONE verdict from BOTH signals, deliberately. Reporting them as separate
    failures was the instrument's own defect: the offset diff and the page-length
    diff see overlapping but different boundary sets, so a fix owner reading
    either list alone splits a revision at some of its boundaries and leaves the
    rest standing — a gate still red, looking like an incomplete fix rather than
    a wrong one. Modelo 303 is the live case: two of its boundaries are visible
    only to the page-length signal.

    The failure text is therefore the split specification. For each revision it
    names every boundary, which signal saw it, and how many revisions the span
    actually needs, so nobody has to union two lists by hand to act on it.
    """
    violations: list[str] = []
    for modelo, revision_id, revision in _filing_revisions():
        boundaries = _boundaries_for(modelo.id, revision)
        if not boundaries:
            continue
        detail = "; ".join(
            # A key whose years are EQUAL is a mid-course split. Rendering it as
            # "2024/2024" reads as a typo and hides the finding the design-file keying
            # exists to surface, so it is named for what it is.
            f"{f'{earlier} mid-year' if earlier == later else f'{earlier}/{later}'}"
            # A boundary resting solely on the description-keyed pass is marked, because
            # that pass runs roughly one false positive in three on individual verdicts.
            # A false positive on a boundary other signals already name costs nothing; one
            # that NAMES a boundary alone is the case a reader must judge rather than act
            # on, and it is invisible unless the verdict says so.
            f"{' ' + _DESCRIPTION_ONLY if len(evidence) == 1 and 'unnumbered slot(s) re-described' in evidence[0] else ''}"
            f" ({' + '.join(evidence)})"
            for (earlier, later), evidence in sorted(boundaries.items())
        )
        violations.append(
            f"modelo {modelo.id} revision {revision_id!r} spans {len(boundaries)} re-layout(s) "
            f"and needs {len(boundaries) + 1} revisions -- {detail}"
        )
    assert not violations, (
        "a revision carries ONE export layout, so a span crossing a re-layout writes prior-year "
        "filings at the wrong byte offsets. Split each revision at every boundary listed; "
        "splitting at only the ones one signal saw leaves the rest live:\n  " + "\n  ".join(violations)
    )


def test_both_occupancy_directions_have_a_positive_case_in_the_corpus() -> None:
    """Neither occupancy direction may be asserted over a corpus that cannot show it.

    This is the companion guard to the reserved-space signal, and it exists
    because the reverse direction was once withheld from the verdict on the
    recorded ground that it "measures zero across the whole bundled corpus, so
    an assertion for it would ship vacuous". That was a reasoned claim, not a
    measured one, and it was wrong by a factor of two in the direction that
    mattered. This test is what makes the same mistake impossible to repeat in
    either direction: if a corpus change ever leaves one of them with no
    instance, the signal becomes unfalsifiable and this fails LOUDLY rather than
    the verdict silently passing over it.

    GATED ON THE PROPERTY, NEVER ON A TALLY. It asserts each direction has AT
    LEAST ONE instance, not how many. The counts move every time AEAT publishes,
    so pinning today's 16 retirements and 32 revivals would encode this moment,
    train the next author to bump two constants, and detect nothing. "The signal
    can still be observed" is the durable property; "the signal is observed
    exactly n times" is a snapshot wearing an assertion's clothes.

    Deliberately spans the WHOLE corpus rather than one revision's claimed span.
    A positive case anywhere proves the signal is live; requiring one inside
    every span would fail on modelos that simply never re-layout, which is not a
    defect.
    """
    retired_seen: list[str] = []
    revived_seen: list[str] = []
    for modelo, _revision_id, revision in _declared_revisions():
        sources = dict(_sources_by_year(modelo.id))
        for earlier, later in pairwise(sorted(_claimed_years(revision, set(sources)))):
            before, after = _occupancy(sources[earlier]), _occupancy(sources[later])
            shared = set(before) & set(after)
            retired_seen.extend(
                f"modelo {modelo.id} {earlier}/{later} {slot[0]} offset {slot[1]}"
                for slot in shared
                if not before[slot] and after[slot]
            )
            revived_seen.extend(
                f"modelo {modelo.id} {earlier}/{later} {slot[0]} offset {slot[1]}"
                for slot in shared
                if before[slot] and not after[slot]
            )
    assert retired_seen, (
        "no slot anywhere in the bundled corpus is RETIRED into reserved space, so that half of the "
        "occupancy signal can no longer fail and its contribution to the verdict is vacuous"
    )
    assert revived_seen, (
        "no slot anywhere in the bundled corpus is REVIVED out of reserved space, so that half of the "
        "occupancy signal can no longer fail and its contribution to the verdict is vacuous"
    )


def test_every_ejercicio_a_design_names_is_attributed_to_it() -> None:
    """A design naming two ejercicios must be attributed to BOTH.

    ``ejercicio-(\\d{4})`` taken as a first match attributed a two-year design to
    its opening year only, so Modelo 303's 2015-y-2016 and 2019-y-2020 designs
    left 2016 and 2020 claimed by nothing -- years the corpus covers and the
    enumeration reported as unmeasured.

    The emptiness guard is deliberate: the subject of this assertion is a map
    built by globbing a directory, and a glob that matches nothing would satisfy
    every ``for`` below vacuously.
    """
    by_year = _designs_by_year("303")
    assert by_year, "no Modelo 303 design was enumerated at all; the assertions below would be vacuous"
    for year in (2015, 2016, 2019, 2020):
        assert year in by_year, f"{year} is covered by a bundled design but attributed to nothing"


def test_a_year_aeat_split_mid_course_keeps_both_of_its_designs() -> None:
    """AEAT split three Modelo 303 ejercicios mid-course; both halves must survive.

    The year-keyed maps keep the first design by filename sort and discard the
    rest, so a consumer deriving an epoch boundary from "the 2024 design" reads
    an arbitrary half of a year that has two incompatible layouts.

    Counted by CONTENT, not by path: the corpus bundles three of these designs
    twice under names differing only by a truncated extension, and a path count
    would report the duplicate as a second design and pass while proving nothing.
    """
    by_year = _designs_by_year("303")
    assert by_year, "no Modelo 303 design was enumerated at all"
    for year in (2018, 2021, 2024):
        distinct = {path.read_bytes() for path in by_year.get(year, ())}
        assert len(distinct) >= 2, (
            f"{year} should carry two distinct Modelo 303 designs (AEAT split it mid-course) "
            f"but {len(distinct)} distinct payload(s) survived enumeration"
        )


def test_the_box_marker_is_the_registry_canonical_one_and_reads_every_modelo() -> None:
    """This module must not hold its own box-number pattern, and must read every modelo.

    ASSERTS IDENTITY WITH THE CANONICAL DEFINITION, NOT A DIGIT WIDTH. Pinning "five
    digits" here would recreate the defect one modelo later and would train the next
    author to bump a literal; worse, it would make this module an independent authority
    on the pattern again, which is what went wrong. The durable property is that there is
    ONE definition and this module uses it.

    The concrete failure it closes: this module's private copy was capped at four digits
    while Modelo 200 numbers its boxes with five, so the box-offset and box-set signals
    read 23 of that modelo's 5561 bracketed tokens and reported nothing amiss. The
    canonical definition had already been widened to five for exactly this reason, and
    its own docstring records the same failure shape -- a matchless sweep reading as
    "0 casillas, 0 gap" for 36 of 38 revisions.

    The second assertion is the one that would have caught it: every modelo whose designs
    bracket a box number at all must yield boxes here. A modelo that parses designs but
    keys zero boxes is not clean, it is unread, and it reports identically to a modelo
    with nothing to find.
    """
    from .._record_design_coverage import _CASILLA_TAG_RE as _CANONICAL_TAG_RE

    assert _BOX_MARKER is _CANONICAL_TAG_RE, (
        "this module re-declared the bracketed box-number pattern instead of using the registry's "
        "canonical one; two definitions of one concept is how the four-digit cap survived while "
        "production already read five"
    )

    unread: list[str] = []
    measured = 0
    for modelo_id in sorted({modelo.id for modelo, _, _ in _filing_revisions()}):
        ordered, _unorderable = _designs_in_publication_order(modelo_id)
        for path in ordered:
            bracketed = any(
                re.search(r"\[\d+\]", field.description) for sheet in _design_sheets(path) for field in sheet.fields
            )
            if not bracketed:
                continue
            measured += 1
            if not _parse_design(path):
                unread.append(f"modelo {modelo_id} design {path.name!r}")
    assert measured, "no bundled design brackets a box number at all; the assertion below would be vacuous"
    assert not unread, (
        "these designs bracket box numbers that this module's marker does not match, so every box "
        "signal is silently switched off for them and reports identically to a design with no "
        "divergence:\n  " + "\n  ".join(sorted(set(unread)))
    )


def test_a_boundary_only_the_description_pass_sees_is_reported_and_marked_for_review() -> None:
    """The least precise signal must still reach the verdict, and must say when it is alone.

    A slot carrying no bracketed box number can change what it declares while keeping its
    offset and its width. Modelo 303 does exactly that between the 2024 halves, where a
    one-byte flag and the reference beside it stop declaring a complementaria and start
    declaring an autoliquidacion rectificativa. No offset check, length check, occupancy
    check or digest detects it, and the box-number key structurally cannot.

    TWO PROPERTIES, NEITHER A COUNT. First, the pass must have a positive case somewhere,
    or it has become unfalsifiable and its silence means nothing. Second, every boundary
    resting SOLELY on it must be marked in the verdict text.

    The marking is the honest part. This pass runs roughly one false positive in three on
    individual verdicts -- a measured example survives at Modelo 303 2014/2015, where a
    leaf goes from ``regimen simplificado`` to ``Regimen Simplificado (RS)``, a rewording
    rather than a meaning change. That costs nothing there, because three other signals
    already name that boundary and a false positive on an already-named boundary adds
    noise to the evidence rather than a wrong split. The case a reader must judge is a
    boundary this pass names ALONE, and that case is invisible unless the verdict says so.
    """
    positive: list[str] = []
    alone: list[tuple[str, str, tuple[int, int]]] = []
    for modelo, revision_id, revision in _declared_revisions():
        for earlier, later in pairwise(_designs_claimed_by(modelo.id, revision)):
            if _description_flip_evidence(earlier, later):
                positive.append(f"modelo {modelo.id} {_boundary_label(earlier, later)}")
        for key, evidence in _boundaries_for(modelo.id, revision).items():
            if len(evidence) == 1 and "unnumbered slot(s) re-described" in evidence[0]:
                alone.append((modelo.id, revision_id, key))

    assert positive, (
        "no design pair anywhere re-describes an unnumbered slot at an unchanged position and "
        "width, so this pass can no longer fail and its silence about the complementaria-to-"
        "rectificativa class means nothing"
    )

    for modelo_id, revision_id, key in alone:
        modelo, revision = next(
            (candidate, current)
            for candidate, current_id, current in _declared_revisions()
            if candidate.id == modelo_id and current_id == revision_id
        )
        boundaries = _boundaries_for(modelo.id, revision)
        rendered = "; ".join(
            f"{f'{a} mid-year' if a == b else f'{a}/{b}'}"
            f"{' ' + _DESCRIPTION_ONLY if len(ev) == 1 and 'unnumbered slot(s) re-described' in ev[0] else ''}"
            for (a, b), ev in sorted(boundaries.items())
        )
        assert _DESCRIPTION_ONLY in rendered, (
            f"modelo {modelo_id} revision {revision_id!r} boundary {key} rests only on the "
            "description-keyed pass, which runs roughly one false positive in three, and the "
            "verdict does not mark it as such -- a reader cannot tell which boundaries to judge "
            "rather than act on"
        )


def _era_ordered_registered_designs(modelo_id: str) -> tuple[Path, ...]:
    """One path per REGISTERED design of a modelo, ordered by the era it declares.

    Deliberately not ``_design_sources``, which answers a different question. That
    walk returns every design FILE, so a design AEAT ships as both ``.xls`` and
    ``.xlsx`` appears twice, and it sorts by filename -- AEAT numbers newest-first
    -- so consecutive entries run backwards through time. Pairing it produces two
    kinds of nonsense: a design compared against its own format twin, and a later
    design read as the earlier one.

    Keyed on the SOURCE ID, which is one per design regardless of how many
    renderings the corpus holds, and ordered on ``applies_from``, which the
    catalogue states rather than a filename implies.
    """
    entries = []
    for source in _authority().catalogues.sources.values():
        if getattr(source, "kind", None) != "record_design" or source.applies_from is None:
            continue
        posix = Path(str(source.corpus_path)).as_posix()
        marker = "disenos_registro/modelo_"
        if marker not in posix:
            continue
        if posix.split(marker, 1)[1].split("/", 1)[0] != modelo_id:
            continue
        path = bundled_path() / source.corpus_path
        if path.is_file() and path.suffix.lower() in _DESIGN_SUFFIXES:
            entries.append((source.applies_from, source.id, path))
    return tuple(path for _, _, path in sorted(entries))


def _membership_only_design_pairs() -> tuple[tuple[str, Path, Path], ...]:
    """Consecutive registered designs whose ONLY difference is which boxes exist."""
    found: list[tuple[str, Path, Path]] = []
    for modelo in _authority().modelos:
        for earlier, later in pairwise(_era_ordered_registered_designs(str(modelo.id))):
            before_boxes, after_boxes = _parse_design(earlier), _parse_design(later)
            if not before_boxes or not after_boxes:
                continue
            shared = set(before_boxes) & set(after_boxes)
            if any(before_boxes[box] != after_boxes[box] for box in shared):
                continue
            if set(before_boxes) == set(after_boxes):
                continue
            before_lengths, after_lengths = _page_lengths(earlier), _page_lengths(later)
            if before_lengths and after_lengths and before_lengths != after_lengths:
                continue
            before_occupancy, after_occupancy = _occupancy(earlier), _occupancy(later)
            if any(
                before_occupancy[slot] != after_occupancy[slot] for slot in set(before_occupancy) & set(after_occupancy)
            ):
                continue
            found.append((str(modelo.id), earlier, later))
    return tuple(found)


def test_a_box_added_or_removed_without_movement_reaches_the_verdict() -> None:
    """A boundary only the box-SET comparison can see must reach the failure text.

    The displacement check iterates the boxes two designs SHARE, so a box present in one
    and absent in the other falls outside its loop. Measured on Modelo 390, that left a
    whole class unreported: 2015 to 2016 adds six boxes and 2016 to 2017 removes twenty,
    both with zero movement, no readable page-length difference and no occupancy
    transition, so no signal in this module named either boundary.

    GATED ON THE MEMBERSHIP PROPERTY, never on the numbers. It does not assert six added
    or twenty removed -- those are today's corpus, and pinning them would train the next
    author to bump two constants and would then detect nothing. The durable property is
    that a pair whose ONLY difference is which boxes exist still produces a boundary.

    THE TWO SIDES ARE DERIVED INDEPENDENTLY, and that is deliberate rather than
    incidental. Availability is measured straight from the parsed designs; the reported
    side comes from the verdict builder. Deriving both from the verdict builder is the
    shape that has already caught this module's author twice: under mutation such a test
    reds on its own vacuity guard, which proves the function changed and nothing about
    whether the signal works.
        WHY THE PAIRS COME FROM THE CATALOGUE AND NOT FROM REVISIONS. This walked the
    designs each REVISION claims, which made its liveness depend on how revisions
    happen to be carved. As the spanning revisions were split, that
    population fell to two across the whole tree, and the assertion below began
    failing for want of an example rather than for want of the signal. The
    property being proved is about the COMPARATOR, so it is now measured over
    consecutive registered designs, a population that does not move when a
    revision is renamed.

    BOX KEYS ARE COMPARED RAW, deliberately. Stripping leading zeros looks like an
    obvious normalisation and is wrong here: 26 bundled designs declare ``001``
    and ``1`` as DISTINCT boxes, so collapsing them would merge real boxes and
    hide the very membership changes this signal exists to see.
    """
    membership_only = _membership_only_design_pairs()

    assert membership_only, (
        "no registered design pair differs ONLY in which boxes it declares, so this assertion "
        "would be vacuous -- the corpus that made the membership signal necessary has changed"
    )
    for modelo_id, earlier, later in membership_only:
        evidence = _compare_design_pair(earlier, later)
        assert any("box SET changed" in item for item in evidence), (
            f"modelo {modelo_id}: {earlier.name} and {later.name} differ only in which boxes they "
            "declare -- no box moved, no page length changed, no slot changed occupancy -- and the "
            "comparison names no membership signal, so a box added or removed is invisible"
        )


def test_no_bundled_design_file_disappears_from_the_inventory() -> None:
    """A design the corpus holds but the inventory does not enumerate must FAIL, not vanish.

    THE FAILURE THIS CLOSES IS SILENT PROGRESS. Every other guard in this module asks
    whether the designs it was given disagree. None of them asks whether it was given
    all of them. Withhold a design file and the boundary it formed simply stops being
    reported: the verdict names fewer violations, which reads as a split landing rather
    than as an instrument going blind. That is the most dangerous direction for this
    gate, because a shorter verdict would then reward the gate's own blindness.

    THE ENUMERATION IS DERIVED INDEPENDENTLY, which is what makes the check possible at
    all. It globs the corpus directory itself rather than asking the inventory under
    test what exists -- a guard built on ``_design_sources`` would be blind to
    ``_design_sources`` dropping a file, which is precisely the defect. Two derivations
    of one fact, deliberately, in the one place where sharing an implementation destroys
    the check.

    GATED ON THE PROPERTY: every accepted-suffix file on disk is either enumerated, or
    named as unparseable by the sibling coverage guard. No count is pinned, so the
    corpus can grow without touching this.
    """
    missing: list[str] = []
    seen_any = 0
    # DELIBERATELY DOES NOT LOAD THE REGISTRY AUTHORITY. Whether a bundled file is
    # enumerated is a fact about the corpus and the inventory; nothing about it depends
    # on a legal catalogue validating or a revision declaring an export layout. An
    # earlier draft derived its modelo list from the exporting revisions and was taken
    # down by an unrelated peer condition -- a legal reference whose corpus sidecar had
    # not been generated yet. That is a blind spot rather than bad luck: a guard that
    # cannot run while another part of the tree is mid-edit is unavailable exactly when
    # a withheld file is most likely to slip in unnoticed.
    design_root = bundled_path(*_DESIGN_ROOT_PARTS)
    for directory in scan_directory(design_root, pattern="modelo_*", select=DirectoryEntryKind.DIRECTORIES):
        modelo_id = directory.name.removeprefix("modelo_")
        # Recursive, matching _design_sources. The two derivations stayed independent
        # in their SET logic while sharing one glob SHAPE, so a design outside
        # ``files/`` was invisible to the inventory AND to the guard that exists to
        # catch the inventory dropping a file. Two derivations of one fact protect
        # nothing where both inherit the same blind spot.
        on_disk = {
            path
            for path in scan_directory(directory, recursive=True, select=DirectoryEntryKind.FILES)
            if path.suffix.lower() in _DESIGN_SUFFIXES
        }
        if not on_disk:
            continue
        seen_any += len(on_disk)
        enumerated = set(_design_sources(modelo_id))
        missing.extend(f"modelo {modelo_id} design {path.name!r}" for path in sorted(on_disk - enumerated))
    assert seen_any, "no bundled design file was found on disk at all; the corpus path has moved"
    assert not missing, (
        "these design files exist in the bundled corpus but the inventory does not enumerate them, so "
        "every boundary they form is silently absent from the verdict and the gate getting shorter "
        "would read as progress:\n  " + "\n  ".join(missing)
    )


def test_the_verdict_names_a_mid_course_boundary_where_aeat_split_an_ejercicio() -> None:
    """A boundary INSIDE one ejercicio must reach the verdict, not just the ones between years.

    This is what the design-file keying buys, and it is the assertion that makes the
    keying provable. The per-signal inventories this replaced kept ONE design per year
    through a ``setdefault`` over a filename sort, so the second half of a mid-course
    ejercicio was discarded before any comparison ran and a boundary inside that year
    could not be reported by any signal. The gate's silence about it was therefore not
    evidence of anything -- and silence that looks like a clean result is this
    instrument's worst failure mode, because a split authored on it would leave the
    boundary live.

    GATED ON THE PROPERTY: the verdict must contain at least one boundary whose two
    years are EQUAL, which is what a mid-course split looks like once the inventory can
    see both halves. It pins no year, no modelo and no count, so it survives AEAT
    splitting a different ejercicio and it cannot be satisfied by a stale constant.

    Guarded against vacuity from the other side too: it first confirms the corpus
    actually HOLDS a mid-split ejercicio inside a gated span, so a corpus that lost one
    fails loudly here instead of passing by having nothing to find.
    """
    mid_split_available: list[str] = []
    mid_course_reported: list[str] = []
    for modelo, revision_id, revision in _filing_revisions():
        # The availability side is derived from the raw publication-order enumeration
        # rather than from the span helper the verdict uses. Deriving both sides from one
        # function makes this test notice only that the function changed, so a defect in
        # it would red the vacuity guard and never reach the assertion that matters.
        ordered, _unorderable = _designs_in_publication_order(modelo.id)
        claimed_years = _claimed_years(revision, {year for path in ordered for year in _design_coverage_years(path)})
        by_year: dict[int, int] = {}
        for path in ordered:
            opening = min(_design_coverage_years(path))
            if set(_design_coverage_years(path)) & claimed_years:
                by_year[opening] = by_year.get(opening, 0) + 1
        mid_split_available.extend(
            f"modelo {modelo.id} revision {revision_id!r} ejercicio {year}"
            for year, count in sorted(by_year.items())
            if count > 1
        )
        mid_course_reported.extend(
            f"modelo {modelo.id} revision {revision_id!r} ejercicio {earlier}"
            for earlier, later in _boundaries_for(modelo.id, revision)
            if earlier == later
        )

    assert mid_split_available, (
        "no gated revision claims an ejercicio carrying two designs, so this assertion would be "
        "vacuous -- the corpus that made the design-file keying necessary has changed"
    )

    # PROVEN ON A CONSTRUCTED SPAN, because no DECLARED revision spans a
    # mid-course boundary any more and that is the tree being right rather than
    # the instrument being blind. Every mid-course split AEAT published is now
    # partitioned into halves, and a half scoped to its own months claims only
    # the design it cites, so it reports nothing -- correctly.
    #
    # Widening one of those halves back across the whole ejercicio reconstructs
    # the case the keying exists for. If an inventory ever returns to keeping one
    # design per year, this reports no boundary and fails, which is the same
    # protection the original assertion gave when the tree still carried a
    # spanning revision.
    widened_reported: list[str] = []
    for modelo, revision_id, revision in _filing_revisions():
        if revision.valid_from is None or revision.valid_to is None:
            continue
        if revision.valid_from.year != revision.valid_to.year:
            continue
        year = revision.valid_from.year
        widened = revision.model_copy(
            update={"valid_from": date(year, 1, 1), "valid_to": date(year, 12, 31)},
        )
        widened_reported.extend(
            f"modelo {modelo.id} revision {revision_id!r} ejercicio {earlier}"
            for earlier, later in _boundaries_for(modelo.id, widened)
            if earlier == later
        )

    assert widened_reported, (
        "the corpus holds a mid-course split inside a gated span "
        f"({sorted(set(mid_split_available))}) but widening a revision across the whole ejercicio "
        "still names no boundary inside a single year, so an inventory is back to keeping one "
        "design per year and its silence about that boundary means nothing"
    )
    assert not mid_course_reported, "a DECLARED revision spans a mid-course boundary: " + ", ".join(
        sorted(set(mid_course_reported))
    )


def test_a_mid_split_ejercicio_orders_its_halves_by_declared_coverage_not_by_filename() -> None:
    """The two halves of a mid-split year order on what AEAT declares, never on the filename.

    GATED ON THE ORDERING PROPERTY, not on today's filename-to-year mapping. It asserts
    that a design bounded ABOVE (``hasta``) precedes one bounded BELOW (``desde`` /
    ``a-partir-de``) for the same ejercicio, which is what "covers the earlier periods"
    means. It pins no prefix, no year and no count, so AEAT renumbering its published
    listing cannot break it and no author is trained to bump a table.

    A year whose designs do not all declare a bound must be REPORTED as unorderable
    rather than ordered on a guess. Modelo 303's 2018 pair is the live case.
    """
    ordered, unorderable = _designs_in_publication_order("303")
    assert ordered, "no Modelo 303 design was enumerated at all; the assertions below would be vacuous"

    grouped: dict[int, list[Path]] = {}
    for path in ordered:
        grouped.setdefault(min(_design_coverage_years(path)), []).append(path)
    multi = {year: paths for year, paths in grouped.items() if len(paths) > 1}
    assert multi, "no ejercicio carries two designs, so this ordering assertion would be vacuous"

    for year, paths in sorted(multi.items()):
        if year in unorderable:
            continue
        declared = [_coverage_start_period(path.name) for path in paths]
        # A year reported orderable must have a bound on every member; otherwise the
        # unorderable report is itself wrong and the comparisons below are meaningless.
        assert all(start is not None for start in declared), (
            f"ejercicio {year} was reported orderable but a design declares no coverage bound: "
            f"{[path.name for path in paths]}"
        )
        starts = [start for start in declared if start is not None]
        assert starts == sorted(starts), f"ejercicio {year} designs are not in declared-coverage order: {starts}"
        assert len(set(starts)) == len(starts), (
            f"ejercicio {year} has two designs declaring the same coverage start {starts}, "
            "so their order is not determined by what AEAT published"
        )
        # The bounded-above half must come first, which is the direction the whole
        # ordering exists to get right.
        assert "hasta" in paths[0].name.lower(), (
            f"ejercicio {year} sorts {paths[0].name!r} first, but the half covering the "
            "earlier periods is the one AEAT bounds with 'hasta'"
        )

    assert 2018 in unorderable, (
        "Modelo 303's 2018 pair declares no period bound on either half ('ejercicio-2018' "
        "and 'ejercicio-2018-salvo-ultimo-periodo-12m-4t'), so it must be reported as "
        "unorderable rather than silently ordered by filename"
    )


def test_the_added_boxes_attach_to_the_epoch_that_introduced_them() -> None:
    """The eight boxes AEAT added mid-2024 belong to the mid-year boundary, not to 2023/2024.

    THIS IS THE ASSERTION THAT CATCHES THE ORDERING DEFECT, and it exists because a
    count-based one structurally cannot. Three consecutive designs yield two boundaries
    in ANY order, so the boundary COUNT is identical whether the two 2024 halves are
    paired early-then-late or late-then-early. Measured: under filename order the eight
    added boxes attributed to the 2023-to-2024 boundary and the mid-year boundary showed
    only occupancy movement; corrected, the eight attach to the mid-year boundary where
    AEAT introduced them and the 2023 transition shows no box-set change at all.

    A split authored against the wrong attribution puts those boxes in the wrong
    revision, and no offset check, length check or digest would detect it.

    Asserts the DIRECTION, never the tally: "no box-set change" against "some box-set
    change". The eight could become nine at the next publication without touching this.
    """

    def numbered(path: Path) -> set[str]:
        found: set[str] = set()
        for sheet in _design_sheets(path):
            for field in sheet.fields:
                boxes = _BOX_MARKER.findall(field.description)
                if boxes:
                    found.add(boxes[-1])
        return found

    # Walk the ORDERED sequence adjacently, exactly as a boundary-deriving consumer
    # does. Re-deriving the order here from the coverage helper would make this test
    # insensitive to the ordering function it exists to guard.
    ordered, _unorderable = _designs_in_publication_order("303")
    window = [path for path in ordered if min(_design_coverage_years(path)) >= 2023]
    assert len(window) >= 3, "fewer than three Modelo 303 designs from 2023 on; this assertion would be vacuous"

    introduced: dict[str, set[str]] = {}
    for earlier, later in pairwise(window):
        left, right = max(_design_coverage_years(earlier)), min(_design_coverage_years(later))
        label = f"{left} mid-year" if left == right else f"{left}/{right}"
        introduced[label] = numbered(later) - numbered(earlier)

    for label in ("2023/2024", "2024 mid-year"):
        assert label in introduced, (
            f"the {label} boundary is absent from the ordered walk, so this assertion would be vacuous"
        )

    assert not introduced["2023/2024"], (
        "the 2023-to-early-2024 transition must introduce NO new numbered box -- boxes "
        f"{sorted(introduced['2023/2024'], key=int)} attributed there instead, which is the "
        "signature of the two 2024 halves being paired in the wrong order"
    )
    assert introduced["2024 mid-year"], (
        "the mid-2024 transition must introduce the numbered boxes AEAT added at periods "
        "09 and 3T, but none attributed there, so the halves are paired in the wrong order"
    )


def test_the_orden_year_in_a_filename_is_not_read_as_a_coverage_year() -> None:
    """``orden-hap-2373-2014-...-ejercicio-2018`` covers 2018, not 2014.

    The negative control for the widened attribution: anchoring on ``ejercicio-``
    rather than scanning for four-digit runs is what keeps a legislative
    instrument's own year out of the coverage map.
    """
    assert _design_years("13-303-orden-hap-2373-2014-de-9-de-diciembre-ejercicio-2018-salvo.xlsx") == (2018,)
    assert _design_years("10-303-orden-hap-2373-2014-ejercicio-2015-y-2016-247-kb-xlsx.xlsx") == (2015, 2016)
    assert _design_years("02-303-ejercicio-2022-y-siguientes-actualizado-27-12-2021.xlsx") == (2022,)
    assert _design_years("07-303-orden-eha-3786-2008-v1-1-36-kb-pdf.pdf") == ()


def test_the_plural_and_range_naming_variants_are_read() -> None:
    """AEAT names an ejercicio four ways and two of them are not two years.

    Matching only the singular ``ejercicio-`` missed 40 of the 209 bundled
    design files across 15 modelos, and reading ``a``/``hasta`` as a pair of
    endpoints drops every year between them. Both failures are silent: an
    unenumerated design is indistinguishable from an absent one.
    """
    # plural, single year
    assert _design_years("01-115-orden-eha-3435-2007-ejercicios-2019-y-siguientes.xlsx") == (2019,)
    # "y" is AND -- two discrete years, nothing between them
    assert _design_years("10-303-ejercicio-2015-y-2016.xlsx") == (2015, 2016)
    # "a" and "hasta" are THROUGH -- an inclusive range
    assert _design_years("02-111-ejercicios-2004-a-2009-49-kb-pdf.pdf") == (2004, 2005, 2006, 2007, 2008, 2009)
    assert _design_years("06-111-ejercicios-2016-hasta-2018.pdf") == (2016, 2017, 2018)
    # the orden year is still never coverage
    assert _design_years("01-111-orden-eha-3127-2009-ejercicios-2019-y-siguientes.xlsx") == (2019,)


def test_a_period_qualified_designation_yields_its_ejercicio() -> None:
    """A pago-fraccionado design names its coverage by PERIOD, and that is still coverage.

    The year-follows-``ejercicio-`` anchor misses these because a period token sits
    between the word and the digits, so three Modelo 202 designs enumerated as
    covering nothing. Widening to the period token is not the orden-year
    relaxation: the anchor is still an explicit coverage designation, which is why
    the negative controls below keep passing.
    """
    assert _design_years("07-202-orden-hap-1552-2016-ejercicio-2p-y-3p-2016-128-kb-xlsx.xlsx") == (2016,)
    assert _design_years("09-202-orden-hap-2214-2013-ejercicios-3p-2013-y-2014-46-kb-pdf.pdf") == (2013, 2014)
    assert _design_years("10-202-orden-hap-523-2015-1p-2016-124-kb-xlsx.xlsx") == (2016,)
    # An orden number carries no period token, so it still supplies no year, and a
    # period-scoped span with no year attached still enumerates nothing rather than
    # borrowing the update year beside it.
    assert _design_years("14-202-orden-eha-664-2010-adaptada-a-la-ultima-normativa-vigente.pdf") == ()
    assert _design_years("01-763-desde-2018-4t-y-siguientes-actualizado-en-2023.xlsx") == ()


def test_a_design_title_is_read_as_coverage_where_the_filename_states_nothing() -> None:
    """The design's own title states the ejercicio the filename withholds.

    THE POSITIVE HALF of the attribution bite proof. Modelo 180's 2014-orden design
    and Modelo 303's 2008-orden design both state their ejercicio in their heading,
    and neither filename carries one, so before this the pair entered no map at all
    and their modelo's ``boundaries`` was empty for want of anything to compare
    rather than for want of a boundary.
    """
    m180 = _design_dir("180") / "files" / "03-180-orden-hap-1732-2014-de-24-de-septiembre-105-kb-pdf.pdf"
    m303 = _design_dir("303") / "files" / "07-303-orden-eha-3786-2008-v1-1-36-kb-pdf.pdf"
    for path in (m180, m303):
        assert path.is_file(), f"corpus anchor moved: {path}"

    # The filename rule refuses both, exactly as it should.
    assert _design_years(m180.name) == ()
    assert _design_years(m303.name) == ()
    # The document states what the filename does not, and the orden year is NOT it:
    # 2014 -> 2021 is a seven-year divergence, which is why the filename cannot be
    # trusted to supply it by proximity or by sequence.
    assert _title_ejercicio_years(m180) == (2021,)
    assert _title_ejercicio_years(m303) == (2009,)
    assert _design_coverage_years(m180) == (2021,)
    assert _design_coverage_years(m303) == (2009,)


def test_a_filename_carrying_only_an_orden_year_is_still_attributed_nothing() -> None:
    """THE NEGATIVE HALF: a design whose content is silent stays unattributed.

    The attribution reads the document, so a document that asserts no ejercicio must
    yield nothing rather than falling back on the orden year the filename carries.
    Modelo 840's design is the case: its filename offers ``2003`` and its content
    offers nothing, and inferring 2003 from the orden is the precise regression the
    filename rule exists to prevent.
    """
    m840 = _design_dir("840") / "files" / "01-840-orden-hac-2572-2003-99-kb-pdf.pdf"
    m720 = _design_dir("720") / "files" / "01-720-599-kb-pdf.pdf"
    for path in (m840, m720):
        assert path.is_file(), f"corpus anchor moved: {path}"
    for path in (m840, m720):
        assert _design_years(path.name) == ()
        assert _content_ejercicio_years(path) == ()
        assert _design_coverage_years(path) == ()


def test_a_constant_ejercicio_slot_is_read_and_a_filer_supplied_one_is_not() -> None:
    """The second content signal, with the discrimination that makes it safe.

    Modelo 714's 2025 design states no ejercicio in its heading and fixes one in a
    field: ``Ejercicio | Constante 2025``. Reading it recovers a design the title
    rule alone leaves unattributed.

    The negative control is the point. Modelo 303 declares an ``Ejercicio de devengo``
    slot on every one of its designs and fixes NONE of them, because that is a value
    the filer writes. A rule keyed on the word alone would attribute all six M303
    designs to whatever year sat beside that slot; keyed on the constant, it
    attributes none of them and their filenames continue to carry the coverage.
    """
    m714 = _design_dir("714") / "files" / "DR714_2025.xls"
    assert m714.is_file(), f"corpus anchor moved: {m714}"
    assert _design_years(m714.name) == ()
    assert _title_ejercicio_years(m714) == ()
    assert _constant_ejercicio_years(m714) == (2025,)
    assert _design_coverage_years(m714) == (2025,)

    m303 = _design_dir("303") / "files" / "01-303-ejercicio-2026-y-siguientes-actualizado-28-01-26-378-kb-xlsx.xlsx"
    assert m303.is_file(), f"corpus anchor moved: {m303}"
    assert _constant_ejercicio_years(m303) == (), (
        "Modelo 303's 'Ejercicio de devengo' slot is filled by the FILER, so reading it as "
        "the design's own coverage would attribute the design to an arbitrary year"
    )
    assert _design_coverage_years(m303) == (2026,)


def test_a_title_naming_one_ejercicio_does_not_shorten_a_filename_naming_a_span() -> None:
    """Content ADDS coverage and never removes it -- the regression proof for the union.

    A design published for one ejercicio and applying through several heads its
    first page with the opening year alone: Modelo 130's
    ``ejercicios-2009-a-2014`` design states ``Ejercicio 2009`` and nothing more.
    Reading the title as the design's coverage rather than as one of its years
    discards 2010 through 2014, and the loss is silent -- those years then form no
    boundary, and a verdict with fewer boundaries reads as a split having landed.

    Four designs in the corpus have this shape and twelve design-years ride on it.
    Pinned on the property (nothing is lost) rather than on the year lists, so it
    survives AEAT republishing any of them.
    """
    spans = {
        "130": "04-130-orden-eha-580-2009-ejercicios-2009-a-2014-36-kb-pdf.pdf",
        "131": "08-131-orden-eha-580-2009-ejercicios-2009-a-2014-26-kb-pdf.pdf",
    }
    checked = 0
    for modelo_id, filename in spans.items():
        path = _design_dir(modelo_id) / "files" / filename
        assert path.is_file(), f"corpus anchor moved: {path}"
        filename_years = set(_design_years(path.name))
        content_years = set(_content_ejercicio_years(path))
        assert len(filename_years) > 1, f"{filename!r} no longer names a span; pick another anchor"
        assert content_years, f"{filename!r} no longer states an ejercicio in its content; pick another anchor"
        assert content_years < filename_years, (
            f"{filename!r} no longer has the shape this guards -- its content used to name FEWER "
            "years than its filename, which is what makes preferring the content lossy"
        )
        assert filename_years <= set(_design_coverage_years(path)), (
            f"attribution DROPPED years for {filename!r}: filename claims {sorted(filename_years)} "
            f"but coverage resolved to {list(_design_coverage_years(path))}. A content signal may "
            "add coverage; it may never take it away."
        )
        checked += 1
    assert checked == len(spans)


def test_a_design_title_never_contradicts_a_trustworthy_filename_year() -> None:
    """Where BOTH signals speak, they must agree -- the title never wins silently.

    This is what makes ranking the title above the filename safe. A bare precedence
    rule would resolve a conflict by preferring one source and saying nothing, which
    would bury exactly the divergence that established the precedence in the first
    place.

    EXERCISED, NOT VACUOUS, and the distinction matters because this module elsewhere
    refuses to ship a signal with no observations. Seventeen designs carry both a
    trustworthy ``ejercicio-`` filename token and a title ejercicio, and all
    seventeen agree; the assertion below runs seventeen real comparisons and finds
    no conflict. That is a live check with a clean result, not a check with nothing
    to look at -- so the population itself is asserted non-empty, and a corpus that
    lost every overlapping design fails here rather than passing by having nothing
    to compare.
    """
    compared = 0
    conflicts: list[str] = []
    design_root = bundled_path(*_DESIGN_ROOT_PARTS)
    for directory in scan_directory(design_root, pattern="modelo_*", select=DirectoryEntryKind.DIRECTORIES):
        modelo_id = directory.name.removeprefix("modelo_")
        for path in _design_sources(modelo_id):
            filename_years = set(_design_years(path.name))
            content_years = set(_content_ejercicio_years(path))
            if not filename_years or not content_years:
                continue
            compared += 1
            if not filename_years & content_years:
                conflicts.append(
                    f"modelo {modelo_id} design {path.name!r}: filename claims "
                    f"{sorted(filename_years)} but the design itself states {sorted(content_years)}"
                )
    assert compared, (
        "no bundled design carries BOTH a filename ejercicio and a title ejercicio, so the "
        "title-over-filename precedence is unchecked. It is ranked higher on the strength of "
        "this comparison; with nothing to compare, the ranking is an unverified assumption."
    )
    assert not conflicts, (
        "a design's title and its filename disagree about which ejercicio it covers. The title "
        "is ranked higher, so the coverage maps have silently taken it -- resolve which is right "
        "rather than letting the precedence decide:\n  " + "\n  ".join(conflicts)
    )


#: Designs whose era IS stated but with an OPEN BOUND, which this module refuses to
#: enumerate for the same reason it refuses ``y siguientes``: turning "everything
#: before 2001" or "from 2018 4T onward" into a year list invents years AEAT did not
#: write. Distinct from :data:`_NON_EJERCICIO_COVERAGE_AXIS`, whose designs are scoped
#: on a different axis entirely -- these two ARE ejercicio-scoped, just unbounded on
#: one side, and conflating the two would misdescribe both.
#:
#: Each reason quotes AEAT's OWN published title, read from the per-modelo corpus
#: manifest rather than inferred from the stored filename.
_OPEN_BOUNDED_ERA_DESIGNS: dict[tuple[str, str], str] = {
    (
        "111",
        "04-111-ejercicios-anteriores-al-2001-65-kb-pdf.pdf",
    ): "AEAT titles it '111 - Ejercicios anteriores al 2001': open below, with no earliest ejercicio stated",
    (
        "763",
        "01-763-desde-2018-4t-y-siguientes-actualizado-en-2023.xlsx",
    ): "AEAT titles it '763 - Desde 2018 4T y siguientes': open above, and period-qualified",
}


#: Designs whose coverage IS stated, on an axis that is not an ejercicio. Each entry
#: names the axis the file itself uses, so the reason is checkable against the
#: filename rather than merely asserted. Keyed by ``(modelo, filename)`` -- never by
#: index or line -- and audited for staleness below.
#:
#: This is NOT a place to park a design whose era is merely unknown. An orden-named
#: design states no coverage at all, and absorbing it here would relabel "nobody
#: knows" as "known on another axis", which is the exact confusion the assertion
#: below exists to prevent.
_NON_EJERCICIO_COVERAGE_AXIS: dict[tuple[str, str], str] = {
    (
        "036",
        "01-036-diseno-de-registro-del-modelo-m036-03-02-2025-y-siguientes-124-kb-xlsx.xlsx",
    ): "censal declaration scoped by the date it comes into force ('03-02-2025-y-siguientes'), not by ejercicio",
    (
        "036",
        "02-036-diseno-de-registro-del-modelo-m036-03-02-2025-y-siguientes-provisional-107-kb-xlsx.xlsx",
    ): "the provisional edition of the same in-force-date scope ('03-02-2025-y-siguientes')",
    (
        "210",
        "01-210-devengos-a-partir-de-2026.xlsx",
    ): "non-resident income scoped by DEVENGO ('devengos-a-partir-de-2026'), an accrual span rather than an ejercicio",
    (
        "210",
        "02-210-devengos-entre-01-06-2022-y-01-01-2026.xls",
    ): "a closed devengo span ('devengos-entre-01-06-2022-y-01-01-2026'), again an accrual axis",
}


def test_a_bundled_design_whose_coverage_cannot_be_read_is_reported_unmeasured() -> None:
    """A design nothing can attribute is UNMEASURED, never absorbed under a guess.

    Same discipline this module already applies to a design it cannot PARSE: a file
    that enters no map is indistinguishable from a file that does not exist, and the
    verdict getting shorter reads as progress. Attribution is the second way a design
    can vanish, and until now it vanished without saying so -- twenty modelos have
    every bundled design named for an orden rather than an ejercicio, so their
    ``boundaries`` was empty because nothing was ever compared, not because nothing
    diverged.

    THE TEMPTATION THIS REFUSES. Reading the title recovers most of them, and the
    near-miss makes the rest look inferrable: an orden-named design plausibly runs
    from promulgation until superseded, so the sequence could supply what the content
    withholds. Measured, that inference is wrong --
    ``03-180-orden-hap-1732-2014-de-24-de-septiembre.pdf`` states ``Ejercicio 2021``,
    seven years from its orden -- so the remainder is genuinely unknown rather than
    merely unstated, and a guess here would put a filing year under another year's
    layout.

    Several of these are not defects at all and are named anyway: Modelo 036 is
    scoped by an in-force DATE and Modelo 210 by a devengo span, so they have real
    coverage expressed on an axis that is not an ejercicio. Enumerating those into
    years would invent years for the same reason ``y-siguientes`` is not expanded.
    Being visible as unattributed is the correct outcome for them; being silently
    absent is not.
    """
    unattributed: list[str] = []
    attributed = 0
    filing_revisions = _filing_revisions()
    filing_modelos = {modelo.id for modelo, _revision_id, _revision in filing_revisions}
    cited_design_sources = {
        str(source.corpus_path): source
        for _modelo, _revision_id, revision in filing_revisions
        for source in _layout_authority_receipts(_modelo.id, revision)
        if source.kind == "record_design"
    }
    design_root = bundled_path(*_DESIGN_ROOT_PARTS)
    for directory in scan_directory(design_root, pattern="modelo_*", select=DirectoryEntryKind.DIRECTORIES):
        modelo_id = directory.name.removeprefix("modelo_")
        if modelo_id not in filing_modelos:
            continue
        for path in _design_sources(modelo_id):
            relative = path.relative_to(bundled_path()).as_posix()
            source_receipt = cited_design_sources.get(relative)
            if source_receipt is None:
                continue
            if _design_coverage_years(path):
                attributed += 1
                continue
            if source_receipt.record_design_epoch is not None:
                attributed += 1
                continue  # catalogue receipt states the epoch the filename omits
            if (modelo_id, path.name) in _NON_EJERCICIO_COVERAGE_AXIS:
                continue  # coverage stated on a declared non-ejercicio axis
            if (modelo_id, path.name) in _OPEN_BOUNDED_ERA_DESIGNS:
                continue  # era stated, but open on one side and so not enumerable
            unattributed.append(f"modelo {modelo_id} design {path.name!r}")
    assert attributed, "no bundled design could be attributed to any year at all; attribution has broken"
    assert not unattributed, (
        "these bundled designs state no ejercicio in their filename OR their own title, so they "
        "enter no comparison and this module's silence about the years they cover means nothing. "
        "They are UNMEASURED, not clean -- attribute them from a source that actually says, or "
        "record why the design has no ejercicio to state:\n  " + "\n  ".join(unattributed)
    )


def _declared_span_is_single_year(revision: ModeloRevision) -> bool:
    """Whether this revision's OWN declared span covers exactly one filing year.

    Distinct from a shortage of comparable CORPUS years -- that conflates "the
    span itself is one year" with "we lack evidence for a wider one". This
    asks only about the DECLARATION: an explicit single-year selector
    (``years=(N,)``) or a closed range whose two ends coincide (``year_from ==
    year_to``). An open-ended span (``year_to is None``) is never single-year
    here, even when today's corpus happens to cover only its opening year --
    it is declared OPEN, so a wider internal comparison remains possible once
    more corpus lands, and it belongs to the multi-year, evidence-acquisition
    branch instead of the neighbour-comparison one.

    A genuinely single-year span cannot cross an internal re-layout boundary
    by construction -- there is no second year inside it to compare against.
    What it needs to prove is different: that its own design differs from the
    design either side of it, which is what warranted splitting it off as its
    own revision in the first place (:func:`_neighbour_divergence`).
    """
    selector = revision.period_selector
    if selector.years:
        return len(set(selector.years)) == 1
    if selector.year_from is None or selector.year_to is None:
        return False
    return selector.year_from == selector.year_to


def _ordered_revisions_by_modelo(
    all_revisions: list[tuple[ModeloDefinition, str, ModeloRevision]],
) -> dict[str, list[tuple[str, ModeloRevision]]]:
    """``{modelo id: [(revision id, revision), ...]}``, ordered by earliest claimed year.

    The ordering key is :func:`_span_years`'s own minimum, the SAME function
    the relayout gate already uses to bound an open-ended span -- one
    definition of "which year does this revision start from", not a second.
    """
    by_modelo: dict[str, list[tuple[str, ModeloRevision]]] = {}
    for modelo, revision_id, revision in all_revisions:
        by_modelo.setdefault(modelo.id, []).append((revision_id, revision))

    def _sort_key(item: tuple[str, ModeloRevision]) -> int:
        _revision_id, revision = item
        years = _span_years(revision)
        return min(years) if years else 0

    return {modelo_id: sorted(revisions, key=_sort_key) for modelo_id, revisions in by_modelo.items()}


def _distinct_orden_documents(modelo: ModeloDefinition) -> set[str]:
    """Every distinct BOE orden document this modelo's revisions cite, tree-wide.

    Keyed by the document token BEFORE the first ``:`` (``orden-eha-3434-2007``
    out of ``orden-eha-3434-2007:art-1``), so two articles of the SAME orden
    count as one document and only a genuinely DIFFERENT orden -- an amending
    or superseding instrument -- counts as legal evidence a revision split
    actually happened. Reads ``orden_aplicabilidad`` AND ``legal_refs`` on
    every revision this modelo declares, not just the one revision under
    test: a later revision's citation is evidence the amending orden exists
    in the catalogue even if it is not (yet) attached to the earlier one.

    This is deliberately NOT a per-revision-span date match. The question
    here is coarser and prior to that one: does ANY second orden exist in
    the record for this modelo at all. A modelo with only ever one cited
    orden, despite corpus-proven design evidence of a relayout, has no
    legal citation to even attempt dating -- that absence is what
    :func:`test_every_modelo_revision_span_is_corpus_proven` reports as its
    own distinct failure reason, never a pass.
    """
    documents: set[str] = set()
    for revision in modelo.revisions.values():
        for ref in (*revision.orden_aplicabilidad, *revision.legal_refs):
            if not ref.startswith("orden-"):
                continue
            documents.add(ref.split(":", 1)[0])
    return documents


def _signal_label(evidence_item: str) -> str:
    """Classify one evidence string by which of the seven signals produced it.

    LIGHTWEIGHT SUBSTRING CLASSIFICATION, matching this module's own established
    convention -- the ``_DESCRIPTION_ONLY`` marking already classifies evidence
    the same way -- rather than a structural refactor of every signal function's
    return shape to carry a label. Seven signals, seven distinguishable phrasings.

    The seventh, ``straddle``, is why this function raises rather than returning
    'unknown': it was added to :func:`_compare_design_pair` without being added
    here, and the loud failure is what surfaced the omission instead of letting a
    live signal classify as nothing.

    Raises rather than falling through to "unknown", deliberately: a seventh
    signal added later without updating this function must fail LOUDLY here,
    the same anti-vacuity posture the rest of the module takes toward a
    detector going quiet without anyone noticing.
    """
    if "shared boxes moved" in evidence_item:
        return "box-displacement"
    if "box SET changed" in evidence_item:
        return "box-SET"
    if "RECORD SET CHANGED" in evidence_item or "page byte-lengths differ" in evidence_item:
        return "page-length/record-count"
    if "RETIRED into reserved space" in evidence_item or "REVIVED out of reserved space" in evidence_item:
        return "occupancy"
    if "field SET changed at these positions" in evidence_item:
        return "position-SET"
    if "unnumbered slot(s) re-described" in evidence_item:
        return "description-flip"
    if "straddle the other design's boundaries" in evidence_item:
        return "straddle"
    raise AssertionError(
        f"evidence string matches none of the seven known signal phrasings -- either a new signal "
        f"was added without updating _signal_label, or an existing one's wording changed under it: "
        f"{evidence_item!r}"
    )


def _neighbour_divergence(
    modelo_id: str,
    revision_id: str,
    ordered: list[tuple[str, ModeloRevision]],
) -> tuple[bool, str]:
    """Whether a single-year revision's design DIFFERS from its adjacent revision(s).

    Reuses :func:`_compare_design_pair`, the SAME comparator
    :func:`_boundaries_for` uses internally -- one instrument, two callers.
    Where that function asks "does a span's OWN internal designs disagree",
    this asks "does this one-year design disagree with what came immediately
    before or after it" -- the question a single-year span can actually
    answer, since it has no internal pair to compare.

    Returns ``(proven, detail)``:

    * ``proven`` is ``True`` when the revision's own design differs from AT
      LEAST ONE neighbour it has. A real difference at either edge is what
      justifies this revision existing as its own split -- a neighbourless
      edge, or an identical one on the OTHER side, is a separate finding
      about that other boundary, not evidence against this revision. ``detail``
      names WHICH signal(s) carried the proof (:func:`_signal_label`), so a
      pass resting on exactly one signal is a property of the returned text,
      not a fact that only exists in whoever ran this by hand -- a real,
      single-signal difference IS a legitimate proof (no corroboration
      threshold is imposed here), but a reader must be able to tell it rests
      on one signal rather than several before trusting it further.
    * ``proven`` is ``False`` either because every available neighbour
      comparison is IDENTICAL (the split introduced no design change and was
      unwarranted) or because no neighbour has a readable design to compare
      against at all (nothing to prove it with, yet). When the revision's OWN
      design is the missing piece, ``detail`` distinguishes ABSENT (no file
      bundled for this year at all -- an acquisition gap) from UNPARSEABLE (a
      file IS bundled but :func:`_design_sheets` cannot read it at all -- an
      extraction-layer defect, never an acquisition gap) via
      :func:`_unparseable_design_sources`. Conflating the two sends a
      fix-owner to acquire a design that is already sitting in the corpus.
    """
    index = next(i for i, (rid, _revision) in enumerate(ordered) if rid == revision_id)
    _own_id, own_revision = ordered[index]
    own_designs = _designs_claimed_by(modelo_id, own_revision)
    if not own_designs:
        own_year = min(_span_years(own_revision)) if _span_years(own_revision) else None
        matching_unparseable = [
            path
            for path in _unparseable_design_sources(modelo_id)
            if own_year is not None and own_year in _design_years(path.name)
        ]
        if matching_unparseable:
            names = ", ".join(path.name for path in matching_unparseable)
            return False, (
                f"a design for {own_year} IS bundled ({names}) but its sheets fail to parse -- "
                "UNPARSEABLE, not absent: this is an extraction-layer defect, not a corpus gap, "
                "and acquiring another copy from AEAT would not help"
            )
        return False, "no design is bundled for its own filing year at all -- ABSENT, so no comparison is possible"

    checks: list[tuple[str, str, list[str]]] = []
    if index > 0:
        neighbour_id, neighbour_revision = ordered[index - 1]
        neighbour_designs = _designs_claimed_by(modelo_id, neighbour_revision)
        if neighbour_designs:
            checks.append(("predecessor", neighbour_id, _compare_design_pair(neighbour_designs[-1], own_designs[0])))
    if index < len(ordered) - 1:
        neighbour_id, neighbour_revision = ordered[index + 1]
        neighbour_designs = _designs_claimed_by(modelo_id, neighbour_revision)
        if neighbour_designs:
            checks.append(("successor", neighbour_id, _compare_design_pair(own_designs[-1], neighbour_designs[0])))

    if not checks:
        return False, "no adjacent revision in this modelo has a readable design to compare against"

    diverging = [(direction, neighbour_id, evidence) for direction, neighbour_id, evidence in checks if evidence]
    if diverging:
        direction, neighbour_id, evidence = diverging[0]
        labels = sorted({_signal_label(item) for item in evidence})
        corroboration = "SINGLE SIGNAL, uncorroborated" if len(labels) == 1 else f"{len(labels)} signals agree"
        return True, (
            f"differs from its {direction} {neighbour_id!r} via {corroboration} [{', '.join(labels)}]: "
            f"{' + '.join(evidence)}"
        )

    names = ", ".join(f"{direction} {neighbour_id!r}" for direction, neighbour_id, _evidence in checks)
    return False, f"identical to its {names} -- the split introduces no design change"


def test_every_modelo_revision_span_is_corpus_proven() -> None:
    """HARD FAIL: every declared revision span must be corpus-PROVEN, tree-wide.

    Operator directive, verbatim in substance: this is a law-derived project.
    AEAT published a record design for every filing year of every modelo that
    has ever existed. Whether this repository has BUNDLED that document is a
    fact about the repository, never a fact about the world -- so a revision
    this module cannot yet compare is NOT UNKNOWABLE, it is NOT YET PROVEN,
    and the correct treatment of "not yet proven" is the same as any other
    unmet requirement: the gate FAILS, by name, naming the one concrete act
    that clears it. There is no third verdict here and no soft status that
    lets a gap flow onward looking like a passed check.

    ONE VERDICT: a revision's span is corpus-proven, or the gate fails on it.
    Corpus-proven requires BOTH:

    1. ``_boundaries_for(modelo.id, revision) == {}`` -- the SAME corpus-diffing
       instrument the sibling tests in this module already prove trustworthy
       (page-length, box-displacement, box-SET-change, reserved-space
       retire/revive, and unnumbered-slot signals, unioned into one verdict)
       finds no re-layout the declared span crosses.
    2. Depends on the SHAPE of the declared span (:func:`_declared_span_is_single_year`):

       - A MULTI-YEAR or open-ended span needs at least TWO comparable
         bundled design years inside it (``_claimed_years`` against
         ``_designs_for``'s output) -- it is a CLAIM that one layout serves
         every year in it, and a claim nobody has checked is not thereby
         true.
       - A SINGLE-YEAR span cannot contain an internal boundary by
         construction, so it proves itself differently
         (:func:`_neighbour_divergence`): its own design must DIFFER from at
         least one adjacent revision's design. That is what justifies the
         split existing at all; an identical neighbour means the split
         introduced no change and was unwarranted, and no neighbour at all
         means there is nothing to prove it with yet. Operator ruling,
         verbatim in substance: there is no undecidable verdict, and a span
         that could NEVER pass regardless of any work anyone does re-creates
         exactly that under a different name -- structural rather than
         evidentiary, and worse for it. Both branches here are answerable
         from corpus: everything is passable.

    Passing on fewer than two comparable years for a wide span, or without a
    genuine neighbour divergence for a narrow one, would mean this gate
    reports "proven" for exactly the modelos it looked hardest at and found
    nothing to compare -- the same shape as the sweep that withdrew nine
    modelos' export layouts behind real-looking legal citations while the
    validator checked only that a citation RESOLVED, never that it PROHIBITED
    anything: the larger the evidence gap, the quieter it was. This gate is
    built to make that shape impossible to reproduce by omission.

    NO ALLOWLIST. NO PER-MODELO EXEMPTION. NO SKIP, XFAIL, OR CONDITIONAL
    GUARD. A failing revision names its own fix in the same line, because the
    remedies are different acts and a fix-owner who reads "split this
    revision" when the real need is "bundle the missing design" wastes a day
    and invents a split with no basis: split the revision at a corpus-proven
    boundary, bundle AEAT's published design for a named year, or merge a
    single-year revision whose split introduced no design change. Every
    remedy has an owner and an action. None is a status to wait out.

    Measured at authoring time, PER REVISION, which is the precision this
    property needs -- a modelo-level tally (design-year COUNT anywhere against
    boundary COUNT anywhere) was tried first and was wrong: it does not check
    whether the comparable years fall INSIDE the specific revision being
    judged. Modelo 100 makes the gap concrete -- its bundled designs run
    2009-2019 while all six of its revisions run 2020-2025, zero overlap --
    and the modelo-level tally called that "11 designs, 0 boundaries, proven"
    by counting the absence of a comparison as the success of one. The
    per-revision check this test runs does not make that mistake.

    97 total revisions: 10 fail the relayout crossing, 64 fail on missing
    corpus for a multi-year span, 11 fail the neighbour-divergence check for a
    single-year span (either no readable neighbour exists yet, or the split
    turned out to introduce no design change), and 12 PASS -- 3 multi-year (131
    revision 2019-2023; 202 revisions 2019-2022 and 2023-2024) plus 9
    single-year revisions the neighbour-comparison branch resolves, all from
    Modelo 131's and Modelo 303's and Modelo 390's declared single-year spans
    (Modelo 100's six single-year revisions do NOT resolve here -- their own
    designs are unreadable/unattributed, a corpus gap the neighbour check
    cannot paper over any more than the multi-year branch could). 85 of 97
    fail overall, down from 94 before the neighbour-comparison branch existed:
    real corpus evidence closed 9 revisions the prior instrument could never
    have proven regardless of any acquisition, because it never asked the
    question a single-year span can actually answer. Modelo 347's
    2008-y-siguientes moved from the multi-year branch (misdiagnosed as
    needing more corpus) to the relayout-crossing branch (its true cause,
    proven) once the fifth signal stopped discarding evidence it had already
    computed -- the fail count did not change, its ACCURACY did.

    WHICH INSTRUMENT EACH PASS RESTS ON, stated explicitly because a pass
    proven by a weaker signal should be legible as such. Of the 9 single-year
    passes: 5 (131/2024, 303/2023, 390/2022, 390/2023, 390/2024) carry at
    least one box-KEYED signal (displacement or box-SET); the other 4
    (131/2025, 131/2026, 303/2025, 390/2025) rest ENTIRELY on box-FREE
    signals -- occupancy, page-length, position-SET, or description-flip.
    None of the 9 rests SOLELY on the newest, least-exercised signal
    (position-SET, added this session) -- every pass position-SET
    contributes to is independently corroborated by at least one other
    signal. This number moves as the bundled
    corpus grows; re-run rather than trust a stale figure.
    """
    ordered_by_modelo = _ordered_revisions_by_modelo(_filing_supported_revisions())

    failures: list[str] = []
    single_signal_passes: list[str] = []
    for modelo, revision_id, revision in _filing_supported_revisions():
        boundaries = _boundaries_for(modelo.id, revision)
        if boundaries:
            detail = "; ".join(
                f"{f'{earlier} mid-year' if earlier == later else f'{earlier}/{later}'} ({' + '.join(evidence)})"
                for (earlier, later), evidence in sorted(boundaries.items())
            )
            failures.append(
                f"modelo {modelo.id} revision {revision_id!r}: spans {len(boundaries)} corpus-evidenced "
                f"re-layout(s), needs {len(boundaries) + 1} revisions -- {detail} -- FIX: split the "
                "revision at the named boundary year(s)",
            )
            orden_documents = _distinct_orden_documents(modelo)
            if len(orden_documents) <= 1:
                # A DISTINCT failure reason from the one above, never a substitute for
                # it and never a path to a pass. The design-evidence failure says WHERE
                # a split is needed; this one says the legal record offers no citation
                # to justify or date it -- the founding orden is the only orden this
                # modelo's entire revision history ever cites, despite corpus evidence
                # a relayout happened somewhere in this revision's span. Absence of a
                # second orden citation is not positive evidence of non-revision, it is
                # evidence the legal catalogue is incomplete -- so this reason can NEVER
                # be cleared by design evidence, and NEVER becomes a pass condition;
                # only a positively-cited amending or superseding orden clears it.
                failures.append(
                    f"modelo {modelo.id} revision {revision_id!r}: NO LEGAL EVIDENCE OF REVISION "
                    f"RECORDED -- the design-evidence failure above proves a relayout crosses this "
                    f"revision's span, but this modelo's entire revision history cites only the "
                    f"founding orden ({sorted(orden_documents)!r}); no amending or superseding orden "
                    "is recorded anywhere in the bundled legal catalogue -- FIX: acquire and cite the "
                    "BOE orden that authorises the later layout; do not attempt to satisfy this with "
                    "design evidence alone, and do not treat the gap as anything other than a failure",
                )
            continue

        receipt_proven, _receipt_detail = _source_epoch_proves_revision_span(modelo.id, revision)
        if receipt_proven:
            # The revision cites an authoritative record-design dependency whose
            # declared epoch covers its entire span. Requiring duplicate annual
            # copies after this point would replace authority with a file count.
            continue

        if _declared_span_is_single_year(revision):
            proven, detail = _neighbour_divergence(modelo.id, revision_id, ordered_by_modelo[modelo.id])
            if not proven:
                # Three distinct remedies for three distinct causes -- naming the wrong one
                # sends a fix-owner to acquire a design that is already bundled, or to
                # wait on AEAT for evidence that is actually an in-tree extraction defect.
                if "UNPARSEABLE, not absent" in detail:
                    fix = "the design IS bundled but unreadable -- fix the extractor for the named file(s); acquiring another copy from AEAT would not help"
                elif "ABSENT" in detail:
                    fix = "no design is bundled for this year at all -- bundle AEAT's published record design for an adjacent ejercicio so the split can be proven"
                else:
                    fix = "identical to a neighbour -- merge this revision into it, the split introduced no design change and was unwarranted"
                failures.append(
                    f"modelo {modelo.id} revision {revision_id!r}: single-year span, {detail} -- FIX: {fix}",
                )
            elif "SINGLE SIGNAL, uncorroborated" in detail:
                # A real, legitimate PASS -- no corroboration threshold is imposed here, one
                # signal finding a genuine difference is proof enough. But a reader of the
                # verdict alone cannot tell this pass rests on one instrument while its
                # neighbours rest on several, and the always-visible failure text below is
                # this gate's only durable output, so recording it there is what makes
                # thinness a property of the gate rather than a fact only a chat message
                # carries. Not a failure; do not add to `failures`.
                single_signal_passes.append(f"modelo {modelo.id} revision {revision_id!r}: PASSES, {detail}")
            continue

        design_years, _unreadable = _designs_for(modelo.id)
        claimed = _claimed_years(revision, set(design_years))
        if len(claimed) < 2:
            # Distinguish ABSENT (no file at all for a needed year -- acquire from AEAT)
            # from UNPARSEABLE (a file is already bundled for that year but the parser
            # returns nothing usable -- an extraction-layer defect, never an acquisition
            # gap) before naming the fix. Conflating the two sends a fix-owner to acquire
            # a design that is already sitting in the corpus.
            unparseable = _unparseable_design_sources(modelo.id)
            unparseable_years = set()
            unparseable_names: list[str] = []
            for path in unparseable:
                matched = _claimed_years(revision, set(_design_years(path.name)))
                if matched:
                    unparseable_years |= matched
                    unparseable_names.append(path.name)
            note = ""
            if unparseable_years:
                note = (
                    f" -- NOTE: {len(unparseable_years)} of the missing year(s) "
                    f"({sorted(unparseable_years)}) already have a BUNDLED design file that fails to "
                    f"parse entirely ({', '.join(unparseable_names)}) -- UNPARSEABLE, not absent: fix "
                    "the extractor for those files before acquiring anything new for those specific years"
                )
            failures.append(
                f"modelo {modelo.id} revision {revision_id!r}: only {len(claimed)} comparable bundled "
                "design year(s) fall inside its claimed span -- FIX: bundle AEAT's published record "
                "design for the missing year(s); do not split this revision on today's evidence, and "
                "do not treat the gap as anything other than a failure" + note,
            )

    single_signal_note = (
        "\n\nSINGLE-SIGNAL PASSES (informational only, NOT failures -- one signal finding a genuine "
        "difference is a legitimate proof and no corroboration threshold is imposed; recorded here "
        "so a reader can tell which passes rest on one instrument rather than several before "
        "trusting them further):\n  " + "\n  ".join(sorted(single_signal_passes))
        if single_signal_passes
        else ""
    )

    assert not failures, (
        "every modelo's declared revision span must be corpus-proven before it may pass: zero "
        "corpus-evidenced re-layout boundaries, AND (a multi-year span) at least two comparable bundled "
        "design years checked inside it, OR (a single-year span) a proven divergence from an adjacent "
        "revision's design. AEAT published a design for every filing year; a gap here is a fact about "
        "this repository's bundled corpus, never about the world -- it is fixed by bundling the design, "
        "splitting the revision, or merging an unwarranted split, never accepted as a standing state:\n  "
        + "\n  ".join(sorted(failures))
        + single_signal_note
    )


def _current_filing_year() -> int:
    """Today's calendar year, the rolling upper bound a coverage sweep must reach.

    Computed at call time rather than pinned as a literal: a coverage hole at the
    tail of a modelo's declared revisions is dated by the CALENDAR, not by when
    this module was last edited. A hardcoded year would itself become a silent
    hole the moment it goes stale -- the exact failure shape this check exists
    to catch, reproduced in the checker.
    """
    return date.today().year


def _covers_year(revision, year: int) -> bool:
    """Whether one revision's period selector resolves for a given filing year.

    Deliberately NOT :func:`_claimed_years` or :func:`_span_years`. Those both
    bound an open-ended (``year_to is None``) span at ``year_from`` alone, on
    purpose, because the relayout-crossing gate above can only ever speak about
    years the bundled CORPUS covers. This function answers a different
    question -- does the revision's own LAW-DETERMINED selector resolve this
    year -- so an open upper bound must extend all the way to the year asked
    about, corpus or no corpus. Collapsing the two meanings into one helper
    would make an open-ended revision that legitimately still covers today read
    as covering only its opening year, inventing a hole that is not there.
    """
    selector = revision.period_selector
    if selector.years:
        return year in selector.years
    if selector.year_from is None:
        return False
    if year < selector.year_from:
        return False
    return selector.year_to is None or year <= selector.year_to


def _period_overlap(id_a: str, periods_a: tuple[str, ...], id_b: str, periods_b: tuple[str, ...]) -> str | None:
    """Evidence that two same-year revisions genuinely collide, or ``None``.

    NOT "more than one revision claims this year" -- that naive check produces
    two confirmed false positives in the bundled corpus. Modelo 303 splits
    2024 mid-course by PERIOD (``2024-hasta-08-y-2t`` declares periods
    ``1T, 2T, 01..08``; ``2024-desde-09-y-3t`` declares ``3T, 4T, 09..12``) --
    two revisions, one year, disjoint periods, zero ambiguity. Modelo 369
    declares three simultaneous 'esquema' revisions from 2021 onward, each
    using its OWN period-token vocabulary (``EXT-1T..EXT-4T`` for the exterior
    scheme, plain ``01..12`` for the import scheme, plain ``1T..4T`` for the
    union scheme) -- a parallel regime axis, not a date collision.

    The revision-selector's own ``periods`` field is therefore the finer-grained
    signal a year-only check cannot see, matching how the production resolver
    actually disambiguates a candidate: two revisions genuinely overlap only
    when their period-token sets share a member, or when either declares NO
    period restriction at all (an empty ``periods`` tuple matches every token,
    so it collides with anything the other side claims for that year).
    """
    if not periods_a or not periods_b:
        return (
            f"revisions {id_a!r} and {id_b!r} both resolve, and at least one declares no "
            "period-level restriction, so nothing distinguishes them"
        )
    shared = sorted(set(periods_a) & set(periods_b))
    if shared:
        return f"revisions {id_a!r} and {id_b!r} both resolve for period token(s) {shared!r}"
    return None


def _earliest_declared_year(revisions: list[tuple[str, ModeloRevision]]) -> int | None:
    """The earliest filing year any of a modelo's revisions declares, or ``None``.

    ``None`` for a modelo whose every revision declares no dateable start at
    all -- there is nothing to sweep a coverage window from, and reporting a
    hole for a modelo with no stated coverage would be inventing a claim the
    registry never made.
    """
    starts: list[int] = []
    for _revision_id, revision in revisions:
        selector = revision.period_selector
        if selector.years:
            starts.append(min(selector.years))
        elif selector.year_from is not None:
            starts.append(selector.year_from)
    return min(starts) if starts else None


def _offset_annual_modelo(revisions: list[tuple[str, ModeloRevision]]) -> bool:
    """Whether every one of a modelo's revisions declares ONLY annual-cadence periods.

    Derived, not declared: :class:`~core.PeriodKind` is already the canonical,
    closed cadence classifier every period token resolves through
    (:func:`~core.registry_period_kind`), and a modelo whose every revision
    declares nothing but the annual token (``0A``) is, by construction, filed
    IN ARREARS -- an annual return cannot be computed before its own ejercicio
    closes, so its filing window necessarily opens the FOLLOWING calendar
    year. A periodic modelo (quarterly, monthly) files within its own
    ejercicio instead.

    Confirmed against the bundled corpus's own declared ``deadline_windows``
    rather than assumed: every revision of modelos 100, 189, 280, 289, 345
    and 390 that declares a deadline window shows ``opens_on.year ==
    filing_year + 1`` (e.g. Modelo 100's 2025 revision opens 2026-04-08); the
    periodic control, Modelo 303, shows ``opens_on.year == filing_year``. No
    new field is declared for this -- the offset is arithmetic on data the
    registry already carries (the period cadence), not a second axis beside
    it, per the operator's instruction not to redeclare what is already
    derivable.

    A modelo with NO declared periods at all, or with even one non-annual
    period on any revision, is treated as NOT offset -- the strict, urgent
    default a permissive read would invert. Silently classifying an
    ambiguous modelo as filed-in-arrears would hide a live gap behind an
    assumption nothing in the registry actually states.
    """
    kinds: set[PeriodKind] = set()
    any_periods = False
    for _revision_id, revision in revisions:
        for token in revision.period_selector.periods:
            any_periods = True
            try:
                kinds.add(registry_period_kind(token))
            except ValueError:
                return False
    return any_periods and kinds == {PeriodKind.ANNUAL}


def test_every_modelo_resolves_exactly_one_revision_for_every_filing_year_through_today() -> None:
    """HARD FAIL: a coverage HOLE or OVERLAP in filing-year resolution, tree-wide.

    Complements, and does not replace, the relayout-crossing gate above. That
    gate asks whether a declared span is too WIDE -- crossing a design boundary
    inside years it already claims. This one asks the orthogonal question: is a
    span too NARROW, does a gap sit BETWEEN two declared spans, or is the most
    recent span CLOSED past the point it should still be open? None of those
    three shapes is visible to boundary-diffing, because boundary-diffing only
    ever compares years a span already claims -- it cannot see a year no span
    claims at all.

    Modelo 390 is why a PAIRWISE abutment check would have been wrong here.
    Measured 2026-08-14 via the raw loader (``bundled_authority()`` itself
    refuses to build, over an unrelated export-layout-completeness gate, so
    this reads the same tier the relayout gate does), every revision is
    closed-ended --

        2022 | valid_from 2022-01-01 | valid_to 2022-12-31
        2023 | valid_from 2023-01-01 | valid_to 2023-12-31
        2024 | valid_from 2024-01-01 | valid_to 2024-12-31
        2025 | valid_from 2025-01-01 | valid_to 2025-12-31

    The revisions abut each other with no gap BETWEEN them, so a check that
    only compares consecutive revisions to EACH OTHER would call this clean
    and never notice the TAIL -- whether the last revision still reaches the
    present. This check does reach the tail: it sweeps every year through
    today. But "today" is not the same question for every modelo, which is
    the second thing this check gets right.

    THE HORIZON IS OFFSET-AWARE, NOT A FLAT "THROUGH TODAY" FOR EVERY MODELO.
    Operator directive, verbatim in substance: an annual return is filed IN
    ARREARS -- ejercicio 2025's Renta is filed in 2026, ejercicio 2026's is
    not filed until 2027 -- while a periodic modelo (quarterly, monthly)
    files WITHIN its own ejercicio. A flat "must resolve through this
    calendar year" horizon is wrong for the first shape: it would demand a
    2026 revision for Modelo 390 (the annual IVA summary) today, when AEAT's
    own filing window for ejercicio 2026 does not open until 2027 and no
    operator could file it even if the revision existed. :func:`_offset_annual_modelo`
    derives which shape a modelo is from its ALREADY-DECLARED period cadence
    -- no new field, an arithmetic ceiling instead of a stored exemption --
    and the sweep's upper bound becomes last year for an arrears modelo,
    this year for a periodic one. Measured 2026-08-14: modelos 100, 189,
    280, 289, 345, 390 and 714 are ALL annual-in-arrears (confirmed against
    their own declared ``deadline_windows``, not assumed), so none of their
    "missing 2026" cases were live gaps -- every one was premature, and the
    gate now correctly does not ask the question until the year the filing
    window could actually open. No periodic modelo carries a coverage hole
    at all; if one ever does, it is live TODAY, not next year, and this
    gate's horizon does not soften that case.

    PASS requires, per modelo, for every filing year from its earliest declared
    coverage (:func:`_earliest_declared_year`) through its own horizon
    (:func:`_current_filing_year`, offset back one year by
    :func:`_offset_annual_modelo`): at least one revision resolves
    (:func:`_covers_year`), and no two resolving revisions genuinely COLLIDE.
    Zero resolving revisions is a HOLE -- this application cannot even attempt
    the calculation for that year. A collision is an OVERLAP -- two revisions
    both claim the SAME period token (or one restricts nothing at all) inside
    that year, so which applies is undefined.

    OVERLAP IS PERIOD-TOKEN AWARE, NOT A BARE "MORE THAN ONE REVISION" COUNT
    (:func:`_period_overlap`), because more than one revision legitimately
    resolving a single year is a real, correct shape here -- AEAT splits some
    modelos mid-year by PERIOD (Modelo 303's 2024) and runs others as parallel
    REGIME tracks that share no period vocabulary at all (Modelo 369's three
    'esquema' revisions). A bare per-year count flags both as false positives;
    checking whether the resolving revisions' period tokens actually intersect
    does not.

    Checked against TODAY, never against the bundled corpus's own design years.
    There does not need to be a published AEAT record design for 2026 for a
    2026 coverage hole to be real -- resolving a revision at all is a
    law-determined prerequisite to attempting the calculation, per
    ``aeat-registry-authority-flow``'s revision-resolution mandate, and is
    upstream of whether a design exists to export it against.

    NO ALLOWLIST. NO PER-MODELO EXEMPTION. The offset horizon is the ONLY
    per-modelo variation this check makes, and it is computed fresh every run
    from the modelo's own declared period cadence -- never a stored flag, a
    disposition, or a list of modelo ids to skip. A closed-ended tail reads
    as tidy, deliberate structure right up until the calendar passes it,
    which is exactly why a static snapshot of this check would rot: it would
    go green the day it is authored and silently start lying every January
    after -- and an arrears modelo's horizon rolls forward with it the SAME
    way, computed from ``date.today()`` on every run, never pinned.

    Measured at authoring time (2026-08-14): PASSES. Zero holes, zero
    overlaps, tree-wide. This is not a weakened result -- the seven prior
    "holes" were re-classified, not deleted: their offset status is verified
    against real declared ``deadline_windows`` data, not asserted, and every
    one of them still resolves its own current-arrears year (2025). No
    periodic modelo has ever failed this check.
    """
    modelos, _catalogues = bundled_registry_tree()
    today_year = _current_filing_year()

    holes: list[str] = []
    overlaps: list[str] = []
    for modelo in modelos:
        revisions = list(modelo.revisions.items())
        earliest = _earliest_declared_year(revisions)
        if earliest is None:
            continue
        # An annual-in-arrears modelo's most recent FILEABLE ejercicio is last
        # year, not this one -- its filing window for this year's ejercicio has
        # not opened yet, arithmetic derived from the declared period cadence
        # (see :func:`_offset_annual_modelo`), never a stored exemption.
        horizon = today_year - 1 if _offset_annual_modelo(revisions) else today_year
        for year in range(earliest, horizon + 1):
            covering = [(rid, rev) for rid, rev in revisions if _covers_year(rev, year)]
            if not covering:
                holes.append(f"modelo {modelo.id}: no revision resolves filing year {year}")
                continue
            for (id_a, rev_a), (id_b, rev_b) in combinations(covering, 2):
                collision = _period_overlap(id_a, rev_a.period_selector.periods, id_b, rev_b.period_selector.periods)
                if collision:
                    overlaps.append(f"modelo {modelo.id} filing year {year}: {collision}")

    assert not holes and not overlaps, (
        "every modelo must resolve EXACTLY ONE revision for every filing year from its earliest "
        "declared coverage through today -- a hole means this application cannot even attempt the "
        "calculation for that year at all; an overlap means two revisions both claim it with no "
        "tie-break. Neither shape is visible to the relayout-crossing gate above, which only "
        "compares years a span already claims.\n"
        "HOLES -- no revision covers this year, most often a closed-ended tail the calendar has now "
        "passed:\n  " + "\n  ".join(sorted(holes)) + "\n"
        "OVERLAPS -- two or more revisions both claim this year:\n  " + "\n  ".join(sorted(overlaps))
    )


def test_every_non_ejercicio_declaration_is_still_earned() -> None:
    """Each declared non-ejercicio design must still exist AND still state no ejercicio.

    Two ways an entry goes stale, and both must fail rather than pass quietly. The
    design can be renamed or dropped from the corpus, leaving an entry that excuses
    nothing. Or the design can BECOME attributable -- a better title read, a widened
    filename pattern -- at which point the entry silently suppresses a design the
    module can now measure, which is precisely the invisibility the assertion above
    was written against.

    Keyed by ``(modelo, filename)`` so a rename fails loudly instead of drifting onto
    whatever file happens to sit at some position.
    """
    design_root = bundled_path(*_DESIGN_ROOT_PARTS)
    on_disk: dict[tuple[str, str], Path] = {}
    for directory in scan_directory(design_root, pattern="modelo_*", select=DirectoryEntryKind.DIRECTORIES):
        modelo_id = directory.name.removeprefix("modelo_")
        for path in _design_sources(modelo_id):
            on_disk[(modelo_id, path.name)] = path

    assert _NON_EJERCICIO_COVERAGE_AXIS, "the declaration is empty; this audit would be vacuous"

    missing = sorted(key for key in _NON_EJERCICIO_COVERAGE_AXIS if key not in on_disk)
    assert not missing, (
        "these designs are declared non-ejercicio-scoped but are no longer bundled under that "
        f"name, so the declaration excuses nothing: {missing}"
    )

    now_attributable = sorted(key for key in _NON_EJERCICIO_COVERAGE_AXIS if _design_coverage_years(on_disk[key]))
    assert not now_attributable, (
        "these designs now yield ejercicio coverage, so the declaration is suppressing a design "
        f"the module can measure -- remove the entry: {now_attributable}"
    )

    unreasoned = sorted(k for k, why in _NON_EJERCICIO_COVERAGE_AXIS.items() if len(why.strip()) < 30)
    assert not unreasoned, f"every entry must state the axis the file itself uses: {unreasoned}"


def test_every_open_bounded_era_declaration_is_still_earned() -> None:
    """Each open-bounded design must still exist AND still yield no year list.

    The same two staleness directions the non-ejercicio audit checks, for the same
    reason: an entry naming a design the corpus no longer holds excuses nothing, and
    an entry whose design BECAME enumerable is suppressing a design this module can
    now measure. Kept separate from that audit rather than folded into it, because
    the two declarations answer different questions and a single audit would let an
    entry drift between them unnoticed.
    """
    design_root = bundled_path(*_DESIGN_ROOT_PARTS)
    on_disk: dict[tuple[str, str], Path] = {}
    for directory in scan_directory(design_root, pattern="modelo_*", select=DirectoryEntryKind.DIRECTORIES):
        modelo_id = directory.name.removeprefix("modelo_")
        for path in _design_sources(modelo_id):
            on_disk[(modelo_id, path.name)] = path

    assert _OPEN_BOUNDED_ERA_DESIGNS, "the declaration is empty; this audit would be vacuous"

    overlap = sorted(set(_OPEN_BOUNDED_ERA_DESIGNS) & set(_NON_EJERCICIO_COVERAGE_AXIS))
    assert not overlap, f"these designs are declared under BOTH classifications, so one of them is wrong: {overlap}"

    missing = sorted(key for key in _OPEN_BOUNDED_ERA_DESIGNS if key not in on_disk)
    assert not missing, (
        "these designs are declared as open-bounded but are no longer bundled under that name, "
        f"so the declaration excuses nothing: {missing}"
    )

    now_attributable = sorted(key for key in _OPEN_BOUNDED_ERA_DESIGNS if _design_coverage_years(on_disk[key]))
    assert not now_attributable, (
        "these designs now yield ejercicio coverage, so the declaration is suppressing a design "
        f"the module can measure -- remove the entry: {now_attributable}"
    )
