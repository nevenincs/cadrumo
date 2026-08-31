"""Record-design corpus discovery and revision-span test support."""

from __future__ import annotations

import re
from functools import cache, lru_cache
from pathlib import Path

from .....core.authority_grade import RegistryAuthorityGrade
from .....core.directory_scan import DirectoryEntryKind, scan_directory
from .....core.external_constants import PDF_EXTENSION as _PDF_EXTENSION
from .....core.external_constants import XLS_EXTENSION as _XLS_EXTENSION
from .....core.resources.bundled_data import bundled_path
from .....tests.registry_tree import bundled_registry_tree
from ..authority import ValidatedRegistryAuthority
from ..record_design import (
    extract_record_design_pdf,
    extract_record_design_workbook,
    extract_record_design_xls_workbook,
)
from ..record_design_coverage import _CASILLA_TAG_RE
from ..record_design_pdf_rows import _clean_pdf_line
from ..record_design_pdf_visual import _extract_pdf_text_lines
from ..record_design_schema import RecordDesignSheet
from ..schema import ModeloDefinition, ModeloRevision
from ..schema_references import SourceReference

_DESIGN_ROOT_PARTS = ("corpus", "aeat_official", "disenos_registro")

#: The bracketed box number AEAT embeds in a design field description, IMPORTED from the
#: registry's canonical definition rather than re-declared here.
#:
#: This module carried its own copy capped at FOUR digits, and Modelo 200 numbers its
#: boxes with FIVE. Measured on its newest bundled design: 5538 of 5561 bracketed tokens
#: are five digits, so the private copy keyed 23 boxes -- under 0.4% of the modelo -- and
#: the box-offset and box-set signals were effectively switched off there while reporting
#: nothing wrong. The description-keyed population is defined as the slots carrying NO
#: box number, so the same cap also mis-classified those 5538 numbered fields as
#: unnumbered.
#:
#: The worst part was not the blindness but that it presented as agreement: an
#: independently derived boundary union and this gate's verdict both said Modelo 200 had
#: exactly one boundary and no gap, because BOTH used a four-digit marker. Agreement
#: between two instruments sharing one blind spot is worth nothing, and unlike a wrong
#: answer it offers nothing to notice.
#:
#: The canonical definition is bounded at five digits deliberately rather than left open:
#: its own rationale records that an unbounded ``\d+`` would admit amounts, NIF fragments
#: and position offsets that appear bracketed in the same columns. That reasoning is not
#: restated here, because restating it is how a third copy begins.
_BOX_MARKER = _CASILLA_TAG_RE
_DESIGN_YEAR = re.compile(r"ejercicios?-(\d{4})(?:-(y|a|hasta)-(\d{4}))?")
#: A design filename may name TWO explicit ejercicios ("ejercicio-2015-y-2016"),
#: and the corpus holds four such spans (2007-y-2008, 2008-y-2009, 2015-y-2016,
#: 2019-y-2020). Taking only the first match attributed the design to its opening
#: year and left the second attributed to NOTHING -- 2016 and 2020 each had a
#: bundled design and no year claiming it. "y-siguientes" is deliberately NOT
#: expanded: its span is open-ended, so enumerating it would invent years.
#: The orden year in a filename ("orden-hap-2373-2014-...-ejercicio-2018") is not
#: a coverage year and must never be read as one, which is why this anchors on
#: "ejercicio-" rather than scanning for any four-digit run.
#: Case-INSENSITIVE on purpose. AEAT writes both `Reservado para la AEAT` and
#: `RESERVADO PARA LA A.E.A.T.`, and a case-sensitive substring test silently
#: misses the uppercase form -- which reads a reserved block's own growth as a
#: moved field and invents a boundary that is not there.
_RESERVED_FIELD = re.compile(r"reservado", re.IGNORECASE)
#: Design SOURCE suffixes, in preference order for a year bundled more than once.
#:
#: ``.xlsm`` is accepted because the corpus bundles two Modelo 220 designs in it
#: (ejercicios 2022 and 2023) and openpyxl reads it exactly as it reads ``.xlsx``.
#: Omitting it dropped both files from every map here for no reason anyone chose:
#: a macro-enabled container is a packaging detail, not a different document.
_DESIGN_SUFFIXES = (".xlsx", ".xlsm", ".xls", ".pdf")
#: The synthetic single-sheet name the PDF backend falls back to when a document
#: declares no per-page sections. Its presence means the parse is FLATTENED, so
#: per-page signals must abstain rather than compare one synthetic page against
#: a spreadsheet's real ones.
_PDF_FLATTENED_SHEET = "PDF record design"

# Each page closes with a "TOTAL <n> POSICIONES" row, in the pipe-delimited
# spreadsheet extractions and the space-delimited PDF ones alike. The sequence of
# page lengths is a SECOND, box-number-free signal: if a page's byte length
# changed, something inside it moved. It reaches designs the box table cannot --
# the older PDF extractions carry these rows while yielding no bracketed boxes.
_PAGE_TOTAL = re.compile(r"TOTAL\s*\|?\s*\|?\s*(\d+|variable)\s*\|?\s*POSICIONES", re.IGNORECASE)

#: The ejercicio a design states in its OWN title text. Bounded to a short window
#: after the word deliberately: "Ejercicio 2023" is a coverage assertion, while
#: "rendimientos obtenidos en ejercicios iniciados antes de 1999" is prose about
#: the filer's data, and a greedy gap between the two admits the second.
_TITLE_EJERCICIO = re.compile(r"(?i)\bejercicio\b[^\n]{0,25}?\b(19[89]\d|20[0-4]\d)\b")
#: A pago-fraccionado designation: a period token immediately followed by its
#: ejercicio, optionally naming a second ("3p-2013-y-2014"). Anchored on the
#: period token so an orden number cannot supply the year -- ``hap-523-2015``
#: carries no period token and contributes nothing.
_PERIOD_QUALIFIED_YEAR = re.compile(r"(?<![0-9a-z])\d+[pt]-(\d{4})(?:-y-(\d{4}))?", re.IGNORECASE)
#: An ``Ejercicio`` slot whose declared value is a CONSTANT year, in either word
#: order AEAT writes it. The constant is what distinguishes the design fixing a
#: year from the design reserving a slot for the filer to write one in.
_CONSTANT_EJERCICIO = re.compile(r"(?i)\bejercicio\b[^|]{0,60}?\bconstante\b[^|]{0,20}?\b(19[89]\d|20[0-4]\d)\b")
_CONSTANT_EJERCICIO_REVERSED = re.compile(
    r"(?i)\bconstante\b[^|]{0,40}?\bejercicio\b[^|]{0,30}?\b(19[89]\d|20[0-4]\d)\b"
)


def _authority() -> ValidatedRegistryAuthority:
    return ValidatedRegistryAuthority.load(bundled_path("registry", "aeat"), source_root=bundled_path())


def _design_dir(modelo_id: str) -> Path:
    return bundled_path(*_DESIGN_ROOT_PARTS, f"modelo_{modelo_id}")


def _sources_by_year(modelo_id: str) -> tuple[tuple[int, Path], ...]:
    """``(design year, source path)`` pairs, first source per year winning."""
    seen: dict[int, Path] = {}
    for path in _design_sources(modelo_id):
        if not _design_sheets(path):
            continue
        for year in _design_coverage_years(path):
            seen.setdefault(year, path)
    return tuple(sorted(seen.items()))


def _design_years(name: str) -> tuple[int, ...]:
    """Every ejercicio a design filename claims, expanded and ordered.

    AEAT names these four ways and they do NOT mean the same thing:

    * ``ejercicio-2024``            -- one year.
    * ``ejercicios-2024``           -- one year; the plural is a naming
      habit, not a range, and it appears on 61 bundled files.
    * ``ejercicio-2015-y-2016``     -- TWO discrete years ("and").
    * ``ejercicios-2004-a-2009`` /
      ``ejercicios-2016-hasta-2018`` -- an INCLUSIVE RANGE ("through"),
      so 2004-a-2009 claims six years, not two endpoints.

    Reading ``a``/``hasta`` as two endpoints silently drops every year
    between them, and matching only the singular ``ejercicio-`` misses 40 of
    the 209 bundled design files across 15 modelos entirely -- a design that
    is invisible to enumeration reads exactly like one that does not exist.

    ``y-siguientes`` is still NOT expanded: its span is open-ended, so
    enumerating it would invent years.

    A PERIOD-QUALIFIED ejercicio counts too: AEAT writes a pago-fraccionado
    design's coverage as ``ejercicio-2p-y-3p-2016``, ``ejercicios-3p-2013-y-2014``
    or a bare ``1p-2016``, and the year-follows-``ejercicio-`` anchor misses all
    three because a period token sits between the word and the digits. This is NOT
    the orden-year anti-pattern relaxed: the year still follows an explicit period
    designation, which is a coverage statement, and the anchor moves from
    ``ejercicio-`` to the period token rather than being dropped.
    """
    matched = _DESIGN_YEAR.search(name)
    if matched is None:
        return _period_qualified_years(name)
    first, connector, second = matched.group(1), matched.group(2), matched.group(3)
    if second is None:
        return (int(first),)
    if connector == "y":
        return (int(first), int(second))
    return tuple(range(int(first), int(second) + 1))


def _period_qualified_years(name: str) -> tuple[int, ...]:
    """Years a period-qualified designation claims, e.g. ``3p-2013-y-2014``."""
    years: set[int] = set()
    for matched in _PERIOD_QUALIFIED_YEAR.finditer(name):
        years.add(int(matched.group(1)))
        if matched.group(2) is not None:
            years.add(int(matched.group(2)))
    return tuple(sorted(years))


@lru_cache(maxsize=512)
def _title_ejercicio_years(path: Path) -> tuple[int, ...]:
    """Every ejercicio the design's own TITLE TEXT states, or empty.

    AEAT prints ``Ejercicio NNNN`` as a title line on a design's first page. That
    is the document asserting which ejercicio it governs, and it is a different
    fact from anything in the filename: measured across the corpus, a filename
    records when AEAT last touched the published page while the title records the
    ejercicio the layout is for, and the two diverge by up to SEVEN years --
    ``03-180-orden-hap-1732-2014-de-24-de-septiembre.pdf`` states ``Ejercicio
    2021``, and ``01-165-diseno-de-registro-actualizado-en-2023.pdf`` states
    ``Ejercicio 2026``.

    WHY THIS IS NOT READ FROM THE PARSED FIELDS. Only eight designs in the whole
    corpus declare an ejercicio as a field constant, so a field-level scan finds
    almost nothing; the assertion lives in the document's heading, above the
    column header the extractor starts at.

    Deliberately reads a BOUNDED prefix -- a PDF's first sixty text lines, a
    workbook's first six rows on its first four sheets -- rather than the whole
    document. Beyond the heading the same phrase appears inside guidance prose
    ("para ejercicios anteriores a 1999"), which is a statement about the DATA a
    filer reports, not about the design's coverage, and admitting it would attribute
    Modelo 193's designs to 1999.
    """
    try:
        lines = _pdf_title_lines(path) if path.suffix.lower() == _PDF_EXTENSION else _workbook_title_lines(path)
    except Exception:
        # An unreadable container attributes nothing, exactly as the sibling parse
        # helper does: a file this cannot open must not be attributed to a year by
        # guesswork, and the coverage guard reports it as unattributable instead.
        return ()
    joined = "\n".join(lines)
    return tuple(sorted({int(matched.group(1)) for matched in _TITLE_EJERCICIO.finditer(joined)}))


def _pdf_title_lines(path: Path) -> list[str]:
    lines = _extract_pdf_text_lines(path.read_bytes(), source_label=str(path))
    return [cleaned for line in lines[:60] if (cleaned := _clean_pdf_line(line))]


def _workbook_title_lines(path: Path) -> list[str]:
    if path.suffix.lower() == _XLS_EXTENSION:
        import xlrd

        book = xlrd.open_workbook(str(path), on_demand=True)
        try:
            return [
                text
                for name in book.sheet_names()[:4]
                for row in range(min(6, book.sheet_by_name(name).nrows))
                if (
                    text := " ".join(
                        str(cell).strip() for cell in book.sheet_by_name(name).row_values(row) if str(cell).strip()
                    )
                )
            ]
        finally:
            book.release_resources()
    from openpyxl import load_workbook

    book = load_workbook(path, read_only=True, data_only=True)
    try:
        return [
            text
            for worksheet in book.worksheets[:4]
            for row in worksheet.iter_rows(max_row=6, values_only=True)
            if (text := " ".join(str(cell).strip() for cell in row if cell is not None and str(cell).strip()))
        ]
    finally:
        book.close()


@lru_cache(maxsize=512)
def _constant_ejercicio_years(path: Path) -> tuple[int, ...]:
    """Every ejercicio a design declares as a FIXED FIELD VALUE, or empty.

    The second place a design states its own coverage: a slot labelled ``Ejercicio``
    whose declared content is a CONSTANT rather than something the filer supplies.
    Modelo 714 writes ``Ejercicio | Constante 2025`` and Modelo 390 writes
    ``2. Devengo - Ejercicio. | Constante "2016"``.

    ``Constante`` is load-bearing, not decoration. Modelo 303 declares ``Ejercicio de
    devengo (EEEE)`` at a four-byte slot with no constant, because that is a value the
    FILER writes; reading it as coverage would attribute every M303 design to whatever
    year happened to appear beside it. Requiring the constant is what separates "the
    design fixes this to 2025" from "the design reserves four bytes for a year".

    Rarer than the title assertion -- eight designs corpus-wide -- but better
    corroborated per instance: seven of the eight also carry a trustworthy filename
    ejercicio and all seven match it.
    """
    years: set[int] = set()
    for sheet in _design_sheets(path):
        for parsed_field in sheet.fields:
            blob = " ".join(
                text for text in (parsed_field.description, parsed_field.content, parsed_field.validation) if text
            )
            for pattern in (_CONSTANT_EJERCICIO, _CONSTANT_EJERCICIO_REVERSED):
                matched = pattern.search(blob)
                if matched is not None:
                    years.add(int(matched.group(1)))
                    break
    return tuple(sorted(years))


def _content_ejercicio_years(path: Path) -> tuple[int, ...]:
    """Every ejercicio the DESIGN ITSELF states, from either content signal."""
    return tuple(sorted(set(_title_ejercicio_years(path)) | set(_constant_ejercicio_years(path))))


@cache
def _catalogue_ejercicio_span() -> dict[str, tuple[int, int]]:
    """Design filename -> the ejercicio span its SOURCE ENTRY declares, where it states one.

    A design that names no ejercicio in its filename or title enters no
    comparison, and that is a hole in boundary detection rather than only a gap
    in a ledger: Modelo 347's ``Orden EHA/3378/2011`` design is cited by its
    revision, carries an authored coverage span in the source catalogue, and is
    still never paired against its neighbour, so the re-layout between them goes
    unreported.

    ADMITTED ONLY FROM A CLOSED, EJERCICIO-ALIGNED SPAN, and both conditions are
    load-bearing. An OPEN span (``applies_to`` unset) cannot be enumerated for
    the same reason ``y-siguientes`` is not expanded -- it would invent every
    year to the end of time. A span whose ends do not fall on 1 January and 31
    December is not describing ejercicios at all: Modelo 180's design runs
    2000-11-28 to 2014-09-26, which are promulgation and repeal dates, and
    Modelo 210's run 2022-06-01 to 2025-12-31, which is a DEVENGO span on an
    axis that is not an ejercicio. Reading either as coverage would repeat the
    update-date-as-governed-period conflation already corrected twice.

    So the rule is checked against the dates themselves rather than against a
    list of modelos, and the designs it admits are exactly those whose catalogue
    entry states a whole number of ejercicios.
    """
    spans: dict[str, tuple[int, int]] = {}
    for source in _authority().catalogues.sources.values():
        if source.kind != "record_design":
            continue
        start, end_date = getattr(source, "applies_from", None), getattr(source, "applies_to", None)
        if start is None or end_date is None:
            continue
        if (start.month, start.day, end_date.month, end_date.day) != (1, 1, 12, 31):
            continue
        spans[str(source.corpus_path).rsplit("/", 1)[-1]] = (start.year, end_date.year)
    return spans


def _design_coverage_years(path: Path) -> tuple[int, ...]:
    """Every ejercicio a design covers, from its content and its filename TOGETHER.

    THE UNION, NOT A PRECEDENCE, and that correction is load-bearing. Ranking the
    content above the filename looks right -- the filename is a publication
    artefact and the title is the document's own assertion -- but the two do not
    describe the same shape of fact. **A design's title names a POINT and its
    filename can name a SPAN.** A design published for ejercicio 2009 and applying
    through 2014 heads its first page ``Ejercicio 2009`` and is filed as
    ``ejercicios-2009-a-2014``; preferring the title there discards 2010 to 2014.

    Measured before this was written that way round: four designs would have lost
    twelve design-years between them, Modelo 130 and Modelo 131 five each. A
    content signal that can only ADD is safe in a way one that can also REPLACE is
    not, and the asymmetry matters here because dropping a year is the QUIET
    failure -- an unattributed design forms no boundary, so the verdict gets
    shorter and reads as progress.

    The union also means the two signals cannot narrow each other by accident,
    which leaves genuine contradiction -- disjoint claims -- as the only thing
    :func:`test_a_design_title_never_contradicts_a_trustworthy_filename_year` has
    to catch, and it is the only thing that check should be asked to judge.

    Returns empty for a design that states no ejercicio anywhere. That is a real
    answer rather than a defect -- Modelo 036 is scoped by an in-force DATE and
    Modelo 210 by a devengo span, neither of which is an ejercicio -- and the
    coverage guard reports such designs as unattributable rather than guessing.
    """
    stated = set(_content_ejercicio_years(path)) | set(_design_years(path.name))
    if stated:
        return tuple(sorted(stated))
    # The document says nothing; fall back to a coverage span its source entry
    # states outright. See _catalogue_ejercicio_span for why only a closed,
    # ejercicio-aligned span qualifies.
    span = _catalogue_ejercicio_span().get(path.name)
    if span is None:
        return ()
    return tuple(range(span[0], span[1] + 1))


#: AEAT's declared period bound for a design that covers only part of an ejercicio.
#: ``hasta-periodo(s)-06`` covers the year UP TO that period, so its coverage STARTS at
#: 01; ``desde-periodo-07`` and ``a-partir-de-periodos-09`` cover FROM that period on, so
#: their coverage starts there. Ordering on the declared coverage start is what puts the
#: two halves of a mid-split ejercicio in publication order.
_COVERAGE_BOUND = re.compile(r"(hasta|a-partir-de|desde)-periodos?-(\d{2})", re.IGNORECASE)


def _coverage_start_period(name: str) -> int | None:
    """The first period a design's own AEAT designation claims, or ``None``.

    WHY THIS IS NOT READ FROM THE FILE BODY, stated because the honest answer is that
    it cannot be. The design workbooks carry NO assertion of the periods they govern:
    measured across Modelo 303's two 2024 halves, both declare the identical sheet set
    and the identical ``Ejercicio de devengo (EEEE)`` and ``Período. (PP)`` header
    fields, because those are slots the FILER fills in, not metadata about the design's
    own coverage. Zero coverage assertions exist inside either file. AEAT's published
    designation is therefore the only place the boundary is stated at all, which is why
    this reads the designation rather than the document.

    WHAT THIS IS NOT. It is not the numeric filename prefix. AEAT numbers its published
    listing NEWEST FIRST -- Modelo 303's ``01`` is the 2026 design and its ``06`` is
    2025 -- so a filename sort is not merely arbitrary, it is close to REVERSED, and
    within one year it silently pairs the LATE half before the EARLY one. Nothing here
    depends on that numbering, so a change to AEAT's listing convention cannot move it.

    Returns ``None`` when the designation declares no period bound, which is a real
    case rather than a defect: Modelo 303's two 2018 designs are ``ejercicio-2018`` and
    ``ejercicio-2018-salvo-ultimo-periodo-12m-4t``, and "except the last period" fixes
    no start while the unqualified sibling asserts nothing at all. Ordering those two
    would be inference by elimination dressed as a measurement, so the caller is told
    the year is unorderable instead.
    """
    matched = _COVERAGE_BOUND.search(name)
    if matched is None:
        return None
    kind, period = matched.group(1).lower(), int(matched.group(2))
    return 1 if kind == "hasta" else period


def _design_fingerprint(path: Path) -> tuple[object, ...]:
    """Format-independent identity of a design: what it DECLARES, not how it is packaged.

    A raw byte hash is the obvious identity and it is the wrong one. AEAT bundles the
    same design twice in two container formats, and an ``.xls`` and an ``.xlsx`` of one
    document are different bytes while being the same design. Measured: Modelo 200
    carries such a twin for every ejercicio from 2015 to 2025, and a byte-keyed
    deduplication reported all ELEVEN of those years as carrying two designs -- which
    reads as eleven mid-course splits on a modelo that has none. Fingerprinting the
    parsed declaration collapses the twins and leaves genuine splits standing: on
    Modelo 303 the 2018, 2021 and 2024 pairs survive, because those really are two
    different designs.

    Byte-keying also has to stay dead rather than merely unused, because it looks
    correct: it DOES collapse the corpus's other duplicate shape, the same file bundled
    twice under a truncated filename, so a reader checking it against that case
    concludes it works.
    """
    return tuple(
        (
            sheet.name,
            sheet.total_positions,
            tuple(
                (field.offset, field.length, field.type_code, " ".join(field.description.split()))
                for field in sheet.fields
            ),
        )
        for sheet in _design_sheets(path)
    )


def _designs_in_publication_order(modelo_id: str) -> tuple[tuple[Path, ...], tuple[int, ...]]:
    """``(designs oldest-first, years whose designs could not be ordered)``.

    Deduplicated by CONTENT, because the corpus bundles several designs twice under
    names differing only by a truncated extension, and a path-keyed pass reports the
    duplicate as a second design.

    A year holding two designs where at least one declares no coverage bound is
    reported in the second element and its members are left in a stable but
    UNASSERTED order, so a consumer can refuse rather than trust it.
    """
    by_year: dict[int, list[Path]] = {}
    seen: set[object] = set()
    for path in _design_sources(modelo_id):
        if not _design_sheets(path):
            continue
        years = _design_coverage_years(path)
        if not years:
            continue
        marker = _design_fingerprint(path)
        if marker in seen:
            continue
        seen.add(marker)
        by_year.setdefault(min(years), []).append(path)

    ordered: list[Path] = []
    unorderable: list[int] = []
    for year in sorted(by_year):
        members = by_year[year]
        starts = [_coverage_start_period(path.name) for path in members]
        if len(members) > 1 and any(start is None for start in starts):
            unorderable.append(year)
            ordered.extend(sorted(members, key=lambda path: path.name))
            continue
        ordered.extend(sorted(members, key=lambda path: (_coverage_start_period(path.name) or 1, path.name)))
    return tuple(ordered), tuple(unorderable)


def _designs_by_year(modelo_id: str) -> dict[int, tuple[Path, ...]]:
    """``{year: every readable design source claiming it}`` -- one to MANY.

    AEAT splits an ejercicio mid-course, so a year can have two designs that differ
    in byte layout. Modelo 303 does it three times: 2018, 2021 and 2024 each ship a
    "hasta periodo N" design and a "desde/a partir de periodo N" one. The year-keyed
    maps below keep the FIRST by filename sort and silently discard the other, so a
    caller asking what a year looked like gets an arbitrary half with no signal a
    second exists. This map keeps both, so a consumer deriving epoch boundaries can
    see the split rather than land a boundary on the wrong side of it.
    """
    grouped: dict[int, list[Path]] = {}
    for path in _design_sources(modelo_id):
        if not _design_sheets(path):
            continue
        for year in _design_coverage_years(path):
            grouped.setdefault(year, []).append(path)
    return {year: tuple(paths) for year, paths in sorted(grouped.items())}


def _design_sources(modelo_id: str) -> list[Path]:
    """Every bundled design SOURCE for one modelo, deterministically ordered.

    Walks the modelo directory RECURSIVELY rather than globbing ``files/*``. Almost
    every modelo keeps its designs in a ``files`` subdirectory, but Modelo 210
    keeps ``dr210_2011.pdf`` directly in the modelo directory, and the fixed-depth
    glob excluded it from every map in this module. Nothing decided that; it is a
    directory-shape assumption that happened to hold for all but one modelo, which
    is the shape of assumption that survives longest unnoticed.
    """
    return sorted(
        (
            path
            for path in scan_directory(_design_dir(modelo_id), recursive=True, select=DirectoryEntryKind.FILES)
            if path.suffix.lower() in _DESIGN_SUFFIXES
        ),
        key=lambda path: (path.name, path.suffix.lower()),
    )


def _design_sheets(path: Path) -> tuple[RecordDesignSheet, ...]:
    """Parse one design SOURCE, dispatching on its suffix.

    The sources are read directly rather than their ``.extracted.md``
    derivatives. The markdown extraction of a PDF drops the offset column and
    the page-total rows, so every PDF-sourced design read that way reports as
    unreadable while the shipped parser reads the same file correctly -- 21 of
    the 24 designs this module called unmeasured were readable all along, six
    of them consecutive Modelo 100 years. Reading the source removes the whole
    class rather than special-casing it.

    Returns an empty tuple for a genuinely unparseable file, so the caller's
    unreadable-reporting path still names it. A parse that fails must stay
    visible as unmeasured; only a parse that never ran should disappear.
    """
    parsers = {
        ".xlsx": extract_record_design_workbook,
        ".xlsm": extract_record_design_workbook,
        ".xls": extract_record_design_xls_workbook,
        ".pdf": extract_record_design_pdf,
    }
    parser = parsers.get(path.suffix.lower())
    if parser is None:
        return ()
    try:
        # ACCEPTS a partial read deliberately. This module compares designs against
        # each other, and a design read in part still carries real evidence about
        # the sheets it did read; refusing it here would replace a comparison that
        # sees most of a boundary with one that sees none of it. The completeness
        # of each read is reported by the coverage guard rather than resolved here.
        return parser(path).accept_partial()
    except Exception:
        return ()


def _unparseable_design_sources(modelo_id: str) -> tuple[Path, ...]:
    """Every bundled design file for one modelo whose sheets fail to parse AT ALL.

    A THIRD SHAPE, distinct from the two this instrument otherwise reports. A
    design that parses (even partially, per :func:`_design_sheets`'s own
    ``accept_partial`` posture) and states no ejercicio anywhere is a real,
    legitimate answer -- some modelos are scoped by something other than an
    ejercicio, and the coverage guard reports that honestly. A year with NO
    bundled file at all is a genuine acquisition gap -- fetching one from AEAT
    is the only fix. A file in THIS set is neither: it physically exists in the
    corpus, so acquisition does nothing, but the parser returns nothing usable
    for it AT ALL, so whatever ejercicio it might state -- filename or content
    -- never reaches :func:`_design_coverage_years`. That is an extraction-layer
    defect, and conflating it with either of the other two shapes sends a
    fix-owner to acquire a design that is already bundled, or to treat a
    legitimately non-ejercicio design as if something were missing from it.

    Measured live: Modelo 036 carries three such files
    (``...-ejercicio-2023-y-siguientes...``, two ``...-ejercicio-2021-y-siguientes...``),
    each with a real filename-derivable ejercicio year that never reaches any
    comparison because :func:`_designs_in_publication_order`'s own first filter
    drops an unparseable design before year-attribution is ever consulted --
    loud at the extraction layer (the parser genuinely fails), but that loudness
    does not reach this module, because the design is dropped before it can be
    reported.

    PROSPECTIVE, NOT SPECULATIVE -- stated because the distinction matters for
    how dormant code is read. As of this writing no revision's declared span
    currently overlaps Modelo 036's three unparseable years, so the UNPARSEABLE
    branch in the two gate failure messages that consult this function is
    proven by construction (bite-proofed synthetically, see the mutation tests
    in this module) but currently unexercised by the live corpus -- correct and
    dormant, not correct and moot. The three unparseable files still exist and
    a future span change (Modelo 036's own, or a currently-absent modelo
    gaining a declared revision over years an unparseable file already claims)
    brings this branch back into range without anyone touching it. A dormant
    branch is still a rot risk: if this signal's wording or matching logic ever
    drifts silently, nothing in today's corpus would catch it until a real case
    arrives. Re-run the synthetic bite proof, not just the live gate, when this
    function changes.
    """
    return tuple(path for path in _design_sources(modelo_id) if not _design_sheets(path))


def _parse_extracted(path: Path) -> dict[str, int]:
    """box number -> offset from a design's ``.extracted.md`` derivative.

    Retained as a FALLBACK, not as the primary read. Measured: switching to the
    sources alone closed 21 PDF-sourced designs but OPENED new blind spots on
    `.xls` files whose markdown parses while the binary does not, so the
    unreadable count moved 24 -> 21 rather than 24 -> 3. Source-first with a
    derivative fallback is strictly better than either alone -- the source
    yields offsets and occupancy the markdown cannot, and the markdown covers
    the binaries the parsers refuse.
    """
    derivative = path.with_name(path.name + ".extracted.md")
    if not derivative.is_file():
        return {}
    table: dict[str, int] = {}
    for line in derivative.read_text(encoding="utf-8", errors="replace").splitlines():
        if "|" not in line:
            continue
        columns = [column.strip() for column in line.split("|")]
        if len(columns) < 5:
            continue
        try:
            offset = int(columns[1])
        except (IndexError, ValueError):
            continue
        boxes = _BOX_MARKER.findall(line)
        if boxes:
            table.setdefault(boxes[-1], offset)
    return table


def _parse_design(path: Path) -> dict[str, int]:
    """box number -> record offset for one design, source first, derivative second."""
    table: dict[str, int] = {}
    for sheet in _design_sheets(path):
        for field in sheet.fields:
            offset = field.offset
            boxes = _BOX_MARKER.findall(field.description)
            if not boxes:
                continue
            # A field's OWN number is the LAST bracket on its row. Formula totals cite
            # their operands first -- "Total cuota devengada ([152]+[167]+...) [27]"
            # -- so first-match keying would file the total under an operand, and
            # requiring a single bracket would drop it entirely. Modelo 303 has
            # eleven such rows, including its three most load-bearing totals; Modelo
            # 390 has none, which is why this only surfaced on the second modelo.
            #
            # A box can also appear once per régimen segment at the same offset; keep
            # the first and never overwrite, so a later segment cannot mask a move.
            #
            # Modelo 390 proves this is not hypothetical: box [114] carries one
            # `Código CNAE` per prorrata row -- once in the 2016 design and FIVE
            # times in 2017. A last-wins map compares 2016's first row against
            # 2017's fifth and reports a +332 "move" that never happened.
            table.setdefault(boxes[-1], offset)
    return table or _parse_extracted(path)


def _page_lengths(path: Path) -> tuple[str, ...]:
    """The per-sheet declared record length, source first, derivative second.

    ABSTAINS on a flattened PDF parse. The PDF backend returns ONE synthetic
    sheet covering the whole document, so its "page lengths" are a one-element
    tuple describing no page at all. Compared against a spreadsheet's nine real
    per-page totals it differs unconditionally, which reports a re-layout
    between every PDF-sourced year and its neighbour -- a boundary manufactured
    by the parser shape rather than by AEAT.

    Measured: with the flattened tuple emitted, Modelo 390 gained a false
    2015->2016 boundary, while a box-keyed comparison of the same two designs
    finds 345 shared boxes and ZERO moved. Returning nothing lets the box signal
    carry those years and keeps this one honest about what it cannot see.
    """
    sheets = _design_sheets(path)
    if len(sheets) == 1 and sheets[0].name == _PDF_FLATTENED_SHEET:
        return ()
    lengths = tuple("variable" if sheet.total_positions is None else str(sheet.total_positions) for sheet in sheets)
    if lengths:
        return lengths
    derivative = path.with_name(path.name + ".extracted.md")
    if not derivative.is_file():
        return ()
    return tuple(str(total) for total in _PAGE_TOTAL.findall(derivative.read_text(encoding="utf-8", errors="replace")))


def _occupancy(path: Path) -> dict[tuple[str, int], bool]:
    """``(sheet, offset) -> is_reserved`` for every field in one design.

    The reserved flag is carried as a VALUE rather than applied as a filter, and
    that is the whole point. Excluding reserved rows from a comparison makes a
    field RETIRED INTO reserved space invisible: it is real on one side and
    reserved on the other, so the exclusion drops one side and the pair is never
    compared. Measured on Modelo 390, three `Reg. Simplificado - Reducción
    aplicable` slots were retired between the 2024 and 2025 designs at offsets
    223, 543 and 1205, and a reserved-excluding diff reported the two years
    identical.

    A filter that drops a category cannot see movement into or out of that
    category. Keying on the position and classifying afterwards makes the
    transition a first-class result.
    """
    return {
        (sheet.name, field.offset): _RESERVED_FIELD.search(field.description) is not None
        for sheet in _design_sheets(path)
        for field in sheet.fields
    }


def _page_lengths_for(modelo_id: str) -> dict[int, tuple[str, ...]]:
    """``{design year: page-length sequence}`` for every readable design."""
    lengths: dict[int, tuple[str, ...]] = {}
    for path in _design_sources(modelo_id):
        found = _page_lengths(path)
        if not found:
            continue
        for year in _design_coverage_years(path):
            lengths.setdefault(year, found)
    return lengths


def _designs_for(modelo_id: str) -> tuple[dict[int, dict[str, int]], dict[int, str]]:
    """Return ``{design year: {box: offset}}`` and ``{year: filename}`` for unreadable ones.

    A design whose year cannot be read off the filename is excluded from both
    maps: it can neither be compared nor attributed to a revision's span, so
    counting it either way would be a guess.
    """
    parsed: dict[int, dict[str, int]] = {}
    unreadable: dict[int, str] = {}
    for path in _design_sources(modelo_id):
        table = _parse_design(path)
        for year in _design_coverage_years(path):
            if table:
                parsed.setdefault(year, table)
            elif year not in parsed:
                unreadable.setdefault(year, path.name)
    return parsed, unreadable


def _span_years(revision: ModeloRevision) -> set[int]:
    """Every filing year the revision's period selector claims."""
    selector = revision.period_selector
    if selector.years:
        return set(selector.years)
    if selector.year_from is None:
        return set()
    upper = selector.year_to
    if upper is None:
        # Open-ended: the gate can only speak about years the corpus covers, so
        # the span is bounded by the newest bundled design rather than by a
        # literal ceiling that would go stale.
        return {selector.year_from}
    return set(range(selector.year_from, upper + 1))


def _filing_revisions() -> list[tuple[ModeloDefinition, str, ModeloRevision]]:
    """Every revision that claims filing support, including incomplete claims.

    Filing grade is the authority boundary. Filtering on populated layouts would
    hide the exact genuine gap this suite must expose: a filing-supported revision
    whose official export evidence or authored layout is still missing.
    """
    return [
        (modelo, revision_id, revision)
        for modelo in _authority().modelos
        for revision_id, revision in modelo.revisions.items()
        if revision.effective_authority_grade is RegistryAuthorityGrade.FILING
    ]


def _declared_revisions() -> list[tuple[ModeloDefinition, str, ModeloRevision]]:
    """Every revision, for detector anti-vacuity controls outside support policy."""
    return [
        (modelo, revision_id, revision)
        for modelo in _authority().modelos
        for revision_id, revision in modelo.revisions.items()
    ]


def _filing_supported_revisions() -> list[tuple[ModeloDefinition, str, ModeloRevision]]:
    """Every revision that explicitly claims filing support.

    Kept as the raw-loader counterpart to :func:`_filing_revisions`: operational
    signal gates use validated authority, while the coverage-completeness gate
    must still run and report the exact structural gap if unrelated validation is
    red elsewhere in the tree.

    Reads through the raw loader (:func:`load_registry_tree`), never
    :func:`_authority`: the coverage-completeness question is structural
    (declared revisions, period selectors, bundled corpus) and needs no
    business-rule or filing-grade validation to answer. Coupling it to
    :class:`ValidatedRegistryAuthority` would make this gate's own
    reliability hostage to an unrelated validation defect anywhere else in
    the tree -- the exact silent-failure shape this gate exists to
    remove, now reproduced as "the coverage gate didn't even run" instead of
    "the coverage gate reported clean".
    """
    modelos, _catalogues = bundled_registry_tree()
    return [
        (modelo, revision_id, revision)
        for modelo in modelos
        for revision_id, revision in modelo.revisions.items()
        if revision.effective_authority_grade is RegistryAuthorityGrade.FILING
    ]


def _claimed_years(revision: ModeloRevision, design_years: set[int]) -> set[int]:
    """Design years the revision claims, honouring an open-ended upper bound."""
    selector = revision.period_selector
    explicit = _span_years(revision)
    if selector.years or selector.year_from is None:
        return explicit & design_years
    if selector.year_to is None:
        return {year for year in design_years if year >= selector.year_from}
    return explicit & design_years


@cache
def _source_reference_by_id() -> dict[str, SourceReference]:
    """Loaded source catalogue, used as the revision's dependency receipts."""
    _modelos, catalogues = bundled_registry_tree()
    return {str(ref): source for ref, source in catalogues.sources.items()}


def _layout_authority_receipts(modelo_id: str, revision: ModeloRevision) -> tuple[SourceReference, ...]:
    """Layout-authority sources explicitly cited by a revision.

    Fixed-width modelos normally cite ``record_design`` sources.  Modelo 100
    is filed from AEAT's published dictionaries and XSD instead; those sources
    carry the same ``layout_authority`` tier and must not be mistaken for an
    absent design merely because their transport grammar is not a workbook.
    """
    sources = _source_reference_by_id()
    return tuple(
        source
        for ref in revision.source_refs
        if (source := sources.get(str(ref))) is not None
        and source.evidence_tier == "layout_authority"
        and f"/modelo_{modelo_id}/" in f"/{str(source.corpus_path).replace('\\', '/')}"
    )


def _receipt_covers_year(source: SourceReference, year: int) -> bool:
    selector = source.period_selector
    if selector is not None:
        return selector.includes_year(year)
    if source.applies_from is None or source.applies_from.year > year:
        return False
    return source.applies_to is None or source.applies_to.year >= year


def _source_epoch_proves_revision_span(modelo_id: str, revision: ModeloRevision) -> tuple[bool, str]:
    """Whether cited record-design receipts cover the revision's declared year span.

    An open revision needs an open receipt; a closed revision needs every claimed
    year covered. This consumes authored source authority and never manufactures
    extra annual designs merely to reach a count. Corpus-detected relayouts are
    checked first by the caller and therefore always override this positive proof.
    """
    receipts = _layout_authority_receipts(modelo_id, revision)
    if not receipts:
        return False, "revision cites no layout-authority source receipt"
    selector = revision.period_selector
    if selector.years:
        years = set(selector.years)
    elif selector.year_from is not None and selector.year_to is not None:
        years = set(range(selector.year_from, selector.year_to + 1))
    elif selector.year_from is not None:
        open_receipts = tuple(
            source
            for source in receipts
            if source.applies_to is None
            and (
                (
                    source.period_selector is not None
                    and source.period_selector.year_from is not None
                    and source.period_selector.year_from <= selector.year_from
                    and source.period_selector.year_to is None
                )
                or (
                    source.period_selector is None
                    and source.applies_from is not None
                    and source.applies_from.year <= selector.year_from
                )
            )
        )
        if open_receipts:
            return True, f"open source epoch(s) {[source.id for source in open_receipts]!r}"
        return False, "open revision has no cited open-ended layout-authority receipt"
    else:
        return False, "revision declares no filing-year span"
    missing = sorted(year for year in years if not any(_receipt_covers_year(source, year) for source in receipts))
    if missing:
        return False, f"cited layout-authority receipts do not cover filing year(s) {missing!r}"
    return True, f"bounded source epoch receipt(s) {[source.id for source in receipts]!r}"
