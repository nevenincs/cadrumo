"""Modelo 390 declares the apartado 10 volumen de operaciones exempt-supply boxes.

The annual resumen modelled seven IVA categories where the quarterly Modelo 303
models ten, and the shortfall was not a binding defect: boxes [103] and [104]
had no casilla at all, so no selector could reach them. A binding cannot select
a category with no target casilla.

What this module pins is threefold. That each casilla's declared official number
is the number AEAT's own record design prints against that description, read
from the bundled Diseno de Registros workbooks rather than restated here. That
the population each box draws is the one the law puts in it -- and, for [103],
that the box's own widened title is not authority to widen the selector. And
that neither box reaches any total, which is what keeps them disclosure
restatements rather than addends.

External authority: the box identities are asserted against the bundled AEAT
workbooks under ``corpus/aeat_official/disenos_registro/modelo_390/files``,
parsed from the source ``.xlsx`` -- not from the ``.extracted`` siblings, which
are the same extraction pass and so are not an independent control. The official
number is read from the ``[NN]`` token inside the field DESCRIPTION; the
workbook's first column is a field sequence index and its second the byte
offset, so a bare match on either reads as a box number while being neither.

Non-tautology: no expected amount is derived from a registry formula. The
identity tests would fail if the registry declared a number AEAT does not print
against that concept; the population tests would fail if a selector admitted a
category the law locates elsewhere; the disjointness test would fail if a later
change wired either box into a total. None of them can be satisfied by
re-deriving a value from the code under test.

Real-behaviour: the committed revision through the real registry authority,
resolved by the real ``ledger_iva_aggregation`` resolver over real
:class:`IvaLedgerObservation` rows. No mocks, stubs, skips or xfail.
"""

from __future__ import annotations

import re
from datetime import date
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook

from .....core.directory_scan import scan_directory
from .....core.resources import resources
from ....iva import IvaCategory, IvaFlowDirection, IvaLedgerObservationRole, IvaRateKind
from .. import (
    IvaLedgerObservation,
    ModeloRevision,
    expression_casilla_refs,
    resolve_ledger_iva_aggregation_binding_values,
)
from .._ledger_bindings import iva_ledger_selector

pytestmark = [pytest.mark.unit, pytest.mark.hex_domain]


_DISENO_DIR = (
    Path(__file__).resolve().parents[4]
    / "_data"
    / "corpus"
    / "aeat_official"
    / "disenos_registro"
    / "modelo_390"
    / "files"
)

# Each entry: casilla id, its declared official box number, and a phrase that
# must appear in the official field description AEAT prints against that number.
# The phrases are deliberately concept words rather than whole titles, because
# box [103]'s printed title changed in the ejercicio-2021 design while its
# identity did not.
#
# Keep every phrase ASCII. The official descriptions carry accented Spanish, and
# a phrase whose accent does not survive the round trip from source to comparand
# matches nothing while looking like a clean absence -- a bare zero that reads as
# "this box is not in the design" when it means "this literal was mistyped".
# Choose a phrase that discriminates without needing an accented character, as
# both entries below do.
_BOXES = (
    ("iva.anual.volumen.entregas-intracomunitarias", "103", "Entregas intracomunitarias"),
    ("iva.anual.volumen.exportaciones-exentas", "104", "Exportaciones"),
)

_ENTREGAS_BASE = Decimal("18400.00")
_EXPORT_THIRD_COUNTRY_BASE = Decimal("9250.00")
_EXPORT_ASSIMILATED_BASE = Decimal("1375.00")
_SERVICE_SUPPLY_BASE = Decimal("6100.00")


def _m390_revision() -> ModeloRevision:
    return resources().modelos.authority.snapshot("390", filing_year=2024, period="0A").revision


def _official_fields(path: Path) -> tuple[tuple[str, str], ...]:
    """Return ``(box_number, description)`` for every numbered field in one design.

    The box number is the ``[NN]`` token inside the description cell. A field
    without one is not a numbered box and is dropped.
    """
    rows: list[tuple[str, str]] = []
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                cells = [str(cell).strip() for cell in row if cell is not None]
                for cell in cells:
                    for number in re.findall(r"\[([A-Z0-9]{1,4})\]", cell):
                        rows.append((number, cell))
    finally:
        workbook.close()
    return tuple(rows)


def _designs() -> tuple[Path, ...]:
    return scan_directory(_DISENO_DIR, pattern="*.xlsx")


def _observation(
    *,
    category: IvaCategory,
    base: Decimal,
) -> IvaLedgerObservation:
    """One exempt repercutido row: a zero-rated supply carrying base but no cuota."""
    return IvaLedgerObservation(
        ledger_id="ledger-m390-volumen",
        transaction_date=date(2024, 5, 20),
        category=category,
        exemption_article=None,
        rate_kind=IvaRateKind.ZERO,
        flow_direction=IvaFlowDirection.REPERCUTIDO,
        base_amount=base,
        iva_amount=Decimal("0"),
        recargo_amount=Decimal("0"),
        applied_rate=Decimal("0.00"),
        observation_role=IvaLedgerObservationRole.SETTLEMENT,
    )


def _rows() -> tuple[IvaLedgerObservation, ...]:
    """Four distinct exempt bases, no two equal and no two summing to a third.

    Distinctness is what makes a mis-routed category visible on value: a
    selector that leaked a neighbour or swapped two boxes cannot land on the
    expected figure by coincidence.
    """
    return (
        _observation(category=IvaCategory.INTRA_COMMUNITY_SUPPLY, base=_ENTREGAS_BASE),
        _observation(category=IvaCategory.EXPORT_THIRD_COUNTRY_ZERO_RATED, base=_EXPORT_THIRD_COUNTRY_BASE),
        _observation(category=IvaCategory.EXPORT_ASSIMILATED_ZERO_RATED, base=_EXPORT_ASSIMILATED_BASE),
        _observation(category=IvaCategory.INTRA_COMMUNITY_SERVICE_SUPPLY, base=_SERVICE_SUPPLY_BASE),
    )


def _resolve() -> dict[str, Decimal]:
    return dict(resolve_ledger_iva_aggregation_binding_values(_m390_revision(), _rows()))


def test_the_volumen_casillas_are_declared() -> None:
    """Both boxes now have a target casilla; before this layer neither existed."""
    declared = {casilla.id: casilla for casilla in _m390_revision().casillas}
    for casilla_id, number, _phrase in _BOXES:
        assert casilla_id in declared, f"{casilla_id} is not declared on the revision"
        assert declared[casilla_id].number == number


@pytest.mark.parametrize(("casilla_id", "number", "phrase"), _BOXES)
def test_each_declared_number_is_the_number_aeat_prints(casilla_id: str, number: str, phrase: str) -> None:
    """The official design prints this box number against this concept.

    Read from AEAT's own workbook, so a registry number invented or transposed
    from a neighbouring box fails here rather than reaching a filed artefact.
    """
    designs = _designs()
    assert designs, "no bundled Modelo 390 record designs found"
    matched = 0
    for design in designs:
        for found_number, description in _official_fields(design):
            if found_number == number and phrase.casefold() in description.casefold():
                matched += 1
                break
    assert matched == len(designs), (
        f"box [{number}] carrying {phrase!r} was found in {matched} of {len(designs)} bundled designs; "
        f"casilla {casilla_id} declares a number whose identity is not stable across this revision's span"
    )


def test_no_other_concept_claims_these_numbers_in_any_design() -> None:
    """Neither number is reused by a second concept inside any one design.

    A box number naming two concepts within a single design makes the declared
    number ambiguous even before the revision's span is considered, and it is
    not hypothetical here: the current design has nine numbers that each name
    more than one distinct concept. For those, any lookup collapsing a design
    to a number-keyed mapping silently keeps one occurrence and discards the
    rest, so a comparison across two designs can end up matching unrelated
    fields while reporting a clean result.

    This is the control for the identity test above: that test can only pass
    because these two numbers are unique in every design, not because its match
    was loose. Checking every design rather than only the newest is what makes
    the control cover the whole span the identity claim is made over.
    """
    for design in _designs():
        rows = _official_fields(design)
        for _casilla_id, number, phrase in _BOXES:
            descriptions = {description for found_number, description in rows if found_number == number}
            assert len(descriptions) == 1, (
                f"box [{number}] names {len(descriptions)} distinct concepts in {design.name}: {descriptions}"
            )
            assert phrase.casefold() in next(iter(descriptions)).casefold()


def test_entregas_intracomunitarias_draws_the_exempt_supply_base() -> None:
    """Box [103] resolves to the art. 25 entrega intracomunitaria base."""
    assert _resolve()["modelo-390-volumen-entregas-intracomunitarias-base"] == _ENTREGAS_BASE


def test_exportaciones_draws_both_zero_rated_export_members() -> None:
    """Box [104] covers art. 21 exports AND art. 22 assimilated operations."""
    assert _resolve()["modelo-390-volumen-exportaciones-exentas-base"] == (
        _EXPORT_THIRD_COUNTRY_BASE + _EXPORT_ASSIMILATED_BASE
    )


def test_a_b2b_eu_service_supply_never_reaches_the_intracomunitarias_box() -> None:
    """The widened box title is not authority to widen the selector.

    The ejercicio-2021 design retitled [103] "Entregas intracomunitarias de
    bienes y servicios", which reads as an invitation to admit a services
    category. A plain B2B service supplied to an EU-established taxable person
    is located outside the TAI by LIVA art. 69.Uno.1, so it is not subject in
    Spain and belongs in the no-sujetas box. Admitting it here would over-declare
    the exempt-supply figure by the service base.
    """
    resolved = _resolve()
    assert resolved["modelo-390-volumen-entregas-intracomunitarias-base"] == _ENTREGAS_BASE
    assert resolved["modelo-390-volumen-entregas-intracomunitarias-base"] != (_ENTREGAS_BASE + _SERVICE_SUPPLY_BASE)


def test_neither_volumen_box_feeds_any_total() -> None:
    """Volumen de operaciones is a restatement, never an addend.

    These boxes disclose turnover; they must not enter the liquidacion. The
    guard matters because box [104] and the Reg. ordinario 0 % rate box can
    legitimately see the same exempt-export rows -- that is only safe while no
    total takes both, so a later change wiring either into a formula would
    create a double-count that files clean.
    """
    revision = _m390_revision()
    volumen_ids = {casilla_id for casilla_id, _number, _phrase in _BOXES}
    seen_operands: set[str] = set()
    for formula in revision.formulas:
        operands = set(expression_casilla_refs(formula.expression))
        seen_operands |= operands
        assert not (operands & volumen_ids), (
            f"formula {formula.id!r} takes volumen casilla(s) {sorted(operands & volumen_ids)} as an operand"
        )
        assert formula.target_casilla_id not in volumen_ids
    # Anti-vacuity: the assertions above are satisfied by an operand set that is
    # merely EMPTY, so an extraction that silently returns nothing would pass
    # them while checking nothing. An earlier revision of this test did exactly
    # that -- it scraped operands from repr() with a pattern matching a mapping
    # form the loaded schema does not use -- so it could never fail. Pin that the
    # walk actually reached the graph, and reached a known operand.
    assert len(seen_operands) >= len(revision.formulas), (
        f"operand walk returned {len(seen_operands)} refs across {len(revision.formulas)} formulas; "
        "the extraction is not reaching the expression graph"
    )
    assert "iva.anual.repercutido.general" in seen_operands


def test_the_volumen_boxes_select_what_the_quarterly_return_selects() -> None:
    """M390 [103]/[104] draw the same categories as M303 casillas 59/60.

    The annual return restates the summed quarters, so a category admitted on
    one side and not the other desynchronises the two returns for the same
    taxpayer and the same year.
    """
    annual = {binding.id: binding for binding in _m390_revision().bindings}
    quarterly_revision = resources().modelos.authority.snapshot("303", filing_year=2024, period="4T").revision
    quarterly = {binding.id: binding for binding in quarterly_revision.bindings}
    pairs = (
        ("modelo-390-volumen-entregas-intracomunitarias-base", "modelo-303-casilla-59-entregas-intracomunitarias-base"),
        ("modelo-390-volumen-exportaciones-exentas-base", "modelo-303-casilla-60-exportaciones-base"),
    )
    for annual_id, quarterly_id in pairs:
        annual_selector = iva_ledger_selector(annual[annual_id])
        quarterly_selector = iva_ledger_selector(quarterly[quarterly_id])
        assert set(annual_selector.categories) == set(quarterly_selector.categories)
