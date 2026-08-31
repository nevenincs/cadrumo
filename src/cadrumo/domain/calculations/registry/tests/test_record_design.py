"""Tests for read-only AEAT record-design PDF extraction."""

from __future__ import annotations

import inspect
from pathlib import Path

import pytest

from .. import record_design_workbook as record_design_workbook_module
from ..corpus_catalogue import resolve_record_design_binary
from ..errors import RegistryValidationError
from ..record_design import (
    extract_record_design,
)
from ..record_design_schema import (
    RecordDesignCompositeRelativeClosing,
    RecordDesignRelativeSuffixMarker,
)
from ._record_design_support import (
    _committed_registry_tree,
    bundled_path,
)

pytestmark = [pytest.mark.unit, pytest.mark.hex_domain]

_WorkbookCell = str | int | None
_WorkbookRow = tuple[_WorkbookCell, ...]

_MODELO_303_DESIGNS = (
    ("aeat-dr-303-2023", 406),
    ("aeat-dr-303-2024-early", 406),
    ("aeat-dr-303-2024-late", 426),
    ("aeat-dr-303-2025", 429),
    ("aeat-dr-303-2026", 430),
)
_MODELO_220_DESIGNS = (
    ("aeat-dr-220-2023", 2023, "2023"),
    ("aeat-dr-220-2024", 2024, "2024"),
    ("aeat-dr-220-2025", 2025, "2025"),
)
_M220_COMPOSITE_CLOSING_ROWS: tuple[_WorkbookRow, ...] = (
    (15, "***", 3, "An", "Constante", None, "</T"),
    (16, "***", 3, "An", "Modelo", None, "220"),
    (17, "***", 1, "An", "Discriminente", None, "(*)[A|E|I|0]"),
    (18, "***", 4, "An", "Ejercicio de devengo", None, None),
    (19, "***", 2, "An", "Período", None, "0A"),
    (20, "***", 5, "An", "Constante", None, "0000>"),
)


def test_modelo_200_workbook_recovers_source_declared_totals_and_variable_envelope() -> None:
    """The pinned binary, rather than terminal inference, owns totals and composition."""
    _modelo, catalogues = _committed_registry_tree()
    source = catalogues.sources["aeat-dr-200-2025"]
    workbook_path = bundled_path() / source.corpus_path

    # The formula half of this evidence lives only in the .xlsx rendering of the
    # same AEAT workbook. xlrd reads the .xls the catalogue pins and returns the
    # CACHED total -- 627.0 -- but no formula text, so "=SUM(C6:C118)" is
    # unavailable from it at all. Reading the sibling keeps the assertion that
    # the total is a sum over the record rows rather than a number that happens
    # to sit beside them, and the parse below still runs against the pinned
    # binary, so the two formats corroborate rather than one replacing the other.
    formula_source = workbook_path.with_suffix(".xlsx")
    declared_totals, formula_anchors = _official_total_rows(formula_source)

    assert declared_totals
    assert formula_anchors["DP200001"] == ("A119", "C119", "=SUM(C6:C118)", 627)
    assert formula_anchors["DP200DID"] == ("A49", "C49", "=SUM(C6:C48)", 774)

    parsed = {sheet.name: sheet for sheet in extract_record_design(workbook_path).accept_partial()}
    assert set(declared_totals) == {name for name, sheet in parsed.items() if sheet.total_positions is not None}
    for name, declared_total in declared_totals.items():
        sheet = parsed[name]
        terminal_extent = max(field.offset + field.length - 1 for field in sheet.fields)
        assert sheet.total_positions == declared_total == terminal_extent

    envelope_sheet = parsed["DP200000"]
    envelope = envelope_sheet.variable_envelope
    assert envelope_sheet.total_positions is None
    assert envelope is not None
    assert isinstance(envelope.closing, RecordDesignRelativeSuffixMarker)
    assert envelope.prefix_extent == 328
    assert max(field.offset + field.length - 1 for field in envelope.prefix_fields) == 328
    assert (envelope.body.row, envelope.body.offset, envelope.body.length) == (14, 329, "Variable")
    assert (
        envelope.closing.row,
        envelope.closing.offset,
        envelope.closing.length,
    ) == (15, "***", 18)
    assert (
        envelope.variable_total.row,
        envelope.variable_total.label,
        envelope.variable_total.length,
    ) == (16, "total", "Variable")


@pytest.mark.parametrize(("source_ref", "expected_field_count"), _MODELO_303_DESIGNS)
def test_modelo_303_workbooks_recognise_the_official_variable_envelope_shape(
    source_ref: str,
    expected_field_count: int,
) -> None:
    """Every pinned M303 binary owns the same explicit DP30300 composition."""
    _modelo, catalogues = _committed_registry_tree()
    source = catalogues.sources[source_ref]

    parsed = extract_record_design(bundled_path() / source.corpus_path).accept_partial()
    fixed = tuple(sheet for sheet in parsed if sheet.variable_envelope is None)
    envelopes = tuple(sheet.variable_envelope for sheet in parsed if sheet.variable_envelope is not None)

    assert len(parsed) == 7
    assert len(fixed) == 6
    assert sum(len(sheet.fields) for sheet in parsed) == expected_field_count
    assert len(envelopes) == 1
    envelope = envelopes[0]
    assert envelope is not None
    assert isinstance(envelope.closing, RecordDesignRelativeSuffixMarker)
    assert envelope.name == "DP30300"
    assert envelope.prefix_extent == 328
    assert (envelope.body.row, envelope.body.ordinal, envelope.body.offset, envelope.body.length) == (
        19,
        14,
        329,
        "Variable",
    )
    assert (
        envelope.closing.row,
        envelope.closing.ordinal,
        envelope.closing.offset,
        envelope.closing.length,
    ) == (20, 15, "***", 18)
    assert (
        envelope.variable_total.row,
        envelope.variable_total.label,
        envelope.variable_total.length,
    ) == (21, "total", "Variable")


@pytest.mark.parametrize(("source_ref", "filing_year", "design_epoch"), _MODELO_220_DESIGNS)
def test_modelo_220_workbooks_preserve_the_exact_composite_relative_closing(
    source_ref: str,
    filing_year: int,
    design_epoch: str,
) -> None:
    """Each SHA-bound M220 design retains all six official closing rows."""
    _modelo, catalogues = _committed_registry_tree()
    resolved = resolve_record_design_binary(
        bundled_path(),
        catalogues.sources,
        source_ref=source_ref,
        filing_year=filing_year,
        design_epoch=design_epoch,
    )

    parsed = extract_record_design(resolved.path).accept_partial()
    envelopes = tuple(sheet.variable_envelope for sheet in parsed if sheet.variable_envelope is not None)

    assert resolved.source.sha256 == catalogues.sources[source_ref].sha256
    assert len(envelopes) == 1
    envelope = envelopes[0]
    assert envelope is not None
    assert envelope.name == "T220000000"
    assert envelope.prefix_extent == 328
    assert (envelope.body.row, envelope.body.ordinal, envelope.body.offset, envelope.body.length) == (
        19,
        14,
        329,
        "Variable",
    )
    assert isinstance(envelope.closing, RecordDesignCompositeRelativeClosing)
    assert tuple(
        (part.row, part.ordinal, part.offset, part.length, part.type_code, part.content)
        for part in envelope.closing.parts
    ) == (
        (20, 15, "***", 3, "An", "</T"),
        (21, 16, "***", 3, "An", "220"),
        (22, 17, "***", 1, "An", "(*)[A|E|I|0]"),
        (23, 18, "***", 4, "An", None),
        (24, 19, "***", 2, "An", "0A"),
        (25, 20, "***", 5, "An", "0000>"),
    )
    assert (envelope.variable_total.row, envelope.variable_total.length) == (26, "Variable")


def test_variable_envelope_recognition_has_no_record_name_selector() -> None:
    """The parser recognises official composition markers, never a known tab name."""
    source = inspect.getsource(record_design_workbook_module._extract_sheet_rows)

    assert "DP200000" not in source
    assert "DP30300" not in source
    assert "T220000000" not in source


def test_workbook_declared_total_must_equal_terminal_parsed_extent(tmp_path: Path) -> None:
    """A cached official total that disagrees with parsed geometry refuses the sheet."""
    from openpyxl import Workbook

    path = tmp_path / "mismatched-total.xlsx"
    workbook = Workbook()
    worksheet = workbook.worksheets[0]
    worksheet.title = "Fixed"
    worksheet.append(("Nº", "Posic.", "Lon", "Tipo", "Descripción"))
    worksheet.append((1, 1, 2, "An", "First"))
    worksheet.append(("Total:", None, 3, None, None))
    workbook.save(path)
    workbook.close()

    with pytest.raises(RegistryValidationError, match="declares 3 total positions but parsed fields fill 2"):
        extract_record_design(path)


def test_workbook_total_recovery_accepts_only_official_labels_and_positive_integers(tmp_path: Path) -> None:
    """Punctuation variants and non-positive values do not become declared totals."""
    from openpyxl import Workbook

    path = tmp_path / "total-labels.xlsx"
    workbook = Workbook()
    labels = (
        ("Total", " Total ", 2),
        ("TotalColon", "Total:", 2),
        ("SpacedColon", "Total :", 2),
        ("NonPositive", "Total:", 0),
    )
    for index, (sheet_name, label, declared) in enumerate(labels):
        worksheet = workbook.worksheets[0] if index == 0 else workbook.create_sheet()
        worksheet.title = sheet_name
        worksheet.append(("Nº", "Posic.", "Lon", "Tipo", "Descripción"))
        worksheet.append((1, 1, 2, "An", "First"))
        worksheet.append((label, None, declared, None, None))
    workbook.save(path)
    workbook.close()

    parsed = {sheet.name: sheet.total_positions for sheet in extract_record_design(path).accept_partial()}
    assert parsed == {
        "Total": 2,
        "TotalColon": 2,
        "SpacedColon": None,
        "NonPositive": None,
    }


@pytest.mark.parametrize(
    ("rows", "message"),
    (
        (
            (
                (1, 1, 328, "An", "Fixed prefix", None, None),
                (2, 329, "Variable", "An", "Variable body", None, None),
                (3, 330, "Variable", "An", "Second body", None, None),
                (4, "***", 18, "An", "Closing suffix", None, '"</T200>"'),
                ("Total", None, "Variable", None, None, None, None),
            ),
            "duplicate variable-body markers",
        ),
        (
            (
                (1, 1, 328, "An", "Fixed prefix", None, None),
                (2, 329, "Variable", "An", "Variable body", None, None),
                (3, "***", 18, "An", "Closing suffix", None, '"</T200>"'),
                (4, "***", 2, "An", "Second suffix", None, "CRLF"),
                ("Total", None, "Variable", None, None, None, None),
            ),
            "incomplete or ambiguous relative closing",
        ),
        (
            (
                (1, 1, 328, "An", "Fixed prefix", None, None),
                (2, 329, "Variable", "An", "Variable body", None, None),
                (3, "***", 18, "An", "Closing suffix", None, '"</T200>"'),
                ("Total", None, "Variable", None, None, None, None),
                ("Total", None, "Variable", None, None, None, None),
            ),
            "duplicate variable totals",
        ),
        (
            (
                (1, 1, 328, "An", "Fixed prefix", None, None),
                ("Total", None, 328, None, None, None, None),
                ("Total", None, 328, None, None, None, None),
            ),
            "duplicate fixed totals",
        ),
        (
            (
                (1, 1, 328, "An", "Fixed prefix", None, None),
                (2, 329, "Variable", "An", "Variable body", None, None),
                (3, "***", 18, "An", "Closing suffix", None, '"</T200>"'),
                ("Total", 328, "Variable", None, None, None, None),
            ),
            "mixes fixed and variable totals",
        ),
        (
            (
                (1, 1, 328, "An", "Fixed prefix", None, None),
                ("Total", 328, "Variable", None, None, None, None),
            ),
            "mixes fixed and variable totals",
        ),
        (
            (
                (1, 1, 328, "An", "Fixed prefix", None, None),
                ("Total:", 328, "Variable", None, None, None, None),
            ),
            "mixes fixed and variable totals",
        ),
        (
            (
                (1, 1, 328, "An", "Fixed prefix", None, None),
                (2, 329, "Variable", "An", "Variable body", None, None),
                (3, "***", 18, "An", "Closing suffix", None, '"</T200>"'),
                ("Total", None, 328, None, None, None, None),
                ("Total", None, "Variable", None, None, None, None),
            ),
            "mixes fixed-total and variable-envelope geometry",
        ),
        (
            (
                (1, 1, 328, "An", "Fixed prefix", None, None),
                (2, 329, "Variable", "An", "Variable body", None, None),
                ("Total", None, "Variable", None, None, None, None),
            ),
            "incomplete variable-envelope composition",
        ),
        (
            (
                (1, 1, 328, "An", "Fixed prefix", None, None),
                (2, 329, "Variable", "An", "Variable body", None, None),
                (3, "***", 18, "An", "Closing suffix", None, '"</T200>"'),
            ),
            "incomplete variable-envelope composition",
        ),
        (
            (
                (1, 1, 328, "An", "Fixed prefix", None, None),
                (None, 329, "Variable", "An", "Malformed variable body", None, None),
            ),
            "malformed variable-envelope marker in row 3",
        ),
        (
            (
                (1, 1, 328, "An", "Fixed prefix", None, None),
                (2, 329, "Variable", "An", "Variable body", None, None),
                (3, "***", 17, "An", "Wrong-length closing suffix", None, '"</T200"'),
                ("Total", None, "Variable", None, None, None, None),
            ),
            "incomplete or ambiguous relative closing",
        ),
        (
            (
                (1, 1, 328, "An", "Fixed prefix", None, None),
                (2, 329, "Variable", "An", "Variable body", None, None),
                (3, "***", 17, "An", "Wrong-length closing suffix", None, '"</T200"'),
                ("Total", None, 328, None, None, None, None),
            ),
            "incomplete variable-envelope composition",
        ),
        (
            (
                (1, 1, 328, "An", "Fixed prefix", None, None),
                (2, 330, "Variable", "An", "Variable body", None, None),
                (3, "***", 18, "An", "Closing suffix", None, '"</T200>"'),
                ("Total", None, "Variable", None, None, None, None),
            ),
            "variable body starts at 330 after fixed prefix extent 328",
        ),
        (
            (
                (1, 1, 328, "An", "Fixed prefix", None, None),
                (3, "***", 18, "An", "Closing suffix", None, '"</T200>"'),
                (2, 329, "Variable", "An", "Variable body", None, None),
                ("Total", None, "Variable", None, None, None, None),
            ),
            "misordered variable-envelope composition markers",
        ),
    ),
)
def test_variable_envelope_rejects_malformed_composition(
    tmp_path: Path,
    rows: tuple[_WorkbookRow, ...],
    message: str,
) -> None:
    """Malformed composition markers refuse through the production workbook parser."""
    from openpyxl import Workbook

    path = tmp_path / "malformed-envelope.xlsx"
    workbook = Workbook()
    worksheet = workbook.worksheets[0]
    worksheet.title = "VARIABLE-ENVELOPE"
    worksheet.append(("Nº", "Posic.", "Lon", "Tipo", "Descripción", "Validación", "Contenido"))
    for row in rows:
        worksheet.append(row)
    workbook.save(path)
    workbook.close()

    with pytest.raises(RegistryValidationError, match=message):
        extract_record_design(path)


@pytest.mark.parametrize(
    ("closing_rows", "message"),
    (
        (_M220_COMPOSITE_CLOSING_ROWS[:-1], "incomplete or ambiguous relative closing"),
        (
            (*_M220_COMPOSITE_CLOSING_ROWS, (21, "***", 1, "An", "Duplicate", None, "!")),
            "incomplete or ambiguous relative closing",
        ),
        (
            (
                _M220_COMPOSITE_CLOSING_ROWS[0],
                _M220_COMPOSITE_CLOSING_ROWS[2],
                _M220_COMPOSITE_CLOSING_ROWS[1],
                *_M220_COMPOSITE_CLOSING_ROWS[3:],
            ),
            "malformed composite relative closing",
        ),
        (
            (
                *_M220_COMPOSITE_CLOSING_ROWS[:4],
                (19, "***", 2, "An", "Período", None, "1T"),
                _M220_COMPOSITE_CLOSING_ROWS[5],
            ),
            "malformed composite relative closing",
        ),
    ),
)
def test_modelo_220_composite_relative_closing_refuses_incomplete_duplicate_reordered_or_ambiguous_rows(
    tmp_path: Path,
    closing_rows: tuple[_WorkbookRow, ...],
    message: str,
) -> None:
    """The six-row contract fails closed without joining or defaulting parts."""
    from openpyxl import Workbook

    path = tmp_path / "malformed-m220-composite.xlsx"
    workbook = Workbook()
    worksheet = workbook.worksheets[0]
    worksheet.title = "OFFICIAL-SHAPE"
    worksheet.append(("Nº", "Posic.", "Lon", "Tipo", "Descripción", "Validación", "Contenido"))
    worksheet.append((13, 1, 328, "An", "Fixed prefix", None, None))
    worksheet.append((14, 329, "Variable", "An", "Variable body", None, None))
    for row in closing_rows:
        worksheet.append(row)
    worksheet.append(("Total", None, "Variable", None, None, None, None))
    workbook.save(path)
    workbook.close()

    with pytest.raises(RegistryValidationError, match=message):
        extract_record_design(path)


def test_closing_and_variable_total_without_a_body_do_not_reclassify_a_fixed_record(tmp_path: Path) -> None:
    """Partial marker facts remain non-envelope facts unless a body or mixed total is present."""
    from openpyxl import Workbook

    path = tmp_path / "partial-marker-fixed-record.xlsx"
    workbook = Workbook()
    worksheet = workbook.worksheets[0]
    worksheet.title = "FIXED"
    worksheet.append(("Nº", "Posic.", "Lon", "Tipo", "Descripción", "Validación", "Contenido"))
    worksheet.append((1, 1, 328, "An", "Fixed record", None, None))
    worksheet.append((2, "***", 18, "An", "Relative marker", None, "</T200020250A0000>"))
    worksheet.append(("Total", None, "Variable", None, None, None, None))
    worksheet.append(("Total", None, 328, None, None, None, None))
    workbook.save(path)
    workbook.close()

    parsed = extract_record_design(path).accept_partial()

    assert len(parsed) == 1
    assert parsed[0].total_positions == 328
    assert parsed[0].variable_envelope is None


@pytest.mark.parametrize(
    ("sheet_name", "rows", "message"),
    (
        (
            "FixedStartsLate",
            ((1, 2, 2, "An", "First"), ("Total", None, 3, None, None)),
            "has a gap before field at row 2: expected offset 1, got 2",
        ),
        (
            "FixedGap",
            ((1, 1, 2, "An", "First"), (2, 4, 2, "An", "Second"), ("Total", None, 5, None, None)),
            "has a gap before field at row 3: expected offset 3, got 4",
        ),
        (
            "FixedOverlap",
            ((1, 1, 3, "An", "First"), (2, 3, 2, "An", "Second"), ("Total", None, 4, None, None)),
            "has an overlap before field at row 3: expected offset 4, got 3",
        ),
        (
            "VariableOverlap",
            (
                (1, 1, 200, "An", "Prefix one"),
                (2, 200, 129, "An", "Overlapping prefix"),
                (3, 329, "Variable", "An", "Variable body"),
                (4, "***", 18, "An", "Closing suffix"),
                ("Total", None, "Variable", None, None),
            ),
            "has an overlap before field at row 3: expected offset 201, got 200",
        ),
        (
            "VariableGap",
            (
                (1, 1, 199, "An", "Prefix one"),
                (2, 201, 128, "An", "Discontinuous prefix"),
                (3, 329, "Variable", "An", "Variable body"),
                (4, "***", 18, "An", "Closing suffix"),
                ("Total", None, "Variable", None, None),
            ),
            "has a gap before field at row 3: expected offset 200, got 201",
        ),
    ),
)
def test_workbook_refuses_noncontiguous_fixed_and_variable_prefix_geometry(
    tmp_path: Path,
    sheet_name: str,
    rows: tuple[_WorkbookRow, ...],
    message: str,
) -> None:
    """Fixed sheets and envelope prefixes require exact source-order geometry."""
    from openpyxl import Workbook

    path = tmp_path / f"{sheet_name}.xlsx"
    workbook = Workbook()
    worksheet = workbook.worksheets[0]
    worksheet.title = sheet_name
    worksheet.append(("Nº", "Posic.", "Lon", "Tipo", "Descripción"))
    for row in rows:
        worksheet.append(row)
    workbook.save(path)
    workbook.close()

    with pytest.raises(RegistryValidationError, match=message):
        extract_record_design(path)


def _official_total_rows(
    workbook_path: Path,
) -> tuple[dict[str, int], dict[str, tuple[str, str, str, int]]]:
    """Read formula and cached views to retain the binary's exact total-row evidence."""
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    formula_book = load_workbook(workbook_path, read_only=True, data_only=False)
    cached_book = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        totals: dict[str, int] = {}
        anchors: dict[str, tuple[str, str, str, int]] = {}
        for formula_sheet in formula_book.worksheets:
            cached_sheet = cached_book[formula_sheet.title]
            for row_number, (formula_row, cached_row) in enumerate(
                zip(formula_sheet.iter_rows(values_only=True), cached_sheet.iter_rows(values_only=True), strict=True),
                start=1,
            ):
                if formula_row[0] != "Total:":
                    continue
                cached_index, cached_total = next(
                    (index, value)
                    for index, value in enumerate(cached_row[1:], start=1)
                    if isinstance(value, int) and not isinstance(value, bool) and value > 0
                )
                formula_value = formula_row[cached_index]
                assert isinstance(formula_value, str) and formula_value.startswith("=")
                totals[formula_sheet.title.strip()] = cached_total
                anchors[formula_sheet.title.strip()] = (
                    f"A{row_number}",
                    f"{get_column_letter(cached_index + 1)}{row_number}",
                    formula_value,
                    cached_total,
                )

        variable_formula = formula_book["DP200000"]
        variable_cached = cached_book["DP200000"]
        assert variable_formula["A16"].value == "Total"
        assert variable_cached["C16"].value == "Variable"
        return totals, anchors
    finally:
        formula_book.close()
        cached_book.close()
