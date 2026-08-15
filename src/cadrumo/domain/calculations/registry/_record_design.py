"""Read-only extraction of official AEAT record-design rows.

Parses official AEAT record-design workbooks (PDF or XLS/XLSX) and derives
coverage casillas from a :class:`ModeloRevision` so that the extracted layout
can be compared against the registry declarations.
"""

from __future__ import annotations

import json
import re
import warnings
from collections.abc import Generator, Iterator, Mapping
from contextlib import contextmanager
from dataclasses import dataclass, field
from functools import lru_cache
from io import BufferedReader, BytesIO
from pathlib import Path
from typing import TYPE_CHECKING, Final

from pydantic import ConfigDict, TypeAdapter, ValidationError

if TYPE_CHECKING:
    # Annotation-only: ``from __future__ import annotations`` above makes every
    # annotation a string, so these never need to exist at runtime. The parser
    # backends themselves (openpyxl, pdfplumber, pypdfium2, xlrd) are among the
    # heaviest third-party imports in the tree and are deferred into the
    # extraction functions that actually call them -- importing the registry
    # must not pay for a PDF/XLS parser stack no calculation touches.
    from openpyxl.worksheet.worksheet import Worksheet
    from pdfplumber.page import Page
    from xlrd.sheet import Sheet as XlrdSheet

from ....core.external_constants import PDF_EXTENSION as _PDF_EXTENSION
from ....core.external_constants import XLS_EXTENSION as _XLS_EXTENSION
from ....core.external_constants import XLSM_EXTENSION as _XLSM_EXTENSION
from ....core.external_constants import XLSX_EXTENSION as _XLSX_EXTENSION
from ....core.logging import get_logger
from ....core.paths import path_stat_fingerprint
from ....core.tabular import coerce_cell_text
from ._errors import RegistryValidationError
from ._record_design_coverage import (
    DerivedDisenoCasilla,
    DisenoCoverageReport,
    build_diseno_coverage_report,
    calculation_closure_casilla_ids,
    calculation_closure_legal_refs,
    calculation_closure_record_design_metadata,
    derive_calculation_completeness_casillas,
    derive_diseno_coverage_casillas,
)
from ._record_design_schema import (
    RecordDesignAuxiliaryEnvelopeHeader,
    RecordDesignAuxiliaryEnvelopeHeaderField,
    RecordDesignAuxiliaryEnvelopeHeaderRole,
    RecordDesignCompositeRelativeClosing,
    RecordDesignCorrection,
    RecordDesignExtraction,
    RecordDesignField,
    RecordDesignFieldTypeCorrection,
    RecordDesignHeaderCellCorrection,
    RecordDesignRelativeSuffixMarker,
    RecordDesignSheet,
    RecordDesignSkippedSheet,
    RecordDesignVariableBodyMarker,
    RecordDesignVariableEnvelope,
    RecordDesignVariableTotalMarker,
)

_log = get_logger(__name__)

_OPENPYXL_HEADER_FOOTER_WARNING = "Cannot parse header or footer so it will be ignored"
_OPENPYXL_PRINT_AREA_WARNING = r"Print area cannot be set to Defined name: .*"
_NUMERIC_TUPLE_ADAPTER: TypeAdapter[tuple[int | float, ...]] = TypeAdapter(
    tuple[int | float, ...], config=ConfigDict(strict=True)
)


@dataclass(frozen=True)
class _WorkbookHeader:
    row_number: int
    ordinal_index: int
    offset_index: int
    length_index: int
    type_index: int
    complementary_index: int | None
    description_index: int
    validation_index: int | None
    content_index: int | None


@dataclass
class _WorkbookSheetRows:
    """Parsed workbook rows retained until their sheet-level shape is known."""

    fields: list[RecordDesignField] = field(default_factory=list)
    total_positions: int | None = None
    variable_bodies: list[RecordDesignVariableBodyMarker] = field(default_factory=list)
    relative_suffixes: list[RecordDesignRelativeSuffixMarker] = field(default_factory=list)
    variable_totals: list[RecordDesignVariableTotalMarker] = field(default_factory=list)
    variable_body_marker_rows: list[int] = field(default_factory=list)
    relative_suffix_marker_rows: list[int] = field(default_factory=list)
    variable_total_marker_rows: list[int] = field(default_factory=list)
    mixed_total_rows: list[int] = field(default_factory=list)
    corrections_applied: list[RecordDesignCorrection] = field(default_factory=list)


type _TypeCorrectionIndex = Mapping[tuple[str, int], RecordDesignFieldTypeCorrection]
type _HeaderCorrectionIndex = Mapping[tuple[str, int, str], RecordDesignHeaderCellCorrection]

_CORRECTION_SUFFIX: Final[str] = ".record-design-correction.json"
_CORRECTION_ADAPTER: Final[TypeAdapter[RecordDesignCorrection]] = TypeAdapter(RecordDesignCorrection)


@dataclass(frozen=True)
class _CorrectionIndex:
    """One binary's declared corrections, split by the row they address.

    ``type_corrections`` keys on ``(sheet, source_row)`` -- one data row.
    ``header_corrections`` keys on ``(sheet, header_row, column_role)`` -- one
    header column, since a header row's blank cell is looked up by ROLE
    (``"length"``) at probe time, not by a data row number.
    """

    type_corrections: _TypeCorrectionIndex
    header_corrections: _HeaderCorrectionIndex


_EMPTY_CORRECTIONS: Final[_CorrectionIndex] = _CorrectionIndex(type_corrections={}, header_corrections={})


def _load_corrections(source_path: Path) -> _CorrectionIndex:
    """Load a hand-authored, per-binary sidecar declaring record-design corrections.

    Colocated with the exact source binary it corrects, named
    ``<binary-name>.record-design-correction.json`` -- a distinct suffix from
    the parser's own generated ``.extracted.json``/``.extracted.md`` cache, so
    a hand-authored grounding declaration is never confused with, or
    overwritten by, machine output. Absent for the overwhelming majority of
    bundled binaries, which read as AEAT published them; this returns empty
    indexes for those, so the parser's behaviour is unchanged unless a sidecar
    is deliberately authored. One file, one discriminated ``corrections`` list
    -- a field-type correction and a header-cell correction may both appear in
    it, per :data:`RecordDesignCorrection`.
    """
    sidecar_path = source_path.with_name(source_path.name + _CORRECTION_SUFFIX)
    if not sidecar_path.is_file():
        return _EMPTY_CORRECTIONS
    payload = json.loads(sidecar_path.read_text(encoding="utf-8"))
    entries = payload.get("corrections") if isinstance(payload, dict) else None
    if not isinstance(entries, list):
        raise RegistryValidationError(f"{sidecar_path}: correction sidecar must declare a 'corrections' list")
    type_corrections: dict[tuple[str, int], RecordDesignFieldTypeCorrection] = {}
    header_corrections: dict[tuple[str, int, str], RecordDesignHeaderCellCorrection] = {}
    for entry in entries:
        # ``strict=False`` here only: JSON has no tuple literal, so the sidecar's
        # ``editions_read`` array arrives as a ``list`` and needs the ordinary
        # list-to-tuple coercion. Every field's own type is still checked --
        # this does not relax ``min_length``, blank-string, discriminator, or shape checks.
        correction = _CORRECTION_ADAPTER.validate_python(entry, strict=False)
        if isinstance(correction, RecordDesignFieldTypeCorrection):
            type_key = (correction.sheet, correction.source_row)
            if type_key in type_corrections:
                raise RegistryValidationError(
                    f"{sidecar_path}: duplicate type correction for sheet {correction.sheet!r} "
                    f"row {correction.source_row}",
                )
            type_corrections[type_key] = correction
        else:
            header_key = (correction.sheet, correction.header_row, correction.column_role)
            if header_key in header_corrections:
                raise RegistryValidationError(
                    f"{sidecar_path}: duplicate header correction for sheet {correction.sheet!r} "
                    f"row {correction.header_row} role {correction.column_role!r}",
                )
            header_corrections[header_key] = correction
    return _CorrectionIndex(type_corrections=type_corrections, header_corrections=header_corrections)


_DECLARED_NON_RECORD_SHEETS_FILENAME: Final[str] = "declared-non-record-sheets.json"


def _load_declared_non_record_sheet_reasons(source_path: Path) -> Mapping[str, str]:
    """Load one modelo's declared, sourced reasons for sheets that are never records.

    Lives once per MODELO directory (sibling to that modelo's own
    ``manifest.json``), not per binary: a legend or lookup tab AEAT republishes
    unchanged across several editions is one judgement, not one per file. This
    never turns a skip into a read -- the sheet stays in
    :attr:`RecordDesignExtraction.skipped` exactly as before -- it only
    replaces the parser's own generic header-probe failure message with the
    grounded reason a reviewer recorded after opening the design. The
    extractor cannot itself tell a lookup tab apart from a dropped record
    body, so that judgement is a registry act, never inferred here.
    """
    modelo_root = source_path.parent.parent
    declaration_path = modelo_root / _DECLARED_NON_RECORD_SHEETS_FILENAME
    if not declaration_path.is_file():
        return {}
    payload = json.loads(declaration_path.read_text(encoding="utf-8"))
    entries = payload.get("declared_non_record_sheets") if isinstance(payload, dict) else None
    if not isinstance(entries, list):
        raise RegistryValidationError(
            f"{declaration_path}: must declare a 'declared_non_record_sheets' list",
        )
    reasons: dict[str, str] = {}
    for entry in entries:
        if (
            not isinstance(entry, dict)
            or not isinstance(entry.get("sheet"), str)
            or not isinstance(
                entry.get("reason"),
                str,
            )
        ):
            raise RegistryValidationError(
                f"{declaration_path}: every entry needs a string 'sheet' and a string 'reason'",
            )
        sheet = entry["sheet"].strip()
        reason = entry["reason"].strip()
        if not sheet or not reason:
            raise RegistryValidationError(f"{declaration_path}: 'sheet' and 'reason' must be non-blank")
        if sheet in reasons:
            raise RegistryValidationError(f"{declaration_path}: duplicate declaration for sheet {sheet!r}")
        reasons[sheet] = reason
    return reasons


def extract_record_design(path: Path) -> RecordDesignExtraction:
    """Return one official record-design source's parsed sheets AND what it could not read.

    Returns:
        The :class:`RecordDesignExtraction` for the source, which names both the
        sheets that parsed and any the extractor had to skip.
    """
    resolved = path.resolve()
    if not resolved.is_file():
        raise FileNotFoundError(f"record-design source not found: {path}")
    return _extract_record_design_cached(*path_stat_fingerprint(resolved))


@lru_cache(maxsize=256)
def _extract_record_design_cached(
    path: str,
    byte_count: int,
    modified_ns: int,
) -> RecordDesignExtraction:
    del byte_count, modified_ns
    source_path = Path(path)
    suffix = source_path.suffix.lower()
    if suffix == _PDF_EXTENSION:
        return extract_record_design_pdf(source_path)
    if suffix in {_XLSX_EXTENSION, _XLSM_EXTENSION}:
        return extract_record_design_workbook(source_path)
    if suffix == _XLS_EXTENSION:
        return extract_record_design_xls_workbook(source_path)
    raise RegistryValidationError(f"unsupported record-design source extension: {source_path.suffix}")


def _extraction(
    source_path: Path,
    sheets: list[RecordDesignSheet],
    skipped: list[RecordDesignSkippedSheet],
) -> RecordDesignExtraction:
    """Assemble one source's result, refusing only when NOTHING could be read.

    A source where every sheet failed is still a hard error -- there is no design
    there to hand back. A source where SOME sheets failed is a partial read, and
    it is returned rather than raised so a caller can decide: refusing outright
    would drop Modelo 232, whose ``TABLAS`` tab is a legitimate lookup table and
    not a record at all. The extractor cannot tell a lookup tab from a lost
    record body, so it reports both and adjudicates neither.
    """
    if not sheets:
        detail = "; ".join(f"{item.name!r}: {item.reason}" for item in skipped) if skipped else "none"
        raise RegistryValidationError(
            f"{source_path}: no record-design sheets found; skipped sheets: {detail}",
        )
    return RecordDesignExtraction(source=str(source_path), sheets=tuple(sheets), skipped=tuple(skipped))


def extract_record_design_workbook(path: Path) -> RecordDesignExtraction:
    """Return the :class:`RecordDesignExtraction` workbook ``path`` describes."""
    resolved = path.resolve()
    if not resolved.is_file():
        raise FileNotFoundError(f"record-design workbook not found: {path}")
    return _extract_record_design_workbook_cached(*path_stat_fingerprint(resolved))


def extract_record_design_xls_workbook(path: Path) -> RecordDesignExtraction:
    """Return a legacy binary XLS workbook's parsed sheets AND any it could not read.

    Returns:
        The :class:`RecordDesignExtraction` for the workbook.
    """
    resolved = path.resolve()
    if not resolved.is_file():
        raise FileNotFoundError(f"record-design XLS workbook not found: {path}")
    return _extract_record_design_xls_workbook_cached(*path_stat_fingerprint(resolved))


@lru_cache(maxsize=256)
def _extract_record_design_workbook_cached(
    path: str,
    byte_count: int,
    modified_ns: int,
) -> RecordDesignExtraction:
    del byte_count, modified_ns
    from openpyxl import load_workbook

    source_path = Path(path)
    corrections = _load_corrections(source_path)
    declared_skip_reasons = _load_declared_non_record_sheet_reasons(source_path)
    with _ignore_openpyxl_header_footer_metadata_warnings():
        workbook = load_workbook(source_path, read_only=True, data_only=True)
        try:
            sheets: list[RecordDesignSheet] = []
            skipped: list[RecordDesignSkippedSheet] = []
            for worksheet in workbook.worksheets:
                try:
                    sheets.append(_extract_sheet(worksheet, corrections))
                except ValueError as exc:
                    if "has no record-design header" not in str(exc):
                        raise
                    sheet_title = worksheet.title.strip()
                    reason = declared_skip_reasons.get(sheet_title, str(exc))
                    skipped.append(RecordDesignSkippedSheet(name=sheet_title, reason=reason))
            return _extraction(source_path, sheets, skipped)
        finally:
            workbook.close()


@lru_cache(maxsize=128)
def _extract_record_design_xls_workbook_cached(
    path: str,
    byte_count: int,
    modified_ns: int,
) -> RecordDesignExtraction:
    del byte_count, modified_ns
    import xlrd

    source_path = Path(path)
    corrections = _load_corrections(source_path)
    declared_skip_reasons = _load_declared_non_record_sheet_reasons(source_path)
    workbook = xlrd.open_workbook(str(source_path), on_demand=True)
    try:
        sheets: list[RecordDesignSheet] = []
        skipped: list[RecordDesignSkippedSheet] = []
        for sheet_name in workbook.sheet_names():
            worksheet = workbook.sheet_by_name(sheet_name)
            try:
                sheets.append(_extract_xls_sheet(worksheet, corrections))
            except ValueError as exc:
                if "has no record-design header" not in str(exc):
                    raise
                stripped_name = sheet_name.strip()
                reason = declared_skip_reasons.get(stripped_name, str(exc))
                skipped.append(RecordDesignSkippedSheet(name=stripped_name, reason=reason))
        return _extraction(source_path, sheets, skipped)
    finally:
        workbook.release_resources()


@contextmanager
def _ignore_openpyxl_header_footer_metadata_warnings() -> Generator[None]:
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message=_OPENPYXL_HEADER_FOOTER_WARNING,
            category=UserWarning,
            module=r"openpyxl\.worksheet\.header_footer",
        )
        warnings.filterwarnings(
            "ignore",
            message=_OPENPYXL_PRINT_AREA_WARNING,
            category=UserWarning,
            module=r"openpyxl\.reader\.workbook",
        )
        yield


def extract_record_design_pdf(path: Path) -> RecordDesignExtraction:
    """Return the :class:`RecordDesignExtraction` read from an official AEAT PDF."""
    resolved = path.resolve()
    if not resolved.is_file():
        raise FileNotFoundError(f"record-design PDF not found: {path}")
    return _extract_record_design_pdf_cached(*path_stat_fingerprint(resolved))


@lru_cache(maxsize=256)
def _extract_record_design_pdf_cached(
    path: str,
    byte_count: int,
    modified_ns: int,
) -> RecordDesignExtraction:
    del byte_count, modified_ns
    source_path = Path(path)
    with source_path.open("rb") as pdf_file:
        return _extract_record_design_pdf_stream(pdf_file, source_label=str(source_path))


def extract_record_design_pdf_bytes(
    pdf_bytes: bytes,
    *,
    source_label: str = "in-memory record-design PDF",
) -> RecordDesignExtraction:
    """Return the record design extracted from PDF bytes.

    Returns:
        The :class:`RecordDesignExtraction` for the PDF content.
    """
    return _extract_record_design_pdf_stream(BytesIO(pdf_bytes), source_label=source_label)


def _extract_record_design_pdf_stream(
    stream: BufferedReader | BytesIO,
    *,
    source_label: str,
) -> RecordDesignExtraction:
    import pdfplumber

    pdf_bytes = stream.read()
    lines = _extract_pdf_text_lines(pdf_bytes, source_label=source_label)
    if _uses_page_record_layout(lines):
        lines = _extract_pdfplumber_text_lines(pdf_bytes, source_label=source_label)
    if not any(line.strip() for line in lines):
        raise RegistryValidationError(f"no text extracted from record-design PDF {source_label}")
    try:
        return _extract_pdf_lines(lines, source_label=source_label)
    except ValueError as pdfium_exc:
        text_fallback_error = pdfium_exc
        try:
            fallback_lines = _extract_pdfplumber_text_lines(pdf_bytes, source_label=source_label)
            return _extract_pdf_lines(fallback_lines, source_label=source_label)
        except ValueError as fallback_exc:
            text_fallback_error = fallback_exc
        if "did not contain parseable field rows" not in str(text_fallback_error):
            raise text_fallback_error from pdfium_exc
        try:
            with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
                pages = tuple(_snapshot_pdf_page(page) for page in pdf.pages)
        except Exception as pdf_exc:  # pragma: no cover - defensive; pdfplumber surface
            raise RegistryValidationError(
                f"pdfplumber could not open record-design PDF {source_label}: {pdf_exc}",
            ) from pdf_exc
        visual_chart = _extract_visual_record_design_chart(pages, source_label=source_label)
        if visual_chart:
            # The geometry reader either reconstructs every charted record or
            # returns nothing, so a result here is complete by construction.
            return RecordDesignExtraction(source=source_label, sheets=visual_chart)
        raise


def _extract_sheet(worksheet: Worksheet, corrections: _CorrectionIndex = _EMPTY_CORRECTIONS) -> RecordDesignSheet:
    header, header_correction = _find_header(worksheet, corrections.header_corrections)
    return _extract_sheet_rows(
        worksheet.title,
        header,
        enumerate(
            worksheet.iter_rows(min_row=header.row_number + 1, values_only=True),
            start=header.row_number + 1,
        ),
        corrections,
        header_correction,
    )


def _extract_xls_sheet(worksheet: XlrdSheet, corrections: _CorrectionIndex = _EMPTY_CORRECTIONS) -> RecordDesignSheet:
    header, header_correction = _find_xls_header(worksheet, corrections.header_corrections)
    return _extract_sheet_rows(
        worksheet.name,
        header,
        ((rowx + 1, tuple(worksheet.row_values(rowx))) for rowx in range(header.row_number, worksheet.nrows)),
        corrections,
        header_correction,
    )


def _extract_sheet_rows(
    sheet_name: str,
    header: _WorkbookHeader,
    rows: Iterator[tuple[int, tuple[object, ...]]],
    corrections: _CorrectionIndex = _EMPTY_CORRECTIONS,
    header_correction: RecordDesignHeaderCellCorrection | None = None,
) -> RecordDesignSheet:
    # AEAT Diseño workbooks occasionally carry surrounding whitespace on a
    # sheet tab (e.g. 'DP200026 '). The sheet name is the record-segment
    # identity that segment-qualified casillas and the calculation-
    # completeness derivation match against, so the raw tab whitespace
    # must not leak into that identity.
    sheet_name = sheet_name.strip()
    parsed_rows = _scan_sheet_rows(sheet_name, header, rows, corrections.type_corrections)
    if header_correction is not None:
        parsed_rows.corrections_applied.insert(0, header_correction)
    terminal_extent = _require_contiguous_field_geometry(sheet_name, parsed_rows.fields)
    if parsed_rows.total_positions is not None and terminal_extent != parsed_rows.total_positions:
        raise RegistryValidationError(
            f"record-design sheet {sheet_name!r} declares {parsed_rows.total_positions} total positions "
            f"but parsed fields fill {terminal_extent}",
        )
    variable_envelope = _variable_envelope(sheet_name, parsed_rows, terminal_extent)
    auxiliary_envelope_header = (
        None
        if variable_envelope is not None
        else _auxiliary_envelope_header(sheet_name, parsed_rows.fields, parsed_rows.total_positions, terminal_extent)
    )
    return RecordDesignSheet(
        name=sheet_name,
        fields=tuple(parsed_rows.fields),
        total_positions=parsed_rows.total_positions,
        variable_envelope=variable_envelope,
        auxiliary_envelope_header=auxiliary_envelope_header,
        corrections=tuple(parsed_rows.corrections_applied),
    )


def _scan_sheet_rows(
    sheet_name: str,
    header: _WorkbookHeader,
    rows: Iterator[tuple[int, tuple[object, ...]]],
    corrections: _TypeCorrectionIndex | None = None,
) -> _WorkbookSheetRows:
    parsed_rows = _WorkbookSheetRows()
    trailing_blank_rows = 0
    for row_number, row in rows:
        values = tuple(row)
        if _is_blank_row(values):
            trailing_blank_rows += 1
            if parsed_rows.fields and trailing_blank_rows >= 25:
                break
            continue
        trailing_blank_rows = 0
        if _consume_total_row(sheet_name, parsed_rows, row_number, values):
            continue
        _consume_field_row(sheet_name, parsed_rows, header, row_number, values, corrections or {})
    return parsed_rows


def _consume_total_row(
    sheet_name: str,
    parsed_rows: _WorkbookSheetRows,
    row_number: int,
    values: tuple[object, ...],
) -> bool:
    total_label_index = _total_label_index(values)
    if total_label_index is None:
        return False
    row_total = _positive_integer_after(values, total_label_index)
    has_variable_total = any(_optional_text(candidate) == "Variable" for candidate in values[total_label_index + 1 :])
    if has_variable_total:
        parsed_rows.variable_total_marker_rows.append(row_number)
    if row_total is not None and has_variable_total:
        parsed_rows.mixed_total_rows.append(row_number)
    if row_total is not None:
        if parsed_rows.total_positions is not None:
            raise RegistryValidationError(f"record-design sheet {sheet_name!r} declares duplicate fixed totals")
        parsed_rows.total_positions = row_total
    elif has_variable_total:
        parsed_rows.variable_totals.append(
            RecordDesignVariableTotalMarker(
                sheet=sheet_name,
                row=row_number,
                label="total",
                length="Variable",
            ),
        )
    return True


def _consume_field_row(
    sheet_name: str,
    parsed_rows: _WorkbookSheetRows,
    header: _WorkbookHeader,
    row_number: int,
    values: tuple[object, ...],
    corrections: _TypeCorrectionIndex,
) -> None:
    # ``RecordDesignField.ordinal`` is the printed LABEL (``ordinal_text``), never
    # an arithmetic value -- it is now representable verbatim, so there is no more
    # "printed but unreadable" case to refuse for it. The two MARKER rows below
    # (variable-body, relative-suffix) are unrelated types whose own ``ordinal``
    # field is still a plain sequential ``int``, so they keep reading the
    # int-or-None form.
    ordinal_int = _int_or_none(_cell(values, header.ordinal_index))
    ordinal_text = _ordinal_text(_cell(values, header.ordinal_index))
    offset = _int_or_none(_cell(values, header.offset_index))
    length = _int_or_none(_cell(values, header.length_index))
    raw_offset = _optional_text(_cell(values, header.offset_index))
    raw_length = _optional_text(_cell(values, header.length_index))
    if raw_length == "Variable":
        parsed_rows.variable_body_marker_rows.append(row_number)
        if ordinal_int is not None and offset is not None:
            parsed_rows.variable_bodies.append(
                _variable_body_marker(sheet_name, header, row_number, values, ordinal_int, offset),
            )
        return
    if raw_offset == "***":
        parsed_rows.relative_suffix_marker_rows.append(row_number)
        if ordinal_int is not None and length is not None:
            parsed_rows.relative_suffixes.append(
                _relative_suffix_marker(sheet_name, header, row_number, values, ordinal_int, length),
            )
        return
    if offset is None or length is None:
        return
    field, applied_correction = _record_design_field(
        sheet_name,
        header,
        row_number,
        values,
        ordinal_text,
        offset,
        length,
        corrections,
    )
    if applied_correction is not None:
        parsed_rows.corrections_applied.append(applied_correction)
    parent_index = _matching_component_parent_index(parsed_rows.fields, ordinal_text, offset, length)
    if parent_index is not None:
        parent = parsed_rows.fields[parent_index]
        parsed_rows.fields[parent_index] = parent.model_copy(update={"components": (*parent.components, field)})
        return
    parsed_rows.fields.append(field)


def _matching_component_parent_index(
    fields: list[RecordDesignField],
    ordinal_text: str | None,
    offset: int,
    length: int,
) -> int | None:
    """Return the index of the field ``ordinal_text``/``offset``/``length`` desglosa from.

    AEAT prints a component with a DOTTED ordinal (``19.1`` under parent ``19``),
    never a bare one -- Modelo 303's ``14bis`` has no dot and is a genuine peer,
    not a component of ``14``, and this is the discriminator that keeps them
    apart. The dotted integer prefix alone is not enough on its own -- an
    unrelated field elsewhere in the same sheet could coincidentally share an
    ordinal -- so BOTH conditions are required together: the immediately
    preceding field's ordinal is the exact dotted prefix, AND this row's byte
    span falls entirely inside that field's own already-declared span. Checked
    against only the immediately preceding field, matching every component's
    observed source position directly beneath its parent, never a scan of the
    whole sheet.
    """
    if ordinal_text is None or "." not in ordinal_text or not fields:
        return None
    prefix = ordinal_text.split(".", 1)[0]
    parent_index = len(fields) - 1
    parent = fields[parent_index]
    if parent.ordinal != prefix:
        return None
    if offset < parent.offset or offset + length > parent.offset + parent.length:
        return None
    return parent_index


def _variable_body_marker(
    sheet_name: str,
    header: _WorkbookHeader,
    row_number: int,
    values: tuple[object, ...],
    ordinal: int | None,
    offset: int,
) -> RecordDesignVariableBodyMarker:
    validation, content, description = _field_texts(sheet_name, header, row_number, values)
    return RecordDesignVariableBodyMarker(
        sheet=sheet_name,
        row=row_number,
        ordinal=ordinal,
        offset=offset,
        length="Variable",
        type_code=_required_text(_cell(values, header.type_index), sheet_name, row_number, "type"),
        description=description,
        validation=validation,
        content=content,
    )


def _relative_suffix_marker(
    sheet_name: str,
    header: _WorkbookHeader,
    row_number: int,
    values: tuple[object, ...],
    ordinal: int,
    length: int,
) -> RecordDesignRelativeSuffixMarker:
    validation, content, description = _field_texts(sheet_name, header, row_number, values)
    return RecordDesignRelativeSuffixMarker(
        sheet=sheet_name,
        row=row_number,
        ordinal=ordinal,
        offset="***",
        length=length,
        type_code=_required_text(_cell(values, header.type_index), sheet_name, row_number, "type"),
        description=description,
        validation=validation,
        content=content,
    )


def _record_design_field(
    sheet_name: str,
    header: _WorkbookHeader,
    row_number: int,
    values: tuple[object, ...],
    ordinal: str | None,
    offset: int,
    length: int,
    corrections: _TypeCorrectionIndex,
) -> tuple[RecordDesignField, RecordDesignFieldTypeCorrection | None]:
    validation, content, description = _field_texts(sheet_name, header, row_number, values)
    type_code, applied_correction = _required_type_code(
        _cell(values, header.type_index),
        sheet_name,
        row_number,
        corrections,
    )
    return (
        RecordDesignField(
            sheet=sheet_name,
            row=row_number,
            ordinal=ordinal,
            offset=offset,
            length=length,
            type_code=type_code,
            complementary=_optional_header_text(values, header.complementary_index),
            description=description,
            validation=validation,
            content=content,
        ),
        applied_correction,
    )


def _field_texts(
    sheet_name: str,
    header: _WorkbookHeader,
    row_number: int,
    values: tuple[object, ...],
) -> tuple[str | None, str | None, str]:
    validation = _optional_header_text(values, header.validation_index)
    content = _optional_header_text(values, header.content_index)
    return (
        validation,
        content,
        _field_description_text(
            values,
            header=header,
            content=content,
            sheet=sheet_name,
            row=row_number,
        ),
    )


def _variable_envelope(
    sheet_name: str,
    parsed_rows: _WorkbookSheetRows,
    terminal_extent: int | None,
) -> RecordDesignVariableEnvelope | None:
    if not (parsed_rows.variable_body_marker_rows or parsed_rows.mixed_total_rows):
        return None
    if parsed_rows.mixed_total_rows:
        raise RegistryValidationError(
            f"record-design sheet {sheet_name!r} mixes fixed and variable totals "
            f"in row {parsed_rows.mixed_total_rows[0]}",
        )
    _require_valid_variable_envelope_markers(sheet_name, parsed_rows)
    variable_body = parsed_rows.variable_bodies[0]
    closing_suffixes, terminator = _split_record_terminator(parsed_rows.relative_suffixes)
    closing = _relative_closing(sheet_name, closing_suffixes)
    closing_parts = _relative_closing_parts(closing)
    variable_total = parsed_rows.variable_totals[0]
    if parsed_rows.total_positions is not None or terminal_extent is None:
        raise RegistryValidationError(
            f"record-design sheet {sheet_name!r} mixes fixed-total and variable-envelope geometry",
        )
    if variable_body.offset != terminal_extent + 1:
        raise RegistryValidationError(
            f"record-design sheet {sheet_name!r} variable body starts at {variable_body.offset} "
            f"after fixed prefix extent {terminal_extent}",
        )
    _require_ordered_variable_envelope(
        sheet_name,
        parsed_rows.fields,
        variable_body,
        closing_parts,
        variable_total,
    )
    _require_terminator_closes_the_record(sheet_name, closing_parts, terminator)
    return RecordDesignVariableEnvelope(
        name=sheet_name,
        prefix_fields=tuple(parsed_rows.fields),
        prefix_extent=terminal_extent,
        body=variable_body,
        closing=closing,
        terminator=terminator,
        variable_total=variable_total,
    )


def _auxiliary_envelope_header(
    sheet_name: str,
    fields: list[RecordDesignField],
    declared_total: int | None,
    terminal_extent: int | None,
) -> RecordDesignAuxiliaryEnvelopeHeader | None:
    """Recognise only the exact total-less Modelo 390 page-zero source shape.

    Fixed sheets otherwise remain unchanged. A no-total 328-byte sheet is not
    permitted to acquire a fixed-record total from terminal extent; only an
    exact thirteen-field M390 header receives this distinct classification.
    """
    if (
        declared_total is not None
        or terminal_extent != 328
        or len(fields) != len(RecordDesignAuxiliaryEnvelopeHeaderRole)
    ):
        return None
    try:
        return RecordDesignAuxiliaryEnvelopeHeader(
            sheet=sheet_name,
            record_identity=sheet_name,
            fields=tuple(
                RecordDesignAuxiliaryEnvelopeHeaderField(role=role, field=field)
                for role, field in zip(RecordDesignAuxiliaryEnvelopeHeaderRole, fields, strict=True)
            ),
            emitted_extent=terminal_extent,
        )
    except ValueError:
        return None


def _require_valid_variable_envelope_markers(sheet_name: str, parsed_rows: _WorkbookSheetRows) -> None:
    malformed_marker_rows = (
        set(parsed_rows.variable_body_marker_rows) - {body.row for body in parsed_rows.variable_bodies}
    ) | (set(parsed_rows.relative_suffix_marker_rows) - {suffix.row for suffix in parsed_rows.relative_suffixes})
    if malformed_marker_rows:
        raise RegistryValidationError(
            f"record-design sheet {sheet_name!r} has malformed variable-envelope marker "
            f"in row {min(malformed_marker_rows)}",
        )
    if len(parsed_rows.variable_bodies) > 1:
        raise RegistryValidationError(f"record-design sheet {sheet_name!r} declares duplicate variable-body markers")
    if len(parsed_rows.variable_totals) > 1:
        raise RegistryValidationError(f"record-design sheet {sheet_name!r} declares duplicate variable totals")
    if not parsed_rows.variable_bodies or not parsed_rows.relative_suffixes or not parsed_rows.variable_totals:
        raise RegistryValidationError(
            f"record-design sheet {sheet_name!r} has an incomplete variable-envelope composition"
        )


def _require_ordered_variable_envelope(
    sheet_name: str,
    fields: list[RecordDesignField],
    variable_body: RecordDesignVariableBodyMarker,
    closing_parts: tuple[RecordDesignRelativeSuffixMarker, ...],
    variable_total: RecordDesignVariableTotalMarker,
) -> None:
    first_closing_part = closing_parts[0]
    last_closing_part = closing_parts[-1]
    # POSITION IS THE ORDERING AUTHORITY, NOT THE ORDINAL.
    #
    # This checked composition order twice, once on source row and once on ordinal,
    # and the ordinal half asserted nothing the row half did not: both ask whether
    # the body sits after the fixed prefix and before the closing. Where the two
    # could disagree, the ordinal is the one that would be wrong, because AEAT's
    # ordinal is a PRINTED LABEL rather than an arithmetic value -- it publishes
    # ``14bis`` to insert a field between 14 and 15 without renumbering, exactly as
    # the law numbers an inserted article. Comparing labels arithmetically assumes a
    # density AEAT never promised.
    #
    # The bytes order themselves and are checked as such elsewhere: the prefix must
    # be contiguous from offset 1, and the variable body must begin at
    # ``prefix_extent + 1``. Those are the real geometric assertions; this one is
    # about which SOURCE ROWS compose the envelope, so it reads rows.
    ordered_rows = max(item.row for item in fields) < variable_body.row < first_closing_part.row
    ordered_rows = ordered_rows and first_closing_part.row <= last_closing_part.row < variable_total.row
    if not ordered_rows:
        raise RegistryValidationError(
            f"record-design sheet {sheet_name!r} has misordered variable-envelope composition markers",
        )


#: How AEAT NAMES the physical end-of-record row, in the wordings it uses across
#: both workbook and PDF designs. ONE definition with two consumers: the workbook
#: closing splitter below, and the PDF compact-row recogniser further down which
#: composes this into its own line pattern.
#:
#: Two independent notions of "what a CRLF row is" is how the two parsers came to
#: disagree about one domain fact -- the PDF path has recognised the terminator
#: since it was written, while the workbook path refused thirty designs across
#: eight modelos for declaring one.
_RECORD_TERMINATOR_PHRASE = r"fin de registro|salto de l[íi]nea|\bCRLF\b"
#: Matched on the declared MEANING rather than on width: a two-byte relative
#: suffix that is not a line terminator is part of the closing identifier and
#: must not be mistaken for one.
_RECORD_TERMINATOR = re.compile(_RECORD_TERMINATOR_PHRASE, re.IGNORECASE)


def _split_record_terminator(
    suffixes: list[RecordDesignRelativeSuffixMarker],
) -> tuple[list[RecordDesignRelativeSuffixMarker], RecordDesignRelativeSuffixMarker | None]:
    """Separate a trailing physical end-of-record row from the closing identifier.

    The closing identifies the record; the terminator ends the line. AEAT declares
    them as adjacent relative-offset rows, and the closing recogniser below reads
    only the first kind, so a design declaring both was refused outright -- thirty
    of them, across eight modelos, every one well formed.

    Split rather than skipped. The terminator is returned to the caller and stored
    on the envelope, because its two bytes are part of the record: discarding it
    would let all thirty parse while every emitted record came out two bytes short,
    which is a clean-looking wrong answer rather than a refusal.
    """
    if not suffixes:
        return suffixes, None
    last = suffixes[-1]
    if last.length == 2 and _RECORD_TERMINATOR.search(last.description):
        return suffixes[:-1], last
    return suffixes, None


def _relative_closing(
    sheet_name: str,
    suffixes: list[RecordDesignRelativeSuffixMarker],
) -> RecordDesignRelativeSuffixMarker | RecordDesignCompositeRelativeClosing:
    if len(suffixes) == 1 and suffixes[0].length == 18:
        return suffixes[0]
    if len(suffixes) != 6:
        raise RegistryValidationError(
            f"record-design sheet {sheet_name!r} has an incomplete or ambiguous relative closing",
        )
    try:
        return RecordDesignCompositeRelativeClosing(
            tag_prefix=suffixes[0],
            modelo=suffixes[1],
            discriminant=suffixes[2],
            filing_year=suffixes[3],
            period=suffixes[4],
            tag_suffix=suffixes[5],
        )
    except ValidationError as exc:
        raise RegistryValidationError(
            f"record-design sheet {sheet_name!r} has a malformed composite relative closing: {exc}",
        ) from exc


def _require_terminator_closes_the_record(
    sheet_name: str,
    closing_parts: tuple[RecordDesignRelativeSuffixMarker, ...],
    terminator: RecordDesignRelativeSuffixMarker | None,
) -> None:
    """Refuse a terminator that does not sit last in the record.

    Without this the split would accept a two-byte line-terminator row appearing
    anywhere among the closing parts and quietly reorder the record's tail. The
    terminator is the LAST thing in a record by definition, so a design declaring
    it earlier is either malformed or is something this parser has misread, and
    both deserve a refusal rather than a rearrangement.
    """
    if terminator is None:
        return
    last_part = closing_parts[-1]
    # Reads the SOURCE ROW only. An earlier version of this also compared ordinals,
    # which was the same mistake the envelope-order check carried: AEAT's ordinal is
    # a printed label, so ordering by it assumes a density the authority never
    # promised. It was harmless while every ordinal happened to be an integer and
    # would have failed the moment a ``bis`` label reached this comparison.
    if terminator.row <= last_part.row:
        raise RegistryValidationError(
            f"record-design sheet {sheet_name!r} declares an end-of-record terminator at row "
            f"{terminator.row} which does not follow its closing identifier at row "
            f"{last_part.row}; a terminator that is not last is not a terminator",
        )


def _relative_closing_parts(
    closing: RecordDesignRelativeSuffixMarker | RecordDesignCompositeRelativeClosing,
) -> tuple[RecordDesignRelativeSuffixMarker, ...]:
    if isinstance(closing, RecordDesignCompositeRelativeClosing):
        return closing.parts
    return (closing,)


def _require_contiguous_field_geometry(
    sheet_name: str,
    fields: list[RecordDesignField],
) -> int | None:
    """Return the exact terminal extent after proving source-order geometry."""
    expected_offset = 1
    for parsed_field in fields:
        if parsed_field.offset != expected_offset:
            defect = "an overlap" if parsed_field.offset < expected_offset else "a gap"
            raise RegistryValidationError(
                f"record-design sheet {sheet_name!r} has {defect} before field at row {parsed_field.row}: "
                f"expected offset {expected_offset}, got {parsed_field.offset}",
            )
        expected_offset = parsed_field.offset + parsed_field.length
    return expected_offset - 1 if fields else None


def _is_blank_row(values: tuple[object, ...]) -> bool:
    return all(value is None or str(value).strip() == "" for value in values)


def _probe_header_row(
    values: tuple[object, ...],
    row_number: int,
    *,
    label: str,
    sheet_name: str,
    header_corrections: _HeaderCorrectionIndex,
) -> tuple[_WorkbookHeader, RecordDesignHeaderCellCorrection | None] | None:
    """Match one row against either recognised AEAT record-design header shape.

    Shape A is AEAT's ordinary header: ``Posic.``/``Lon``/``Tipo``/``Descripcion``,
    with an optional ``Validacion`` and an optional ``Com.`` (complementary
    indicator) column. Its ``Lon`` cell may be recovered by a declared
    ``RecordDesignHeaderCellCorrection`` when AEAT's own publication leaves it
    blank -- applied ONLY when a sidecar names this exact
    ``(sheet, row, "length")``, never inferred from column position alone.

    Shape B is a second, real AEAT shape -- confirmed only in bundled Modelo
    151 annex sheets (``M15100000``, ``M15102000``-``M15109000``) -- carrying
    the same ``Posic.``/``Lon``/``Tipo`` columns, but where the description
    column holds the sheet's own topical caption instead of the literal word
    ``Descripcion``, sitting directly after a recognised ``Com.`` column, with
    no separate ``Validacion`` column at all. This is resilience to a real
    AEAT format variation, not a widened match: Shape B still requires the
    ``Com.`` token by its own recognised alias AND the very next column to
    carry real text, so a sheet lacking either -- no ``Descripcion`` token, no
    ``Com.`` token, or a blank column following ``Com.`` -- matches neither
    shape and still refuses.
    """
    if not _is_ordinal_header(_cell(values, 0)):
        return None
    try:
        offset_index = _required_header_index(values, "posic.")
        type_index = _required_header_index(values, "tipo")
    except ValueError as header_exc:
        _log.debug(
            "record-design header probe (%s): row %d missing required columns (%s); trying next",
            label,
            row_number,
            header_exc,
        )
        return None
    length_correction: RecordDesignHeaderCellCorrection | None = None
    try:
        length_index = _required_header_index(values, "lon")
    except ValueError:
        length_correction = header_corrections.get((sheet_name, row_number, "length"))
        if length_correction is None:
            _log.debug(
                "record-design header probe (%s): row %d missing required columns ('lon'); trying next",
                label,
                row_number,
            )
            return None
        length_index = length_correction.column_index
    complementary_index = _optional_header_index(values, "com", "comp")
    description_index = _optional_header_index(values, "descripcion")
    validation_index = _optional_header_index(values, "validacion", "oblig.")
    if description_index is None:
        if complementary_index is None:
            _log.debug(
                "record-design header probe (%s): row %d matches neither header shape "
                "(no 'descripcion' and no 'com'/'comp'); trying next",
                label,
                row_number,
            )
            return None
        caption_index = complementary_index + 1
        if caption_index >= len(values) or not coerce_cell_text(_cell(values, caption_index)):
            _log.debug(
                "record-design header probe (%s): row %d has 'com'/'comp' but no caption in the "
                "following column; trying next",
                label,
                row_number,
            )
            return None
        description_index = caption_index
        validation_index = None
    header = _WorkbookHeader(
        row_number=row_number,
        ordinal_index=0,
        offset_index=offset_index,
        length_index=length_index,
        type_index=type_index,
        complementary_index=complementary_index,
        description_index=description_index,
        validation_index=validation_index,
        content_index=_optional_header_index(values, "contenido"),
    )
    return header, length_correction


def _find_header(
    worksheet: Worksheet,
    header_corrections: _HeaderCorrectionIndex | None = None,
) -> tuple[_WorkbookHeader, RecordDesignHeaderCellCorrection | None]:
    sheet_name = worksheet.title.strip()
    for row_number, row in enumerate(worksheet.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        matched = _probe_header_row(
            tuple(row),
            row_number,
            label=f"xlsx {worksheet.title}",
            sheet_name=sheet_name,
            header_corrections=header_corrections or {},
        )
        if matched is not None:
            return matched
    raise RegistryValidationError(f"{worksheet.title!r} has no record-design header")


def _find_xls_header(
    worksheet: XlrdSheet,
    header_corrections: _HeaderCorrectionIndex | None = None,
) -> tuple[_WorkbookHeader, RecordDesignHeaderCellCorrection | None]:
    sheet_name = worksheet.name.strip()
    header_corrections = header_corrections or {}
    for rowx in range(min(10, worksheet.nrows)):
        matched = _probe_header_row(
            tuple(worksheet.row_values(rowx)),
            rowx + 1,
            label="xls",
            sheet_name=sheet_name,
            header_corrections=header_corrections,
        )
        if matched is not None:
            return matched
    raise RegistryValidationError(f"{worksheet.name!r} has no record-design header")


def _is_ordinal_header(value: object | None) -> bool:
    normalized = _normalise_header_cell(value)
    return normalized in {"no", "n"} or re.fullmatch(r"version \d+(?:\.\d+)*", normalized) is not None


def _cell(values: tuple[object, ...], index: int) -> object | None:
    return values[index] if index < len(values) else None


def _optional_text(value: object | None) -> str | None:
    cleaned = coerce_cell_text(value)
    return cleaned or None


def _ordinal_text(value: object | None) -> str | None:
    """Render a printed ordinal cell verbatim, including a non-numeric label.

    ``integral_floats_as_int=True`` because openpyxl hands back a whole-number
    ordinal cell as a ``float`` (``19.0``); rendering that as ``"19.0"`` would
    break BOTH the M390 auxiliary-header ordinal sequence (which compares
    against plain ``"1"``..``"13"``) and the dotted-component prefix match
    (``"19.1"``'s prefix must equal parent ``"19"``, not ``"19.0"``).
    """
    cleaned = coerce_cell_text(value, integral_floats_as_int=True)
    return cleaned or None


def _optional_header_text(values: tuple[object, ...], index: int | None) -> str | None:
    if index is None:
        return None
    return _optional_text(_cell(values, index))


def _required_text(value: object | None, sheet: str, row: int, field: str) -> str:
    cleaned = coerce_cell_text(value)
    if not cleaned:
        raise RegistryValidationError(f"{sheet!r} row {row} missing {field}")
    return cleaned


def _required_type_code(
    value: object | None,
    sheet: str,
    row: int,
    corrections: _TypeCorrectionIndex,
) -> tuple[str, RecordDesignFieldTypeCorrection | None]:
    """Return the row's ``Tipo`` value, applying a declared correction for a blank cell.

    A correction fires ONLY when the cell is genuinely blank AND a sidecar
    declares one for this exact ``(sheet, row)`` -- never as a fallback for a
    present-but-unreadable value, which stays a hard refusal exactly as
    before. This is the sole read path that can turn a "missing type" refusal
    into a read.
    """
    cleaned = coerce_cell_text(value)
    if cleaned:
        return cleaned, None
    correction = corrections.get((sheet, row))
    if correction is not None:
        return correction.corrected_type, correction
    raise RegistryValidationError(f"{sheet!r} row {row} missing type")


def _field_description_text(
    values: tuple[object, ...],
    *,
    header: _WorkbookHeader,
    content: str | None,
    sheet: str,
    row: int,
) -> str:
    description = _optional_text(_cell(values, header.description_index))
    if description is not None:
        return description
    if content is not None:
        return content
    raise RegistryValidationError(f"{sheet!r} row {row} missing description")


def _int_or_none(value: object | None) -> int | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return None


def _normalise_header_cell(value: object | None) -> str:
    return (
        coerce_cell_text(value)
        .casefold()
        .replace("º", "o")
        .replace("ó", "o")
        .replace("í", "i")
        .replace("á", "a")
        .replace("é", "e")
        .replace("ú", "u")
    )


def _required_header_index(values: tuple[object, ...], header_name: str) -> int:
    index = _optional_header_index(values, header_name)
    if index is None:
        raise RegistryValidationError(f"missing workbook header {header_name!r}")
    return index


def _header_token(value: str) -> str:
    """Return a header cell's identity, ignoring AEAT's abbreviating full stop.

    AEAT abbreviates header labels inconsistently WITHIN one workbook: Modelo
    115 writes ``Lon`` on its first sheet and ``Lon.`` on its second. Exact
    membership therefore matched one sheet and missed the other, and because a
    sheet failing header detection is skipped, the entire record body was
    dropped while the file looked healthy.

    The stop is typography, not identity, so it is ignored on BOTH sides rather
    than by enrolling each spelling separately -- which is what the accepted set
    already does ad hoc for ``posic.`` and ``oblig.``, and what would have to be
    remembered again for every future label.
    """
    return value.rstrip(".")


def _header_names_one_column(cell: str, expected: str) -> bool:
    """Whether a header cell names the column ``expected`` names.

    AEAT ABBREVIATES BY TRUNCATION, and inconsistently within a single workbook:
    the length column is ``Lon`` on one Modelo 115 sheet and ``Lon.`` on the next,
    and ``Long.`` throughout Modelo 100 and on Modelo 714's header sheet. Matching
    on truncation-compatibility rather than against an enrolled list of spellings
    is what stops the next abbreviation needing a code change -- and the enrolled
    pairs already in this module (``com``/``comp``, ``validacion``/``oblig.``) are
    that treadmill visible in progress.

    Measured across every bundled workbook design before shipping: 117 sources
    unchanged, 6 improved, ZERO degraded. Modelo 714's four editions go from a
    silently dropped sheet to a complete read, and Modelo 100's 2019 design from
    refusing outright to 41 sheets.

    The three-character floor stops a one- or two-letter cell prefix-matching its
    way onto a column it does not name; without it a stray ``N`` or ``No`` cell
    would claim whichever column happened to start with it.
    """
    if cell == expected:
        return True
    shorter, longer = sorted((cell, expected), key=len)
    return len(shorter) >= 3 and longer.startswith(shorter)


def _optional_header_index(values: tuple[object, ...], *header_names: str) -> int | None:
    expected = [_header_token(name) for name in header_names]
    for index, value in enumerate(values):
        cell = _header_token(_normalise_header_cell(value))
        if cell and any(_header_names_one_column(cell, candidate) for candidate in expected):
            return index
    return None


def _total_label_index(values: tuple[object, ...]) -> int | None:
    for index, value in enumerate(values):
        if _normalise_header_cell(value) in {"total", "total:"}:
            return index
    return None


def _positive_integer_after(values: tuple[object, ...], label_index: int) -> int | None:
    for candidate in values[label_index + 1 :]:
        total = _int_or_none(candidate)
        if total is not None and total > 0:
            return total
    return None


_COMPACT_PDF_ROW_RE = re.compile(
    r"^\s*(?P<ordinal>\d+)\s+(?P<offset>\d+)\s+(?P<length>\d+)\s+(?P<type>An|Num|N|A)\s+(?P<text>.+)$",
    re.IGNORECASE,
)
#: A PDF row declaring the physical end of record. Its DESCRIPTION half composes
#: the SHARED terminator phrase rather than carrying a second private spelling, so
#: the workbook splitter and this recogniser cannot drift on what a CRLF row is.
#: They already had: this one has known the terminator since it was written, while
#: the workbook path refused every design that declared one.
_COMPACT_PDF_CRLF_ROW_RE = re.compile(
    r"^\s*(?P<ordinal>\d+)\s+(?P<offset>\d+)\s+(?P<type>An|Num|N|A)\s+"
    rf"(?P<text>(?:{_RECORD_TERMINATOR_PHRASE}).*)$",
    re.IGNORECASE,
)
_NARRATIVE_PDF_ROW_RE = re.compile(
    r"^\s*(?P<start>\d+)(?:\s*[-\u2013]\s*(?P<end>\d+))?\s+"
    r"(?P<type>Alfanum[eé]rico|Alfab[eé]tico|Num[eé]rico|Alphanumeric|Alphabetic|Numeric|Blank|[-\u2013]+)\s*"
    r"(?P<text>.*)$",
    re.IGNORECASE,
)
_PDF_PAGE_RECORD_RE = re.compile(r"^P[áa]g\s+(?P<page>\d+)\s+DISE[ÑN]O DE REGISTRO\b", re.IGNORECASE)
_PDF_RECORD_HEADING_RE = re.compile(
    r"^(?:[A-Z]\.?\s*-?\s*)?(?:TIPO DE REGISTRO|Tipo de registro|RECORD TYPE|Record [Tt]ype)\s+"
    r"(?P<record>\d+)\s*:\s*(?P<title>.+)$",
    re.IGNORECASE,
)
#: English AEAT record-design translations also head a record with the reversed
#: word order "TYPE <n> RECORD: <title>" (observed in the modelo 604 ATF English
#: appendix), distinct from the "RECORD TYPE <n>: <title>" order above.
_PDF_RECORD_HEADING_REVERSED_RE = re.compile(
    r"^(?:[A-Z]\.?\s*-?\s*)?TYPE\s+(?P<record>\d+)\s+RECORD\s*:\s*(?P<title>.+)$",
    re.IGNORECASE,
)
#: A THIRD Spanish word order -- "REGISTRO DE TIPO <n>: <title>" -- which AEAT
#: uses at least as often as the two above: modelos 165, 180, 182, 184, 187, 188,
#: 193, 296 and 345 all head their perceptor/declarado record with it, several
#: separating number from title with a full stop rather than a colon.
#:
#: DELIBERATELY NOT A SPLITTING HEADING. The same phrase occurs inside ordinary
#: field prose ("Consignar lo contenido en estas mismas posiciones del registro
#: de tipo 1.", "... del registro de tipo 2. Registro de perceptor, toma el valor
#: 1) ..."), so treating a match as a record boundary on the text alone would
#: manufacture records that do not exist -- the opposite defect, and a worse one,
#: because an invented record inflates every coverage denominator derived from
#: the design. A match here only STAGES a name; whether a record actually starts
#: is decided by geometry, in :meth:`_PdfParseState._begins_a_new_record_body`.
_PDF_RECORD_HEADING_TYPE_LAST_RE = re.compile(
    r"^(?:[A-Z]\.?\s*-?\s*)?REGISTRO DE TIPO\s+(?P<record>\d+)\s*[:.]\s*(?P<title>\S.*)$",
    re.IGNORECASE,
)


@dataclass(slots=True)
class _PdfFieldDraft:
    sheet: str
    row: int
    ordinal: str | None
    offset: int
    length: int
    type_code: str
    description_parts: list[str] = field(default_factory=list)
    content_parts: list[str] = field(default_factory=list)

    def append_continuation(self, line: str) -> None:
        if not self.description_parts or (not self.content_parts and _looks_like_title_continuation(line)):
            self.description_parts.append(line)
            return
        self.content_parts.append(line)

    def finish(self) -> RecordDesignField:
        description = _join_pdf_parts(self.description_parts)
        if not description:
            raise RegistryValidationError(f"{self.sheet!r} PDF row {self.row} missing description")
        return RecordDesignField(
            sheet=self.sheet,
            row=self.row,
            ordinal=self.ordinal,
            offset=self.offset,
            length=self.length,
            type_code=self.type_code,
            complementary=None,
            description=description,
            validation=None,
            content=_join_pdf_parts(self.content_parts) or None,
        )


@dataclass(slots=True)
class _PdfSheetDraft:
    name: str
    fields: list[RecordDesignField] = field(default_factory=list)
    current: _PdfFieldDraft | None = None
    #: Whether the source NAMED this record body. ``False`` marks a body the
    #: geometry proved exists -- its positions restart at 1 -- whose heading the
    #: parser did not recognise, so its identity is unknown. Such a body is
    #: reported as a skipped sheet rather than returned, because handing back a
    #: record under a name nobody read would be an invented record identity.
    identified: bool = True
    #: The source row at which this body's first position was seen, used to name
    #: an unidentified body by where it is rather than by what it might be.
    opened_at_row: int | None = None

    def has_started(self) -> bool:
        """Whether any field of this record body has been seen yet."""
        return bool(self.fields) or self.current is not None

    def start_field(self, row: _PdfRow) -> None:
        self.finish_current()
        self.current = _PdfFieldDraft(
            sheet=self.name,
            row=row.source_row,
            ordinal=row.ordinal if row.ordinal is not None else str(len(self.fields) + 1),
            offset=row.offset,
            length=row.length,
            type_code=row.type_code,
            description_parts=[row.description] if row.description else [],
        )

    def finish_current(self) -> None:
        if self.current is None:
            return
        self.fields.append(self.current.finish())
        self.current = None

    def finish(self, *, source_label: str) -> RecordDesignSheet:
        self.finish_current()
        total_positions = max((field.offset + field.length - 1 for field in self.fields), default=None)
        sheet = RecordDesignSheet(name=self.name, fields=tuple(self.fields), total_positions=total_positions)
        _validate_pdf_sheet(sheet, source_label=source_label)
        return sheet


@dataclass(frozen=True, slots=True)
class _PdfRow:
    source_row: int
    ordinal: str | None
    offset: int
    length: int
    type_code: str
    description: str


@dataclass(frozen=True, slots=True)
class _PdfWord:
    text: str
    x0: float
    x1: float
    top: float
    bottom: float


@dataclass(frozen=True, slots=True)
class _PdfRect:
    x0: float
    x1: float
    top: float
    bottom: float
    width: float
    height: float
    fill: object | None


@dataclass(frozen=True, slots=True)
class _PdfPageSnapshot:
    lines: tuple[str, ...]
    words: tuple[_PdfWord, ...]
    rects: tuple[_PdfRect, ...]


@dataclass(frozen=True, slots=True)
class _VisualChartFragment:
    start: int
    end: int
    description: str


def _extract_pdf_text_lines(pdf_bytes: bytes, *, source_label: str) -> tuple[str, ...]:
    import pypdfium2 as pdfium

    try:
        document = pdfium.PdfDocument(pdf_bytes)
    except Exception as exc:  # pragma: no cover - pdfium parser surface
        raise RegistryValidationError(f"pypdfium2 could not open record-design PDF {source_label}: {exc}") from exc
    try:
        lines: list[str] = []
        for page in document:
            text_page = page.get_textpage()
            try:
                lines.extend(text_page.get_text_range().splitlines())
            finally:
                text_page.close()
                page.close()
        return tuple(lines)
    finally:
        document.close()


def _extract_pdfplumber_text_lines(pdf_bytes: bytes, *, source_label: str) -> tuple[str, ...]:
    import pdfplumber

    try:
        with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
            return tuple(line for page in pdf.pages for line in _extract_pdf_page_lines(page))
    except Exception as exc:  # pragma: no cover - defensive; pdfplumber surface
        raise RegistryValidationError(f"pdfplumber could not open record-design PDF {source_label}: {exc}") from exc


def _uses_page_record_layout(lines: tuple[str, ...]) -> bool:
    return any(_pdf_page_name(_clean_pdf_line(line)) is not None for line in lines)


def _snapshot_pdf_page(page: Page) -> _PdfPageSnapshot:
    return _PdfPageSnapshot(
        lines=_extract_pdf_page_lines(page),
        words=tuple(
            _PdfWord(
                text=str(word["text"]),
                x0=float(word["x0"]),
                x1=float(word["x1"]),
                top=float(word["top"]),
                bottom=float(word["bottom"]),
            )
            for word in page.extract_words()
        ),
        rects=tuple(
            _PdfRect(
                x0=float(rect["x0"]),
                x1=float(rect["x1"]),
                top=float(rect["top"]),
                bottom=float(rect["bottom"]),
                width=float(rect["width"]),
                height=float(rect["height"]),
                fill=rect.get("non_stroking_color"),
            )
            for rect in page.rects
        ),
    )


def _extract_pdf_page_lines(page: Page) -> tuple[str, ...]:
    text = page.extract_text() or ""
    return tuple(text.splitlines())


@dataclass(frozen=True, slots=True)
class _PdfSheetResult:
    """One finished record body and whether the source named it."""

    sheet: RecordDesignSheet
    identified: bool
    opened_at_row: int | None


def _unidentified_record_body_name(row_number: int) -> str:
    """Name an unnamed record body by WHERE it is, never by what it might be."""
    return f"<unidentified record body beginning at source row {row_number}>"


class _PdfParseState:
    """Mutable state for the PDF record-design line parser.

    Encapsulates the locals (``current`` draft sheet, ``in_table`` flag,
    ``pending_name`` carried across page-name boundaries, ``pending_record_name``
    staged by a candidate record heading) so the per-line dispatch can mutate
    them without threading out-parameters through every helper.
    """

    __slots__ = ("current", "in_table", "pending_name", "pending_record_name", "results", "source_label")

    def __init__(self, *, source_label: str) -> None:
        self.results: list[_PdfSheetResult] = []
        self.current: _PdfSheetDraft | None = None
        self.in_table: bool = False
        self.pending_name: str | None = None
        self.pending_record_name: str | None = None
        self.source_label = source_label

    def finalise(self) -> RecordDesignExtraction:
        """Return the parsed records, naming every one the parser did not read.

        Three things land in ``skipped`` rather than being dropped, and each was
        previously invisible:

        * a record heading the parser recognised but found no field rows under;
        * a record body whose existence geometry proves -- its positions restart
          at 1 -- but whose heading the parser did not recognise, so it has no
          identity to return it under;
        * both together.

        The rule is one sentence: A READ THAT RETURNS FEWER RECORDS THAN THE
        DOCUMENT CONTAINS MUST NEVER REPORT COMPLETE. Every one of these makes
        :attr:`RecordDesignExtraction.is_complete` false, so
        :meth:`RecordDesignExtraction.require_complete` -- the guard that exists
        precisely to catch an incomplete read -- can finally see them.
        """
        self._close_current_body()
        read = tuple(result.sheet for result in self.results if result.identified and result.sheet.fields)
        if not read:
            raise RegistryValidationError("record-design PDF did not contain parseable field rows")
        return RecordDesignExtraction(
            source=self.source_label,
            sheets=read,
            skipped=tuple(
                RecordDesignSkippedSheet(name=result.sheet.name, reason=_skipped_record_reason(result))
                for result in self.results
                if not (result.identified and result.sheet.fields)
            ),
        )

    def feed(self, line: str, row_number: int) -> None:
        if not line or _is_pdf_footer(line):
            return
        if self._consume_page_name(line):
            return
        if self._consume_record_heading(line):
            return
        self._stage_candidate_record_name(line)
        if self._consume_table_header(line):
            return
        if self._consume_title_continuation(line):
            return
        if _is_pdf_page_heading(line):
            return
        if self._consume_field_row(line, row_number):
            return
        self._consume_field_continuation(line)

    def _close_current_body(self) -> None:
        if self.current is None:
            return
        self.results.append(
            _PdfSheetResult(
                sheet=self.current.finish(source_label=self.source_label),
                identified=self.current.identified,
                opened_at_row=self.current.opened_at_row,
            ),
        )
        self.current = None

    def _open_body(self, name: str, *, identified: bool = True) -> None:
        self._close_current_body()
        self.current = _PdfSheetDraft(name, identified=identified)
        self.pending_record_name = None

    def _consume_page_name(self, line: str) -> bool:
        page_name = _pdf_page_name(line)
        if page_name is None:
            return False
        self.pending_name = page_name
        if self.current is not None and self.current.name != page_name:
            self._open_body(page_name)
        return True

    def _consume_record_heading(self, line: str) -> bool:
        heading_name = _pdf_record_heading_name(line)
        if heading_name is None:
            return False
        self._open_body(heading_name)
        self.in_table = False
        return True

    def _consume_table_header(self, line: str) -> bool:
        if not _is_pdf_header(line):
            return False
        if self.current is None:
            self.current = _PdfSheetDraft(self.pending_name or "PDF record design")
        self.in_table = True
        return True

    def _consume_title_continuation(self, line: str) -> bool:
        if self.in_table or self.current is None or self.current.fields:
            return False
        if not _looks_like_title_continuation(line):
            return False
        self.current.name = _normalise_pdf_sheet_name(_join_pdf_parts([self.current.name, line]))
        return True

    def _stage_candidate_record_name(self, line: str) -> None:
        """Remember a candidate record name WITHOUT acting on it.

        Staged rather than consumed on both counts: the line stays in the
        parser's ordinary pipeline exactly as before (so no field description
        loses text it used to carry), and no record boundary is created from
        the text. Only :meth:`_begins_a_new_record_body`, reading position
        geometry, decides a record actually starts -- at which point this name
        is used if one was staged since the last field row, and discarded
        otherwise. A candidate matched inside field prose is therefore inert.
        """
        candidate = _pdf_candidate_record_name(line)
        if candidate is not None:
            self.pending_record_name = candidate

    def _begins_a_new_record_body(self, row: _PdfRow) -> bool:
        """Whether ``row`` starts a record body distinct from the one being read.

        POSITION 1 OCCURS EXACTLY ONCE PER RECORD. A fixed-width record is
        contiguous from its first byte, so a row declaring position 1 while the
        body under construction already holds fields is not a continuation of
        that body under any reading -- it is the next record. This is geometry
        AEAT itself declares, not a text heuristic, so it holds for every
        heading spelling, every word order and every design that heads its
        records with no recognisable line at all.

        Modelo 180 is the worked case: AEAT heads its perceptor record
        ``REGISTRO DE TIPO 2: REGISTRO DE PERCEPTOR.`` -- a word order the
        heading recogniser did not know -- so seventeen perceptor positions
        were appended to the declarante record, the extraction returned ONE
        sheet for a two-record document, and it reported itself complete.
        """
        return self.current is not None and row.offset == 1 and self.current.has_started()

    def _consume_field_row(self, line: str, row_number: int) -> bool:
        row = _parse_pdf_row(line, row_number)
        if row is None:
            return False
        if self.current is None:
            self.current = _PdfSheetDraft(self.pending_name or "PDF record design")
        elif self._begins_a_new_record_body(row):
            staged = self.pending_record_name
            self._open_body(
                staged if staged is not None else _unidentified_record_body_name(row_number),
                identified=staged is not None,
            )
        if self.current.opened_at_row is None:
            self.current.opened_at_row = row_number
        self.current.start_field(row)
        self.in_table = True
        self.pending_record_name = None
        return True

    def _consume_field_continuation(self, line: str) -> None:
        if self.in_table and self.current is not None and self.current.current is not None:
            self.current.current.append_continuation(line)


def _skipped_record_reason(result: _PdfSheetResult) -> str:
    if not result.sheet.fields:
        return "record heading recognised but no field rows parsed under it"
    return (
        f"a distinct record body begins here -- its positions restart at 1 at source row "
        f"{result.opened_at_row} -- but the source's heading for it was not recognised, so this "
        f"record has no read identity. It is reported unread rather than merged into the record "
        f"above it, because merging understates the document by a whole record while still "
        f"reporting the read complete"
    )


def _extract_pdf_lines(lines: tuple[str, ...], *, source_label: str) -> RecordDesignExtraction:
    state = _PdfParseState(source_label=source_label)
    for row_number, raw_line in enumerate(lines, start=1):
        state.feed(_clean_pdf_line(raw_line), row_number)
    return state.finalise()


def _validate_pdf_sheet(sheet: RecordDesignSheet, *, source_label: str) -> None:
    if not sheet.fields:
        return
    first_field = sheet.fields[0]
    if first_field.offset != 1:
        raise RegistryValidationError(
            f"{source_label} {sheet.name!r} first field starts at position {first_field.offset}; expected 1",
        )
    for parsed_field in sheet.fields:
        if parsed_field.offset < 1:
            raise RegistryValidationError(
                f"{source_label} {sheet.name!r} field ordinal {parsed_field.ordinal} has invalid "
                f"position {parsed_field.offset}",
            )
        if parsed_field.length < 1:
            raise RegistryValidationError(
                f"{source_label} {sheet.name!r} field ordinal {parsed_field.ordinal} has invalid "
                f"length {parsed_field.length}",
            )
    terminal_position = max(parsed_field.offset + parsed_field.length - 1 for parsed_field in sheet.fields)
    if sheet.total_positions is not None and terminal_position != sheet.total_positions:
        raise RegistryValidationError(
            f"{source_label} {sheet.name!r} declares {sheet.total_positions} total positions "
            f"but parsed fields fill {terminal_position}",
        )


def _parse_pdf_row(line: str, source_row: int) -> _PdfRow | None:
    compact = _COMPACT_PDF_ROW_RE.match(line)
    if compact is not None:
        return _PdfRow(
            source_row=source_row,
            ordinal=compact.group("ordinal"),
            offset=int(compact.group("offset")),
            length=int(compact.group("length")),
            type_code=compact.group("type"),
            description=compact.group("text").strip(),
        )

    crlf = _COMPACT_PDF_CRLF_ROW_RE.match(line)
    if crlf is not None:
        return _PdfRow(
            source_row=source_row,
            ordinal=crlf.group("ordinal"),
            offset=int(crlf.group("offset")),
            length=2,
            type_code=crlf.group("type"),
            description=crlf.group("text").strip(),
        )

    narrative = _NARRATIVE_PDF_ROW_RE.match(line)
    if narrative is None:
        return None

    start = int(narrative.group("start"))
    end_group = narrative.group("end")
    end = int(end_group) if end_group is not None else start
    if end < start:
        raise RegistryValidationError(f"PDF row {source_row} has inverted position range {start}-{end}")
    return _PdfRow(
        source_row=source_row,
        ordinal=None,
        offset=start,
        length=end - start + 1,
        type_code=_normalise_pdf_type_code(narrative.group("type")),
        description=narrative.group("text").strip(),
    )


def _normalise_pdf_type_code(value: str) -> str:
    normalised = value.strip(" .").lower()
    if set(normalised) <= {"-", "\u2013"} or normalised == "blank":
        return "Blancos"
    if normalised.startswith(("alfanum", "alphanum")):
        return "Alfanumérico"
    if normalised.startswith(("alfab", "alphab")):
        return "Alfabético"
    if normalised.startswith("num"):
        return "Numérico"
    return value.strip()


def _pdf_page_name(line: str) -> str | None:
    match = _PDF_PAGE_RECORD_RE.match(line)
    if match is None:
        return None
    return f"Pág. {match.group('page')}"


def _pdf_record_heading_name(line: str) -> str | None:
    match = _PDF_RECORD_HEADING_RE.match(line) or _PDF_RECORD_HEADING_REVERSED_RE.match(line)
    if match is None:
        return None
    title = _normalise_pdf_sheet_name(match.group("title"))
    return f"Tipo {match.group('record')} - {title}"


def _pdf_candidate_record_name(line: str) -> str | None:
    """Return the record name a line MIGHT be heading, for geometry to confirm.

    Separate from :func:`_pdf_record_heading_name` precisely because it is not
    trusted on its own: see :data:`_PDF_RECORD_HEADING_TYPE_LAST_RE` for why
    this word order cannot split a record on the text alone.
    """
    match = _PDF_RECORD_HEADING_TYPE_LAST_RE.match(line)
    if match is None:
        return None
    title = _normalise_pdf_sheet_name(match.group("title"))
    return f"Tipo {match.group('record')} - {title}"


# The column-header spellings AEAT prints across diseño-de-registro PDFs.
# Each variant is a list of required column groups; a group holds the
# interchangeable spellings of one column, so a line matches a variant when
# every group contributes at least one token.
_PDF_HEADER_VARIANTS: Final[tuple[tuple[tuple[str, ...], ...], ...]] = (
    (("POSICIONES", "POSICIÓN"), ("NATURALEZA",), ("DESCRIPCI",)),
    (("Nº POSIC",), ("LON",), ("TIPO",), ("DESCRIPCI",)),
    (("POSITIONS",), ("NATURE",), ("DESCRIPTION",)),
)


def _is_pdf_header(line: str) -> bool:
    normalised = line.upper()
    return any(
        all(any(token in normalised for token in column) for column in variant) for variant in _PDF_HEADER_VARIANTS
    )


def _is_pdf_footer(line: str) -> bool:
    return bool(
        re.match(r"^P[áa]gina\s+\d+\s+de\s+\d+$", line, re.IGNORECASE)
        or re.match(r"^Ejercicio\s+\d{4}(?:\s+\d+)?$", line, re.IGNORECASE)
        or re.match(r"^\d+$", line),
    )


def _is_pdf_page_heading(line: str) -> bool:
    return bool(
        line.startswith("Modelo ")
        or line.startswith("Agencia Tributaria")
        or line.startswith("Declaración Informativa")
        or line.startswith("Declaración informativa")
        or line.startswith("determinados ")
        or line.startswith("determinadas ")
        or line == "Resumen anual"
        or line == "MODELO 193"
        or line == "MODELO 190"
        or line == "DISEÑOS DE REGISTRO",
    )


def _looks_like_title_continuation(line: str) -> bool:
    letters = [char for char in line if char.isalpha()]
    if not letters:
        return False
    return not any(char.islower() for char in letters)


def _clean_pdf_line(line: str) -> str:
    return " ".join(line.strip().split())


def _join_pdf_parts(parts: list[str]) -> str:
    return " ".join(part.strip() for part in parts if part.strip())


def _normalise_pdf_sheet_name(value: str) -> str:
    return _join_pdf_parts([value.replace(".", " ").strip()]).strip(". ").title()


def _extract_visual_record_design_chart(
    pages: tuple[_PdfPageSnapshot, ...],
    *,
    source_label: str,
) -> tuple[RecordDesignSheet, ...]:
    pages_by_sheet: dict[str, list[_PdfPageSnapshot]] = {}
    for page in pages:
        sheet_name = _visual_chart_page_sheet_name(page)
        if sheet_name is not None:
            pages_by_sheet.setdefault(sheet_name, []).append(page)
    if not pages_by_sheet:
        return ()

    sheets = tuple(
        _extract_visual_chart_sheet(sheet_name, tuple(sheet_pages), source_label=source_label)
        for sheet_name, sheet_pages in pages_by_sheet.items()
    )
    return sheets if all(sheet.fields for sheet in sheets) else ()


def _visual_chart_page_sheet_name(page: _PdfPageSnapshot) -> str | None:
    for line in page.lines:
        match = _VISUAL_CHART_HEADER_RE.match(_clean_pdf_line(line))
        if match is not None:
            title = _normalise_pdf_sheet_name(match.group("title"))
            return f"Tipo {match.group('record')} - {title}"
    return None


def _extract_visual_chart_sheet(
    name: str,
    pages: tuple[_PdfPageSnapshot, ...],
    *,
    source_label: str,
) -> RecordDesignSheet:
    fragments: list[_VisualChartFragment] = []
    for page in pages:
        fragments.extend(_extract_visual_chart_fragments(page))

    merged = _merge_visual_chart_fragments(sorted(fragments, key=lambda fragment: fragment.start))
    fields = tuple(
        RecordDesignField(
            sheet=name,
            row=ordinal,
            ordinal=str(ordinal),
            offset=fragment.start,
            length=fragment.end - fragment.start + 1,
            type_code=_VISUAL_CHART_TYPE_CODE,
            description=fragment.description,
            content="Extracted from visual record-design chart geometry.",
        )
        for ordinal, fragment in enumerate(merged, start=1)
    )
    total_positions = max((field.offset + field.length - 1 for field in fields), default=None)
    sheet = RecordDesignSheet(name=name, fields=fields, total_positions=total_positions)
    _validate_pdf_sheet(sheet, source_label=source_label)
    return sheet


def _extract_visual_chart_fragments(page: _PdfPageSnapshot) -> list[_VisualChartFragment]:
    grid = _visual_chart_grid(page)
    if grid is None:
        return []
    left, cell_width, horizontal_rules = grid
    fragments: list[_VisualChartFragment] = []
    number_rows = _visual_chart_number_rows(page)
    for index, (number_top, first_position) in enumerate(number_rows):
        row_rules = _visual_chart_rules_for_number_row(horizontal_rules, number_top)
        if not row_rules:
            continue
        region_top = number_rows[index - 1][0] + 8 if index else 20
        for rule in row_rules:
            start = first_position - 1 + round((rule.x0 - left) / cell_width) + 1
            end = first_position - 1 + round((rule.x1 - left) / cell_width)
            if start > end:
                continue
            fragments.append(
                _VisualChartFragment(
                    start=start,
                    end=end,
                    description=_visual_chart_description(page, rule, region_top=region_top),
                ),
            )
    return fragments


def _is_black_fill(fill: object | None) -> bool:
    """Return whether a pdfplumber rect ``fill`` value is opaque black.

    pdfplumber reports a rect's ``non_stroking_color`` in whatever colour
    space the source PDF declared: a bare grayscale float (``0.0`` is
    black), an RGB triple (``(0.0, 0.0, 0.0)``), or a named/indexed
    colourspace string the geometry-only chart extractor does not attempt
    to resolve. Both numeric shapes are checked so a rule ruled in RGB
    (observed in the AEAT modelo 038 record-design PDF) is not silently
    excluded from the visual-chart grid the way a bare ``fill == 0.0``
    comparison would exclude it.
    """
    if isinstance(fill, int | float):
        return fill == 0.0
    if isinstance(fill, tuple):
        try:
            components = _NUMERIC_TUPLE_ADAPTER.validate_python(fill)
        except ValidationError:
            return False
        return bool(components) and all(component == 0.0 for component in components)
    return False


def _visual_chart_grid(page: _PdfPageSnapshot) -> tuple[float, float, tuple[_PdfRect, ...]] | None:
    horizontal_rules = tuple(
        rect for rect in page.rects if _is_black_fill(rect.fill) and rect.height <= 2.0 and rect.width >= 8.0
    )
    full_width_rules = tuple(rect for rect in horizontal_rules if rect.width > 700.0)
    if not full_width_rules:
        return None
    left = min(rect.x0 for rect in full_width_rules)
    right = max(rect.x1 for rect in full_width_rules)
    return left, (right - left) / 65, horizontal_rules


def _visual_chart_number_rows(page: _PdfPageSnapshot) -> list[tuple[float, int]]:
    grouped_words: dict[float, list[_PdfWord]] = {}
    for word in page.words:
        if _visual_chart_number_values(word.text):
            grouped_words.setdefault(round(word.top, 1), []).append(word)

    rows: list[tuple[float, int]] = []
    for top, words in grouped_words.items():
        values = [
            value
            for word in sorted(words, key=lambda current: current.x0)
            for value in _visual_chart_number_values(word.text)
        ]
        if len(values) >= 20 and max(values) - min(values) >= 30:
            rows.append((top, min(values)))
    return sorted(rows)


def _visual_chart_number_values(text: str) -> tuple[int, ...]:
    if not text.isdigit():
        return ()
    if len(text) <= 3:
        return (int(text),)
    if len(text) % 3 == 0:
        return tuple(int(text[index : index + 3]) for index in range(0, len(text), 3))
    return ()


def _visual_chart_rules_for_number_row(
    horizontal_rules: tuple[_PdfRect, ...],
    number_top: float,
) -> tuple[_PdfRect, ...]:
    grouped_rules: dict[float, list[_PdfRect]] = {}
    for rule in horizontal_rules:
        if 0 < number_top - rule.top <= 30:
            grouped_rules.setdefault(round(rule.top, 1), []).append(rule)
    if not grouped_rules:
        return ()
    rule_top = max(grouped_rules)
    return tuple(sorted(grouped_rules[rule_top], key=lambda rule: rule.x0))


def _visual_chart_description(
    page: _PdfPageSnapshot,
    rule: _PdfRect,
    *,
    region_top: float,
) -> str:
    words = [
        word
        for word in page.words
        if rule.x0 - 2 <= (word.x0 + word.x1) / 2 <= rule.x1 + 2
        and region_top <= word.top <= rule.top - 1
        and not _is_visual_chart_number_text(word.text)
    ]
    description = _normalise_visual_chart_description(words)
    return description or "BLANCOS."


def _normalise_visual_chart_description(words: list[_PdfWord]) -> str:
    tokens = [word.text for word in sorted(words, key=lambda word: (word.top, word.x0))]
    tokens = [
        _REVERSED_VISUAL_CHART_TOKENS.get(visual_word, visual_word) for visual_word in tokens if visual_word != "D"
    ]
    if not tokens:
        return ""
    if any(token.strip(".").upper() in _REVERSED_VISUAL_CHART_WORDS for token in tokens):
        tokens = [token[::-1] for token in reversed(tokens)]
    return _clean_visual_chart_description(_dedupe_visual_chart_tokens(tokens))


def _dedupe_visual_chart_tokens(tokens: list[str]) -> str:
    deduped: list[str] = []
    for token in tokens:
        if not deduped or deduped[-1] != token:
            deduped.append(token)
    return " ".join(deduped).strip()


def _clean_visual_chart_description(description: str) -> str:
    replacements = {
        "DEL DECLARANTE N.I.F. DEL DECLARANTE": "N.I.F. DEL DECLARANTE",
        "DECLARANTE N.I.F. DECLARANTE": "N.I.F. DECLARANTE",
        "PROVINCIA AICNIVORP OGIDOC": "CODIGO PROVINCIA",
        "CODIGO PAIS SIAP": "CODIGO PAIS",
        "DECIMAL ED": "DECIMAL",
        "I TIPO DE HOJA": "TIPO DE HOJA",
        "REFERENCIA CATASTRAL REFERENCIA CATASTRAL": "REFERENCIA CATASTRAL",
    }
    for before, after in replacements.items():
        description = description.replace(before, after)
    return description


def _is_visual_chart_number_text(text: str) -> bool:
    return bool(re.fullmatch(r"\d+", text) or re.fullmatch(r"\d{3}(?:\d{3})+", text))


def _merge_visual_chart_fragments(
    fragments: list[_VisualChartFragment],
) -> tuple[_VisualChartFragment, ...]:
    merged: list[_VisualChartFragment] = []
    for fragment in fragments:
        if (
            merged
            and fragment.start == merged[-1].end + 1
            and merged[-1].end % 65 == 0
            and _visual_chart_fragments_should_merge(merged[-1], fragment)
        ):
            previous = merged[-1]
            description = _merge_visual_chart_descriptions(previous.description, fragment.description)
            merged[-1] = _VisualChartFragment(start=previous.start, end=fragment.end, description=description)
            continue
        merged.append(fragment)
    return tuple(merged)


def _visual_chart_fragments_should_merge(
    previous: _VisualChartFragment,
    current: _VisualChartFragment,
) -> bool:
    return not (previous.description == "BLANCOS." and current.description != "BLANCOS.")


def _merge_visual_chart_descriptions(previous: str, current: str) -> str:
    parts = [description for description in (previous, current) if description != "BLANCOS."]
    return _clean_visual_chart_description(_join_pdf_parts(parts)) or "BLANCOS."


_VISUAL_CHART_HEADER_RE = re.compile(
    r"^MODELO\s+\d+\s+REGISTRO DE TIPO\s+(?P<record>\d+)\.?\s+(?P<title>REGISTRO DE .+)$",
    re.IGNORECASE,
)
_VISUAL_CHART_TYPE_CODE = "No consta en gráfico"
_REVERSED_VISUAL_CHART_WORDS = {
    "AICNIVORP",
    "AJOH",
    "DNERRA",
    "EVALC",
    "ETROPOS",
    "LACOL",
    "LAMICED",
    "NÓICAREPO",
    "OPIT",
    "ORTSIGER",
}
_REVERSED_VISUAL_CHART_TOKENS = {
    "AIRATNEMELPMOC.CED": "DEC. COMPLEMENTARIA",
    "AVITUTITSUS.CED": "DEC. SUSTITUTIVA",
    ".REPO": "OPER.",
    "ORUGES": "SEGURO",
    "ELBEUMNI": "INMUEBLE",
    ".CAUTIS": "SITUAC.",
    "ARELACSE": "ESCALERA",
    "IMNUEBLE": "INMUEBLE",
}


__all__ = [
    "DerivedDisenoCasilla",
    "DisenoCoverageReport",
    "RecordDesignCompositeRelativeClosing",
    "RecordDesignCorrection",
    "RecordDesignExtraction",
    "RecordDesignField",
    "RecordDesignFieldTypeCorrection",
    "RecordDesignHeaderCellCorrection",
    "RecordDesignRelativeSuffixMarker",
    "RecordDesignSheet",
    "RecordDesignSkippedSheet",
    "RecordDesignVariableBodyMarker",
    "RecordDesignVariableEnvelope",
    "RecordDesignVariableTotalMarker",
    "build_diseno_coverage_report",
    "calculation_closure_casilla_ids",
    "calculation_closure_legal_refs",
    "calculation_closure_record_design_metadata",
    "derive_calculation_completeness_casillas",
    "derive_diseno_coverage_casillas",
    "extract_record_design",
    "extract_record_design_pdf",
    "extract_record_design_pdf_bytes",
    "extract_record_design_workbook",
]
