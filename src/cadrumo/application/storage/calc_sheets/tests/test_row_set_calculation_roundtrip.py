"""Real local worksheet → pull → M720 calculation encrypted-roundtrip proof."""

from __future__ import annotations

from datetime import UTC, date, datetime
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell

from .....adapters.outbound.google.calc_sheets_pull import _decode_row_set_block
from .....adapters.outbound.google.calc_sheets_pull_records import RowSetEdit
from .....adapters.persistence.profile.invoices import InvoiceCatalogueRepository
from .....adapters.persistence.profile.modelos_calculation import CalculationRevisionCatalogueRepository
from .....adapters.persistence.profile.modelos_work_units import WorkUnitCatalogueRepository
from .....adapters.persistence.profile.transactions import TransactionCatalogueRepository
from .....core.aggregation import BindingSourceKind
from .....core.hashing import content_hash_hex
from .....core.period import Period
from .....domain.calculations.registry.authority import bundled_authority
from .....domain.calculations.registry.detail_record_bindings import Modelo720RowObservation
from .....domain.user_profile.values import ProfileSetupState, UserProfileFact, UserProfileRecord
from .....tests.profile_capsule import seed_test_profile_record
from .....tests.secure_sql import isolated_runtime_profile
from ....modelo.calculation_actions import calculate_modelo_revision_from_bucket_aggregation_with_diagnostics
from ....modelo.work_lifecycle import create_work_unit
from ..row_set_assembly import assemble_row_sets_for_snapshot
from .._styling import StyleRole, _data_tab_ranges
from ..engine import build_export_plan
from ..records import TabName
from ..workbook_export import serialize_offline_workbook

pytestmark = [pytest.mark.unit, pytest.mark.hex_application]

_BUCKET_ID = "72000000-0000-4000-8000-000000000720"
_T0 = datetime(2026, 1, 10, 10, 0, tzinfo=UTC)
_T1 = datetime(2026, 1, 10, 11, 0, tzinfo=UTC)


def _seed_ready_profile() -> None:
    seed_test_profile_record(
        UserProfileRecord(
            setup_state=ProfileSetupState.COMPLETE,
            profile_id=_BUCKET_ID,
            facts=(
                UserProfileFact(path="identity.tax_id", value="12345678Z"),
                UserProfileFact(path="identity.name", value="Ready"),
                UserProfileFact(path="identity.surnames", value="Operator"),
                UserProfileFact(path="activities.description", value="worksheet source"),
                UserProfileFact(path="tax_residence.ccaa", value="madrid"),
                UserProfileFact(path="tax_residence.jurisdiction_scope", value="common_regime"),
                UserProfileFact(path="iva.regime", value="GENERAL"),
                UserProfileFact(path="iva.m303_regime_composition", value="general"),
                UserProfileFact(path="iva.redeme_enrolled", value="false"),
                UserProfileFact(path="iva.cash_accounting_regime_enrolled", value="false"),
                UserProfileFact(path="iva.voluntary_sii_enrolled", value="false"),
                UserProfileFact(path="iva.hydrocarbon_deposit_advance_payment_deduction_entitled", value="false"),
                UserProfileFact(path="taxpayer_type.entity_type", value="natural_person"),
                UserProfileFact(path="taxpayer_type.irpf_income_categories", value="actividad_economica"),
                UserProfileFact(path="irpf.estimation_regime", value="directa_normal"),
                UserProfileFact(path="censo.activity_start_date", value=date(2020, 1, 1)),
            ),
            created_at=_T0,
            updated_at=_T0,
        )
    )


def test_empty_calculation_tab_omits_invalid_data_body_style_range() -> None:
    """Empty M720 calculation layouts still export a valid header-only tab."""
    ranges = _data_tab_ranges(
        tab=TabName.CALCULOS,
        last_row=1,
        value_role=StyleRole.COMPUTED,
        section_headers=(),
    )
    assert len(ranges) == 1
    assert ranges[0].start_row == ranges[0].end_row == 1


def test_real_worksheet_pull_calculation_roundtrips_m720_row_source_through_encrypted_revision(tmp_path: Path) -> None:
    """A real XLSX and canonical pull decoder retain every row-identity axis."""
    snapshot = bundled_authority().snapshot("720", filing_year=2025, period="0A")
    plan = build_export_plan(snapshot)
    row_set = next(row_set for row_set in plan.row_sets if row_set.grouping == "per_foreign_asset")
    values = {
        "modelo-720-asset-row-class": "C",
        "modelo-720-asset-row-country": "CH",
        "modelo-720-asset-row-currency": "CHF",
        "modelo-720-asset-row-identifier": "worksheet-canary-001",
        "modelo-720-asset-row-acquisition-date": "2020-01-15",
        "modelo-720-asset-row-valuation": "120000.00",
    }
    workbook = load_workbook(BytesIO(serialize_offline_workbook(plan)))
    worksheet = workbook[row_set.tab.value]
    for column in row_set.columns:
        cell = worksheet.cell(row=row_set.first_data_row, column=column.header_address.column)
        assert isinstance(cell, Cell)
        cell.value = values[str(column.binding)]
    rows: list[list[object]] = []
    for offset in range(50):
        row: list[object] = []
        for column in range(1, len(row_set.columns) + 1):
            cell = worksheet.cell(row=row_set.first_data_row + offset, column=column)
            assert isinstance(cell, Cell)
            row.append(cell.value)
        rows.append(row)
    cells, cells_read = _decode_row_set_block(rows, row_set)
    assert cells_read == len(values), "real workbook pull must expose every exported row binding"
    assembled = assemble_row_sets_for_snapshot((RowSetEdit(grouping=row_set.grouping, cells=cells),), snapshot)
    source_kind, observations = assembled[0]
    assert source_kind == "foreign_asset"
    assert len(observations) == 1
    typed_observations: list[Modelo720RowObservation] = []
    for candidate in observations:
        assert isinstance(candidate, Modelo720RowObservation)
        typed_observations.append(candidate)
    observation = typed_observations[0]
    expected_fingerprint = content_hash_hex(observation.model_dump(mode="json"))

    with isolated_runtime_profile(tmp_path=tmp_path, bucket_id=_BUCKET_ID) as profile:
        _seed_ready_profile()
        objects = profile.repository
        work_units = WorkUnitCatalogueRepository(objects=objects)
        calculations = CalculationRevisionCatalogueRepository(objects=objects)
        transactions = TransactionCatalogueRepository(bucket_id=_BUCKET_ID, objects=objects)
        invoices = InvoiceCatalogueRepository(objects=objects)
        work_unit = create_work_unit(
            bucket_id=_BUCKET_ID,
            modelo="720",
            filing_year=2025,
            period=Period.from_year_and_code(2025, "0A"),
            revision_id="2013-y-siguientes",
            repository=work_units,
            clock=_T0,
        )
        calculated = calculate_modelo_revision_from_bucket_aggregation_with_diagnostics(
            work_unit.work_unit_id,
            work_unit_repository=work_units,
            calculation_repository=calculations,
            transaction_repository=transactions,
            invoice_repository=invoices,
            foreign_asset_row_observations=tuple(typed_observations),
            clock=_T1,
        ).revision
        stored = calculations.load().get(calculated.calculation_revision_id)

    assert stored == calculated
    assert stored is not None
    key = ("modelo-720-asset-row-valuation", 1)
    identity = stored.row_source_identities[key]
    assert identity.source_kind is BindingSourceKind.FOREIGN_ASSET
    assert identity.source_row_identity == "detalle:per_foreign_asset:row-1"
    assert identity.row_set_grouping == "per_foreign_asset"
    assert identity.fingerprint == expected_fingerprint
    assert stored.row_binding_values["modelo-720-asset-row-valuation"] == {"1": "120000"}
    assert stored.source_provenance[0].source_ref == "worksheet:detalle:per_foreign_asset:row-1"
    assert stored.source_provenance[0].fingerprint == expected_fingerprint
