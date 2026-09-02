"""Ledger evidence reaches the workbook's ``Evidencia`` tab and JSON sidecar.

Three things are proven here against real objects -- the real bundled registry
tree, a real :class:`LedgerFilingEvidence` bundle, the real engine, and the real
offline materialiser:

* :func:`sheet_evidence_from_ledger_filing` projects a contributor-oriented
  ledger bundle into the casilla-oriented workbook facet, and refuses an
  unattributed contributor rather than guessing its casilla.
* :func:`build_export_plan` threads a supplied bundle through that projection,
  so the plan a renderer receives carries the evidence.
* The materialised workbook and its sidecar actually show the contributing
  row -- amount, IVA rate, counterparty -- which is the whole point of the tab:
  an operator, an asesor, or AEAT in a comprobacion opens it to see why a
  casilla holds the number it holds.

The defect case is the last test: a plan built WITHOUT a bundle yields an empty
facet, so a regression that stops threading the evidence stays visible as an
empty ``Evidencia`` tab rather than as a silently passing export.
"""

from __future__ import annotations

import json
from datetime import UTC, date, datetime
from decimal import Decimal
from functools import cache
from io import BytesIO

import pytest
from openpyxl import load_workbook

from .....core.authority_grade import RegistryAuthorityGrade
from .....core.casilla_id import CasillaId
from .....core.resources.bundled_data import bundled_path
from .....domain.calculations.registry.schema import RegistrySnapshot
from .....domain.modelos.ledger_filing_snapshot import (
    LedgerEvidenceRow,
    LedgerFilingEvidence,
    ManualFactBasisEntry,
)
from .....tests.registry_snapshot import build_snapshot
from .....tests.registry_tree import bundled_registry_tree
from ..engine import build_export_plan
from ..errors import CalcSheetsEngineError
from ..evidence import sheet_evidence_from_ledger_filing
from ..records import TabName
from ..workbook_export import serialize_offline_export

pytestmark = [pytest.mark.unit, pytest.mark.hex_application]

_CONTRIBUTOR_ID = "a" * 64
_UNATTRIBUTED_CONTRIBUTOR_ID = "b" * 64
_SNAPSHOT_FINGERPRINT = "c" * 64
_COUNTERPARTY = "Suministros Iberia SL"


@cache
def _snapshot() -> RegistrySnapshot:
    """The bundled modelo 303 calculation snapshot every test here exports."""
    modelos, catalogues = bundled_registry_tree()
    modelo = next(candidate for candidate in modelos if candidate.id == "303")
    return build_snapshot(
        modelo,
        catalogues,
        source_root=bundled_path(),
        filing_year=2025,
        period="1T",
        on=date(2025, 4, 1),
        grade=RegistryAuthorityGrade.CALCULATION,
    )


def _casilla_ids() -> tuple[CasillaId, CasillaId, CasillaId]:
    """Three real casilla ids from the bundled modelo 303 revision."""
    ids = tuple(casilla.id for casilla in _snapshot().revision.casillas)
    return (ids[0], ids[1], ids[2])


def _evidence_row(transaction_id: str) -> LedgerEvidenceRow:
    return LedgerEvidenceRow(
        transaction_id=transaction_id,
        fingerprint=_SNAPSHOT_FINGERPRINT,
        booked_date="2025-02-14",
        amount=Decimal("121.00"),
        currency="EUR",
        direction="outflow",
        business_classification="business",
        taxable_base=Decimal("100.00"),
        iva_rate=Decimal("0.21"),
        iva_amount=Decimal("21.00"),
        lifecycle_state="verified",
        counterparty=_COUNTERPARTY,
        attachment_ids=("attachment-1",),
        document_link_ids=("drive-doc-1",),
        legal_refs=("ley-37-1992:art-99",),
        source_refs=("boe-a-2026-1",),
    )


def _ledger_evidence(*, transaction_id: str = _CONTRIBUTOR_ID) -> LedgerFilingEvidence:
    _, _, manual_casilla = _casilla_ids()
    return LedgerFilingEvidence(
        snapshot_fingerprint=_SNAPSHOT_FINGERPRINT,
        rows=(_evidence_row(transaction_id),),
        manual_entries=(
            ManualFactBasisEntry(
                casilla_id=manual_casilla,
                value="140000.00",
                kind="casilla_input",
                note="operator supplied taxable base",
                legal_refs=("ley-37-1992:art-78",),
                source_refs=("operator-manual-evidence",),
            ),
        ),
        captured_at=datetime(2026, 6, 3, 15, 0, tzinfo=UTC),
    )


def test_projection_maps_a_contributor_onto_every_attributed_casilla() -> None:
    first, second, manual_casilla = _casilla_ids()

    facet = sheet_evidence_from_ledger_filing(
        _ledger_evidence(),
        casilla_ids_by_contributor_id={_CONTRIBUTOR_ID: (first, second)},
    )

    assert facet.snapshot_fingerprint == _SNAPSHOT_FINGERPRINT
    assert tuple(row.casilla_id for row in facet.contributor_rows) == (first, second)
    assert {row.transaction_id for row in facet.contributor_rows} == {_CONTRIBUTOR_ID}
    assert facet.contributor_rows[0].iva_rate == Decimal("0.21")
    assert facet.contributor_rows[0].counterparty == _COUNTERPARTY
    assert facet.manual_entries[0].casilla_id == manual_casilla
    assert facet.manual_entries[0].value == "140000.00"


def test_projection_refuses_a_contributor_with_no_casilla_attribution() -> None:
    """The adapter never guesses modelo-specific tax meaning from row contents."""
    first, _, _ = _casilla_ids()

    with pytest.raises(CalcSheetsEngineError):
        sheet_evidence_from_ledger_filing(
            _ledger_evidence(transaction_id=_UNATTRIBUTED_CONTRIBUTOR_ID),
            casilla_ids_by_contributor_id={_CONTRIBUTOR_ID: (first,)},
        )


def test_build_export_plan_threads_supplied_ledger_evidence_into_the_plan() -> None:
    first, _, _ = _casilla_ids()

    plan = build_export_plan(
        _snapshot(),
        ledger_filing_evidence=_ledger_evidence(),
        casilla_ids_by_contributor_id={_CONTRIBUTOR_ID: (first,)},
    )

    assert plan.evidence.snapshot_fingerprint == _SNAPSHOT_FINGERPRINT
    assert len(plan.evidence.contributor_rows) == 1
    assert plan.evidence.contributor_rows[0].casilla_id == first
    assert len(plan.evidence.manual_entries) == 1


def test_the_exported_workbook_and_sidecar_carry_the_threaded_evidence() -> None:
    first, _, _ = _casilla_ids()
    plan = build_export_plan(
        _snapshot(),
        ledger_filing_evidence=_ledger_evidence(),
        casilla_ids_by_contributor_id={_CONTRIBUTOR_ID: (first,)},
    )

    export = serialize_offline_export(plan)
    workbook = load_workbook(BytesIO(export.workbook_payload), data_only=False)
    evidencia = workbook[TabName.EVIDENCIA.value]

    assert evidencia["B1"].value == _SNAPSHOT_FINGERPRINT
    assert evidencia["A4"].value == "ledger"
    assert evidencia["B4"].value == first
    assert evidencia["C4"].value == _CONTRIBUTOR_ID
    assert evidencia["D4"].value == "121.00"
    assert evidencia["E4"].value == "EUR"
    assert evidencia["F4"].value == "100.00"
    assert evidencia["G4"].value == "0.21"
    assert evidencia["H4"].value == "21.00"
    assert evidencia["I4"].value == _COUNTERPARTY
    assert evidencia["A5"].value == "manual"

    sidecar = json.loads(export.evidence_sidecar_payload.decode("utf-8"))
    assert sidecar["evidence"]["snapshot_fingerprint"] == _SNAPSHOT_FINGERPRINT
    contributor = sidecar["evidence"]["contributor_rows"][0]
    assert contributor["transaction_id"] == _CONTRIBUTOR_ID
    assert contributor["casilla_id"] == first
    assert contributor["iva_rate"] == "0.21"
    assert contributor["counterparty"] == _COUNTERPARTY
    assert sidecar["evidence"]["manual_entries"][0]["kind"] == "casilla_input"


def test_a_plan_built_without_ledger_evidence_yields_an_empty_facet() -> None:
    """Detector teeth: the empty facet is exactly the defect this change closed.

    A workbook exported from a plan carrying no bundle ships an empty
    ``Evidencia`` tab. If a future change stops threading the bundle through
    :func:`build_export_plan`, the populated cases above fail and this is the
    state they collapse into.
    """
    plan = build_export_plan(_snapshot())

    assert plan.evidence.snapshot_fingerprint is None
    assert plan.evidence.contributor_rows == ()
    assert plan.evidence.manual_entries == ()

    export = serialize_offline_export(plan)
    evidencia = load_workbook(BytesIO(export.workbook_payload), data_only=False)[TabName.EVIDENCIA.value]
    assert evidencia["A4"].value is None
