"""Offline workbook evidence-tab coverage for ``SheetExportPlan``."""

from __future__ import annotations

import hashlib
import json
from datetime import UTC, datetime
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook

from .....core import Period
from .....core.casilla_id import CasillaId, validated_casilla_id
from .._records import (
    SheetCellAddress,
    SheetEvidenceContributorRow,
    SheetEvidenceFacet,
    SheetEvidenceManualEntry,
    SheetExportMetadata,
    SheetExportPlan,
    SheetFormulaCell,
    SheetGuideContent,
    SheetProtectedRange,
    SheetValueCell,
    TabName,
)
from .._workbook_export import serialize_offline_export, serialize_offline_workbook

pytestmark = [pytest.mark.unit, pytest.mark.hex_application]
_BASE_CASILLA: CasillaId = validated_casilla_id("base", surface="_BASE_CASILLA")
_CUOTA_CASILLA: CasillaId = validated_casilla_id("cuota", surface="_CUOTA_CASILLA")
_RESULTADO_CONTABLE_CASILLA: CasillaId = validated_casilla_id(
    "resultado.contable",
    surface="_RESULTADO_CONTABLE_CASILLA",
)


def _metadata() -> SheetExportMetadata:
    return SheetExportMetadata(
        modelo_id="303",
        revision_id="2022",
        filing_year=2026,
        period=Period.from_year_and_code(2026, "1T"),
        engine_version="test",
        registry_sha="abcd1234",
        exported_at=datetime(2026, 6, 3, 15, 0, tzinfo=UTC),
    )


def _evidence_plan() -> SheetExportPlan:
    return SheetExportPlan(
        metadata=_metadata(),
        value_cells=(
            SheetValueCell(
                address=SheetCellAddress.at(TabName.ENTRADAS, 2, 4),
                value=Decimal("100.00"),
                role="operator_input",
                casilla_id=_BASE_CASILLA,
            ),
        ),
        formula_cells=(
            SheetFormulaCell(
                address=SheetCellAddress.at(TabName.CALCULOS, 2, 4),
                formula="'Entradas'!D2*0.21",
                casilla_id=_CUOTA_CASILLA,
                rounding_rule="money",
                rounding_scale=2,
            ),
        ),
        guide=SheetGuideContent(title="Modelo 303", paragraphs=("Use Entradas.",)),
        # Evidencia protection is now derived from the plan rather than applied
        # unconditionally by the materialiser, so a plan that wants the tab
        # locked has to say so -- which is the point: the plan is the contract
        # both transports read.
        protected_ranges=(
            SheetProtectedRange(
                tab=TabName.EVIDENCIA,
                start_row=1,
                end_row=1000,
                start_column=1,
                end_column=16,
                description="Generated evidence: read-only",
            ),
        ),
        evidence=SheetEvidenceFacet(
            snapshot_fingerprint="f" * 64,
            contributor_rows=(
                SheetEvidenceContributorRow(
                    casilla_id=_CUOTA_CASILLA,
                    transaction_id="c" * 64,
                    amount=Decimal("-121.00"),
                    currency="EUR",
                    taxable_base=Decimal("100.00"),
                    iva_rate=Decimal("0.21"),
                    iva_amount=Decimal("21.00"),
                    counterparty="Proveedor SL",
                    attachment_ids=("attachment-1",),
                    document_link_ids=("drive-doc-1",),
                    legal_refs=("ley-37-1992:art-99",),
                    source_refs=("boe-a-2026-1",),
                ),
            ),
            manual_entries=(
                SheetEvidenceManualEntry(
                    casilla_id=_RESULTADO_CONTABLE_CASILLA,
                    value="140000.00",
                    kind="casilla_input",
                    note="operator supplied accounting result",
                    legal_refs=("ley-27-2014:art-10",),
                    source_refs=("operator-manual-evidence",),
                ),
            ),
        ),
    )


def test_offline_workbook_renders_evidencia_tab_from_plan_evidence() -> None:
    plan = _evidence_plan()

    payload = serialize_offline_workbook(plan)
    workbook = load_workbook(BytesIO(payload), data_only=False)

    assert workbook.sheetnames == [tab.value for tab in TabName]
    assert workbook[TabName.ENTRADAS.value]["D2"].value == "100.00"
    assert workbook[TabName.CALCULOS.value]["D2"].value == "='Entradas'!D2*0.21"

    evidencia = workbook[TabName.EVIDENCIA.value]
    assert evidencia["A1"].value == "Snapshot fingerprint"
    assert evidencia["B1"].value == "f" * 64
    assert evidencia["A3"].value == "Tipo"
    assert evidencia["A4"].value == "ledger"
    assert evidencia["B4"].value == "cuota"
    assert evidencia["C4"].value == "c" * 64
    assert evidencia["D4"].value == "-121.00"
    assert evidencia["F4"].value == "100.00"
    assert evidencia["M4"].value == "attachment-1"
    assert evidencia["O4"].value == "ley-37-1992:art-99"
    assert evidencia["A5"].value == "manual"
    assert evidencia["B5"].value == "resultado.contable"
    assert evidencia["J5"].value == "140000.00"
    assert evidencia["K5"].value == "casilla_input"
    assert evidencia["L5"].value == "operator supplied accounting result"
    assert evidencia.protection.sheet is True


def test_offline_export_emits_machine_readable_evidence_sidecar() -> None:
    plan = _evidence_plan()

    export = serialize_offline_export(plan)
    sidecar = json.loads(export.evidence_sidecar_payload.decode("utf-8"))

    assert export.workbook_media_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert export.evidence_sidecar_media_type == "application/json"
    assert export.workbook_sha256 == hashlib.sha256(export.workbook_payload).hexdigest()
    assert export.evidence_sidecar_sha256 == hashlib.sha256(export.evidence_sidecar_payload).hexdigest()
    assert sidecar["schema_version"] == "calc-sheets-evidence-sidecar/v1"
    assert sidecar["workbook_sha256"] == export.workbook_sha256
    assert sidecar["metadata"]["modelo_id"] == "303"
    assert sidecar["metadata"]["revision_id"] == "2022"
    assert sidecar["metadata"]["registry_sha"] == "abcd1234"
    assert sidecar["evidence"]["snapshot_fingerprint"] == "f" * 64
    assert sidecar["evidence"]["contributor_rows"][0]["transaction_id"] == "c" * 64
    assert sidecar["evidence"]["contributor_rows"][0]["amount"] == "-121.00"
    assert sidecar["evidence"]["manual_entries"][0]["kind"] == "casilla_input"
