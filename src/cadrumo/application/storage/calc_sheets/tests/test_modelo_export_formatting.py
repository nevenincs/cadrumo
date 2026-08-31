"""Section-header + start/final anchor formatting facets.

The export must mirror the official AEAT workbook presentation: casilla sections
rendered as bold headers and an explicit labelled start (Entradas) / final
(resultado) anchor orienting the inputs→result flow. This asserts the engine
emits the typed facets and the offline workbook renders them (bold) — the online
apply adapter mirrors the same facets (see the conformance test).
"""

from __future__ import annotations

from datetime import date
from io import BytesIO

import pytest
from openpyxl import load_workbook

from .....domain.calculations.registry.authority import bundled_authority
from ..engine import build_export_plan
from ..records import TabName
from ..workbook_export import serialize_offline_workbook

pytestmark = [pytest.mark.unit, pytest.mark.hex_application]


def _m130_plan():
    snapshot = bundled_authority().snapshot("130", filing_year=2025, period="1T", on=date(2025, 4, 1))
    return build_export_plan(snapshot)


def test_engine_emits_section_headers_and_start_final_anchors() -> None:
    plan = _m130_plan()

    # Section headers: at least one, each pointing at a column-A label cell.
    assert plan.section_headers
    assert all(header.address.column == 1 and header.text for header in plan.section_headers)

    # Anchors: exactly one labelled start (Entradas) and one final (resultado).
    kinds = {anchor.kind: anchor for anchor in plan.anchors}
    assert set(kinds) == {"start", "final"}
    assert kinds["start"].address.tab is TabName.ENTRADAS
    assert kinds["final"].address.tab is TabName.CALCULOS
    assert kinds["start"].label and kinds["final"].label

    # The anchor labels are also rendered as value cells (so both transports
    # write them via the value batch).
    value_addresses = {cell.address.qualified() for cell in plan.value_cells}
    assert all(anchor.address.qualified() in value_addresses for anchor in plan.anchors)


def test_offline_workbook_bolds_section_headers_and_anchors() -> None:
    plan = _m130_plan()
    workbook = load_workbook(BytesIO(serialize_offline_workbook(plan)), data_only=False)

    for header in plan.section_headers:
        cell = workbook[header.address.tab.value].cell(row=header.address.row, column=header.address.column)
        assert cell.font.bold is True, header.text
    for anchor in plan.anchors:
        cell = workbook[anchor.address.tab.value].cell(row=anchor.address.row, column=anchor.address.column)
        assert cell.value == anchor.label
        assert cell.font.bold is True, anchor.kind
