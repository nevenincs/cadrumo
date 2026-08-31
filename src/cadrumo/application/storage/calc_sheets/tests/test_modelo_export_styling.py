"""Design-system facets: font, role fills, freezes, filters, widths, wrap.

The export must mirror the official AEAT modelo look uniformly across both
transports per the official-structure contract. This locks the engine
emitting the typed styling facets and the offline openpyxl workbook rendering
them — the slate header band, blue-grey section banners, pale-yellow operator
inputs, grey computed cells, green result accent, the monospace family, frozen
header rows, basic filters, sized columns, and wrapped body text. The online
apply adapter mirrors the same facets (see the offline/online conformance test).
"""

from __future__ import annotations

from datetime import date
from io import BytesIO

import pytest
from openpyxl import load_workbook

from .....domain.calculations.registry.authority import bundled_authority
from .._engine import build_export_plan
from .._records import TabName
from .._theme import ROLE_STYLES, WORKBOOK_FONT_FAMILY, StyleRole, openpyxl_argb
from .._workbook_export import serialize_offline_workbook

pytestmark = [pytest.mark.unit, pytest.mark.hex_application]


def _m130_plan():
    snapshot = bundled_authority().snapshot("130", filing_year=2025, period="1T", on=date(2025, 4, 1))
    return build_export_plan(snapshot)


def test_engine_emits_design_facets() -> None:
    plan = _m130_plan()

    assert plan.font_family == WORKBOOK_FONT_FAMILY

    roles = {styled.role for styled in plan.styled_ranges}
    # The full role vocabulary the official look relies on is present.
    assert {StyleRole.HEADER, StyleRole.SECTION_BANNER, StyleRole.INPUT, StyleRole.COMPUTED, StyleRole.RESULT} <= roles

    # A header band on row 1 of both data tabs.
    header_tabs = {
        styled.tab for styled in plan.styled_ranges if styled.role is StyleRole.HEADER and styled.start_row == 1
    }
    assert TabName.ENTRADAS in header_tabs
    assert TabName.CALCULOS in header_tabs

    # The input column lives on Entradas; the computed column on Cálculos.
    input_tabs = {styled.tab for styled in plan.styled_ranges if styled.role is StyleRole.INPUT}
    computed_tabs = {styled.tab for styled in plan.styled_ranges if styled.role is StyleRole.COMPUTED}
    assert input_tabs == {TabName.ENTRADAS}
    assert computed_tabs == {TabName.CALCULOS}

    # Frozen header row + basic filter on the data tabs; sized columns.
    frozen = {view.tab: view for view in plan.frozen_views}
    assert frozen[TabName.ENTRADAS].frozen_rows == 1
    assert frozen[TabName.CALCULOS].frozen_rows == 1
    assert {flt.tab for flt in plan.auto_filters} >= {TabName.ENTRADAS, TabName.CALCULOS}
    assert {width.tab for width in plan.column_widths} >= {TabName.ENTRADAS, TabName.CALCULOS}


def test_result_accent_is_ordered_after_the_computed_column() -> None:
    # The green result accent must win over the broad grey computed column on
    # overlap, so it is emitted later in declaration order.
    plan = _m130_plan()
    roles_in_order = [styled.role for styled in plan.styled_ranges]
    assert StyleRole.RESULT in roles_in_order
    assert StyleRole.COMPUTED in roles_in_order
    assert roles_in_order.index(StyleRole.RESULT) > roles_in_order.index(StyleRole.COMPUTED)


def test_offline_workbook_renders_the_palette_and_font() -> None:
    plan = _m130_plan()
    workbook = load_workbook(BytesIO(serialize_offline_workbook(plan)), data_only=False)

    entradas = workbook[TabName.ENTRADAS.value]
    # Header band: slate fill, white bold, monospace.
    header = entradas.cell(row=1, column=1)
    header_style = ROLE_STYLES[StyleRole.HEADER]
    assert header.font.bold is True
    assert header.font.name == WORKBOOK_FONT_FAMILY
    assert header.fill.fgColor.rgb == openpyxl_argb(header_style.fill_hex or "")

    # Frozen header row + installed basic filter.
    assert entradas.freeze_panes == "A2"
    assert entradas.auto_filter.ref is not None

    # An operator-input value cell carries the pale-yellow fill.
    input_style = ROLE_STYLES[StyleRole.INPUT]
    first_input = next(
        cell for cell in plan.value_cells if cell.role == "operator_input" and cell.address.tab is TabName.ENTRADAS
    )
    input_cell = entradas.cell(row=first_input.address.row, column=first_input.address.column)
    assert input_cell.fill.fgColor.rgb == openpyxl_argb(input_style.fill_hex or "")
    assert input_cell.alignment.horizontal == "right"

    # A computed value cell on Cálculos carries the grey fill.
    calculos = workbook[TabName.CALCULOS.value]
    computed_style = ROLE_STYLES[StyleRole.COMPUTED]
    first_formula = plan.formula_cells[0]
    computed_cell = calculos.cell(row=first_formula.address.row, column=first_formula.address.column)
    assert computed_cell.fill.fgColor.rgb == openpyxl_argb(computed_style.fill_hex or "")

    # Concepto column wraps so long labels do not clip.
    concepto = entradas.cell(row=2, column=3)
    assert concepto.alignment.wrap_text is True


def test_offline_workbook_applies_print_setup_to_every_tab() -> None:
    # The print-setup step in _apply_styling: landscape, fit-to-one-page-width,
    # repeat the header row on every printed page, on every tab.
    plan = _m130_plan()
    workbook = load_workbook(BytesIO(serialize_offline_workbook(plan)), data_only=False)

    for tab in TabName:
        worksheet = workbook[tab.value]
        assert worksheet.page_setup.orientation == "landscape"
        assert worksheet.page_setup.fitToWidth == 1
        assert worksheet.page_setup.fitToHeight == 0
        assert worksheet.sheet_properties.pageSetUpPr is not None
        assert worksheet.sheet_properties.pageSetUpPr.fitToPage is True
        # openpyxl normalises the repeat-rows defined name to absolute on save.
        assert worksheet.print_title_rows == "$1:$1"
