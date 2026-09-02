"""Offline XLSX materializer for :class:`~application.storage.calc_sheets.SheetExportPlan` records.

This module serializes the shared workbook plan into operator-directed
plaintext export bytes: an XLSX workbook and an adjacent machine-readable JSON
evidence sidecar. It returns bytes and digests to the caller; it does not choose
paths, persist secure-object state, or make the export file canonical.

See Also:
    :class:`~application.storage.calc_sheets.SheetExportPlan`
        Renderer-neutral workbook contract emitted by the calc-sheets engine.
    :class:`~application.storage.calc_sheets.SheetEvidenceFacet`
        Evidence facet rendered into the Evidencia tab and JSON sidecar.
"""

from __future__ import annotations

import json
from collections.abc import Iterable, Sequence
from decimal import Decimal
from io import BytesIO
from typing import TYPE_CHECKING, Literal

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties
from pydantic import BaseModel

from ....core.decimal.formatting import format_decimal
from ....core.external_constants import UTF_8_ENCODING
from ....core.hashing import sha256_hex
from ....core.identity import ContentDigest
from ....core.models import STRICT_FROZEN_CONFIG as _STRICT_FROZEN
from .records import (
    SheetAutoFilter,
    SheetColumnWidth,
    SheetEvidenceContributorRow,
    SheetEvidenceFacet,
    SheetEvidenceManualEntry,
    SheetExportMetadata,
    SheetExportPlan,
    SheetFormulaCell,
    SheetFrozenView,
    SheetProtectedRange,
    SheetRowSet,
    SheetStyledRange,
    SheetValueCell,
    TabName,
)
from .theme import ROLE_STYLES, STYLED_RANGE_VERTICAL_ALIGN, WORKBOOK_FONT_FAMILY, openpyxl_argb

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet


_EVIDENCE_SIDECAR_SCHEMA_VERSION: Literal["calc-sheets-evidence-sidecar/v1"] = "calc-sheets-evidence-sidecar/v1"
_XLSX_MEDIA_TYPE: Literal["application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"] = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
_JSON_MEDIA_TYPE: Literal["application/json"] = "application/json"

_EVIDENCE_HEADERS: tuple[str, ...] = (
    "Tipo",
    "Casilla",
    "Transaction ID",
    "Amount",
    "Currency",
    "Taxable base",
    "IVA rate",
    "IVA amount",
    "Counterparty",
    "Value",
    "Kind",
    "Note",
    "Attachment IDs",
    "Document link IDs",
    "Legal refs",
    "Source refs",
)


class OfflineWorkbookEvidenceSidecar(BaseModel):
    """Machine-readable evidence sidecar emitted beside an offline workbook.

    The sidecar binds
    :class:`~application.storage.calc_sheets.SheetExportMetadata`, the
    workbook SHA-256, and
    :class:`~application.storage.calc_sheets.SheetEvidenceFacet` so
    external review can inspect the fact basis without parsing XLSX cells.
    """

    model_config = _STRICT_FROZEN

    schema_version: Literal["calc-sheets-evidence-sidecar/v1"] = _EVIDENCE_SIDECAR_SCHEMA_VERSION
    metadata: SheetExportMetadata
    workbook_sha256: ContentDigest
    evidence: SheetEvidenceFacet


class OfflineWorkbookExportResult(BaseModel):
    """Serialized offline workbook plus its machine-readable evidence sidecar.

    The payloads are operator-directed export bytes. Callers decide where to
    write them and remain responsible for protecting the plaintext files after
    export.
    """

    model_config = _STRICT_FROZEN

    workbook_payload: bytes
    workbook_media_type: Literal["application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"] = _XLSX_MEDIA_TYPE
    workbook_filename_extension: Literal["xlsx"] = "xlsx"
    workbook_sha256: ContentDigest
    evidence_sidecar_payload: bytes
    evidence_sidecar_media_type: Literal["application/json"] = _JSON_MEDIA_TYPE
    evidence_sidecar_filename_extension: Literal["evidence.json"] = "evidence.json"
    evidence_sidecar_sha256: ContentDigest


def evidence_table(plan: SheetExportPlan) -> tuple[str, tuple[str, ...], tuple[tuple[str, ...], ...]]:
    """Return ``(snapshot_fingerprint, header_row, body_rows)`` for the Evidencia surface.

    The single source of truth both transports render: the offline openpyxl
    workbook and the online Google-Sheets apply adapter consume this so the
    Evidencia surface is byte-identical across offline and online exports.
    Every cell is a string — contributor facts are pre-formatted by the same
    helpers ``_write_evidence`` uses.
    """
    body = tuple(
        tuple(str(value) for value in _contributor_values(row)) for row in plan.evidence.contributor_rows
    ) + tuple(tuple(str(value) for value in _manual_values(row)) for row in plan.evidence.manual_entries)
    return (plan.evidence.snapshot_fingerprint or "", _EVIDENCE_HEADERS, body)


def guide_stamps(plan: SheetExportPlan) -> tuple[tuple[str, str], ...]:
    """Return the Guide tab's ``(label, value)`` export stamps for ``plan``.

    The sibling of :func:`evidence_table`, and for the same reason: both
    transports claim to render one :class:`SheetExportPlan`, so anything they
    both draw has to come from one place. Each kept its own hand-maintained
    copy of this table and the two had already drifted — the offline XLSX
    wrote ``Revision`` and ``Periodo`` while the Google Sheets value builder
    wrote ``Revisión`` and ``Período``. The metadata VALUES matched, so
    nothing looked wrong, but the two exports of one plan disagreed about the
    row labels a reader or a label-keyed instruction would look for.

    The accented spellings win: these are Spanish operator-facing labels and
    both transports carry UTF-8 text, so the unaccented pair was a
    transliteration with nothing to recommend it.
    """
    metadata = plan.metadata
    return (
        ("Modelo", metadata.modelo_id),
        ("Revisión", metadata.revision_id),
        ("Período", f"{metadata.period.registry_token} / {metadata.filing_year}"),
        ("Motor", metadata.engine_version),
        ("Registry SHA", metadata.registry_sha),
        ("Exportado", metadata.exported_at.isoformat()),
    )


def build_offline_workbook(plan: SheetExportPlan) -> Workbook:
    """Materialise a :class:`~application.storage.calc_sheets.SheetExportPlan` as an offline openpyxl workbook."""
    workbook = Workbook()
    default = workbook.active
    assert default is not None
    default.title = TabName.ENTRADAS.value
    for tab in TabName:
        if tab.value not in workbook.sheetnames:
            workbook.create_sheet(tab.value)

    _write_value_cells(workbook, plan.value_cells)
    _write_formula_cells(workbook, plan.formula_cells)
    _write_row_set_headers(workbook, plan.row_sets)
    _write_guide(workbook[TabName.GUIDE.value], plan)
    _write_evidence(workbook[TabName.EVIDENCIA.value], plan)
    _apply_styling(workbook, plan)
    return workbook


def _apply_styling(workbook: Workbook, plan: SheetExportPlan) -> None:
    """Apply the design system — font, role fills, widths, freezes, filters.

    The single offline materialisation of the shared
    :mod:`~application.storage.calc_sheets.theme` palette: it sets the
    monospace family on every populated cell, tints each styled range by its role
    (header band, section banner, pale-yellow inputs, grey computed, green
    result), wraps the body columns, sizes the columns, freezes the header rows,
    and installs the basic filters — mirroring exactly what the online apply
    adapter emits from the same
    :class:`~application.storage.calc_sheets.SheetExportPlan` facets. The
    phases run in the same order as before: base font first, then styled
    overrides, widths, freezes, filters, and finally print setup.
    """
    family = plan.font_family or WORKBOOK_FONT_FAMILY
    _apply_base_font(workbook, family)
    _apply_styled_ranges(workbook, family, plan.styled_ranges)
    _apply_column_widths(workbook, plan.column_widths)
    _apply_frozen_views(workbook, plan.frozen_views)
    _apply_auto_filters(workbook, plan.auto_filters)
    _apply_protected_ranges(workbook, plan.protected_ranges)
    _apply_print_setup(workbook)


def _apply_base_font(workbook: Workbook, family: str) -> None:
    """Set the monospace family on every populated cell across all tabs."""
    base_font = Font(name=family)
    for tab in TabName:
        worksheet = workbook[tab.value]
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.font = base_font


def _apply_styled_ranges(workbook: Workbook, family: str, styled_ranges: Sequence[SheetStyledRange]) -> None:
    """Tint each styled range by its role, in declaration order so later ranges win."""
    for styled in styled_ranges:
        style = ROLE_STYLES[styled.role]
        font = Font(
            name=family,
            bold=style.bold,
            color=openpyxl_argb(style.font_hex) if style.font_hex else None,
        )
        fill = (
            PatternFill(fill_type="solid", fgColor=openpyxl_argb(style.fill_hex))
            if style.fill_hex is not None
            else None
        )
        alignment = Alignment(
            horizontal=style.align,
            vertical=STYLED_RANGE_VERTICAL_ALIGN,
            wrap_text=styled.wrap,
        )
        worksheet = workbook[styled.tab.value]
        for row_index in range(styled.start_row, styled.end_row + 1):
            for column_index in range(styled.start_column, styled.end_column + 1):
                cell = worksheet.cell(row=row_index, column=column_index)
                cell.font = font
                cell.alignment = alignment
                if fill is not None:
                    cell.fill = fill


def _apply_column_widths(workbook: Workbook, column_widths: Sequence[SheetColumnWidth]) -> None:
    """Size each declared column."""
    for width in column_widths:
        worksheet = workbook[width.tab.value]
        worksheet.column_dimensions[get_column_letter(width.column)].width = width.width


def _apply_frozen_views(workbook: Workbook, frozen_views: Sequence[SheetFrozenView]) -> None:
    """Freeze the header rows/columns on each declared tab."""
    for frozen in frozen_views:
        worksheet = workbook[frozen.tab.value]
        worksheet.freeze_panes = f"{get_column_letter(frozen.frozen_columns + 1)}{frozen.frozen_rows + 1}"


def _apply_auto_filters(workbook: Workbook, auto_filters: Sequence[SheetAutoFilter]) -> None:
    """Install the basic filter over each declared range."""
    for filter_range in auto_filters:
        worksheet = workbook[filter_range.tab.value]
        start = f"{get_column_letter(filter_range.start_column)}{filter_range.start_row}"
        end = f"{get_column_letter(filter_range.end_column)}{filter_range.end_row}"
        worksheet.auto_filter.ref = f"{start}:{end}"


def _apply_protected_ranges(
    workbook: Workbook,
    protected_ranges: Sequence[SheetProtectedRange],
) -> None:
    """Materialise the plan's read-only contract in the offline workbook.

    ``SheetExportPlan.protected_ranges`` is the declared read-only surface and
    the online adapter emits an ``addProtectedRange`` per entry, but the
    offline materialiser consumed none of them: every planned range shipped
    editable in the XLSX while the same plan shipped locked in Sheets. The two
    transports disagreed in both directions at once, which is why neither
    looked broken on its own.

    XLSX has no per-range protection primitive -- it has a sheet-level flag
    plus a per-cell ``locked`` bit that only takes effect once that flag is on.
    So the range set is expressed the only way the format allows: on a tab the
    plan protects, every cell is unlocked first and the planned ranges are
    locked back. That order is what keeps the editable-input policy explicit
    rather than incidental. Cells default to ``locked=True`` in openpyxl, so
    turning the sheet flag on without the unlock pass would freeze the entire
    tab -- protecting far more than the plan asked for, and silently.

    A tab the plan does not name is left alone, so ``Entradas`` stays fully
    editable, matching the online transport.
    """
    by_tab: dict[str, list[SheetProtectedRange]] = {}
    for protected in protected_ranges:
        by_tab.setdefault(protected.tab.value, []).append(protected)

    for tab, ranges in by_tab.items():
        worksheet = workbook[tab]
        for row in worksheet.iter_rows():
            for cell in row:
                cell.protection = Protection(locked=False)
        for protected in ranges:
            for row_index in range(protected.start_row, protected.end_row + 1):
                for column_index in range(protected.start_column, protected.end_column + 1):
                    worksheet.cell(row=row_index, column=column_index).protection = Protection(locked=True)
        worksheet.protection.sheet = True


def _apply_print_setup(workbook: Workbook) -> None:
    """Landscape, fit all columns to one page width, repeat the header row.

    A printed filing artefact then stays readable across page breaks.
    """
    for tab in TabName:
        worksheet = workbook[tab.value]
        worksheet.page_setup.orientation = "landscape"
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0
        worksheet.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        worksheet.print_title_rows = "1:1"


def build_evidence_sidecar(
    plan: SheetExportPlan,
    *,
    workbook_payload: bytes,
) -> OfflineWorkbookEvidenceSidecar:
    """Build the machine-readable evidence sidecar for one workbook export.

    The sidecar is keyed to the workbook payload digest, so a reviewer can pair
    the JSON evidence with the exact XLSX bytes produced by
    :func:`~application.storage.calc_sheets.serialize_offline_workbook`.

    Takes the workbook BYTES rather than a digest of them. Accepting the
    digest asked the caller to assert the one fact the sidecar exists to
    carry, and nothing on this side could check it: a sidecar claiming to
    bind bytes it had never seen validated exactly as well as a correct one.
    Deriving the digest here makes the binding true by construction, which is
    what "keyed to the workbook payload" has to mean if a reviewer is to rely
    on it.

    Returns:
        :class:`~application.storage.calc_sheets.OfflineWorkbookEvidenceSidecar`:
            The evidence sidecar.
    """
    return OfflineWorkbookEvidenceSidecar(
        metadata=plan.metadata,
        workbook_sha256=sha256_hex(workbook_payload),
        evidence=plan.evidence,
    )


def serialize_evidence_sidecar(sidecar: OfflineWorkbookEvidenceSidecar) -> bytes:
    """Serialize an evidence sidecar as canonical UTF-8 JSON bytes."""
    payload = sidecar.model_dump(mode="json")
    text = json.dumps(payload, ensure_ascii=True, sort_keys=True, separators=(",", ":"))
    return text.encode(UTF_8_ENCODING)


def serialize_offline_workbook(plan: SheetExportPlan) -> bytes:
    """Serialize an offline workbook plan to XLSX bytes."""
    workbook = build_offline_workbook(plan)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def serialize_offline_export(plan: SheetExportPlan) -> OfflineWorkbookExportResult:
    """Serialize an offline workbook and its adjacent evidence sidecar.

    The return value is a pure byte payload plus integrity metadata. No file is
    written here; this keeps the plaintext-export exception at the caller's
    explicit output path boundary.

    Returns:
        :class:`~application.storage.calc_sheets.OfflineWorkbookExportResult`:
            The export result.
    """
    workbook_payload = serialize_offline_workbook(plan)
    sidecar = build_evidence_sidecar(plan, workbook_payload=workbook_payload)
    evidence_sidecar_payload = serialize_evidence_sidecar(sidecar)
    return OfflineWorkbookExportResult(
        workbook_payload=workbook_payload,
        workbook_sha256=sidecar.workbook_sha256,
        evidence_sidecar_payload=evidence_sidecar_payload,
        evidence_sidecar_sha256=sha256_hex(evidence_sidecar_payload),
    )


def _write_value_cells(workbook: Workbook, cells: Iterable[SheetValueCell]) -> None:
    for cell in cells:
        worksheet = workbook[cell.address.tab.value]
        target = worksheet.cell(row=cell.address.row, column=cell.address.column, value=coerce_cell_value(cell.value))
        if cell.note is not None:
            assert isinstance(target, Cell)
            target.comment = Comment(cell.note, "AEAT")


def _write_formula_cells(workbook: Workbook, cells: Iterable[SheetFormulaCell]) -> None:
    for cell in cells:
        worksheet = workbook[cell.address.tab.value]
        worksheet.cell(row=cell.address.row, column=cell.address.column, value=f"={cell.formula}")


def _write_row_set_headers(workbook: Workbook, row_sets: Iterable[SheetRowSet]) -> None:
    for row_set in row_sets:
        worksheet = workbook[row_set.tab.value]
        for column in row_set.columns:
            worksheet.cell(
                row=column.header_address.row,
                column=column.header_address.column,
                value=column.header_label,
            )


def _write_guide(worksheet: Worksheet, plan: SheetExportPlan) -> None:
    worksheet["A1"] = plan.guide.title
    for row, paragraph in enumerate(plan.guide.paragraphs, start=3):
        worksheet.cell(row=row, column=1, value=paragraph)

    base_row = 3 + len(plan.guide.paragraphs) + 2
    for offset, (label, value) in enumerate(guide_stamps(plan)):
        worksheet.cell(row=base_row + offset, column=1, value=label)
        worksheet.cell(row=base_row + offset, column=2, value=value)


def _write_evidence(worksheet: Worksheet, plan: SheetExportPlan) -> None:
    worksheet["A1"] = "Snapshot fingerprint"
    worksheet["B1"] = plan.evidence.snapshot_fingerprint or ""
    for column, header in enumerate(_EVIDENCE_HEADERS, start=1):
        worksheet.cell(row=3, column=column, value=header)

    row_index = 4
    for row in plan.evidence.contributor_rows:
        _write_evidence_row(worksheet, row_index, _contributor_values(row))
        row_index += 1
    for row in plan.evidence.manual_entries:
        _write_evidence_row(worksheet, row_index, _manual_values(row))
        row_index += 1

    worksheet.freeze_panes = "A4"


def _write_evidence_row(worksheet: Worksheet, row_index: int, values: Sequence[str]) -> None:
    for column, value in enumerate(values, start=1):
        worksheet.cell(row=row_index, column=column, value=value)


def _contributor_values(row: SheetEvidenceContributorRow) -> tuple[str, ...]:
    return (
        "ledger",
        row.casilla_id,
        row.transaction_id,
        format_decimal(row.amount),
        row.currency,
        _format_optional_decimal(row.taxable_base),
        _format_optional_decimal(row.iva_rate),
        _format_optional_decimal(row.iva_amount),
        row.counterparty or "",
        "",
        "",
        "",
        _join(row.attachment_ids),
        _join(row.document_link_ids),
        _join(row.legal_refs),
        _join(row.source_refs),
    )


def _manual_values(row: SheetEvidenceManualEntry) -> tuple[str, ...]:
    return (
        "manual",
        row.casilla_id,
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        row.value,
        row.kind,
        row.note,
        "",
        "",
        _join(row.legal_refs),
        _join(row.source_refs),
    )


def coerce_cell_value(value: Decimal | str | bool | None) -> str | bool:
    if value is None:
        return ""
    if isinstance(value, Decimal):
        return format_decimal(value)
    return value


def _format_optional_decimal(value: Decimal | None) -> str:
    return format_decimal(value, none_value="")


def _join(values: tuple[str, ...]) -> str:
    return ";".join(values)


__all__ = [
    "OfflineWorkbookEvidenceSidecar",
    "OfflineWorkbookExportResult",
    "build_evidence_sidecar",
    "build_offline_workbook",
    "serialize_evidence_sidecar",
    "serialize_offline_export",
    "serialize_offline_workbook",
]
