"""Tests for the bulk CSV/XLSX invoice import application service.

:func:`~application.invoices.import_invoices_from_rows` delegates every row's
write to :func:`~application.invoices.create_catalogue_invoice` -- the sole
sanctioned :class:`~domain.invoices.Invoice` writer
(``aeat-architecture-boundaries``) -- and never persists a row
itself. These tests exercise it against the real encrypted
:class:`~adapters.persistence.profile.invoices.InvoiceCatalogueRepository` (real
master-key provider, real engine) -- no mocks.

See Also:
    :class:`~application.invoices.BulkInvoiceImportRow`
        Typed per-row boundary the tests validate before persistence.
    :func:`~application.invoices.read_bulk_invoice_import_rows`
        CSV/XLSX reader whose extension and column validation are covered here.
    :func:`~entrypoints.cli._ledger_business_invoice_cli.catalogue_import`
        CLI wrapper that feeds operator files into this application service.
"""

from __future__ import annotations

import zipfile
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook
from pydantic import ValidationError

from ....adapters.persistence.profile.invoices import InvoiceCatalogueRepository
from ....domain.invoices.errors import InvoiceValidationError
from ....domain.iva.classification import InvoiceKind
from ....tests.secure_sql import isolated_runtime_profile
from .._bulk_import import BulkInvoiceImportRow, import_invoices_from_rows, read_bulk_invoice_import_source

pytestmark = [pytest.mark.integration, pytest.mark.hex_application]

_BUCKET_ID = "29292929-2929-4292-8292-292929292929"
_CIF = "A58818501"


def _csv_source(text: str, tmp_path: Path):
    """Read *text* through the real reader, so tests exercise the production path."""
    csv_path = tmp_path / "bulk.csv"
    csv_path.write_text(text, encoding="utf-8")
    return read_bulk_invoice_import_source(csv_path)


def test_bulk_invoice_row_model_requires_all_mandatory_fields() -> None:
    """The typed row model enforces every required field via pydantic, not a bare dict."""
    with pytest.raises(ValidationError):
        BulkInvoiceImportRow.model_validate({})


def test_import_invoices_from_rows_persists_through_create_catalogue_invoice(tmp_path: Path) -> None:
    """Each valid row creates a real, reloadable catalogue invoice."""
    with isolated_runtime_profile(tmp_path=tmp_path, bucket_id=_BUCKET_ID):
        rows = _csv_source(
            "counterparty_nif,counterparty_name,invoice_number,invoice_date,taxable_base,iva_rate\n"
            f"{_CIF},Papeleria Sol SL,BULK-A-001,2026-05-01,100.00,21\n"
            f"{_CIF},Papeleria Sol SL,BULK-A-002,2026-05-02,50.00,10\n",
            tmp_path,
        )
        result = import_invoices_from_rows(rows, bucket_id=_BUCKET_ID, kind=InvoiceKind.RECEIVED, declared_country="ES")

        assert result.rows == 2
        assert result.created == 2
        assert result.skipped_duplicate == 0
        assert result.refused == ()
        assert len(result.created_invoice_ids) == 2

        catalogue = InvoiceCatalogueRepository(bucket_id=_BUCKET_ID).load()
        stored_numbers = {invoice.invoice_number for invoice in catalogue.invoices.values()}
        assert stored_numbers == {"BULK-A-001", "BULK-A-002"}
        for invoice_id in result.created_invoice_ids:
            assert invoice_id in catalogue.invoices


def test_import_invoices_from_rows_reimport_is_idempotent_no_op(tmp_path: Path) -> None:
    """Re-running the identical rows a second time skips every row as a duplicate."""
    with isolated_runtime_profile(tmp_path=tmp_path, bucket_id=_BUCKET_ID):
        rows = _csv_source(
            "counterparty_nif,counterparty_name,invoice_number,invoice_date,taxable_base,iva_rate\n"
            f"{_CIF},Papeleria Sol SL,BULK-B-001,2026-05-01,100.00,21\n",
            tmp_path,
        )
        first = import_invoices_from_rows(rows, bucket_id=_BUCKET_ID, kind=InvoiceKind.RECEIVED, declared_country="ES")
        assert first.created == 1

        second = import_invoices_from_rows(rows, bucket_id=_BUCKET_ID, kind=InvoiceKind.RECEIVED, declared_country="ES")
        assert second.created == 0
        assert second.skipped_duplicate == 1
        assert second.refused == ()

        catalogue = InvoiceCatalogueRepository(bucket_id=_BUCKET_ID).load()
        matching = [inv for inv in catalogue.invoices.values() if inv.invoice_number == "BULK-B-001"]
        assert len(matching) == 1


def test_import_invoices_from_rows_refuses_malformed_row_names_field(tmp_path: Path) -> None:
    """A malformed row (bad date) is refused naming its row number and field; valid rows still import."""
    with isolated_runtime_profile(tmp_path=tmp_path, bucket_id=_BUCKET_ID):
        rows = _csv_source(
            "counterparty_nif,counterparty_name,invoice_number,invoice_date,taxable_base,iva_rate\n"
            f"{_CIF},Papeleria Sol SL,BULK-C-001,2026-05-01,100.00,21\n"
            f"{_CIF},Papeleria Sol SL,BULK-C-002,not-a-date,50.00,10\n",
            tmp_path,
        )
        result = import_invoices_from_rows(rows, bucket_id=_BUCKET_ID, kind=InvoiceKind.RECEIVED, declared_country="ES")

        assert result.created == 1
        assert len(result.refused) == 1
        failure = result.refused[0]
        assert failure.row_number == 3
        assert failure.field == "invoice_date"


def test_import_invoices_from_rows_refuses_unsupported_iva_rate(tmp_path: Path) -> None:
    """An IVA percentage outside the closed slot taxonomy refuses that row, not the whole batch."""
    with isolated_runtime_profile(tmp_path=tmp_path, bucket_id=_BUCKET_ID):
        rows = _csv_source(
            "counterparty_nif,counterparty_name,invoice_number,invoice_date,taxable_base,iva_rate\n"
            f"{_CIF},Papeleria Sol SL,BULK-D-001,2026-05-01,100.00,13\n",
            tmp_path,
        )
        result = import_invoices_from_rows(rows, bucket_id=_BUCKET_ID, kind=InvoiceKind.RECEIVED, declared_country="ES")
        assert result.created == 0
        assert len(result.refused) == 1
        assert result.refused[0].field == "invoice"


def test_import_refuses_the_spanish_thousands_amount_instead_of_reading_it_as_cents(tmp_path: Path) -> None:
    """``1.234`` written by a gestor must refuse, not silently become one euro twenty-three.

    This is the shape a bare ``Decimal`` reads without complaint, because
    ``1.234`` is legal syntax for a different number. A gestor writing Spanish
    means one thousand two hundred and thirty-four euros; the old parser
    stored ``Decimal('1.234')`` and the invoice reached Modelo 303/390 a
    thousandfold light, as a wrong number rather than as any kind of failure.

    The row is refused naming its own field so the operator fixes the source.
    Reinterpreting it as ``1234`` is deliberately NOT the behaviour: the same
    text is a valid fractional amount under the grammar the file might
    equally be written in, and guessing between them would trade a loud
    refusal for a quiet reinterpretation of somebody's tax base.
    """
    with isolated_runtime_profile(tmp_path=tmp_path, bucket_id=_BUCKET_ID):
        rows = _csv_source(
            "counterparty_nif,counterparty_name,invoice_number,invoice_date,taxable_base,iva_rate\n"
            f"{_CIF},Papeleria Sol SL,BULK-ES-001,2026-05-01,1.234,21\n",
            tmp_path,
        )
        result = import_invoices_from_rows(rows, bucket_id=_BUCKET_ID, kind=InvoiceKind.RECEIVED, declared_country="ES")

        assert result.created == 0
        assert len(result.refused) == 1
        assert result.refused[0].field == "taxable_base"


@pytest.mark.parametrize(
    "taxable_base",
    (
        pytest.param("1.234,56", id="grouped-with-comma-decimal"),
        pytest.param("1234,56", id="bare-comma-decimal"),
        pytest.param("1.234", id="dot-grouped-thousands"),
    ),
)
def test_import_refuses_every_spanish_amount_grammar(taxable_base: str, tmp_path: Path) -> None:
    """No Spanish-written amount is silently accepted under the euro grammar.

    The three spellings a Spanish spreadsheet actually produces are asserted
    together because only one of them ever misbehaved: the comma forms always
    raised, and testing those alone would have proved nothing about the one
    that did not. That asymmetry is exactly how the defect shipped.
    """
    with isolated_runtime_profile(tmp_path=tmp_path, bucket_id=_BUCKET_ID):
        rows = _csv_source(
            "counterparty_nif,counterparty_name,invoice_number,invoice_date,taxable_base,iva_rate\n"
            f'{_CIF},Papeleria Sol SL,BULK-ES-002,2026-05-01,"{taxable_base}",21\n',
            tmp_path,
        )
        result = import_invoices_from_rows(rows, bucket_id=_BUCKET_ID, kind=InvoiceKind.RECEIVED, declared_country="ES")

        assert result.created == 0
        assert [failure.field for failure in result.refused] == ["taxable_base"]


def test_the_amount_refusal_teaches_the_grammar_it_wants(tmp_path: Path) -> None:
    """The refusal names the accepted form, not just that the value was rejected.

    An operator who typed ``1.234`` meaning one thousand two hundred and
    thirty-four euros has not made a typo — they wrote their own language
    correctly into a field that wanted a different grammar. Telling them only
    that the value is invalid leaves them to guess which of the two readings
    the field disagreed with, and the likeliest guess is that the amount is
    wrong rather than the notation.
    """
    with isolated_runtime_profile(tmp_path=tmp_path, bucket_id=_BUCKET_ID):
        rows = _csv_source(
            "counterparty_nif,counterparty_name,invoice_number,invoice_date,taxable_base,iva_rate\n"
            f"{_CIF},Papeleria Sol SL,BULK-ES-005,2026-05-01,1.234,21\n",
            tmp_path,
        )
        result = import_invoices_from_rows(rows, bucket_id=_BUCKET_ID, kind=InvoiceKind.RECEIVED, declared_country="ES")

        reason = result.refused[0].reason
        assert "1234.56" in reason, "the refusal must show a correctly-written amount"
        assert "decimal separator" in reason
        assert "1.234" in reason, "the refusal must echo what the operator actually wrote"


def test_import_still_accepts_the_canonical_euro_amount(tmp_path: Path) -> None:
    """The refusal above is specific: a canonical dot-decimal amount still imports.

    Without this, a parser that refused every amount would pass the Spanish
    tests while breaking every real import.
    """
    with isolated_runtime_profile(tmp_path=tmp_path, bucket_id=_BUCKET_ID):
        rows = _csv_source(
            "counterparty_nif,counterparty_name,invoice_number,invoice_date,taxable_base,iva_rate\n"
            f"{_CIF},Papeleria Sol SL,BULK-ES-003,2026-05-01,1234.56,21\n",
            tmp_path,
        )
        result = import_invoices_from_rows(rows, bucket_id=_BUCKET_ID, kind=InvoiceKind.RECEIVED, declared_country="ES")

        assert result.refused == ()
        assert result.created == 1


def test_import_keeps_an_already_numeric_workbook_cell_unjudged(tmp_path: Path) -> None:
    """A numeric XLSX cell carries the workbook's representation, not the operator's grammar.

    The euro-grammar cap applies to text a person wrote. A spreadsheet
    formula can yield a float with more precision than a euro carries, and
    that precision is the workbook's doing; holding it to the typed-text
    grammar would refuse rows that never passed through anyone's keyboard.
    """
    xlsx_path = tmp_path / "invoices.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.append(
        ["counterparty_nif", "counterparty_name", "invoice_number", "invoice_date", "taxable_base", "iva_rate"],
    )
    sheet.append([_CIF, "Papeleria Sol SL", "BULK-ES-004", "2026-05-01", 1000 / 3, 21])
    workbook.save(xlsx_path)

    with isolated_runtime_profile(tmp_path=tmp_path, bucket_id=_BUCKET_ID):
        rows = read_bulk_invoice_import_source(xlsx_path)
        result = import_invoices_from_rows(rows, bucket_id=_BUCKET_ID, kind=InvoiceKind.RECEIVED, declared_country="ES")

        assert result.refused == ()
        assert result.created == 1


def test_read_bulk_invoice_import_rows_reads_csv_and_xlsx_identically(tmp_path: Path) -> None:
    """The CSV and XLSX readers yield the same row content for the same data."""
    csv_path = tmp_path / "invoices.csv"
    csv_path.write_text(
        "counterparty_nif,counterparty_name,invoice_number,invoice_date,taxable_base,iva_rate\n"
        f"{_CIF},Papeleria Sol SL,BULK-E-001,2026-05-01,100.00,21\n",
        encoding="utf-8",
    )
    xlsx_path = tmp_path / "invoices.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.append(
        ["counterparty_nif", "counterparty_name", "invoice_number", "invoice_date", "taxable_base", "iva_rate"],
    )
    sheet.append([_CIF, "Papeleria Sol SL", "BULK-E-001", "2026-05-01", 100.00, 21])
    workbook.save(xlsx_path)

    csv_source = read_bulk_invoice_import_source(csv_path)
    xlsx_source = read_bulk_invoice_import_source(xlsx_path)

    assert len(csv_source.rows) == 1
    assert len(xlsx_source.rows) == 1
    csv_values = csv_source.rows[0].values
    xlsx_values = xlsx_source.rows[0].values
    assert csv_values["invoice_number"] == xlsx_values["invoice_number"] == "BULK-E-001"
    # The two readers legitimately hand back different CELL types for the same
    # figure -- csv yields the text "100.00", openpyxl yields the number 100 --
    # and the claim here is that the VALUES agree regardless. Narrowed to what
    # Decimal accepts rather than to one type, so the comparison stays the
    # cross-format one and does not quietly become a type assertion.
    csv_base, xlsx_base = csv_values["taxable_base"], xlsx_values["taxable_base"]
    assert isinstance(csv_base, str | int | float | Decimal)
    assert isinstance(xlsx_base, str | int | float | Decimal)
    assert Decimal(csv_base) == Decimal(xlsx_base)


def test_read_bulk_invoice_import_rows_rejects_unknown_extension(tmp_path: Path) -> None:
    """An unsupported file extension refuses before any row is read."""
    bad_path = tmp_path / "invoices.txt"
    bad_path.write_text("counterparty_nif\n", encoding="utf-8")
    with pytest.raises(InvoiceValidationError):
        read_bulk_invoice_import_source(bad_path)


def test_an_unrecognised_column_is_reported_and_the_file_still_imports(tmp_path: Path) -> None:
    """An unknown column is reported, never a refusal.

    This assertion was inverted. The importer used to refuse the whole file for
    one column it did not know, so a book carrying every required field plus one
    extra imported nothing. The column must be reported and the rows must land.
    """
    csv_path = tmp_path / "invoices.csv"
    csv_path.write_text(
        "counterparty_nif,counterparty_name,invoice_number,invoice_date,taxable_base,bogus\n"
        f"{_CIF},Papeleria Sol SL,BULK-F-001,2026-05-01,100.00,xyz\n",
        encoding="utf-8",
    )
    source = read_bulk_invoice_import_source(csv_path)

    assert [column.header for column in source.resolution.unmapped_columns] == ["bogus"]
    assert len(source.rows) == 1
    assert "bogus" not in source.rows[0].values

    with isolated_runtime_profile(tmp_path=tmp_path, bucket_id=_BUCKET_ID):
        result = import_invoices_from_rows(
            source, bucket_id=_BUCKET_ID, kind=InvoiceKind.RECEIVED, declared_country="ES"
        )
        assert result.created == 1
        assert result.refused == ()


def test_a_file_with_no_column_for_a_required_field_still_refuses(tmp_path: Path) -> None:
    """Positive control for the refusal that remains: no taxable base, no import.

    Removing the refuse-whole on unknown columns must not remove the refusal for
    a file that genuinely cannot supply a required field. Without this, the test
    above would pass equally against an importer that never refuses anything.
    """
    csv_path = tmp_path / "invoices.csv"
    csv_path.write_text(
        "counterparty_nif,counterparty_name,invoice_number,invoice_date\n"
        f"{_CIF},Papeleria Sol SL,BULK-F-002,2026-05-01\n",
        encoding="utf-8",
    )
    with pytest.raises(InvoiceValidationError, match="taxable_base"):
        read_bulk_invoice_import_source(csv_path)


def _workbook_with_stale_cached_formula(tmp_path: Path) -> Path:
    """Write a book whose taxable_base cell states ``=500*2`` but caches ``7.77``.

    This is the shape a spreadsheet application saves when a sheet is edited
    somewhere else and stored without recalculating: the formula and the number
    beside it disagree. openpyxl cannot author the pair, so the cached value is
    injected into the sheet XML directly -- the file is a real xlsx either way,
    and what it exercises is exactly what a real stale book would.
    """
    authored = tmp_path / "authored.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.append(
        ["counterparty_nif", "counterparty_name", "invoice_number", "invoice_date", "taxable_base", "iva_rate"],
    )
    sheet.append([_CIF, "Papeleria Sol SL", "BULK-ES-005", "2026-05-01", "=500*2", 21])
    workbook.save(authored)

    stale = tmp_path / "stale.xlsx"
    with zipfile.ZipFile(authored) as source, zipfile.ZipFile(stale, "w", zipfile.ZIP_DEFLATED) as target:
        for item in source.infolist():
            payload = source.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                payload = payload.replace(b"<f>500*2</f>", b"<f>500*2</f><v>7.77</v>")
            target.writestr(item, payload)
    return stale


def test_import_refuses_a_workbook_whose_cached_value_disagrees_with_its_formula(tmp_path: Path) -> None:
    """A stale cached value must be refused, never read as the taxable base.

    The discriminating case. A book whose formulas were never computed reads as
    a blank and would be refused by the amount grammar even with no formula
    guard at all, so it cannot tell a working guard from a missing one. Here the
    cached number is present, well-formed, and wrong: ``7.77`` sits beside a
    formula stating ``1000``. Read, it becomes the invoice's taxable base and
    reaches Modelo 303/390 as a wrong figure with nothing anywhere reporting a
    failure -- which is the whole reason a cached value is not an accepted
    financial figure.
    """
    stale = _workbook_with_stale_cached_formula(tmp_path)

    with pytest.raises(InvoiceValidationError) as exc_info:
        read_bulk_invoice_import_source(stale)

    assert "formula cached values are not accepted" in str(exc_info.value)
    assert "7.77" not in str(exc_info.value)


def test_import_refuses_an_uncomputed_formula_naming_the_formula_not_the_amount(tmp_path: Path) -> None:
    """An uncomputed formula is refused for being a formula, not for being blank.

    openpyxl authors formulas with no cached value, so this book is the one a
    non-spreadsheet tool produces. Before the formula guard the cell read as
    ``None`` and the row was refused by the decimal grammar, telling the
    operator to write the amount with a dot as the decimal separator while they
    looked at a perfectly good ``=500*2``. The guard has to fire first for the
    refusal to name the real cause.
    """
    xlsx_path = tmp_path / "uncomputed.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.append(
        ["counterparty_nif", "counterparty_name", "invoice_number", "invoice_date", "taxable_base", "iva_rate"],
    )
    sheet.append([_CIF, "Papeleria Sol SL", "BULK-ES-006", "2026-05-01", "=500*2", 21])
    workbook.save(xlsx_path)

    with pytest.raises(InvoiceValidationError) as exc_info:
        read_bulk_invoice_import_source(xlsx_path)

    message = str(exc_info.value)
    assert "formula cached values are not accepted" in message
    assert "decimal separator" not in message


# ---------------------------------------------------------------------------
# Counterparty country: recordable-but-blank refuses, unanswerable is declared
# ---------------------------------------------------------------------------

# A minimal accounting export: the shape every fixture above already uses, and
# the shape a book that predates any country column has. It can state no
# country for any row, which is what makes it the unanswerable case rather
# than a blank-cell one.
_EXPORT_WITHOUT_COUNTRY_COLUMN = (
    "counterparty_nif,counterparty_name,invoice_number,invoice_date,taxable_base,iva_rate\n"
    f"{_CIF},Papeleria Sol SL,BULK-CTY-001,2026-05-01,100.00,21\n"
)

# The same export from a book that DOES carry the column, with one EU supplier
# whose country the export left empty -- the row-level omission.
_EXPORT_WITH_COUNTRY_COLUMN_ONE_BLANK = (
    "counterparty_nif,counterparty_name,invoice_number,invoice_date,taxable_base,iva_rate,country_code\n"
    f"{_CIF},Papeleria Sol SL,BULK-CTY-010,2026-05-01,100.00,21,ES\n"
    "DE811907980,Papier Nord GmbH,BULK-CTY-011,2026-05-02,200.00,21,\n"
)


def test_import_refuses_a_file_that_can_state_no_country_at_all(tmp_path: Path) -> None:
    """A book with no country column is refused once, before any row is read.

    Inferring Spain here would silently reclassify a foreign supplier as
    domestic, and the invoice source resolver routes IVA on exactly that
    comparison. Refusing per-row instead would emit one identical failure per
    row and bury the single fact the operator needs.
    """
    with isolated_runtime_profile(tmp_path=tmp_path, bucket_id=_BUCKET_ID):
        source = _csv_source(_EXPORT_WITHOUT_COUNTRY_COLUMN, tmp_path)

        with pytest.raises(InvoiceValidationError, match="country"):
            import_invoices_from_rows(source, bucket_id=_BUCKET_ID, kind=InvoiceKind.RECEIVED)


def test_a_declared_country_lets_a_book_without_the_column_import(tmp_path: Path) -> None:
    """Positive control: the same file imports once the operator states a country.

    Without this the refusal above is indistinguishable from an importer that
    refuses this file for some unrelated reason, and the recourse the refusal
    names would be unproven.
    """
    with isolated_runtime_profile(tmp_path=tmp_path, bucket_id=_BUCKET_ID):
        source = _csv_source(_EXPORT_WITHOUT_COUNTRY_COLUMN, tmp_path)

        result = import_invoices_from_rows(
            source,
            bucket_id=_BUCKET_ID,
            kind=InvoiceKind.RECEIVED,
            declared_country="ES",
        )

        assert result.created == 1
        assert result.refused == ()


def test_a_blank_country_cell_refuses_only_its_own_row(tmp_path: Path) -> None:
    """A file that HAS the column is not refused whole; the blank row is.

    This is the recordable-but-blank class: the book demonstrably can carry the
    fact, so an empty cell is an omission specific to that row. Partial-success
    semantics keep every row that did state a country.
    """
    with isolated_runtime_profile(tmp_path=tmp_path, bucket_id=_BUCKET_ID):
        source = _csv_source(_EXPORT_WITH_COUNTRY_COLUMN_ONE_BLANK, tmp_path)

        result = import_invoices_from_rows(source, bucket_id=_BUCKET_ID, kind=InvoiceKind.RECEIVED)

        assert result.created == 1
        assert [(f.row_number, f.field) for f in result.refused] == [(3, "country_code")]


def test_a_declared_country_never_overrides_a_row_that_states_one(tmp_path: Path) -> None:
    """The whole-import country is a fallback for an unanswerable file, not an override.

    If it silently won over a stated cell, declaring ES to get a legacy book
    moving would rewrite every foreign counterparty in a book that had them
    right -- the same reclassification the default caused, reintroduced through
    the recourse.
    """
    with isolated_runtime_profile(tmp_path=tmp_path, bucket_id=_BUCKET_ID):
        source = _csv_source(_EXPORT_WITH_COUNTRY_COLUMN_ONE_BLANK, tmp_path)

        result = import_invoices_from_rows(
            source,
            bucket_id=_BUCKET_ID,
            kind=InvoiceKind.RECEIVED,
            declared_country="DE",
        )

        assert result.created == 2
        assert result.refused == ()

        # The stated ES row kept ES rather than taking the declared DE, and
        # the blank row took DE rather than refusing. Read back through the
        # real repository, so this is what was persisted and not what the
        # in-memory result happened to carry.
        catalogue = InvoiceCatalogueRepository(bucket_id=_BUCKET_ID).load()
        by_number = {inv.invoice_number: inv.counterparty_country for inv in catalogue.invoices.values()}
        assert by_number["BULK-CTY-010"] == "ES"
        assert by_number["BULK-CTY-011"] == "DE"
