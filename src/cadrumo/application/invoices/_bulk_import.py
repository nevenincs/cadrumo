"""Bulk CSV/XLSX transport for creating catalogue :class:`Invoice` records.

The accountant/gestor batch case: a spreadsheet of invoice rows (counterparty
NIF, invoice number, date, taxable base, IVA rate) is turned into one
:class:`~domain.invoices.Invoice` per row. This module is a typed transport
over :func:`~application.invoices.create_catalogue_invoice` -- the sole
sanctioned :class:`Invoice` writer (``aeat-architecture-boundaries``);
it never persists a row itself.

Each row's identity is the same content-derived
:attr:`~domain.invoices.Invoice.invoice_id` hash the single-invoice
``catalogue create`` verb and the evidence-confirm slice use, so a re-import of
an unchanged file is a guarded no-op per row
(``aeat-cli-contract``): an already-catalogued
identical row is reported ``skipped_duplicate``, never re-written and never
raised as an error. A malformed row (missing required field, invalid NIF,
unsupported IVA rate) is collected as a ``refused`` row carrying its 1-based
CSV row number and the failing field, and the remaining valid rows are still
applied -- partial-success semantics matching the ledger CSV import and bulk
classify pattern (``no-silent-under-declaration``: a bad row is reported, never
silently dropped).

See Also:
    :func:`~application.invoices.import_invoices_from_rows`
        Public application facade for applying validated bulk rows.
    :func:`~application.invoices.create_catalogue_invoice`
        Single catalogue writer invoked for every accepted row.
    :func:`~application.invoices.create_invoice_via_wizard`
        Manual single-invoice path with the same writer and idempotent identity.
    :func:`~application.ledger.invoice_confirmation.confirm_invoice_draft_from_evidence`
        Evidence-confirm path that also delegates the final invoice write to
        the catalogue writer.
"""

from __future__ import annotations

from collections.abc import Iterable, Mapping
from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import ClassVar, Literal

from pydantic import BaseModel, Field, NonNegativeInt, ValidationError

from ...adapters.persistence.profile.invoices import InvoiceCatalogueRepository
from ...core.country_code import CountryCodeAlpha2
from ...core.decimal._coerce import coerce_decimal, normalize_decimal_separators
from ...core.decimal._grammar import try_parse_canonical_decimal
from ...core.external_constants import DEFAULT_CURRENCY
from ...core.models import STRICT_FROZEN_CONFIG
from ...core.parsing import IsoCurrencyCode, parse_iso8601_date
from ...core.tabular import TabularSourceError, coerce_cell_text, normalize_tabular_bytes
from ...core.workbook import FORMULA_CELL_REFUSAL, WorkbookCell, first_formula_cell_column
from ...domain.invoices.errors import InvoiceValidationError
from ...domain.invoices.protocols import InvoiceCatalogueRepositoryProtocol
from ...domain.iva.classification import InvoiceKind
from ._bulk_import_columns import (
    BulkImportColumnResolution,
    ColumnRoleMapper,
    resolve_bulk_import_columns,
)
from ._creation import build_catalogue_invoice, create_catalogue_invoice

__all__ = [
    "BULK_INVOICE_IMPORT_ALLOWED_COLUMNS",
    "BULK_INVOICE_IMPORT_REQUIRED_COLUMNS",
    "BulkImportSourceRow",
    "BulkInvoiceImportResult",
    "BulkInvoiceImportRow",
    "BulkInvoiceImportRowFailure",
    "BulkInvoiceImportSource",
    "import_invoices_from_rows",
    "read_bulk_invoice_import_source",
]

BULK_INVOICE_IMPORT_REQUIRED_COLUMNS: frozenset[str] = frozenset(
    {
        "counterparty_nif",
        "counterparty_name",
        "invoice_number",
        "invoice_date",
        "taxable_base",
    },
)

BULK_INVOICE_IMPORT_OPTIONAL_COLUMNS: frozenset[str] = frozenset(
    {
        "iva_rate",
        "retencion_amount",
        "currency",
        "country_code",
        "notes",
    },
)

BULK_INVOICE_IMPORT_ALLOWED_COLUMNS: frozenset[str] = (
    BULK_INVOICE_IMPORT_REQUIRED_COLUMNS | BULK_INVOICE_IMPORT_OPTIONAL_COLUMNS
)


class BulkImportSourceRow(BaseModel):
    """One source row, its cells already keyed by the importer field they feed.

    Attributes:
        row_number: 1-based row number in the source file, header included, so
            a refusal names the row an operator counts in a spreadsheet.
        values: One cell per importer field, in the source's own
            representation. A delimited file yields text exactly as printed; a
            workbook yields the type the workbook itself chose, so a numeric
            cell is never stringified into the operator's typed-text grammar.
    """

    model_config = STRICT_FROZEN_CONFIG

    row_number: int = Field(ge=1)
    values: dict[str, object]


class BulkInvoiceImportSource(BaseModel):
    """One bulk-import file read, resolved and ready to apply.

    Attributes:
        rows: Every non-blank data row.
        resolution: How the file's own headers resolved onto importer fields,
            carrying the columns that resolved to nothing for reporting.
        decimal_separator: The convention the file writes amounts in, detected
            once for the whole file. Amounts are converted with it at parse
            time rather than being rewritten in ``rows``.
    """

    model_config = STRICT_FROZEN_CONFIG

    rows: tuple[BulkImportSourceRow, ...]
    resolution: BulkImportColumnResolution
    decimal_separator: Literal[",", "."] = "."


class BulkInvoiceImportRow(BaseModel):
    """One parsed row from a bulk invoice import CSV/XLSX file.

    Mirrors the operator fields ``aeat app ledger invoice add``
    accepts one at a time; ``taxable_base`` and ``iva_rate`` synthesise the
    single line item exactly as :func:`build_catalogue_invoice` does for the
    single-invoice verb, so a bulk row produces an identical
    :class:`~domain.invoices.Invoice` shape.
    """

    model_config = STRICT_FROZEN_CONFIG

    counterparty_nif: str = Field(min_length=1)
    counterparty_name: str = Field(min_length=1)
    invoice_number: str = Field(min_length=1)
    invoice_date: date
    taxable_base: Decimal
    iva_rate: Decimal | None = None
    retencion_amount: Decimal | None = None
    currency: IsoCurrencyCode = DEFAULT_CURRENCY
    # Required, and deliberately not defaulted to Spain. The counterparty's
    # country decides whether the invoice is treated as domestic or as an
    # intra-community operation, so inferring ES for a row that never stated
    # one silently reclassifies a foreign supplier as domestic and drops the
    # treatment that classification carries. A file that cannot state a
    # country per row makes the operator declare one for the whole import
    # instead; the inference is then theirs and explicit, never ours and
    # silent.
    country_code: CountryCodeAlpha2
    notes: str = ""


class BulkInvoiceImportRowFailure(BaseModel):
    """One row that could not be parsed or persisted during a bulk import."""

    model_config = STRICT_FROZEN_CONFIG

    row_number: int = Field(ge=1)
    field: str = Field(min_length=1)
    reason: str = Field(min_length=1)


class BulkInvoiceImportResult(BaseModel):
    """Aggregate outcome of one bulk invoice import run.

    Uses the same partial-success semantics as the ledger CSV import and bulk
    classify surfaces: every parseable, non-duplicate row is created; rows that
    fail validation are collected in ``refused``; rows whose derived
    ``invoice_id`` already exists in the catalogue are counted in
    ``skipped_duplicate`` (the idempotent-guarded re-import no-op) rather than
    raised as an error.
    """

    model_config = STRICT_FROZEN_CONFIG

    rows: NonNegativeInt
    created: NonNegativeInt
    skipped_duplicate: NonNegativeInt
    refused: tuple[BulkInvoiceImportRowFailure, ...] = ()
    created_invoice_ids: tuple[str, ...] = ()


def _parse_row_date(raw: str, *, row_number: int, field: str) -> date:
    try:
        value = parse_iso8601_date(raw)
    except ValueError as exc:
        raise _RowParseError(row_number=row_number, field=field, reason=f"invalid ISO-8601 date: {raw!r}") from exc
    if value is None:
        raise _RowParseError(row_number=row_number, field=field, reason=f"invalid ISO-8601 date: {raw!r}")
    return value


def _parse_row_decimal(raw: object, *, row_number: int, field: str) -> Decimal:
    """Parse one spreadsheet cell into a euro-grammar :class:`~decimal.Decimal`.

    A cell that already arrives numeric — an XLSX float, int, or Decimal — is
    coerced as-is, exactly as the bank-statement importer's
    ``_already_numeric_amount`` does. The workbook, not the operator, chose
    that representation, so its precision is not the operator's grammar to
    judge; a float is routed through :func:`coerce_decimal` so the printed
    precision survives instead of the binary expansion.

    TEXT is the operator's own writing and is held to the canonical euro
    grammar. That matters for one shape specifically. A gestor writing
    Spanish types ``1.234`` for one thousand two hundred and thirty-four
    euros, and a bare ``Decimal`` reads it as ``1.234`` — one euro
    twenty-three, a thousandfold under-declaration of an invoice's taxable
    base, reaching Modelo 303/390 as a wrong number rather than as any kind
    of failure. The comma spellings (``1.234,56``, ``1234,56``) were always
    refused loudly; this dot-grouped one was the only one that passed
    silently, because it happens to be legal syntax for a different number.
    The two-fractional-digit cap is what separates them: no euro amount
    carries three decimals, so the shape that means "thousands" in Spanish
    can be refused without guessing which grammar was intended.

    Refusing rather than reinterpreting is the deliberate posture — the row
    is reported with its number and field so the operator corrects the
    source, and no amount is inferred on their behalf.
    """
    if isinstance(raw, Decimal):
        numeric: Decimal | None = raw
    elif isinstance(raw, bool):
        # bool is an int subclass; a TRUE cell is not an amount.
        numeric = None
    elif isinstance(raw, int):
        numeric = Decimal(raw)
    elif isinstance(raw, float):
        numeric = coerce_decimal(raw)
    else:
        numeric = try_parse_canonical_decimal(coerce_cell_text(raw), max_fraction_digits=2)
    if numeric is None or not numeric.is_finite():
        raise _RowParseError(
            row_number=row_number,
            field=field,
            reason=(
                f"invalid decimal amount: {coerce_cell_text(raw)!r} — write the amount with a dot as the "
                "decimal separator, no thousands separator, and at most two decimals, so one "
                "thousand two hundred and thirty-four euros fifty-six is '1234.56'. The Spanish "
                "forms '1.234,56' and '1.234' are refused rather than guessed at, because '1.234' "
                "is also a valid way to write one euro twenty-three."
            ),
        )
    return numeric


class _RowParseError(Exception):
    """Internal control-flow exception carrying one row's failure detail."""

    __bare_base_rationale__: ClassVar[str] = (
        "private row parser control-flow carrier; converted to BulkInvoiceImportRowFailure before leaving the module"
    )

    def __init__(self, *, row_number: int, field: str, reason: str) -> None:
        super().__init__(reason)
        self.row_number = row_number
        self.field = field
        self.reason = reason


def _canonicalise_amount_text(raw: object, *, decimal_separator: Literal[",", "."]) -> object:
    """Rewrite a comma-convention amount into canonical form before the grammar check.

    Only text is touched, and only when the file's own detected convention says
    the comma is its decimal mark. That detection is what makes the rewrite safe:
    the ambiguity :func:`_parse_row_decimal` refuses — a bare ``1.234`` that could
    be one thousand or one point two three four — is resolved by evidence from the
    whole file rather than guessed at per cell. A dot-convention file is passed
    through untouched, so the strict grammar still governs it.

    A trailing percent sign is dropped for the same reason: a rate column printed
    ``21%`` states a rate, not a different number.
    """
    if not isinstance(raw, str):
        return raw
    text = raw.strip().removesuffix("%").strip()
    if decimal_separator != ",":
        return text
    return normalize_decimal_separators(text, strip_thousands="." in text and "," in text)


def _parse_bulk_invoice_row(
    raw_row: Mapping[str, object],
    *,
    row_number: int,
    decimal_separator: Literal[",", "."] = ".",
    declared_country: str | None = None,
) -> BulkInvoiceImportRow:
    """Return a validated :class:`BulkInvoiceImportRow`, or raise :class:`_RowParseError`.

    Every field failure is attributed to its originating column name so a
    refusal names both the row number and the field that failed
    (``no-silent-under-declaration``) rather than a bare "row invalid".

    ``declared_country`` is the whole-import country the caller supplies when
    the source carries no country column at all. It is a fallback for a file
    that cannot express the fact, never for a row that simply left the cell
    empty: a blank cell in a file that HAS the column is an omission specific
    to that row, and is refused as one.
    """
    missing = [column for column in BULK_INVOICE_IMPORT_REQUIRED_COLUMNS if not coerce_cell_text(raw_row.get(column))]
    if missing:
        raise _RowParseError(row_number=row_number, field=missing[0], reason="required field is missing or blank")

    invoice_date_raw = coerce_cell_text(raw_row.get("invoice_date"))
    invoice_date = _parse_row_date(invoice_date_raw, row_number=row_number, field="invoice_date")

    # The raw cell is handed over unstringified so an already-numeric workbook
    # value keeps its own representation; only operator-written TEXT is held to
    # the euro grammar.
    taxable_base = _parse_row_decimal(
        _canonicalise_amount_text(raw_row.get("taxable_base"), decimal_separator=decimal_separator),
        row_number=row_number,
        field="taxable_base",
    )

    iva_rate_raw = coerce_cell_text(raw_row.get("iva_rate"))
    iva_rate = (
        _parse_row_decimal(
            _canonicalise_amount_text(raw_row.get("iva_rate"), decimal_separator=decimal_separator),
            row_number=row_number,
            field="iva_rate",
        )
        if iva_rate_raw
        else None
    )

    retencion_raw = coerce_cell_text(raw_row.get("retencion_amount"))
    retencion_amount = (
        _parse_row_decimal(
            _canonicalise_amount_text(raw_row.get("retencion_amount"), decimal_separator=decimal_separator),
            row_number=row_number,
            field="retencion_amount",
        )
        if retencion_raw
        else None
    )

    currency_raw = coerce_cell_text(raw_row.get("currency")) or DEFAULT_CURRENCY
    country_code_raw = coerce_cell_text(raw_row.get("country_code")) or declared_country or ""
    if not country_code_raw:
        raise _RowParseError(
            row_number=row_number,
            field="country_code",
            reason="counterparty country is missing or blank; state it in the row or declare one for the import",
        )

    try:
        return BulkInvoiceImportRow(
            counterparty_nif=coerce_cell_text(raw_row.get("counterparty_nif")),
            counterparty_name=coerce_cell_text(raw_row.get("counterparty_name")),
            invoice_number=coerce_cell_text(raw_row.get("invoice_number")),
            invoice_date=invoice_date,
            taxable_base=taxable_base,
            iva_rate=iva_rate,
            retencion_amount=retencion_amount,
            currency=currency_raw,
            country_code=country_code_raw,
            notes=coerce_cell_text(raw_row.get("notes")),
        )
    except ValidationError as exc:
        first = exc.errors()[0] if exc.errors() else {"loc": ("row",), "msg": "invalid row"}
        field = str(first["loc"][0]) if first.get("loc") else "row"
        raise _RowParseError(row_number=row_number, field=field, reason=str(first.get("msg", "invalid row"))) from exc


def _assert_required_fields_present(resolution: BulkImportColumnResolution) -> None:
    """Refuse a file that cannot supply a required field under any column.

    This is the one refusal the header can still cause, and it is not the
    refuse-whole this Step removed: an unknown column is reported and the file
    still imports, but a book carrying no taxable base at all has no row to
    create. The refusal names the fields, and lists what the file did carry, so
    the operator can see whether a column was simply not recognised.
    """
    missing = BULK_INVOICE_IMPORT_REQUIRED_COLUMNS - resolution.fields_present
    if not missing:
        return
    raise InvoiceValidationError(
        "bulk invoice import file supplies no column for required field(s): " + ", ".join(sorted(missing)),
        translated_message="application.invoices.bulk_import.errors.missing_columns",
        context={
            "missing_columns": ", ".join(sorted(missing)),
            "unmapped_columns": ", ".join(column.header for column in resolution.unmapped_columns) or "none",
        },
    )


def _assert_country_is_answerable(
    resolution: BulkImportColumnResolution,
    *,
    declared_country: str | None,
) -> None:
    """Refuse a file that can state no counterparty country, before any row is read.

    A book with no country column cannot answer the question for any row, so
    letting the import run would produce one identical refusal per row and
    bury the single fact the operator needs. This raises once, up front, and
    names the recourse: add the column, or declare one country for the whole
    import. A file that HAS the column is not refused here even when cells are
    blank -- those are per-row omissions and stay per-row refusals, so the
    rows that do state a country still import.
    """
    if declared_country or "country_code" in resolution.fields_present:
        return
    raise InvoiceValidationError(
        "bulk invoice import file supplies no counterparty country column and no country was declared "
        "for the import; add a country_code column or declare one country for the whole file",
        translated_message="application.invoices.bulk_import.errors.country_unanswerable",
        context={
            "unmapped_columns": ", ".join(column.header for column in resolution.unmapped_columns) or "none",
        },
    )


def _read_delimited_source(path: Path, *, mapper: ColumnRoleMapper | None) -> BulkInvoiceImportSource:
    """Read a delimited invoice book of any dialect into a resolved source."""
    try:
        source_bytes = path.read_bytes()
    except OSError as exc:
        raise InvoiceValidationError(
            "bulk invoice import file could not be read",
            translated_message="application.invoices.bulk_import.errors.file_read_failed",
            context={"path_name": path.name, "error_type": type(exc).__name__},
        ) from exc
    try:
        table = normalize_tabular_bytes(source_bytes)
    except TabularSourceError as exc:
        raise InvoiceValidationError(
            "bulk invoice import file carries no readable table",
            translated_message="application.invoices.bulk_import.errors.unreadable_table",
            context={"path_name": path.name, "detail": str(exc)},
        ) from exc
    resolution = resolve_bulk_import_columns(
        table.headers, mapper=mapper, required_fields=BULK_INVOICE_IMPORT_REQUIRED_COLUMNS
    )
    _assert_required_fields_present(resolution)
    field_by_index = resolution.field_by_index
    rows = tuple(
        BulkImportSourceRow(
            row_number=row.source_line_number,
            values={field: row.cells[index] for index, field in field_by_index.items() if index < len(row.cells)},
        )
        for row in table.rows
    )
    # A file written in the product's OWN column names is the operator's
    # template, and its amounts stay under the canonical euro grammar: a bare
    # ``1.234`` there is genuinely ambiguous and must refuse rather than be
    # guessed at. Only a book the mapping lane had to interpret -- foreign
    # headers, its own conventions -- is read under its detected separator,
    # where the whole file's evidence settles what a comma means.
    return BulkInvoiceImportSource(
        rows=rows,
        resolution=resolution,
        decimal_separator=table.dialect.decimal_separator if resolution.consulted_mapping_lane else ".",
    )


def _refuse_formula_cells(cells: Iterable[WorkbookCell], *, path: Path, row_number: int) -> None:
    """Refuse the book when any cell of *cells* states a formula instead of a value."""
    column_number = first_formula_cell_column(cells)
    if column_number is None:
        return
    raise InvoiceValidationError(
        f"bulk invoice import file contains formula cell at row {row_number}, "
        f"column {column_number}; {FORMULA_CELL_REFUSAL}",
        translated_message="application.invoices.bulk_import.errors.formula_cell",
        context={"path_name": path.name, "row": str(row_number), "column": str(column_number)},
    )


def _read_workbook_source(path: Path, *, mapper: ColumnRoleMapper | None) -> BulkInvoiceImportSource:
    """Read a workbook invoice book into a resolved source.

    A workbook states its own cell types, so there is no dialect to detect: a
    numeric cell arrives numeric and keeps the representation the workbook
    chose. Only the header resolution is shared with the delimited path.

    **A formula cell is refused rather than read.** A cached value is not the
    formula's value: a sheet saved without recalculation carries whatever number
    was last computed into it, and reading that would enter a figure nobody
    typed and the formula does not produce as an invoice's taxable base --
    reaching Modelo 303/390 as a wrong number rather than as any kind of
    failure, exactly as :func:`_parse_row_decimal` refuses rather than
    reinterprets a Spanish ``1.234``. A book whose formulas were never computed
    at all is the same defect wearing a blank: refusing on the formula names the
    real cause, where reading its empty cached value would report a missing
    amount and send the operator looking at their decimal separator.

    The casilla-value spreadsheet reader and the bank-statement workbook
    provider refuse on this ground already, in these words; this is the third
    member of that family, not a new posture.
    """
    from openpyxl import load_workbook

    workbook = load_workbook(filename=path, read_only=True, data_only=False)
    try:
        worksheet = workbook.worksheets[0]
        rows_iter = worksheet.iter_rows()
        try:
            header_cells = next(rows_iter)
        except StopIteration:
            return BulkInvoiceImportSource(rows=(), resolution=resolve_bulk_import_columns((), mapper=None))
        _refuse_formula_cells(header_cells, path=path, row_number=1)
        headers = [coerce_cell_text(cell.value) for cell in header_cells]
        resolution = resolve_bulk_import_columns(
            headers, mapper=mapper, required_fields=BULK_INVOICE_IMPORT_REQUIRED_COLUMNS
        )
        _assert_required_fields_present(resolution)
        field_by_index = resolution.field_by_index
        rows: list[BulkImportSourceRow] = []
        for row_number, row_cells in enumerate(rows_iter, start=2):  # header is row 1
            _refuse_formula_cells(row_cells, path=path, row_number=row_number)
            row = [cell.value for cell in row_cells]
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue
            values: dict[str, object] = {}
            for index, field in field_by_index.items():
                cell_value = row[index] if index < len(row) else None
                values[field] = cell_value.isoformat() if isinstance(cell_value, date) else cell_value
            rows.append(BulkImportSourceRow(row_number=row_number, values=values))
        return BulkInvoiceImportSource(rows=tuple(rows), resolution=resolution)
    finally:
        workbook.close()


def read_bulk_invoice_import_source(
    path: Path,
    *,
    mapper: ColumnRoleMapper | None = None,
) -> BulkInvoiceImportSource:
    """Read a CSV, TSV or XLSX invoice book into rows keyed by importer field.

    A column whose header already names an importer field binds to it outright.
    Anything left over is put to ``mapper`` once for the whole file, and a column
    that still resolves to nothing is carried on the returned resolution for
    reporting — **the file is never refused for carrying a column the importer
    does not know**.

    ``row_number`` is 1-based against the file's own rows, so a refusal names
    exactly the row an operator would count in a spreadsheet application.

    Args:
        path: The invoice book to read.
        mapper: Establishes roles for columns exact matching did not resolve.

    Returns:
        The resolved source, ready to apply.

    Raises:
        InvoiceValidationError: The extension is unsupported, the file cannot be
            read, or no column supplies a required field.
    """
    suffix = path.suffix.lower()
    if suffix in {".csv", ".tsv"}:
        return _read_delimited_source(path, mapper=mapper)
    if suffix in {".xlsx", ".xlsm"}:
        return _read_workbook_source(path, mapper=mapper)
    raise InvoiceValidationError(
        "bulk invoice import file must be .csv, .tsv or .xlsx",
        translated_message="application.invoices.bulk_import.errors.unsupported_extension",
        context={"path_name": path.name, "extension": suffix},
    )


def import_invoices_from_rows(
    source: BulkInvoiceImportSource,
    *,
    bucket_id: str,
    kind: InvoiceKind,
    declared_country: str | None = None,
    repository: InvoiceCatalogueRepositoryProtocol | None = None,
) -> BulkInvoiceImportResult:
    """Create one catalogue :class:`Invoice` per valid row in *rows*.

    Every accepted row is handed to
    :func:`~application.invoices.create_catalogue_invoice` -- this
    function never persists a row itself
    (``aeat-architecture-boundaries``). Because
    :class:`~domain.invoices.Invoice` identity is a content-derived hash of
    ``(kind, invoice_number, issued_at, counterparty_tax_id, currency,
    grand_total)``, re-importing the identical file a second time resolves
    every row to its already-catalogued ``invoice_id`` and reports it in
    ``skipped_duplicate`` -- no second write, no raised error
    (``aeat-cli-contract``). A row whose resolved
    fields genuinely differ (a corrected amount, a new invoice number) mints a
    distinct record rather than overwriting one filer's data with another's.

    A malformed row (missing/blank required field, invalid date, unsupported
    IVA rate percentage, invalid NIF) is collected in ``refused`` with its row
    number and the failing field name; the remaining valid rows still import
    (partial-success semantics).

    ``declared_country`` states one counterparty country for the whole import.
    It is required only when the source carries no country column at all, and
    a source that carries the column ignores it row by row: an operator whose
    book cannot express the fact declares it once and consciously, rather than
    having Spain inferred for a foreign counterparty.
    """
    _assert_country_is_answerable(source.resolution, declared_country=declared_country)
    repo = repository or InvoiceCatalogueRepository(bucket_id=bucket_id)
    catalogue = repo.load()
    existing_ids = set(catalogue.invoices)

    refused: list[BulkInvoiceImportRowFailure] = []
    created = 0
    skipped_duplicate = 0
    created_ids: list[str] = []

    for source_row in source.rows:
        row_number = source_row.row_number
        try:
            parsed = _parse_bulk_invoice_row(
                source_row.values,
                row_number=row_number,
                decimal_separator=source.decimal_separator,
                declared_country=declared_country,
            )
        except _RowParseError as exc:
            refused.append(BulkInvoiceImportRowFailure(row_number=exc.row_number, field=exc.field, reason=exc.reason))
            continue

        try:
            candidate = build_catalogue_invoice(
                bucket_id=bucket_id,
                kind=kind,
                counterparty_name=parsed.counterparty_name,
                counterparty_tax_id=parsed.counterparty_nif,
                counterparty_country=parsed.country_code,
                invoice_number=parsed.invoice_number,
                issued_at=parsed.invoice_date,
                taxable_base=parsed.taxable_base,
                iva_rate=parsed.iva_rate,
                retention_amount=parsed.retencion_amount,
                currency=parsed.currency,
                notes=parsed.notes,
            )
        except (InvoiceValidationError, ValidationError) as exc:
            reason = str(exc.errors()[0].get("msg", str(exc))) if isinstance(exc, ValidationError) else str(exc)
            refused.append(BulkInvoiceImportRowFailure(row_number=row_number, field="invoice", reason=reason))
            continue

        if candidate.invoice_id in existing_ids:
            # Guarded idempotent retry (aeat-cli-contract):
            # a re-import of the same file resolves every row to an
            # already-catalogued identity -- report it, do not re-write or raise.
            skipped_duplicate += 1
            continue

        result = create_catalogue_invoice(
            invoice=candidate,
            repository=repo,
        )
        existing_ids.add(result.invoice.invoice_id)
        created_ids.append(result.invoice.invoice_id)
        created += 1

    return BulkInvoiceImportResult(
        rows=len(source.rows),
        created=created,
        skipped_duplicate=skipped_duplicate,
        refused=tuple(refused),
        created_invoice_ids=tuple(created_ids),
    )
