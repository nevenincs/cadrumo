"""Unit tests for the casilla-value spreadsheet parser."""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from openpyxl import Workbook, load_workbook

from ..action_errors import ModeloLocalObservationError
from ..local_observation_spreadsheet import parse_casilla_value_spreadsheet

pytestmark = [pytest.mark.unit, pytest.mark.hex_application]


def _write_xlsx_with_stale_formula_cache(path: Path) -> None:
    """Create a real workbook whose formula and cached amount contradict."""
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.append(["casilla_code", "value"])
    sheet.append(["01", 1])
    workbook.save(path)
    workbook.close()

    worksheet_member = "xl/worksheets/sheet1.xml"
    original_cell = b'<c r="B2" t="n"><v>1</v></c>'
    formula_cell_with_stale_cache = b'<c r="B2"><f>1+1</f><v>9999</v></c>'
    with ZipFile(path) as source_archive:
        members = {member.filename: source_archive.read(member) for member in source_archive.infolist()}
    assert members[worksheet_member].count(original_cell) == 1
    members[worksheet_member] = members[worksheet_member].replace(original_cell, formula_cell_with_stale_cache)
    with ZipFile(path, mode="w", compression=ZIP_DEFLATED) as destination_archive:
        for filename, contents in members.items():
            destination_archive.writestr(filename, contents)


def test_parse_csv_with_header_row(tmp_path: Path) -> None:
    """A CSV with an explicit ``casilla_code,value`` header parses by column name."""
    path = tmp_path / "m130-2025q4.csv"
    path.write_text(
        "casilla_code,value\n01,1234.56\n03,789.10\n",
        encoding="utf-8",
    )

    values = parse_casilla_value_spreadsheet(path)

    assert values == {"01": Decimal("1234.56"), "03": Decimal("789.10")}


def test_parse_csv_positional_headerless(tmp_path: Path) -> None:
    """A headerless two-column CSV falls back to positional column mapping."""
    path = tmp_path / "sheet.csv"
    path.write_text("01,100\n02,200.5\n", encoding="utf-8")

    values = parse_casilla_value_spreadsheet(path)

    assert values == {"01": Decimal("100"), "02": Decimal("200.5")}


def test_parse_csv_with_alias_headers(tmp_path: Path) -> None:
    """Header aliases (casilla / valor) are recognised, case-insensitively."""
    path = tmp_path / "sheet.csv"
    path.write_text("Casilla,Valor\n05,10.00\n06,-3.25\n", encoding="utf-8")

    values = parse_casilla_value_spreadsheet(path)

    assert values == {"05": Decimal("10.00"), "06": Decimal("-3.25")}


def test_parse_xlsx_with_header_row(tmp_path: Path) -> None:
    """An XLSX workbook with a header row parses identically to CSV."""
    path = tmp_path / "sheet.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.append(["casilla_code", "value"])
    sheet.append(["01", 1234.56])
    sheet.append(["03", 789.10])
    workbook.save(path)

    values = parse_casilla_value_spreadsheet(path)

    assert values == {"01": Decimal("1234.56"), "03": Decimal("789.1")}


def test_parse_xlsx_refuses_stale_formula_cache_before_value_materialisation(tmp_path: Path) -> None:
    """A cached amount that contradicts its formula cannot become prefill evidence."""
    path = tmp_path / "stale-formula.xlsx"
    _write_xlsx_with_stale_formula_cache(path)

    formula_workbook = load_workbook(filename=path, read_only=True, data_only=False)
    cached_workbook = load_workbook(filename=path, read_only=True, data_only=True)
    try:
        formula_sheet = formula_workbook.active
        cached_sheet = cached_workbook.active
        assert formula_sheet is not None
        assert cached_sheet is not None
        assert formula_sheet["B2"].value == "=1+1"
        assert cached_sheet["B2"].value == 9999
    finally:
        formula_workbook.close()
        cached_workbook.close()

    with pytest.raises(ModeloLocalObservationError) as exc_info:
        parse_casilla_value_spreadsheet(path)

    # The refusal renders from the catalogue; its detail rides the context.
    assert exc_info.value.context
    assert exc_info.value.context == {"path": str(path), "row": "2", "column": "2"}


def test_parse_csv_rejects_duplicate_casilla_rows_before_last_writer_wins(tmp_path: Path) -> None:
    """Two amounts for one casilla are ambiguous; row order cannot select one."""
    path = tmp_path / "duplicate.csv"
    path.write_text("casilla_code,value\n01,1000.00\n01,1.00\n", encoding="utf-8")

    with pytest.raises(ModeloLocalObservationError) as exc_info:
        parse_casilla_value_spreadsheet(path)

    # The refusal renders from the catalogue; its detail rides the context.
    assert exc_info.value.context
    context = exc_info.value.context
    assert context is not None
    assert context["path"] == str(path)
    assert context["malformed_row_count"] == "1"
    # The refused row is named, not merely counted.
    assert "duplicate casilla_code" in str(context["malformed_rows"])


def test_parse_xlsx_rejects_duplicate_casilla_rows_even_when_values_match(tmp_path: Path) -> None:
    """Coordinate uniqueness applies independently of coincidentally equal values."""
    path = tmp_path / "duplicate.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.append(["casilla_code", "value"])
    sheet.append(["01", 1000])
    sheet.append(["01", 1000])
    workbook.save(path)

    with pytest.raises(ModeloLocalObservationError) as exc_info:
        parse_casilla_value_spreadsheet(path)

    # The refusal renders from the catalogue; its detail rides the context.
    assert exc_info.value.context
    context = exc_info.value.context
    assert context is not None
    assert context["path"] == str(path)
    assert context["malformed_row_count"] == "1"
    # The refused row is named, not merely counted.
    assert "duplicate casilla_code" in str(context["malformed_rows"])


def test_parse_rejects_non_numeric_value(tmp_path: Path) -> None:
    """A non-numeric value raises, naming the offending row and casilla."""
    path = tmp_path / "sheet.csv"
    path.write_text("casilla_code,value\n01,not-a-number\n", encoding="utf-8")

    with pytest.raises(ModeloLocalObservationError):
        parse_casilla_value_spreadsheet(path)


def test_parse_rejects_incomplete_row(tmp_path: Path) -> None:
    """A row with a casilla code but no value raises, naming the row."""
    path = tmp_path / "sheet.csv"
    path.write_text("casilla_code,value\n01,\n02,50\n", encoding="utf-8")

    with pytest.raises(ModeloLocalObservationError):
        parse_casilla_value_spreadsheet(path)


def test_parse_rejects_missing_file(tmp_path: Path) -> None:
    """A nonexistent path raises a typed refusal rather than an OS error."""
    with pytest.raises(ModeloLocalObservationError):
        parse_casilla_value_spreadsheet(tmp_path / "missing.csv")


def test_parse_rejects_empty_file(tmp_path: Path) -> None:
    """A spreadsheet with no rows at all raises."""
    path = tmp_path / "sheet.csv"
    path.write_text("", encoding="utf-8")

    with pytest.raises(ModeloLocalObservationError):
        parse_casilla_value_spreadsheet(path)


def test_parse_rejects_header_only_file(tmp_path: Path) -> None:
    """A spreadsheet carrying only a header row (no data) raises."""
    path = tmp_path / "sheet.csv"
    path.write_text("casilla_code,value\n", encoding="utf-8")

    with pytest.raises(ModeloLocalObservationError):
        parse_casilla_value_spreadsheet(path)


def test_parse_skips_blank_rows(tmp_path: Path) -> None:
    """Fully-blank rows between data rows are skipped, not treated as malformed."""
    path = tmp_path / "sheet.csv"
    path.write_text("casilla_code,value\n01,100\n\n02,200\n", encoding="utf-8")

    values = parse_casilla_value_spreadsheet(path)

    assert values == {"01": Decimal("100"), "02": Decimal("200")}


def test_parse_accepts_quoted_comma_decimal_separator(tmp_path: Path) -> None:
    """A quoted Spanish comma-decimal value (e.g. ``"1234,56"``) is coerced correctly."""
    path = tmp_path / "sheet.csv"
    path.write_text('casilla_code,value\n01,"1234,56"\n', encoding="utf-8")

    values = parse_casilla_value_spreadsheet(path)

    assert values == {"01": Decimal("1234.56")}


def test_parse_accepts_european_thousands_and_decimal_separators(tmp_path: Path) -> None:
    """A Spanish-formatted amount remains its full amount, not a malformed Decimal."""
    path = tmp_path / "sheet.csv"
    path.write_text('casilla_code,value\n01,"1.234,56"\n', encoding="utf-8")

    values = parse_casilla_value_spreadsheet(path)

    assert values == {"01": Decimal("1234.56")}


@pytest.mark.parametrize("raw_value", ("NaN", "Infinity", "-Infinity", "1e5", "+1", "1_000", "abc"))
def test_parse_rejects_what_no_one_writes_on_a_return(tmp_path: Path, raw_value: str) -> None:
    """Spellings a bare Decimal accepts but a figure on a return never uses.

    The non-finite pair is the dangerous half -- a NaN casilla value compares
    False to every threshold, so a guard keyed on ``> 0`` never fires for it --
    but ``1e5`` is the quiet one: it is legal input that silently becomes a
    hundred thousand. The canonical grammar refuses all of them in one place.
    """
    path = tmp_path / "sheet.csv"
    path.write_text(f"casilla_code,value\n01,{raw_value}\n", encoding="utf-8")

    with pytest.raises(ModeloLocalObservationError):
        parse_casilla_value_spreadsheet(path)


def test_parse_refuses_the_two_way_readable_thousands_amount(tmp_path: Path) -> None:
    """``1.234`` on a casilla row is refused, not read as one euro twenty-three.

    The parser already reads every unambiguous Spanish spelling correctly --
    ``1.234,56`` and ``12.345.678,90`` both land right -- because a comma tells
    it which mark is which. Without one it was choosing the English reading in
    silence, so a gestor typing the ordinary Spanish form for 1234 put 1.234
    on the return. These are casilla values: the figures a human submits.
    """
    path = tmp_path / "m303-2025q4.csv"
    path.write_text("casilla_code,value\n01,1.234\n", encoding="utf-8")

    with pytest.raises(ModeloLocalObservationError) as caught:
        parse_casilla_value_spreadsheet(path)

    context = caught.value.context
    assert context is not None
    message = str(context["malformed_rows"])
    assert "1.234" in message, "the refusal must echo what the operator wrote"
    assert "01" in message, "and name the casilla, so the row is findable"
    assert "1234" in message, "and show the thousands reading as a number"
    assert "1,234" in message, "and the fractional one, so the operator picks between them"
    assert "thousands" in message, "and say plainly what the ambiguity is"


@pytest.mark.parametrize(
    ("raw", "expected"),
    (
        pytest.param("1.234,56", Decimal("1234.56"), id="spanish-grouped-with-comma"),
        pytest.param("1234,56", Decimal("1234.56"), id="spanish-bare-comma"),
        pytest.param("12.345.678,90", Decimal("12345678.90"), id="spanish-fully-grouped"),
        pytest.param("1234.56", Decimal("1234.56"), id="canonical-dot-decimal"),
        pytest.param("0.333", Decimal("0.333"), id="three-decimal-coefficient"),
        pytest.param("1000.000", Decimal("1000.000"), id="long-lead-decimal"),
        pytest.param("100", Decimal("100"), id="plain-integer"),
    ),
)
def test_parse_still_reads_every_unambiguous_spelling(raw: str, expected: Decimal, tmp_path: Path) -> None:
    """The refusal is narrow: everything carrying its own evidence still parses.

    ``0.333`` is the case that matters most here. This parser runs before
    registry canonicalisation, so it does not yet know which casillas are euro
    amounts and which are coefficients -- a rule that refused three decimals
    outright would reject valid coefficient rows to catch amount rows.
    """
    path = tmp_path / "sheet.csv"
    path.write_text(f'casilla_code,value\n01,"{raw}"\n', encoding="utf-8")

    assert parse_casilla_value_spreadsheet(path) == {"01": expected}


def test_parse_casilla_value_spreadsheet_reads_a_semicolon_delimited_csv(tmp_path: Path) -> None:
    """A semicolon CSV is the ordinary Spanish export, not an edge case.

    Without delimiter detection every row parses as one column, the header
    aliases never match, and the whole sheet reads as a single unusable field.
    """
    source = tmp_path / "casillas-semicolon.csv"
    source.write_text(
        "casilla_code;value\n01;1234.56\n02;-78.90\n",
        encoding="utf-8",
    )

    parsed = parse_casilla_value_spreadsheet(source)

    assert parsed == {"01": Decimal("1234.56"), "02": Decimal("-78.90")}


def test_parse_casilla_value_spreadsheet_decodes_a_bom_prefixed_csv(tmp_path: Path) -> None:
    """A UTF-8 BOM must decide the codec rather than be read as a header character."""
    source = tmp_path / "casillas-bom.csv"
    source.write_bytes("casilla_code,value\n01,1234.56\n".encode("utf-8-sig"))

    parsed = parse_casilla_value_spreadsheet(source)

    assert parsed == {"01": Decimal("1234.56")}
