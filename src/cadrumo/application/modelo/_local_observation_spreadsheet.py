"""Parse a two-column casilla-value spreadsheet into a decimal mapping.

An operator keeps a hand-authored CSV or XLSX spreadsheet of ``casilla_code, value``
rows — a cert-free reconstruction path for a past filing when neither the
justificante PDF (:mod:`application.filing._import`) nor a live AEAT pull
is available. This module owns exactly the tabular-to-mapping projection: read
the two declared columns, coerce every value to :class:`~decimal.Decimal`, and
hand the caller a plain ``{casilla_code: Decimal}`` mapping keyed by the raw
spreadsheet token (not yet validated against any registry revision).

Casilla-id canonicalisation, registry-membership validation, and
:class:`~domain.calculations.registry.CasillaObservation` construction
remain owned by :func:`~application.modelo._local_observation_actions.record_operator_local_observation`,
which this module's CLI caller feeds directly — there is no second casilla
validation path here (``aeat-calculation-aggregation`` companion: one
validation authority, not two).

See Also:
    :func:`~application.modelo._local_observation_actions.record_operator_local_observation`:
        Consumes the parsed mapping, validates every casilla id against the
        law-determined :class:`~domain.calculations.registry.RegistrySnapshot`,
        and persists the non-official observation.
    :mod:`adapters.inbound.financial.providers._csv`:
        Sibling tabular-ingest module for bank-statement rows; this module is
        deliberately smaller — a casilla-value sheet has two logical columns
        and no bank-layout detection, date parsing, or currency handling.
"""

from __future__ import annotations

import csv
import io
from decimal import Decimal
from pathlib import Path
from typing import Final

from openpyxl import load_workbook

from ...core.decimal._coerce import normalize_decimal_separators
from ...core.decimal._grammar import european_thousands_reading_is_ambiguous, try_parse_canonical_decimal
from ...core.external_constants import XLSX_EXTENSION
from ...core.tabular import (
    TabularSourceError,
    coerce_cell_text,
    decode_tabular_bytes,
    detect_tabular_delimiter,
)
from ...core.workbook import first_formula_cell_column
from ._action_errors import ModeloLocalObservationError

CSV_EXTENSIONS: Final[frozenset[str]] = frozenset({".csv", ".txt"})
"""Extensions routed to the CSV reader; anything else is routed to XLSX."""

_CASILLA_CODE_HEADER_ALIASES: Final[frozenset[str]] = frozenset(
    {"casilla_code", "casilla", "casilla_id", "code", "id", "box", "casilla.id"},
)
_VALUE_HEADER_ALIASES: Final[frozenset[str]] = frozenset({"value", "valor", "amount", "importe"})


def _normalise_header(value: object) -> str:
    return str(value if value is not None else "").strip().lower().replace(" ", "_")


def parse_casilla_value_spreadsheet(path: Path) -> dict[str, Decimal]:
    """Parse a ``casilla_code, value`` spreadsheet into a raw code-to-Decimal mapping.

    Accepts CSV (``.csv`` / ``.txt``) or XLSX (``.xlsx``). The first
    non-blank row is treated as the header; a header row naming both a
    casilla-code column (``casilla_code`` / ``casilla`` / ``casilla_id`` /
    ``code`` / ``id`` / ``box``) and a value column (``value`` / ``valor`` /
    ``amount`` / ``importe``, case-insensitive) selects those columns by
    name. A headerless two-column sheet falls back positionally: column A is
    the casilla code, column B is the value.

    Every value is coerced to :class:`~decimal.Decimal`; a non-numeric value
    raises :class:`ModeloLocalObservationError` naming the offending row. A
    blank row is skipped. A row that omits the casilla code but carries a
    value (or vice versa) raises, naming the row.

    Returns:
        A ``{raw_casilla_code: Decimal}`` mapping in row order. Keys are the
        spreadsheet's literal cell text — not yet canonicalised or validated
        against any registry revision.

    Raises:
        ModeloLocalObservationError: The file cannot be opened, carries no
            data rows, or a data row cannot be parsed into a
            ``(casilla_code, Decimal)`` pair.
    """
    if not path.exists() or not path.is_file():
        raise ModeloLocalObservationError(
            translated_message="errors.error.error_modelos",
            context={"path": str(path)},
        )
    suffix = path.suffix.lower()
    rows = _read_xlsx_rows(path) if suffix == XLSX_EXTENSION else _read_csv_rows(path)
    if not rows:
        raise ModeloLocalObservationError(
            translated_message="errors.error.error_modelos",
            context={"path": str(path)},
        )

    code_index, value_index, data_rows = _locate_columns(rows)
    values, malformed = _parse_value_rows(data_rows, code_index=code_index, value_index=value_index)
    if malformed:
        raise ModeloLocalObservationError(
            translated_message="errors.error.error_modelos",
            context={
                "path": str(path),
                "malformed_row_count": str(len(malformed)),
                # The operator has to see WHICH cell was refused and what they
                # wrote in it; a bare count cannot be acted on.
                "malformed_rows": " | ".join(malformed),
            },
        )
    if not values:
        raise ModeloLocalObservationError(
            translated_message="errors.error.error_modelos",
            context={"path": str(path)},
        )
    return values


def parse_casilla_lexical_spreadsheet(path: Path) -> dict[str, str]:
    """Read the complete spreadsheet casilla manifest without rewriting value tokens.

    The same structural and numeric validation as
    :func:`parse_casilla_value_spreadsheet` applies. CSV value cells are returned
    byte-for-text exactly as decoded, while casilla ids remain trimmed identifiers.
    """
    if not path.exists() or not path.is_file():
        raise ModeloLocalObservationError(
            translated_message="errors.error.error_modelos",
            context={"path": str(path)},
        )
    rows = _read_xlsx_rows(path) if path.suffix.lower() == XLSX_EXTENSION else _read_csv_rows(path)
    if not rows:
        raise ModeloLocalObservationError(
            translated_message="errors.error.error_modelos",
            context={"path": str(path)},
        )
    code_index, value_index, data_rows = _locate_columns(rows)
    decimal_values, malformed = _parse_value_rows(data_rows, code_index=code_index, value_index=value_index)
    if malformed or not decimal_values:
        raise ModeloLocalObservationError(
            translated_message="errors.error.error_modelos",
            context={"path": str(path), "malformed_rows": " | ".join(malformed)},
        )
    return {_cell(row, code_index): row[value_index] for row in data_rows if _cell(row, code_index)}


def _parse_value_rows(
    data_rows: list[list[str]],
    *,
    code_index: int,
    value_index: int,
) -> tuple[dict[str, Decimal], list[str]]:
    """Project data rows into a code-to-Decimal mapping plus one message per unusable row.

    Accumulates rather than raising on the first bad row, so the caller can name
    every offending row at once instead of making the operator fix a sheet one
    refusal at a time. A wholly blank row is skipped and reported as neither.
    """
    values: dict[str, Decimal] = {}
    first_row_by_code: dict[str, int] = {}
    malformed: list[str] = []
    for row_number, row in enumerate(data_rows, start=1):
        code = _cell(row, code_index)
        raw_value = _cell(row, value_index)
        if not code and not raw_value:
            continue
        if not code or not raw_value:
            malformed.append(f"row {row_number}: incomplete (casilla_code={code!r}, value={raw_value!r})")
            continue
        first_row = first_row_by_code.get(code)
        if first_row is not None:
            malformed.append(
                f"row {row_number}: duplicate casilla_code {code!r} (first declared on row {first_row})",
            )
            continue
        first_row_by_code[code] = row_number
        value = _parse_row_amount(raw_value, row_number=row_number, code=code, malformed=malformed)
        if value is None:
            continue
        values[code] = value
    return values, malformed


def _parse_row_amount(
    raw_value: str,
    *,
    row_number: int,
    code: str,
    malformed: list[str],
) -> Decimal | None:
    """Coerce one cell to :class:`~decimal.Decimal`, appending a named refusal instead when it cannot."""
    # Thousands are stripped only when a comma is present, which reads every
    # unambiguous Spanish spelling correctly and leaves exactly one that it
    # cannot: a lone dot before three digits. ``1.234`` is one thousand two
    # hundred and thirty-four to the operator who typed it and one point two
    # three four to this parser, and nothing in the token decides between
    # them. These are casilla values -- the figures that go on the return --
    # so the row is refused and named rather than read a thousandfold light.
    if european_thousands_reading_is_ambiguous(raw_value):
        # Name both numbers rather than the abstract ambiguity: the operator
        # knows which one they meant, and seeing them side by side is what
        # lets them pick the spelling that says so.
        written = raw_value.strip()
        malformed.append(
            f"row {row_number}: value {raw_value!r} for casilla {code!r} could mean "
            f"{written.replace('.', '')} (a Spanish thousands separator) or "
            f"{written.replace('.', ',')} (a decimal fraction), and nothing in the value "
            f"decides which; write {written.replace('.', '')} if you meant the first, "
            f"or {written.replace('.', ',')} if you meant the second",
        )
        return None
    # Separators are normalised first, because the operator's grammar is
    # Spanish and the canonical one is not: "1.234,56" has to become
    # "1234.56" before anything can judge its shape. What survives that is
    # then held to the canonical grammar rather than handed to a bare
    # Decimal, which accepts a good deal that no one writes in a
    # spreadsheet cell -- "1e5" (silently a hundred thousand), a leading
    # "+", "1_000", and the non-finite "NaN"/"Infinity". Fraction digits
    # are deliberately left unconstrained: this runs before registry
    # canonicalisation, so the casilla's type is unknown here, and a
    # two-decimal euro cap would refuse a coefficient row like "0.333" to
    # catch an amount row. Shape is decidable without knowing the type;
    # precision is not, and belongs where the declared type is known.
    normalized = normalize_decimal_separators(raw_value, strip_thousands="," in raw_value)
    value = try_parse_canonical_decimal(normalized)
    if value is None:
        malformed.append(
            f"row {row_number}: value {raw_value!r} for casilla {code!r} is not a plain number; "
            f'write digits with an optional decimal comma (for example "1234,56"). Scientific '
            f"notation, a leading '+', underscore separators, 'NaN' and 'Infinity' are refused "
            f"because they are not what a figure on a return looks like",
        )
        return None
    return value


def _locate_columns(rows: list[list[str]]) -> tuple[int, int, list[list[str]]]:
    """Return ``(code_column_index, value_column_index, data_rows)``.

    Selects by header-alias name when the first row matches; otherwise falls
    back to the positional convention (column 0 = code, column 1 = value)
    and treats every row as data.
    """
    header = rows[0]
    normalised = [_normalise_header(cell) for cell in header]
    code_index = next((i for i, cell in enumerate(normalised) if cell in _CASILLA_CODE_HEADER_ALIASES), None)
    value_index = next((i for i, cell in enumerate(normalised) if cell in _VALUE_HEADER_ALIASES), None)
    if code_index is not None and value_index is not None:
        return code_index, value_index, rows[1:]
    return 0, 1, rows


def _cell(row: list[str], index: int) -> str:
    if index >= len(row):
        return ""
    return row[index].strip()


def _read_csv_rows(path: Path) -> list[list[str]]:
    source_bytes = path.read_bytes()
    text = _decode_bytes(source_bytes, path=path)
    delimiter = detect_tabular_delimiter(text) or ","
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows = [list(row) for row in reader]
    return [row for row in rows if any(cell.strip() for cell in row)]


def _decode_bytes(source_bytes: bytes, *, path: Path) -> str:
    try:
        text, _, _ = decode_tabular_bytes(source_bytes)
    except TabularSourceError as exc:
        raise ModeloLocalObservationError(
            translated_message="errors.error.error_modelos",
            context={"path": str(path)},
        ) from exc
    return text


def _read_xlsx_rows(path: Path) -> list[list[str]]:
    try:
        workbook = load_workbook(filename=path, read_only=True, data_only=False)
    except Exception as exc:  # openpyxl raises multiple unrelated types (OSError/KeyError/...) on a bad workbook
        raise ModeloLocalObservationError(
            translated_message="errors.error.error_modelos",
            context={"path": str(path)},
        ) from exc
    try:
        worksheet = workbook.worksheets[0]
        rows: list[list[str]] = []
        for row_index, cells in enumerate(worksheet.iter_rows(), start=1):
            column_index = first_formula_cell_column(cells)
            if column_index is not None:
                raise ModeloLocalObservationError(
                    translated_message="errors.error.error_modelos",
                    context={
                        "path": str(path),
                        "row": str(row_index),
                        "column": str(column_index),
                    },
                )
            rows.append([coerce_cell_text(cell.value, integral_floats_as_int=True) for cell in cells])
        return [row for row in rows if any(cell for cell in row)]
    finally:
        workbook.close()


__all__ = [
    "CSV_EXTENSIONS",
    "parse_casilla_lexical_spreadsheet",
    "parse_casilla_value_spreadsheet",
]
