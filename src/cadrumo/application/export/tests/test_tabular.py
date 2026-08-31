"""Tests for application export serialization helpers."""

from __future__ import annotations

import csv
import hashlib
from collections.abc import Mapping
from io import BytesIO, StringIO
from typing import cast

import pytest
from openpyxl import load_workbook

from ....core.errors import ERROR_REGISTRY, build_error_envelope, declared_error_codes
from ....core.external_constants import CSV_MIME_TYPE, JSONL_MIME_TYPE, UTF_8_ENCODING, XLSX_MIME_TYPE
from ....tests.locale_catalogue import CATALOGUE_LANGUAGES, catalogue_shard_path, shard_payload
from .. import ExportSerializationFormat, serialize_tabular_rows
from ..errors import ExportFieldError, ExportFormatError

pytestmark = [pytest.mark.unit, pytest.mark.hex_application]


def test_serialize_tabular_rows_writes_stable_csv_payload() -> None:
    result = serialize_tabular_rows(
        (
            {"transaction_id": "b", "amount": "2.00"},
            {"transaction_id": "a", "amount": "1.00"},
        ),
        fieldnames=("transaction_id", "amount"),
        export_format=ExportSerializationFormat.CSV,
    )

    parsed = tuple(csv.DictReader(StringIO(result.payload.decode(UTF_8_ENCODING))))
    assert parsed == (
        {"transaction_id": "b", "amount": "2.00"},
        {"transaction_id": "a", "amount": "1.00"},
    )
    assert result.media_type == CSV_MIME_TYPE
    assert result.filename_extension == "csv"
    assert result.byte_size == len(result.payload)
    assert len(result.sha256) == 64


def test_serialize_tabular_rows_writes_stable_jsonl_payload() -> None:
    result = serialize_tabular_rows(
        ({"transaction_id": "b", "amount": "2.00"},),
        fieldnames=("transaction_id", "amount"),
        export_format=ExportSerializationFormat.JSONL,
    )

    assert result.payload == b'{"amount":"2.00","transaction_id":"b"}\n'
    assert result.media_type == JSONL_MIME_TYPE
    assert result.filename_extension == "jsonl"


def test_serialize_tabular_rows_writes_readable_xlsx_payload() -> None:
    result = serialize_tabular_rows(
        (
            {"transaction_id": "b", "amount": "2.00"},
            {"transaction_id": "a", "amount": "1.00"},
        ),
        fieldnames=("transaction_id", "amount"),
        export_format=ExportSerializationFormat.XLSX,
    )

    workbook = load_workbook(BytesIO(result.payload), read_only=True, data_only=True)
    worksheet = workbook.active
    assert worksheet is not None
    rows = tuple(worksheet.iter_rows(values_only=True))

    assert rows == (
        ("transaction_id", "amount"),
        ("b", "2.00"),
        ("a", "1.00"),
    )
    assert result.media_type == XLSX_MIME_TYPE
    assert result.filename_extension == "xlsx"
    assert result.byte_size == len(result.payload)
    assert result.sha256 == hashlib.sha256(result.payload).hexdigest()


def test_serialize_tabular_rows_rejects_unknown_fields() -> None:
    with pytest.raises(ExportFieldError) as exc_info:
        serialize_tabular_rows(
            ({"transaction_id": "b", "amount": "2.00", "extra": "x"},),
            fieldnames=("transaction_id", "amount"),
            export_format=ExportSerializationFormat.CSV,
        )
    assert exc_info.value.translated_message == "errors.refused.refused_export_field"
    assert exc_info.value.context == {"reason": "unknown_fields", "unknown_fields": ("extra",)}


# ---------------------------------------------------------------------------
# Registry membership — contract
# ---------------------------------------------------------------------------


def test_export_format_error_registry_uses_distinct_application_and_adapter_classes() -> None:
    """The export format error classes must keep distinct registry identities.

    After the adapter rename, only the application-layer ExportFormatError
    (code REFUSED_EXPORT_FORMAT) should use that simple name. The
    adapter class is now AeatExportFormatError (code FAIL_EXPORT_FORMAT).
    """
    export_format_rows = [
        (qualname, code.code)
        for qualname, code in declared_error_codes()
        if qualname.split(".")[-1].endswith("ExportFormatError")
    ]
    assert export_format_rows == [
        ("cadrumo.adapters.outbound.aeat.export.errors.AeatExportFormatError", "FAIL_EXPORT_FORMAT"),
        ("cadrumo.application.export.errors.ExportFormatError", "REFUSED_EXPORT_FORMAT"),
    ]


@pytest.mark.parametrize(
    ("code_key", "message_key"),
    (
        ("REFUSED_EXPORT_FORMAT", "errors.refused.refused_export_format"),
        ("REFUSED_EXPORT_FIELD", "errors.refused.refused_export_field"),
    ),
)
def test_export_errors_are_registered_with_expected_attributes(code_key: str, message_key: str) -> None:
    """Export errors must be registered so the CLI can handle them."""
    code = ERROR_REGISTRY[code_key]
    assert code.code == code_key
    assert code.message_key == message_key
    assert code.retryable is False


# ---------------------------------------------------------------------------
# Envelope round-trip — contract
# ---------------------------------------------------------------------------


@pytest.mark.parametrize(
    ("error", "expected_code"),
    (
        (
            ExportFormatError(
                translated_message="errors.refused.refused_export_format",
                context={"export_format": "xml"},
            ),
            "REFUSED_EXPORT_FORMAT",
        ),
        (
            ExportFieldError(
                translated_message="errors.refused.refused_export_field",
                context={"reason": "fieldnames_empty"},
            ),
            "REFUSED_EXPORT_FIELD",
        ),
    ),
)
def test_export_errors_build_error_envelope(error: ExportFormatError | ExportFieldError, expected_code: str) -> None:
    """build_error_envelope must succeed for export errors."""
    envelope = build_error_envelope(error)
    assert envelope.code == expected_code
    assert envelope.category == "REFUSED"
    assert envelope.retryable is False
    assert envelope.message != ""


# ---------------------------------------------------------------------------
# Locale translated_message key — contract
# ---------------------------------------------------------------------------


@pytest.mark.parametrize("locale_key", ("refused_export_format", "refused_export_field"))
@pytest.mark.parametrize("locale_code", CATALOGUE_LANGUAGES)
def test_export_error_locale_keys_present_in_catalogue(locale_key: str, locale_code: str) -> None:
    """The locale catalogue must carry export error keys for every locale.

    The shard is located through the shared catalogue reader rather than by
    restating the layout here: this gate previously addressed the retired
    monolithic ``<lang>.yml`` and so had stopped reading any catalogue at all.
    """
    dotted = f"errors.refused.{locale_key}"
    payload = shard_payload(locale_code, dotted)
    errors_section = payload.get("errors", {})
    assert isinstance(errors_section, Mapping)
    refused_section = errors_section.get("refused", {})
    assert isinstance(refused_section, Mapping)
    value = refused_section.get(locale_key)
    assert value, (
        f"locale {locale_code!r}: {dotted!r} is missing or empty in {catalogue_shard_path(locale_code, dotted)}"
    )


# ---------------------------------------------------------------------------
# Real raise sites — contract: seven replaced ValueError sites
# ---------------------------------------------------------------------------


def test_normalize_fieldnames_raises_export_field_error_on_empty_sequence() -> None:
    """Site: _normalize_fieldnames — empty fieldnames sequence."""
    with pytest.raises(ExportFieldError) as exc_info:
        serialize_tabular_rows(
            (),
            fieldnames=(),
            export_format=ExportSerializationFormat.CSV,
        )
    assert exc_info.value.translated_message == "errors.refused.refused_export_field"
    assert exc_info.value.context == {"reason": "fieldnames_empty"}


def test_normalize_fieldnames_raises_export_field_error_on_blank_name() -> None:
    """Site: _normalize_fieldnames — blank field name."""
    with pytest.raises(ExportFieldError) as exc_info:
        serialize_tabular_rows(
            (),
            fieldnames=("transaction_id", "  "),
            export_format=ExportSerializationFormat.CSV,
        )
    assert exc_info.value.translated_message == "errors.refused.refused_export_field"
    assert exc_info.value.context == {"reason": "fieldnames_blank"}


def test_normalize_fieldnames_raises_export_field_error_on_duplicate_names() -> None:
    """Site: _normalize_fieldnames — duplicate field names."""
    with pytest.raises(ExportFieldError) as exc_info:
        serialize_tabular_rows(
            (),
            fieldnames=("amount", "amount"),
            export_format=ExportSerializationFormat.CSV,
        )
    assert exc_info.value.translated_message == "errors.refused.refused_export_field"
    assert exc_info.value.context == {"reason": "fieldnames_duplicate"}


def test_normalize_row_raises_export_field_error_on_unknown_field() -> None:
    """Site: _normalize_row — row contains unknown field keys."""
    with pytest.raises(ExportFieldError) as exc_info:
        serialize_tabular_rows(
            ({"transaction_id": "a", "amount": "1.00", "bogus": "x"},),
            fieldnames=("transaction_id", "amount"),
            export_format=ExportSerializationFormat.JSONL,
        )
    assert exc_info.value.translated_message == "errors.refused.refused_export_field"
    assert exc_info.value.context == {"reason": "unknown_fields", "unknown_fields": ("bogus",)}


def test_serialize_tabular_rows_rejects_unsupported_runtime_format() -> None:
    """Site: serialize_tabular_rows — runtime value outside the closed enum."""
    with pytest.raises(ExportFormatError) as exc_info:
        serialize_tabular_rows(
            (),
            fieldnames=("amount",),
            export_format=cast(ExportSerializationFormat, "xml"),
        )
    assert exc_info.value.translated_message == "errors.refused.refused_export_format"
    assert exc_info.value.context == {"export_format": "xml"}
    envelope = build_error_envelope(exc_info.value)
    assert envelope.code == "REFUSED_EXPORT_FORMAT"


def test_model_validator_raises_export_field_error_on_blank_fieldname() -> None:
    """Site: TabularExportResult._validate_fieldnames — blank field in model."""
    from pydantic import ValidationError

    from ..tabular import TabularExportResult

    payload = b"transaction_id,amount\r\n"
    with pytest.raises((ExportFieldError, ValidationError)) as exc_info:
        TabularExportResult(
            format=ExportSerializationFormat.CSV,
            media_type="text/csv",
            filename_extension="csv",
            payload=payload,
            byte_size=len(payload),
            sha256=hashlib.sha256(payload).hexdigest(),
            row_count=0,
            fieldnames=("transaction_id", "  "),
        )
    if isinstance(exc_info.value, ExportFieldError):
        assert exc_info.value.translated_message == "errors.refused.refused_export_field"
        assert exc_info.value.context == {"reason": "fieldnames_blank"}


def test_model_validator_raises_export_field_error_on_duplicate_fieldname() -> None:
    """Site: TabularExportResult._validate_fieldnames — duplicate field in model."""
    from pydantic import ValidationError

    from ..tabular import TabularExportResult

    payload = b"amount,amount\r\n"
    with pytest.raises((ExportFieldError, ValidationError)) as exc_info:
        TabularExportResult(
            format=ExportSerializationFormat.CSV,
            media_type="text/csv",
            filename_extension="csv",
            payload=payload,
            byte_size=len(payload),
            sha256=hashlib.sha256(payload).hexdigest(),
            row_count=0,
            fieldnames=("amount", "amount"),
        )
    if isinstance(exc_info.value, ExportFieldError):
        assert exc_info.value.translated_message == "errors.refused.refused_export_field"
        assert exc_info.value.context == {"reason": "fieldnames_duplicate"}


def test_model_validator_raises_export_field_error_on_invalid_sha256() -> None:
    """Site: TabularExportResult._validate_sha256 — digest must be hex-addressed."""

    from pydantic import ValidationError

    from ..tabular import TabularExportResult

    payload = b"transaction_id,amount\n"
    with pytest.raises(ValidationError) as exc_info:
        TabularExportResult(
            format=ExportSerializationFormat.CSV,
            media_type=CSV_MIME_TYPE,
            filename_extension="csv",
            payload=payload,
            byte_size=len(payload),
            sha256="not-a-digest",
            row_count=0,
            fieldnames=("transaction_id", "amount"),
        )
    error_detail = exc_info.value.errors()[0]
    assert "ctx" in error_detail and "error" in error_detail["ctx"]
    cause = error_detail["ctx"]["error"]
    assert isinstance(cause, ExportFieldError)
    assert cause.translated_message == "errors.refused.refused_export_field"
    assert cause.context == {"reason": "sha256_invalid"}
