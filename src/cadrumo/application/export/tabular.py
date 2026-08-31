"""Serialize tabular rows into :class:`~application.export.TabularExportResult` payloads.

Rows are rendered through the closed
:class:`~application.export.ExportSerializationFormat` surface, with
:class:`~application.export.errors.ExportFieldError` and
:class:`~application.export.errors.ExportFormatError` preserving
validation failures as structured application errors.

This module is a pure in-memory serializer. It returns bytes, media type,
extension, field metadata, row count, and SHA-256 digest to the calling
export service; it does not choose paths, write files, emit bucket events,
or mutate canonical storage.
"""

from __future__ import annotations

import csv
import json
from collections.abc import Mapping, Sequence
from enum import StrEnum
from io import BytesIO, StringIO
from types import MappingProxyType
from xml.etree.ElementTree import ParseError
from zipfile import BadZipFile

from pydantic import BaseModel, Field, NonNegativeInt, field_validator, model_validator

from ...core.external_constants import CSV_MIME_TYPE as _CSV_MIME_TYPE
from ...core.external_constants import JSONL_MIME_TYPE as _JSONL_MIME_TYPE
from ...core.external_constants import UTF_8_ENCODING as _UTF_8_ENCODING
from ...core.external_constants import XLSX_MIME_TYPE as _XLSX_MIME_TYPE
from ...core.hashing import sha256_hex
from ...core.models import STRICT_FROZEN_CONFIG
from .errors import ExportFieldError, ExportFormatError


class ExportSerializationFormat(StrEnum):
    """Closed set of backend tabular export serialization formats."""

    CSV = "csv"
    JSONL = "jsonl"
    XLSX = "xlsx"


#: The one declaration of what each serialization format is called on the wire.
#: ``serialize_tabular_rows`` reads it to stamp a result and
#: :meth:`TabularExportResult._metadata_describes_the_payload` reads it to refuse
#: a result whose stamp contradicts its own format, so the producer and the
#: verifier cannot drift apart.
_FORMAT_WIRE_NAMES: Mapping[ExportSerializationFormat, tuple[str, str]] = MappingProxyType(
    {
        ExportSerializationFormat.CSV: (_CSV_MIME_TYPE, "csv"),
        ExportSerializationFormat.JSONL: (_JSONL_MIME_TYPE, "jsonl"),
        ExportSerializationFormat.XLSX: (_XLSX_MIME_TYPE, "xlsx"),
    },
)

_REFUSED_EXPORT_FIELD_MESSAGE = "errors.refused.refused_export_field"
_REFUSED_EXPORT_FORMAT_MESSAGE = "errors.refused.refused_export_format"
_FIELDNAMES_EMPTY_REASON = "fieldnames_empty"
_FIELDNAMES_BLANK_REASON = "fieldnames_blank"
_FIELDNAMES_DUPLICATE_REASON = "fieldnames_duplicate"
_UNKNOWN_FIELDS_REASON = "unknown_fields"
_SHA256_INVALID_REASON = "sha256_invalid"
_BYTE_SIZE_MISMATCH_REASON = "byte_size_mismatch"
_SHA256_MISMATCH_REASON = "sha256_mismatch"
_MEDIA_TYPE_MISMATCH_REASON = "media_type_mismatch"
_EXTENSION_MISMATCH_REASON = "filename_extension_mismatch"
_ROW_COUNT_MISMATCH_REASON = "row_count_mismatch"
_PAYLOAD_DECODE_INVALID_REASON = "payload_decode_invalid"
_JSONL_RECORD_INVALID_REASON = "jsonl_record_invalid"


class TabularExportResult(BaseModel):
    """Serialized tabular payload produced by :func:`~application.export.serialize_tabular_rows`.

    The result carries the raw payload plus operator-facing metadata:
    :class:`~application.export.ExportSerializationFormat`, media type,
    filename extension, byte count, SHA-256 digest, row count, and normalized
    field names.
    """

    model_config = STRICT_FROZEN_CONFIG

    format: ExportSerializationFormat
    media_type: str = Field(min_length=1)
    filename_extension: str = Field(min_length=1)
    payload: bytes
    byte_size: NonNegativeInt
    sha256: str
    row_count: NonNegativeInt
    fieldnames: tuple[str, ...]

    @field_validator("fieldnames")
    @classmethod
    def _validate_fieldnames(cls, value: tuple[str, ...]) -> tuple[str, ...]:
        normalized = tuple(field.strip() for field in value)
        if any(not field for field in normalized):
            raise _export_field_error(_FIELDNAMES_BLANK_REASON)
        if len(set(normalized)) != len(normalized):
            raise _export_field_error(_FIELDNAMES_DUPLICATE_REASON)
        return normalized

    @field_validator("sha256")
    @classmethod
    def _validate_sha256(cls, value: str) -> str:
        normalized = value.strip().lower()
        if len(normalized) != 64 or any(char not in "0123456789abcdef" for char in normalized):
            raise _export_field_error(_SHA256_INVALID_REASON)
        return normalized

    @model_validator(mode="after")
    def _metadata_describes_the_payload(self) -> TabularExportResult:
        """Refuse a result whose metadata contradicts the bytes it carries.

        Every one of these fields is a pure function of ``payload`` and
        ``format``, so a result that disagrees with its own bytes is not a
        different opinion, it is false. Validating only the digest *shape*
        meant a caller could publish a 29-byte payload as ``byte_size=0``
        under an all-zero digest with JSON metadata for CSV -- and the ledger
        export action anchors exactly these values into a durable
        ``LEDGER_TRANSACTION_EXPORTED`` bucket event, where the lie outlives
        the payload.

        Raises:
            ExportFieldError: A metadata field disagrees with the payload.
        """
        verify_export_metadata(
            payload=self.payload,
            export_format=self.format,
            byte_size=self.byte_size,
            sha256=self.sha256,
            media_type=self.media_type,
            filename_extension=self.filename_extension,
            row_count=self.row_count,
        )
        return self


def serialize_tabular_rows(
    rows: Sequence[Mapping[str, str]],
    *,
    fieldnames: Sequence[str],
    export_format: ExportSerializationFormat,
) -> TabularExportResult:
    """Serialize string-keyed rows as deterministic CSV, JSON Lines, or XLSX.

    Field order follows ``fieldnames`` and row order follows ``rows``.
    Values are coerced to strings, missing fields become empty strings,
    and unknown fields raise
    :class:`~application.export.errors.ExportFieldError`. Returns a
    :class:`~application.export.TabularExportResult` with encoded bytes,
    media type, filename extension, row count, field metadata, and payload
    digest.
    """
    normalized_fields = _normalize_fieldnames(fieldnames)
    normalized_rows = tuple(_normalize_row(row, fieldnames=normalized_fields) for row in rows)
    if export_format is ExportSerializationFormat.CSV:
        payload = _serialize_csv(normalized_rows, fieldnames=normalized_fields)
    elif export_format is ExportSerializationFormat.JSONL:
        payload = _serialize_jsonl(normalized_rows, fieldnames=normalized_fields)
    elif export_format is ExportSerializationFormat.XLSX:
        payload = _serialize_xlsx(normalized_rows, fieldnames=normalized_fields)
    else:
        raise ExportFormatError(
            translated_message=_REFUSED_EXPORT_FORMAT_MESSAGE,
            context={"export_format": str(export_format)},
        )
    media_type, extension = _FORMAT_WIRE_NAMES[export_format]
    return TabularExportResult(
        format=export_format,
        media_type=media_type,
        filename_extension=extension,
        payload=payload,
        byte_size=len(payload),
        sha256=sha256_hex(payload),
        row_count=len(normalized_rows),
        fieldnames=normalized_fields,
    )


def _normalize_fieldnames(fieldnames: Sequence[str]) -> tuple[str, ...]:
    normalized = tuple(field.strip() for field in fieldnames)
    if not normalized:
        raise _export_field_error(_FIELDNAMES_EMPTY_REASON)
    if any(not field for field in normalized):
        raise _export_field_error(_FIELDNAMES_BLANK_REASON)
    if len(set(normalized)) != len(normalized):
        raise _export_field_error(_FIELDNAMES_DUPLICATE_REASON)
    return normalized


def _normalize_row(row: Mapping[str, str], *, fieldnames: tuple[str, ...]) -> dict[str, str]:
    unknown = sorted(set(row).difference(fieldnames))
    if unknown:
        raise _export_field_error(_UNKNOWN_FIELDS_REASON, unknown_fields=tuple(unknown))
    return {field: str(row.get(field, "")) for field in fieldnames}


def _export_field_error(reason: str, **context: object) -> ExportFieldError:
    return ExportFieldError(
        translated_message=_REFUSED_EXPORT_FIELD_MESSAGE,
        context={"reason": reason, **context},
    )


def _payload_decode_error(export_format: ExportSerializationFormat) -> ExportFieldError:
    return _export_field_error(_PAYLOAD_DECODE_INVALID_REASON, export_format=export_format.value)


def verify_export_metadata(
    *,
    payload: bytes,
    export_format: ExportSerializationFormat,
    byte_size: int,
    sha256: str,
    media_type: str,
    filename_extension: str,
    row_count: int,
) -> None:
    """Refuse export metadata that contradicts the payload it describes.

    The single check behind every serialized-payload result. Each argument is a
    pure function of ``payload`` and ``export_format``, so a value that
    disagrees with the bytes is false rather than merely different --
    and :class:`~application.ledger.models.LedgerExportResult` redeclares the same
    seven fields independently, which is exactly why the verification lives
    here once instead of at each declaration.

    Raises:
        ExportFieldError: A metadata value disagrees with the payload.
    """
    if byte_size != len(payload):
        raise _export_field_error(_BYTE_SIZE_MISMATCH_REASON)
    if sha256 != sha256_hex(payload):
        raise _export_field_error(_SHA256_MISMATCH_REASON)
    expected_media_type, expected_extension = _FORMAT_WIRE_NAMES[export_format]
    if media_type != expected_media_type:
        raise _export_field_error(_MEDIA_TYPE_MISMATCH_REASON)
    if filename_extension != expected_extension:
        raise _export_field_error(_EXTENSION_MISMATCH_REASON)
    if row_count != _payload_row_count(payload, export_format=export_format):
        raise _export_field_error(_ROW_COUNT_MISMATCH_REASON)


def _payload_row_count(payload: bytes, *, export_format: ExportSerializationFormat) -> int:
    """Return the data-row count the serialized ``payload`` actually carries.

    The inverse of the three serializers, and the reason ``row_count`` can be
    checked rather than trusted. Each branch mirrors exactly what its writer
    emits: CSV and XLSX carry one header row above the data, JSONL carries one
    object per line and no header. CSV is read back through :mod:`csv` rather
    than by counting newlines so a quoted field containing a line break counts
    as the one row it is.
    """
    if export_format is ExportSerializationFormat.CSV:
        return _csv_payload_row_count(payload)
    if export_format is ExportSerializationFormat.JSONL:
        return _jsonl_payload_row_count(payload)
    if export_format is ExportSerializationFormat.XLSX:
        return _xlsx_payload_row_count(payload)
    raise ExportFormatError(
        translated_message=_REFUSED_EXPORT_FORMAT_MESSAGE,
        context={"export_format": str(export_format)},
    )


def _csv_payload_row_count(payload: bytes) -> int:
    """Count parsed CSV records after the serializer's one header record."""
    try:
        rows = tuple(csv.reader(StringIO(payload.decode(_UTF_8_ENCODING))))
    except (UnicodeDecodeError, csv.Error) as exc:
        raise _payload_decode_error(ExportSerializationFormat.CSV) from exc
    return max(len(rows) - 1, 0)


def _jsonl_payload_row_count(payload: bytes) -> int:
    """Count nonblank JSON Lines object records, retaining their source line on refusal."""
    try:
        lines = payload.decode(_UTF_8_ENCODING).splitlines()
    except UnicodeDecodeError as exc:
        raise _payload_decode_error(ExportSerializationFormat.JSONL) from exc

    row_count = 0
    for line_number, line in enumerate(lines, start=1):
        if not line.strip():
            continue
        try:
            row = json.loads(line)
        except json.JSONDecodeError as exc:
            raise _payload_decode_error(ExportSerializationFormat.JSONL) from exc
        if not isinstance(row, dict):
            raise _export_field_error(_JSONL_RECORD_INVALID_REASON, line_number=line_number)
        row_count += 1
    return row_count


def _xlsx_payload_row_count(payload: bytes) -> int:
    """Count active-sheet rows after the serializer's one header record."""
    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import InvalidFileException

    try:
        workbook = load_workbook(BytesIO(payload), read_only=True)
        try:
            worksheet = workbook.active
            if worksheet is None:
                return 0
            return max((worksheet.max_row or 0) - 1, 0)
        finally:
            workbook.close()
    except (BadZipFile, InvalidFileException, KeyError, OSError, ParseError, ValueError) as exc:
        raise _payload_decode_error(ExportSerializationFormat.XLSX) from exc


def _serialize_csv(rows: tuple[dict[str, str], ...], *, fieldnames: tuple[str, ...]) -> bytes:
    buffer = StringIO()
    writer = csv.DictWriter(buffer, fieldnames=fieldnames, lineterminator="\n")
    writer.writeheader()
    writer.writerows(rows)
    return buffer.getvalue().encode(_UTF_8_ENCODING)


def _serialize_jsonl(rows: tuple[dict[str, str], ...], *, fieldnames: tuple[str, ...]) -> bytes:
    del fieldnames
    text = "".join(json.dumps(row, ensure_ascii=True, sort_keys=True, separators=(",", ":")) + "\n" for row in rows)
    return text.encode(_UTF_8_ENCODING)


def _serialize_xlsx(rows: tuple[dict[str, str], ...], *, fieldnames: tuple[str, ...]) -> bytes:
    """Serialize rows into a single-worksheet XLSX workbook (header + data rows).

    Every cell is written as text so a deterministic, locale-independent
    round-trip is preserved; the workbook re-reads through
    :class:`~adapters.inbound.financial.providers._xlsx.XlsxProvider`,
    which shares the CSV bank-layout catalogue.
    """
    from openpyxl import Workbook

    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None, "Workbook.active must not be None on a fresh Workbook"
    worksheet.title = "ledger"
    worksheet.append(list(fieldnames))
    for row in rows:
        worksheet.append([row.get(field, "") for field in fieldnames])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
