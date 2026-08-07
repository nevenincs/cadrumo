"""The shared workbook-reading policy, against real openpyxl cells.

The three readers that refuse formula cells -- the bank-statement workbook
provider, the casilla-value spreadsheet, and the bulk invoice book -- resolve
what counts as a formula through this one predicate. These tests use real cells
from a real saved workbook rather than a stand-in with a ``data_type``
attribute: the whole value of a structural protocol is that openpyxl's own cell
types satisfy it, and a hand-rolled double would assert nothing about that.
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from ..workbook import FORMULA_CELL_REFUSAL, first_formula_cell_column

pytestmark = [pytest.mark.unit, pytest.mark.hex_core]


def _rows_of(tmp_path: Path, values: list[list[object]], *, read_only: bool) -> list[list[object]]:
    """Save *values* as a real workbook and read its rows back as cells."""
    path = tmp_path / f"cells-{read_only}.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    for row in values:
        sheet.append(row)
    workbook.save(path)

    reopened = load_workbook(path, read_only=read_only, data_only=False)
    try:
        return [list(cells) for cells in reopened.worksheets[0].iter_rows()]
    finally:
        reopened.close()


@pytest.mark.parametrize("read_only", [True, False])
def test_the_formula_column_is_found_in_both_workbook_read_modes(tmp_path: Path, read_only: bool) -> None:
    """Both read modes must resolve alike; the three readers do not all use the same one.

    ``read_only=True`` yields ``ReadOnlyCell`` and ``EmptyCell`` while the
    default mode yields ``Cell``. The predicate is structural precisely so all
    three satisfy it, and a policy that held in one mode only would refuse a
    formula for one reader and read it for another.
    """
    rows = _rows_of(tmp_path, [["a", "b", "=1+1", "d"]], read_only=read_only)

    assert first_formula_cell_column(rows[0]) == 3


def test_a_row_of_plain_values_carries_no_formula_column(tmp_path: Path) -> None:
    """The negative control: a literal row must resolve to ``None``, not to column 1.

    Without this, a predicate that answered "formula" for every row would pass
    the positive case above and refuse every workbook in the product.
    """
    rows = _rows_of(tmp_path, [["a", 12, "2026-05-01", None]], read_only=True)

    assert first_formula_cell_column(rows[0]) is None


def test_the_first_formula_wins_when_a_row_carries_several(tmp_path: Path) -> None:
    """The reported column is the leftmost formula, so the operator is sent to the first."""
    rows = _rows_of(tmp_path, [["a", "=1+1", "=2+2"]], read_only=True)

    assert first_formula_cell_column(rows[0]) == 2


def test_an_empty_row_resolves_to_no_formula(tmp_path: Path) -> None:
    """A blank row must not refuse: ``EmptyCell`` carries a data type too."""
    rows = _rows_of(tmp_path, [["a", "b"], []], read_only=True)

    assert first_formula_cell_column(rows[-1]) is None


def test_the_refusal_clause_is_the_sentence_all_three_readers_end_with() -> None:
    """The shared clause is the operator-visible half of the policy.

    Pinned because the three refusals are recognisable to an operator by this
    sentence; rewording it in one place while the constant kept its old value
    is exactly the drift the shared constant exists to prevent.
    """
    assert FORMULA_CELL_REFUSAL == "formula cached values are not accepted"
