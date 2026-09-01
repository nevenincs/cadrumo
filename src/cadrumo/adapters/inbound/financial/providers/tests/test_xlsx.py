"""Unit tests for the XLSX financial ingestion provider.

Verifies :class:`cadrumo.adapters.inbound.financial.providers.XlsxProvider`
correctly auto-detects the header row of a synthetic worksheet and that
malformed workbooks fail validation cleanly without leaking pikepdf /
openpyxl exceptions.
"""

from __future__ import annotations

from pathlib import Path
from shutil import copyfile
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from openpyxl import load_workbook

from ......domain.transactions.enums import TransactionDirection
from ......tests import FIXTURES_DIR
from .._xlsx import XlsxProvider
from ..base import InvalidFinancialSourceError

pytestmark = [pytest.mark.unit, pytest.mark.hex_inbound_adapter]

_FIXTURES = FIXTURES_DIR / "financial"


def _copy_fixture_with_stale_formula_cache(destination: Path) -> None:
    """Create a real workbook whose formula result cache disagrees with its formula."""
    copyfile(_FIXTURES / "synthetic-transactions.xlsx", destination)
    worksheet_member = "xl/worksheets/sheet1.xml"
    original_cell = b'<c r="D4" t="inlineStr"><is><t>1120,00</t></is></c>'
    formula_cell_with_stale_cache = b'<c r="D4"><f>1+1</f><v>9000</v></c>'
    with ZipFile(destination) as source_archive:
        members = {member.filename: source_archive.read(member) for member in source_archive.infolist()}
    assert members[worksheet_member].count(original_cell) == 1
    members[worksheet_member] = members[worksheet_member].replace(original_cell, formula_cell_with_stale_cache)
    with ZipFile(destination, mode="w", compression=ZIP_DEFLATED) as destination_archive:
        for filename, contents in members.items():
            destination_archive.writestr(filename, contents)


def test_xlsx_provider_ingests_header_detected_worksheet() -> None:
    """XlsxProvider should locate the header row and ingest transactions."""
    provider = XlsxProvider()
    fixture = _FIXTURES / "synthetic-transactions.xlsx"
    validation = provider.validate_source(fixture)
    assert validation.is_valid, validation.warnings
    parsed_rows = tuple(provider.ingest(fixture))
    assert len(parsed_rows) == 2
    assert parsed_rows[0].raw.description == "Cobro factura F-2026-021"
    # The second source row is a debit: magnitude stored, OUTGOING direction.
    assert parsed_rows[1].raw.amount >= 0
    assert parsed_rows[1].direction is TransactionDirection.OUTGOING


def test_xlsx_provider_validation_handles_unopenable_workbook(tmp_path: Path) -> None:
    """Validation should fail cleanly for malformed workbooks without cleanup errors."""
    source = tmp_path / "broken.xlsx"
    source.write_text("not a workbook", encoding="utf-8")
    validation = XlsxProvider().validate_source(source)
    assert not validation.is_valid
    assert "could not open workbook" in validation.warnings[0].lower()


def test_xlsx_provider_refuses_stale_formula_cache_before_financial_amount_parsing(tmp_path: Path) -> None:
    """A stale cached formula result cannot become an imported transaction amount."""
    source = tmp_path / "stale-formula.xlsx"
    _copy_fixture_with_stale_formula_cache(source)

    cached_values_workbook = load_workbook(filename=source, read_only=True, data_only=True)
    try:
        worksheet = cached_values_workbook.active
        assert worksheet is not None
        cached_amount = next(
            worksheet.iter_rows(min_row=4, max_row=4, min_col=4, max_col=4, values_only=True),
        )[0]
    finally:
        cached_values_workbook.close()
    assert cached_amount == 9000

    provider = XlsxProvider()
    validation = provider.validate_source(source)
    assert not validation.is_valid
    assert "formula cached values are not accepted" in validation.warnings[0]
    with pytest.raises(InvalidFinancialSourceError, match=r"formula cell at row 4, column 4"):
        tuple(provider.ingest(source))
