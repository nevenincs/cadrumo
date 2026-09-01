"""XLSX financial provider with header-row detection.

Implements :class:`XlsxProvider`, an
:class:`~adapters.inbound.financial.providers.FinancialProvider`
backed by ``openpyxl``. Reuses the bank-layout catalogue and scoring
helpers from :mod:`adapters.inbound.financial.providers._csv` so
the same alias rules apply to spreadsheet exports as to the
matching CSV downloads.

The selected worksheet emits
:class:`~adapters.inbound.financial.providers.ParsedLedgerRow` records
with the same magnitude/direction split as the CSV provider, while preserving
typed workbook cell values for dates and amounts until the parse boundary.
"""

from __future__ import annotations

from collections.abc import Iterator, Sequence
from dataclasses import dataclass
from pathlib import Path
from typing import Any, override

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .....core.logging import get_logger
from .....core.workbook import FORMULA_CELL_REFUSAL, first_formula_cell_column
from .....domain.transactions.raw_transaction import SourceFormat
from ._constants import XLSX_EXTENSION
from .base import (
    FinancialProvider,
    FinancialValidationError,
    InvalidFinancialSourceError,
    ParsedLedgerRow,
    ProviderValidation,
    archive_cell_text,
    default_currency,
)
from .csv import (
    CSV_LAYOUTS,
    CsvBankLayout,
    _find_column,
    _header_lookup,
    _layout_score,
    _parse_tabular_transaction_row,
    _row_is_blank,
    build_provider_row,
)

_logger = get_logger(__name__)


class XlsxProvider(FinancialProvider):
    """Ingest raw transactions from ``.xlsx`` bank statement exports.

    Iterates every worksheet, scoring the first ten rows of each
    against :data:`adapters.inbound.financial.providers._csv.CSV_LAYOUTS`
    and selecting the worksheet/row pair with the highest match
    score. Numeric and date cell values are read directly from the
    cell types (rather than coerced through their printed strings)
    so locale-formatted ``Decimal`` and ``date`` parsing stays
    accurate. Formula cells in selected data rows are refused rather than
    trusting a workbook's potentially stale cached values.

    Like :class:`~adapters.inbound.financial.providers.CsvProvider`, this
    provider shares the tabular alias catalogue and stores every parsed row as
    a raw transaction with explicit flow direction.

    Attributes:
        _last_sheet_name: Name of the worksheet selected by the most
            recent :meth:`_locate_sheet` call; surfaced via
            ``ProviderValidation.detected_dialect``.
        _last_header_index: 1-based index of the header row in that
            worksheet.
    """

    name = "XLSX provider"
    supported_extensions = frozenset({XLSX_EXTENSION})
    source_format = SourceFormat.XLSX
    # Corpus fixture is a synthetic XLSX generated from the standard bank
    # export column schema; layout fidelity confirmed against the spec.
    verification_source = "synthetic_from_bank_published_text"
    provisional_pending_specimen = False

    def __init__(self) -> None:
        """Initialise the validation metadata placeholders."""
        self._last_sheet_name = "Sheet1"
        self._last_header_index = 1

    @override
    def validate_source(self, path: Path) -> ProviderValidation:
        """Validate workbook accessibility and header detection.

        Returns:
            A :class:`ProviderValidation` with the validation outcome.
        """
        workbook: Workbook | None = None
        try:
            workbook, rows, _, layout, header_row, _, _ = self._locate_sheet(path)
        except InvalidFinancialSourceError as exc:
            return ProviderValidation(is_valid=False, warnings=(str(exc),))
        finally:
            if workbook is not None:
                # BROAD-EXCEPT-RATIONALE-XLSX-TEARDOWN:
                # openpyxl raises OSError/ValueError/KeyError/IndexError/TypeError;
                # teardown must run unconditionally.
                try:
                    workbook.close()
                except Exception as close_exc:
                    _logger.debug(
                        "xlsx provider: workbook.close() after validate_source failed (%s)",
                        close_exc,
                        exc_info=True,
                    )
        warnings: list[str] = []
        if layout is None:
            return ProviderValidation(
                is_valid=False,
                warnings=("Workbook does not contain a supported bank-statement header row",),
            )
        if header_row is None or len(rows) <= self._last_header_index:
            return ProviderValidation(
                is_valid=False,
                warnings=(f"{layout.bank_name} worksheet has no data rows after the header",),
            )
        if not _find_column(_header_lookup(header_row), layout.columns.currency):
            warnings.append(
                f"{layout.bank_name} worksheet has no currency column; falling back to {default_currency()}",
            )
        return ProviderValidation(
            is_valid=True,
            warnings=tuple(warnings),
            detected_encoding=None,
            detected_dialect=f"worksheet={self._last_sheet_name},header_row={self._last_header_index}",
        )

    @override
    def ingest(self, path: Path) -> Iterator[ParsedLedgerRow]:
        """Yield :class:`ParsedLedgerRow` records (magnitude + direction) from the first matching worksheet."""
        source_bytes = self._read_source_bytes(path)
        source_sha256 = self._compute_sha256(source_bytes)
        workbook, rows, sheet_name, layout, headers, lookup, header_index = self._locate_sheet(path)
        if layout is None or headers is None or lookup is None:
            workbook.close()
            raise InvalidFinancialSourceError("Workbook does not contain a supported bank-statement header row")
        try:
            for source_row_index, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
                raw_fields = _row_to_mapping(headers, row)
                cell_lookup = _row_to_cells(headers, row)
                if _row_is_blank(raw_fields):
                    continue
                try:
                    parsed = _parse_tabular_transaction_row(
                        layout=layout,
                        lookup=lookup,
                        raw_fields=raw_fields,
                        typed_fields=cell_lookup,
                        synthetic_provider_name=f"{layout.bank_name}-{sheet_name}",
                        source_sha256=source_sha256,
                        source_row_index=source_row_index,
                        required_field_context="worksheet row",
                    )
                except (ValueError, FinancialValidationError) as exc:
                    _logger.warning(
                        "xlsx_provider: parse error row=%d file=%s",
                        source_row_index,
                        path.name,
                        exc_info=True,
                    )
                    raise InvalidFinancialSourceError(
                        f"worksheet row {source_row_index} could not be parsed: {exc}",
                    ) from exc
                yield build_provider_row(
                    provider=self,
                    path=path,
                    source_sha256=source_sha256,
                    source_row_index=source_row_index,
                    parsed=parsed,
                    raw_fields=raw_fields,
                )
        finally:
            workbook.close()

    def _locate_sheet(
        self,
        path: Path,
    ) -> tuple[Workbook, list[list[Any]], str, CsvBankLayout | None, list[str] | None, dict[str, str] | None, int]:
        """Return the first worksheet that matches a known bank layout."""
        workbook = _open_workbook_or_refuse(path)
        try:
            best = _select_best_layout_across_worksheets(workbook)
            best_rows = _materialize_selected_rows_or_refuse_formula_cells(
                best.worksheet,
                sheet_name=best.sheet_name,
                header_index=best.header_index,
            )
            self._last_sheet_name = best.sheet_name
            self._last_header_index = best.header_index + 1
            if best.score < _MIN_LAYOUT_SCORE:
                return workbook, best_rows, best.sheet_name, None, None, None, best.header_index
            return (
                workbook,
                best_rows,
                best.sheet_name,
                best.layout,
                best.headers,
                best.lookup,
                best.header_index,
            )
        # BROAD-EXCEPT-RATIONALE-XLSX-TEARDOWN:
        # openpyxl raises OSError (file I/O), ValueError (invalid cell values),
        # KeyError (missing sheet/named range), IndexError (out-of-range
        # row/column access), and TypeError (unexpected cell type);
        # `_close_workbook_during_teardown` must run unconditionally on any
        # failure before re-raising.
        except Exception:
            _close_workbook_during_teardown(workbook)
            raise


_MIN_LAYOUT_SCORE = 3


@dataclass(frozen=True, slots=True)
class _BestLayoutMatch:
    """Best (worksheet, layout) match across every worksheet in the workbook.

    Carries the per-worksheet score + selected layout + header row /
    lookup so the caller can both report the picked sheet (via
    name / index) and trigger the layout-not-supported short-circuit
    when the best score falls below the minimum.
    """

    worksheet: Worksheet | None
    sheet_name: str
    layout: CsvBankLayout | None
    headers: list[str] | None
    lookup: dict[str, str] | None
    header_index: int
    score: int


def _select_best_layout_across_worksheets(workbook: Workbook) -> _BestLayoutMatch:
    """Iterate every worksheet and keep the highest-scoring layout match.

    ``best_worksheet`` defaults to the first sheet in the workbook
    so a workbook whose every sheet scores below the minimum still
    returns a deterministic fallback (the caller emits an
    "unsupported" error envelope keyed on that sheet's identity).
    """
    fallback = workbook.worksheets[0] if workbook.worksheets else None
    best = _BestLayoutMatch(
        worksheet=fallback,
        sheet_name=fallback.title if fallback is not None else "Sheet1",
        layout=None,
        headers=None,
        lookup=None,
        header_index=0,
        score=-1,
    )
    for worksheet in workbook.worksheets:
        candidate = _best_layout_match_for_worksheet(worksheet)
        if candidate is None or candidate[0] <= best.score:
            continue
        score, index, row, lookup, layout = candidate
        best = _BestLayoutMatch(
            worksheet=worksheet,
            sheet_name=worksheet.title,
            layout=layout,
            headers=row,
            lookup=lookup,
            header_index=index,
            score=score,
        )
    return best


def _open_workbook_or_refuse(path: Path) -> Workbook:
    """Open ``path`` with formulas visible or re-wrap the parse failure."""
    try:
        return load_workbook(filename=path, read_only=True, data_only=False)
    except Exception as exc:  # pragma: no cover - exercised via validation path
        raise InvalidFinancialSourceError(f"could not open workbook: {path}") from exc


def _close_workbook_during_teardown(workbook: Workbook) -> None:
    """Best-effort ``workbook.close()`` after a parse error; never raise.

    Broad ``except Exception`` because the upstream parse can raise
    openpyxl / xlrd errors, KeyError, ValueError, OSError,
    IndexError, or TypeError depending on file shape. The close()
    must run uniformly. The caller re-raises the original cause —
    this helper only owns the teardown side effect.
    """
    try:
        workbook.close()
    except Exception as close_exc:
        _logger.debug(
            "xlsx provider: workbook.close() during parse-error teardown failed (%s)",
            close_exc,
            exc_info=True,
        )


def _materialize_selected_rows_or_refuse_formula_cells(
    worksheet: Worksheet | None,
    *,
    sheet_name: str,
    header_index: int,
) -> list[list[Any]]:
    """Materialize a selected worksheet after refusing formula-bearing data rows."""
    if worksheet is None:
        return []
    rows: list[list[Any]] = []
    for row_index, cells in enumerate(worksheet.iter_rows(), start=1):
        if row_index > header_index + 1:
            column_index = first_formula_cell_column(cells)
            if column_index is not None:
                raise InvalidFinancialSourceError(
                    f"worksheet {sheet_name!r} contains formula cell at row {row_index}, column {column_index}; "
                    f"{FORMULA_CELL_REFUSAL}",
                )
        rows.append([cell.value for cell in cells])
    return rows


def _best_layout_match_for_worksheet(
    worksheet: Worksheet,
) -> tuple[int, int, list[str], dict[str, str], CsvBankLayout] | None:
    """Return the best (score, header_index, row, lookup, layout) the worksheet matches, or ``None``."""
    sample_rows = [
        [archive_cell_text(cell) for cell in row]
        for row in worksheet.iter_rows(min_row=1, max_row=10, values_only=True)
    ]
    best: tuple[int, int, list[str], dict[str, str], CsvBankLayout] | None = None
    for index, row in enumerate(sample_rows):
        if not any(cell.strip() for cell in row):
            continue
        lookup = _header_lookup(row)
        for layout in CSV_LAYOUTS:
            score = _layout_score(lookup, layout)
            if best is None or score > best[0]:
                best = (score, index, row, lookup, layout)
    return best


def _row_to_mapping(headers: Sequence[str], row: Sequence[object]) -> dict[str, str]:
    """Convert one worksheet row into the stored raw-field mapping."""
    return {header: archive_cell_text(row[index]) if index < len(row) else "" for index, header in enumerate(headers)}


def _row_to_cells(headers: Sequence[str], row: Sequence[object]) -> dict[str, object]:
    """Map worksheet headers to the original cell values for typed parsing."""
    return {header: row[index] if index < len(row) else "" for index, header in enumerate(headers)}
