"""Structural tests for registry-backed fichero-BOE record specs.

The accepted registry direction is that concrete fixed-width export layouts are
authored as registry TOML, not as per-modelo Python ``modelo_*`` modules under
``_formats``. These tests therefore validate the real registry-backed layouts
that the production serialiser consumes.
"""

from __future__ import annotations

from functools import cache
from pathlib import Path

import pytest
from openpyxl import load_workbook

from .......core import ExportLayoutFormat
from .......core.resources import bundled_path, resources
from .......domain.calculations.registry import (
    CasillaFieldKind,
    CasillaId,
    ExportLayoutDefinition,
    ExportRecordDefinition,
    ModeloDefinition,
    ModeloRevision,
    casillas_by_id,
    derive_export_layouts_from_bindings,
)
from .......domain.calculations.registry._export import _verify_record_offsets

pytestmark = [pytest.mark.unit, pytest.mark.hex_outbound_adapter]

_RequiredLayout = tuple[str, str]
_LayoutCase = tuple[ModeloDefinition, ModeloRevision, ExportLayoutDefinition]

_REQUIRED_LAYOUTS: frozenset[_RequiredLayout] = frozenset(
    {
        ("130", "modelo-130-fichero-boe"),
        ("303", "modelo-303-fichero-boe"),
    },
)

_M303_2025_RECORD_DESIGN = (
    bundled_path("corpus", "aeat_official")
    / "disenos_registro"
    / "modelo_303"
    / "files"
    / "06-303-ejercicio-2025-actualizado-04-12-2025-380-kb-xlsx.xlsx"
)

_M390_2025_RECORD_DESIGN = (
    bundled_path("corpus", "aeat_official")
    / "disenos_registro"
    / "modelo_390"
    / "files"
    / "01-390-ejercicio-2025-actualizado-05-12-2025-544-kb-xlsx.xlsx"
)


@cache
def _fixed_width_layout_cases() -> tuple[_LayoutCase, ...]:
    """Return all registry-authored fixed-width export layouts."""
    cases: list[_LayoutCase] = []
    for modelo in resources().modelos.authority.modelos:
        for revision in modelo.revisions.values():
            for layout in derive_export_layouts_from_bindings(revision):
                if layout.format is ExportLayoutFormat.FIXED_WIDTH:
                    cases.append((modelo, revision, layout))
    return tuple(cases)


def _layout_case_id(case: _LayoutCase) -> str:
    modelo, revision, layout = case
    return f"M{modelo.id}:{revision.id}:{layout.id}"


def _workbook_record_lengths(workbook_path: Path) -> dict[str, int]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    lengths: dict[str, int] = {}
    try:
        for worksheet in workbook.worksheets:
            last_position: int | None = None
            last_length: int | None = None
            for row in worksheet.iter_rows(min_row=1, max_col=3, values_only=True):
                position, length = row[1], row[2]
                if isinstance(position, int | float) and isinstance(length, int | float):
                    last_position = int(position)
                    last_length = int(length)
            if last_position is not None and last_length is not None:
                lengths[worksheet.title] = last_position + last_length - 1
    finally:
        workbook.close()
    return lengths


def _record_length(record: ExportRecordDefinition) -> int:
    return max((field.offset or 0) + (field.length or 0) - 1 for field in record.fields)


def _page_lengths(workbook_path: Path) -> dict[str, int]:
    """Map each design page label ("0", "2 bis", "5") to its official record length.

    The M390 design sheets are titled by page rather than by the record
    mnemonics the M303 design uses, so the label is taken from the sheet title.
    """
    lengths: dict[str, int] = {}
    for title, length in _workbook_record_lengths(workbook_path).items():
        normalised = title.replace("\xa0", " ").strip()
        _, separator, label = normalised.partition(".")
        if separator and label.strip():
            lengths[label.strip().lower()] = length
    return lengths


def _design_slot(workbook_path: Path, sheet_label: str, box_marker: str) -> tuple[int, int]:
    """Return the ``(offset, length)`` the design assigns to ``box_marker`` on a page."""
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        for worksheet in workbook.worksheets:
            normalised = worksheet.title.replace("\xa0", " ").strip()
            _, separator, label = normalised.partition(".")
            if not separator or label.strip().lower() != sheet_label:
                continue
            for row in worksheet.iter_rows(min_row=1, max_col=6, values_only=True):
                position, length, description = row[1], row[2], row[5]
                if not isinstance(position, int | float) or not isinstance(length, int | float):
                    continue
                if isinstance(description, str) and box_marker in description:
                    return int(position), int(length)
    finally:
        workbook.close()
    raise AssertionError(f"{box_marker} not found on sheet {sheet_label!r} of {workbook_path.name}")


def _modelo_390_fichero_boe_layout() -> ExportLayoutDefinition:
    modelo = next(item for item in resources().modelos.authority.modelos if item.id == "390")
    revision = modelo.revisions["2010-y-siguientes"]
    return next(
        item for item in derive_export_layouts_from_bindings(revision) if item.id == "modelo-390-fichero-boe"
    )


def test_registry_fixed_width_export_layout_surface_is_not_empty() -> None:
    """The fixed-width exporter must be backed by real registry layouts."""
    cases = _fixed_width_layout_cases()

    assert cases, "no registry-backed fixed-width export layouts were discovered"
    present = {(modelo.id, layout.id) for modelo, _revision, layout in cases}
    assert present >= _REQUIRED_LAYOUTS


def test_modelo_303_2025_record_lengths_match_official_workbook() -> None:
    """M303 registry record lengths are grounded in the official 2025 DR workbook."""
    official_lengths = _workbook_record_lengths(_M303_2025_RECORD_DESIGN)
    expected = {
        "modelo-303-page-01": official_lengths["DP30301"],
        "modelo-303-page-02": official_lengths["DP30302"],
        "modelo-303-page-03": official_lengths["DP30303"],
        "modelo-303-page-04": official_lengths["DP30304"],
        "modelo-303-page-05": official_lengths["DP30305"],
        "modelo-303-page-did": official_lengths["DP303DID"],
    }
    modelo = next(item for item in resources().modelos.authority.modelos if item.id == "303")
    revision = modelo.revisions["2023-y-siguientes"]
    layout = next(item for item in derive_export_layouts_from_bindings(revision) if item.id == "modelo-303-fichero-boe")
    actual = {record.id: _record_length(record) for record in layout.records if record.id in expected}

    assert actual == expected


def test_modelo_390_record_lengths_match_official_workbook() -> None:
    """M390 registry record lengths are grounded in the official DR workbook.

    A record's length is the extent of its end-of-record identifier, so a field
    authored past that identifier silently lengthens the emitted record. This
    gate is what makes such a field visible.
    """
    official = _page_lengths(_M390_2025_RECORD_DESIGN)

    assert official, "no page lengths were read from the official M390 design"

    expected = {
        "modelo-390-envelope-header": official["0"],
        "modelo-390-page-01": official["1"],
        "modelo-390-page-02": official["2"],
        "modelo-390-page-02b": official["2 bis"],
        "modelo-390-page-03": official["3"],
        "modelo-390-page-04": official["4"],
        "modelo-390-page-05": official["5"],
        "modelo-390-page-06": official["6"],
    }
    layout = _modelo_390_fichero_boe_layout()
    actual = {record.id: _record_length(record) for record in layout.records if record.id in expected}

    assert actual == expected


def test_modelo_390_simplificado_total_occupies_its_official_page_05_slot() -> None:
    """Box [79] is exported from the Pag. 5 record at the offset the design assigns it.

    The record-length gate alone cannot see this box go missing, because the
    end-of-record identifier fixes the length independently of the casilla
    fields in front of it.
    """
    offset, length = _design_slot(_M390_2025_RECORD_DESIGN, "5", "[79]")
    layout = _modelo_390_fichero_boe_layout()
    record = next(item for item in layout.records if item.id == "modelo-390-page-05")
    casilla_fields = [field for field in record.fields if field.kind is CasillaFieldKind.CASILLA]

    assert casilla_fields, "the Pag. 5 record declares no casilla field"
    slots = {(field.casilla_id, field.offset, field.length) for field in casilla_fields}

    assert ("iva.anual.reconciliacion.devengada-simplificado-303", offset, length) in slots


@pytest.mark.parametrize("case", _fixed_width_layout_cases(), ids=_layout_case_id)
def test_registry_fixed_width_layout_offsets_and_casilla_ids_are_valid(case: _LayoutCase) -> None:
    """Production offset guard accepts each layout and every casilla field is declared."""
    modelo, revision, layout = case

    _verify_record_offsets(layout)

    declared_ids: frozenset[CasillaId] = frozenset(casillas_by_id(revision))

    missing: list[str] = []
    for record in layout.records:
        for field in record.fields:
            if field.kind != CasillaFieldKind.CASILLA:
                continue
            if field.casilla_id not in declared_ids:
                missing.append(
                    f"{record.id}.{field.id}: casilla_id={field.casilla_id!r} "
                    f"offset={field.offset} length={field.length}",
                )

    assert not missing, (
        f"modelo {modelo.id} revision {revision.id} layout {layout.id} has export fields "
        "that do not reference declared canonical casilla.id values:\n" + "\n".join(missing)
    )
