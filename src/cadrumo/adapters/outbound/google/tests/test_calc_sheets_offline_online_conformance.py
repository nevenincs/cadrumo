"""Offline/online renderer conformance.

One ``SheetExportPlan`` must render to structurally identical grids whether
materialised offline (openpyxl xls) or online (Google-Sheets apply). This test
compares, cell-for-cell, the offline workbook against the online value/formula/
evidence writes the apply adapter would send — proving the single-builder /
two-transport invariant with no network.
"""

from __future__ import annotations

from datetime import UTC, datetime
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook

from .....application.storage.calc_sheets.records import (
    SheetCellAddress,
    SheetEvidenceContributorRow,
    SheetEvidenceFacet,
    SheetExportMetadata,
    SheetExportPlan,
    SheetFormulaCell,
    SheetGuideContent,
    SheetNumberFormat,
    SheetValueCell,
    TabName,
)
from .....application.storage.calc_sheets.workbook_export import guide_stamps, serialize_offline_workbook
from .....core.casilla_id import CasillaId, validated_casilla_id
from .....core.period import Period
from .._calc_sheets_apply_formatting import (
    build_emphasis_format_requests,
    build_number_format_requests,
)
from .._calc_sheets_apply_values import (
    build_evidence_value_data,
    build_formula_data,
    build_guide_value_data,
    build_value_data,
)

pytestmark = [pytest.mark.unit, pytest.mark.hex_outbound_adapter]
_BASE_CASILLA: CasillaId = validated_casilla_id("base", surface="_BASE_CASILLA")
_CUOTA_CASILLA: CasillaId = validated_casilla_id("cuota", surface="_CUOTA_CASILLA")


def _plan() -> SheetExportPlan:
    return SheetExportPlan(
        metadata=SheetExportMetadata(
            modelo_id="303",
            revision_id="2022",
            filing_year=2026,
            period=Period.from_year_and_code(2026, "1T"),
            engine_version="test",
            registry_sha="abcd1234",
            exported_at=datetime(2026, 6, 3, 15, 0, tzinfo=UTC),
        ),
        value_cells=(
            SheetValueCell(
                address=SheetCellAddress.at(TabName.ENTRADAS, 2, 4),
                value=Decimal("100.00"),
                role="operator_input",
                casilla_id=_BASE_CASILLA,
            ),
            SheetValueCell(
                address=SheetCellAddress.at(TabName.ENTRADAS, 1, 1),
                value="Sección",
                role="label",
            ),
        ),
        formula_cells=(
            SheetFormulaCell(
                address=SheetCellAddress.at(TabName.CALCULOS, 2, 4),
                formula="'Entradas'!D2*0.21",
                casilla_id=_CUOTA_CASILLA,
                rounding_rule="money",
                rounding_scale=2,
            ),
        ),
        guide=SheetGuideContent(title="Modelo 303", paragraphs=("Use Entradas.",)),
        evidence=SheetEvidenceFacet(
            snapshot_fingerprint="f" * 64,
            contributor_rows=(
                SheetEvidenceContributorRow(
                    casilla_id=_CUOTA_CASILLA,
                    transaction_id="c" * 64,
                    amount=Decimal("-121.00"),
                    currency="EUR",
                    taxable_base=Decimal("100.00"),
                    iva_rate=Decimal("0.21"),
                    iva_amount=Decimal("21.00"),
                    counterparty="Proveedor SL",
                    legal_refs=("ley-37-1992:art-99",),
                    source_refs=("boe-a-2026-1",),
                ),
            ),
        ),
    )


def _offline_value(workbook, address: SheetCellAddress) -> object:
    cell = workbook[address.tab.value].cell(row=address.row, column=address.column).value
    return cell if cell is not None else ""


def test_value_and_formula_cells_render_identically_offline_and_online() -> None:
    plan = _plan()
    workbook = load_workbook(BytesIO(serialize_offline_workbook(plan)), data_only=False)

    online_values = {entry["range"]: entry["values"][0][0] for entry in build_value_data(plan.value_cells)}
    online_formulas = {entry["range"]: entry["values"][0][0] for entry in build_formula_data(plan.formula_cells)}

    for cell in plan.value_cells:
        addr = cell.address.qualified()
        assert online_values[addr] == _offline_value(workbook, cell.address), addr
    for cell in plan.formula_cells:
        addr = cell.address.qualified()
        assert online_formulas[addr] == _offline_value(workbook, cell.address), addr
        # Both transports emit a live spreadsheet formula (leading '=').
        assert str(online_formulas[addr]).startswith("=")


def test_apply_adapter_emits_number_format_and_emphasis_requests() -> None:
    # contract: the online apply renders number formats + start/final + section
    # headers, not just values. A plan with a money number format yields a
    # NUMBER repeatCell; section headers / anchors yield bold repeatCells.
    from .....application.storage.calc_sheets.records import SheetAnchor, SheetSectionHeader

    plan = _plan().model_copy(
        update={
            "number_formats": (
                SheetNumberFormat(
                    address=SheetCellAddress.at(TabName.CALCULOS, 2, 4),
                    casilla_id=_CUOTA_CASILLA,
                    data_type="money",
                    pattern="#,##0.00",
                ),
            ),
            "section_headers": (
                SheetSectionHeader(address=SheetCellAddress.at(TabName.ENTRADAS, 2, 1), text="Sección A"),
            ),
            "anchors": (
                SheetAnchor(address=SheetCellAddress.at(TabName.CALCULOS, 2, 6), kind="final", label="RESULTADO"),
            ),
        },
    )
    sheet_id_by_tab = {tab.value: index for index, tab in enumerate(TabName)}
    number_requests = build_number_format_requests(plan, sheet_id_by_tab=sheet_id_by_tab)
    assert number_requests, "money casilla must yield a numberFormat request"
    fmt = number_requests[0]["repeatCell"]["cell"]["userEnteredFormat"]["numberFormat"]
    assert fmt == {"type": "NUMBER", "pattern": "#,##0.00"}

    emphasis = build_emphasis_format_requests(plan, sheet_id_by_tab=sheet_id_by_tab)
    assert emphasis, "section headers / anchors must yield bold requests"
    assert emphasis[0]["repeatCell"]["cell"]["userEnteredFormat"]["textFormat"]["bold"] is True


def test_guide_stamps_conform_offline_and_online() -> None:
    """The Guide export stamps render label-for-label across both transports.

    Each transport used to carry its own hand-written copy of the stamp
    table, and the two had drifted: offline wrote ``Revision`` / ``Periodo``,
    online wrote ``Revisión`` / ``Período``. The metadata VALUES matched, so
    nothing surfaced, while two exports of one plan disagreed about the row
    labels a reader or a label-keyed instruction would look for.

    Comparing the offline cells against the online writes — rather than
    against a restated expected table — is what makes this catch a future
    divergence instead of just pinning today's spelling.
    """
    plan = _plan()
    workbook = load_workbook(BytesIO(serialize_offline_workbook(plan)), data_only=False)
    sheet = workbook[TabName.GUIDE.value]
    online = {entry["range"]: entry["values"][0] for entry in build_guide_value_data(plan)}

    tab = TabName.GUIDE.value
    base_row = 3 + len(plan.guide.paragraphs) + 2
    stamps = guide_stamps(plan)
    assert stamps, "the plan must produce stamps or this test proves nothing"

    for offset in range(len(stamps)):
        row = base_row + offset
        offline_row = [(sheet.cell(row=row, column=column).value or "") for column in (1, 2)]
        assert online[f"'{tab}'!A{row}"] == offline_row, f"guide stamp row {row} diverged"


def test_the_guide_stamp_labels_are_the_accented_spanish_forms() -> None:
    """Pin which side of the drift won, so a future edit is a deliberate one.

    The conformance assertion above holds for ANY consistent pair, including
    a regression to the transliterated labels. These are Spanish
    operator-facing labels and both transports carry UTF-8, so the accented
    forms are the ones that stay.
    """
    labels = [label for label, _ in guide_stamps(_plan())]

    assert "Revisión" in labels
    assert "Período" in labels
    assert "Revision" not in labels
    assert "Periodo" not in labels


def test_evidence_surface_conforms_offline_and_online() -> None:
    plan = _plan()
    workbook = load_workbook(BytesIO(serialize_offline_workbook(plan)), data_only=False)
    sheet = workbook[TabName.EVIDENCIA.value]
    online = {entry["range"]: entry["values"][0] for entry in build_evidence_value_data(plan)}

    tab = TabName.EVIDENCIA.value
    # Fingerprint banner + header + the one contributor row conform cell-for-cell.
    assert online[f"'{tab}'!A1"] == [(sheet.cell(row=1, column=c).value or "") for c in (1, 2)]
    header_width = len(online[f"'{tab}'!A3"])
    assert online[f"'{tab}'!A3"] == [(sheet.cell(row=3, column=c).value or "") for c in range(1, header_width + 1)]
    assert online[f"'{tab}'!A4"] == [(sheet.cell(row=4, column=c).value or "") for c in range(1, header_width + 1)]
