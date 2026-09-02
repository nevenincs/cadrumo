"""Online Sheets Evidencia rendering + offline/online evidence parity.

The apply adapter must write the Evidencia surface to Google Sheets identically
to the offline xls workbook. Both transports consume the single ``evidence_table``
source, so the surfaces are byte-identical by construction; this test pins that
offline (no network) by comparing the apply adapter's Evidencia value-writes
against the offline workbook's Evidencia cells.
"""

from __future__ import annotations

from datetime import UTC, datetime
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook

from .....application.storage.calc_sheets.records import (
    SheetCellAddress,
    SheetEvidenceContributorRow,
    SheetEvidenceFacet,
    SheetEvidenceManualEntry,
    SheetExportMetadata,
    SheetExportPlan,
    SheetGuideContent,
    SheetValueCell,
    TabName,
)
from .....application.storage.calc_sheets.workbook_export import serialize_offline_workbook
from .....core.casilla_id import CasillaId, validated_casilla_id
from .....core.period import Period
from ..calc_sheets_apply import build_evidence_value_data

pytestmark = [pytest.mark.unit, pytest.mark.hex_outbound_adapter]
_BASE_CASILLA: CasillaId = validated_casilla_id("base", surface="_BASE_CASILLA")
_CUOTA_CASILLA: CasillaId = validated_casilla_id("cuota", surface="_CUOTA_CASILLA")
_RESULTADO_CONTABLE_CASILLA: CasillaId = validated_casilla_id(
    "resultado.contable",
    surface="_RESULTADO_CONTABLE_CASILLA",
)


def _evidence_plan() -> SheetExportPlan:
    return SheetExportPlan(
        metadata=SheetExportMetadata(
            modelo_id="303",
            revision_id="2022",
            filing_year=2026,
            period=Period.from_year_and_code(2026, "1T"),
            engine_version="test",
            registry_sha="abcd1234",
            exported_at=datetime(2026, 6, 3, 15, 0, tzinfo=UTC),
        ),
        value_cells=(
            SheetValueCell(
                address=SheetCellAddress.at(TabName.ENTRADAS, 2, 4),
                value=Decimal("100.00"),
                role="operator_input",
                casilla_id=_BASE_CASILLA,
            ),
        ),
        guide=SheetGuideContent(title="Modelo 303", paragraphs=("Use Entradas.",)),
        evidence=SheetEvidenceFacet(
            snapshot_fingerprint="f" * 64,
            contributor_rows=(
                SheetEvidenceContributorRow(
                    casilla_id=_CUOTA_CASILLA,
                    transaction_id="c" * 64,
                    amount=Decimal("-121.00"),
                    currency="EUR",
                    taxable_base=Decimal("100.00"),
                    iva_rate=Decimal("0.21"),
                    iva_amount=Decimal("21.00"),
                    counterparty="Proveedor SL",
                    attachment_ids=("attachment-1",),
                    document_link_ids=("drive-doc-1",),
                    legal_refs=("ley-37-1992:art-99",),
                    source_refs=("boe-a-2026-1",),
                ),
            ),
            manual_entries=(
                SheetEvidenceManualEntry(
                    casilla_id=_RESULTADO_CONTABLE_CASILLA,
                    value="140000.00",
                    kind="casilla_input",
                    note="operator supplied accounting result",
                    legal_refs=("ley-27-2014:art-10",),
                    source_refs=("operator-manual-evidence",),
                ),
            ),
        ),
    )


def _online_cell_grid(plan: SheetExportPlan) -> dict[str, list[object]]:
    """Map A1 cell anchor -> row values from the apply adapter's evidence writes."""
    grid: dict[str, list[object]] = {}
    for entry in build_evidence_value_data(plan):
        grid[entry["range"]] = entry["values"][0]
    return grid


def test_apply_adapter_renders_evidencia_surface() -> None:
    plan = _evidence_plan()
    grid = _online_cell_grid(plan)
    tab = TabName.EVIDENCIA.value
    # Fingerprint banner + header row + one contributor + one manual row.
    assert grid[f"'{tab}'!A1"] == ["Snapshot fingerprint", "f" * 64]
    assert grid[f"'{tab}'!A3"][0] == "Tipo"
    assert grid[f"'{tab}'!A4"][:3] == ["ledger", "cuota", "c" * 64]
    assert grid[f"'{tab}'!A5"][:2] == ["manual", "resultado.contable"]


def test_online_evidencia_is_byte_identical_to_offline() -> None:
    plan = _evidence_plan()
    online = _online_cell_grid(plan)

    workbook = load_workbook(BytesIO(serialize_offline_workbook(plan)), data_only=False)
    sheet = workbook[TabName.EVIDENCIA.value]

    def _offline_row(row: int, width: int) -> list[object]:
        # openpyxl reads an empty cell back as None; the online transport writes
        # "" for the same blank. Normalise that representation difference — the
        # logical evidence content is what must match.
        return [(sheet.cell(row=row, column=col).value or "") for col in range(1, width + 1)]

    tab = TabName.EVIDENCIA.value
    # A1 banner (2 cols), header row 3, contributor row 4, manual row 5 — the
    # online value-writes must equal the offline cell values cell-for-cell.
    assert online[f"'{tab}'!A1"] == _offline_row(1, 2)
    assert online[f"'{tab}'!A3"] == _offline_row(3, len(online[f"'{tab}'!A3"]))
    assert online[f"'{tab}'!A4"] == _offline_row(4, len(online[f"'{tab}'!A4"]))
    assert online[f"'{tab}'!A5"] == _offline_row(5, len(online[f"'{tab}'!A5"]))
