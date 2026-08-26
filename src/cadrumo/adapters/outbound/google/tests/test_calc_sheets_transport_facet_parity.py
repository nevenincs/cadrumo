"""Real-behavior tests: both calc-sheets transports render the same plan facets.

The module documentation says the offline XLSX materialiser and the Google
Sheets request builder mirror one style plan. Two facets did not.

Vertical alignment lived as a literal inside the offline materialiser, so the
online builder never emitted it and Sheets kept its bottom-aligned default.
Neither side looked wrong on its own: no palette entry was missing, and the
omission was of a key nobody had written down.

Protection diverged in *both* directions at once, which is why it survived
review. ``SheetExportPlan.protected_ranges`` is the declared read-only
contract; the online adapter emitted one ``addProtectedRange`` per entry
while the offline materialiser consumed none of them and instead locked
Evidencia -- a tab the plan never mentioned. So offline left every planned
range editable and locked something unplanned, and online did the reverse.

These tests read a real Modelo 130 plan and compare the two rendered
artefacts, so a facet added to one transport and not the other fails here.
"""

from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import load_workbook

from .....application.storage.calc_sheets import (
    STYLED_RANGE_VERTICAL_ALIGN,
    SheetExportPlan,
    build_export_plan,
    serialize_offline_workbook,
)
from .....domain.calculations.registry.authority import bundled_authority
from .._calc_sheets_apply import (
    _build_protected_range_requests,
    _build_styled_range_requests,
)

pytestmark = [pytest.mark.unit, pytest.mark.hex_outbound_adapter]


def _plan() -> SheetExportPlan:
    snapshot = bundled_authority().snapshot("130", filing_year=2026, period="1T")
    return build_export_plan(snapshot)


def _sheet_ids(plan: SheetExportPlan) -> dict[str, int]:
    tabs = {styled.tab.value for styled in plan.styled_ranges}
    tabs |= {protected.tab.value for protected in plan.protected_ranges}
    return {tab: index for index, tab in enumerate(sorted(tabs))}


def _require_concrete_bounds(
    *,
    start_row: int | None,
    end_row: int | None,
    start_column: int | None,
    end_column: int | None,
    tab: str,
) -> tuple[int, int, int, int]:
    """Return the concrete bounds an offline-protected range must carry."""
    if start_row is None or end_row is None or start_column is None or end_column is None:
        raise AssertionError(f"protected range {tab!r} must have concrete bounds")
    return start_row, end_row, start_column, end_column


def test_every_styled_range_carries_vertical_alignment_online() -> None:
    plan = _plan()

    requests = _build_styled_range_requests(plan, sheet_id_by_tab=_sheet_ids(plan))

    assert requests, "no styled repeatCell requests were generated"
    for request in requests:
        cell_format = request["repeatCell"]["cell"]["userEnteredFormat"]
        assert cell_format["verticalAlignment"] == "TOP"
        assert "userEnteredFormat.verticalAlignment" in request["repeatCell"]["fields"]


def test_offline_vertical_alignment_matches_the_shared_constant() -> None:
    plan = _plan()

    workbook = load_workbook(BytesIO(serialize_offline_workbook(plan)))
    styled = plan.styled_ranges[0]
    cell = workbook[styled.tab.value].cell(row=styled.start_row, column=styled.start_column)

    assert cell.alignment.vertical == STYLED_RANGE_VERTICAL_ALIGN


def test_the_two_transports_protect_the_same_tabs() -> None:
    plan = _plan()
    sheet_ids = _sheet_ids(plan)
    tab_by_sheet_id = {sheet_id: tab for tab, sheet_id in sheet_ids.items()}

    requests = _build_protected_range_requests(plan.protected_ranges, sheet_id_by_tab=sheet_ids)
    online_tabs = {
        tab_by_sheet_id[request["addProtectedRange"]["protectedRange"]["range"]["sheetId"]] for request in requests
    }
    workbook = load_workbook(BytesIO(serialize_offline_workbook(plan)))
    offline_tabs = {name for name in workbook.sheetnames if workbook[name].protection.sheet}

    planned_tabs = {protected.tab.value for protected in plan.protected_ranges}
    assert planned_tabs
    assert online_tabs == planned_tabs
    assert offline_tabs == planned_tabs


def test_a_tab_the_plan_does_not_protect_stays_editable_offline() -> None:
    """The editable-input policy has to be explicit, not a side effect."""
    plan = _plan()
    planned_tabs = {protected.tab.value for protected in plan.protected_ranges}

    workbook = load_workbook(BytesIO(serialize_offline_workbook(plan)))

    unplanned = [name for name in workbook.sheetnames if name not in planned_tabs]
    assert unplanned
    for name in unplanned:
        assert workbook[name].protection.sheet is False


def test_every_planned_range_cell_is_locked_offline() -> None:
    """XLSX has no per-range primitive, so the range set is expressed per cell.

    The observable contract stops at the tab's written surface. A cell that
    was never written does not exist in the file, so it carries no ``locked``
    bit of its own and simply inherits the sheet flag -- there is nothing to
    assert about it, and asserting anyway would be reading openpyxl's
    materialisation rather than the plan.
    """
    plan = _plan()

    workbook = load_workbook(BytesIO(serialize_offline_workbook(plan)))

    for protected in plan.protected_ranges:
        worksheet = workbook[protected.tab.value]
        corner = worksheet.cell(row=protected.start_row, column=protected.start_column)
        assert corner.protection.locked is True, protected.tab.value


def test_written_cells_outside_every_planned_range_stay_editable() -> None:
    """The editable-input policy, asserted where the format can express it."""
    plan = _plan()
    planned_tabs = {protected.tab.value for protected in plan.protected_ranges}

    workbook = load_workbook(BytesIO(serialize_offline_workbook(plan)))

    for protected in plan.protected_ranges:
        worksheet = workbook[protected.tab.value]
        bounds = _require_concrete_bounds(
            start_row=protected.start_row,
            end_row=protected.end_row,
            start_column=protected.start_column,
            end_column=protected.end_column,
            tab=protected.tab.value,
        )
        start_row, end_row, start_column, end_column = bounds
        assert isinstance(start_row, int)
        assert isinstance(end_row, int)
        assert isinstance(start_column, int)
        assert isinstance(end_column, int)
        for row in worksheet.iter_rows():
            for cell in row:
                row_index = cell.row
                column_index = cell.column
                assert isinstance(row_index, int)
                assert isinstance(column_index, int)
                inside_rows = start_row <= row_index <= end_row
                inside_columns = start_column <= column_index <= end_column
                if inside_rows and inside_columns:
                    continue
                assert cell.protection.locked is False, f"{protected.tab.value}!{cell.coordinate}"

    # Entradas carries no planned range at all, so it is the tab that proves
    # the policy end to end: untouched by the protection pass entirely, and
    # therefore fully editable rather than locked-by-default.
    assert "Entradas" not in planned_tabs
    assert workbook["Entradas"].protection.sheet is False
