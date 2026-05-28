"""Read-only extraction of official AEAT record-design rows."""

from __future__ import annotations

import re
import warnings
from collections.abc import Iterator, Mapping
from contextlib import contextmanager
from dataclasses import dataclass, field
from functools import lru_cache
from io import BufferedReader, BytesIO
from pathlib import Path

import pdfplumber
import pypdfium2 as pdfium
import xlrd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from pdfplumber.page import Page
from pydantic import ConfigDict
from xlrd.sheet import Sheet as XlrdSheet

from ....core.logging import get_logger
from ._errors import RegistryValidationError
from ._runtime_graph import expression_casilla_refs
from ._schema import CasillaDefinition, ModeloRevision, RegistryModel

_log = get_logger(__name__)

_OPENPYXL_HEADER_FOOTER_WARNING = "Cannot parse header or footer so it will be ignored"
_OPENPYXL_PRINT_AREA_WARNING = r"Print area cannot be set to Defined name: .*"


class RecordDesignField(RegistryModel):
    """One fixed-width field described by an AEAT record-design sheet."""

    sheet: str
    row: int
    ordinal: int
    offset: int
    length: int
    type_code: str
    complementary: str | None = None
    description: str
    validation: str | None = None
    content: str | None = None


class RecordDesignSheet(RegistryModel):
    """Parsed field rows and declared total length for one workbook sheet."""

    model_config = ConfigDict(frozen=True, extra="forbid")

    name: str
    fields: tuple[RecordDesignField, ...]
    total_positions: int | None = None


@dataclass(frozen=True)
class _WorkbookHeader:
    row_number: int
    ordinal_index: int
    offset_index: int
    length_index: int
    type_index: int
    complementary_index: int | None
    description_index: int
    validation_index: int | None
    content_index: int | None


def extract_record_design(path: Path) -> tuple[RecordDesignSheet, ...]:
    """Return fixed-width field rows from a supported official record-design source."""

    resolved = path.resolve()
    if not resolved.is_file():
        raise FileNotFoundError(f"record-design source not found: {path}")
    stat = resolved.stat()
    return _extract_record_design_cached(str(resolved), stat.st_size, stat.st_mtime_ns)


@lru_cache(maxsize=256)
def _extract_record_design_cached(
    path: str,
    byte_count: int,
    modified_ns: int,
) -> tuple[RecordDesignSheet, ...]:
    del byte_count, modified_ns
    source_path = Path(path)
    suffix = source_path.suffix.lower()
    if suffix == ".pdf":
        return extract_record_design_pdf(source_path)
    if suffix in {".xlsx", ".xlsm"}:
        return extract_record_design_workbook(source_path)
    if suffix == ".xls":
        return extract_record_design_xls_workbook(source_path)
    raise RegistryValidationError(f"unsupported record-design source extension: {source_path.suffix}")


def extract_record_design_workbook(path: Path) -> tuple[RecordDesignSheet, ...]:
    """Return the official fixed-width field rows described by workbook ``path``."""

    resolved = path.resolve()
    if not resolved.is_file():
        raise FileNotFoundError(f"record-design workbook not found: {path}")
    stat = resolved.stat()
    return _extract_record_design_workbook_cached(str(resolved), stat.st_size, stat.st_mtime_ns)


def extract_record_design_xls_workbook(path: Path) -> tuple[RecordDesignSheet, ...]:
    """Return official fixed-width field rows from a legacy binary XLS workbook."""

    resolved = path.resolve()
    if not resolved.is_file():
        raise FileNotFoundError(f"record-design XLS workbook not found: {path}")
    stat = resolved.stat()
    return _extract_record_design_xls_workbook_cached(str(resolved), stat.st_size, stat.st_mtime_ns)


@lru_cache(maxsize=256)
def _extract_record_design_workbook_cached(
    path: str,
    byte_count: int,
    modified_ns: int,
) -> tuple[RecordDesignSheet, ...]:
    del byte_count, modified_ns
    source_path = Path(path)
    with _ignore_openpyxl_header_footer_metadata_warnings():
        workbook = load_workbook(source_path, read_only=True, data_only=True)
        try:
            sheets: list[RecordDesignSheet] = []
            skipped: list[str] = []
            for worksheet in workbook.worksheets:
                try:
                    sheets.append(_extract_sheet(worksheet))
                except ValueError as exc:
                    if "has no record-design header" not in str(exc):
                        raise
                    skipped.append(worksheet.title)
            if not sheets:
                skipped_sheets = ", ".join(skipped) if skipped else "none"
                raise RegistryValidationError(
                    f"{source_path}: no record-design sheets found; skipped sheets: {skipped_sheets}"
                )
            return tuple(sheets)
        finally:
            workbook.close()


@lru_cache(maxsize=128)
def _extract_record_design_xls_workbook_cached(
    path: str,
    byte_count: int,
    modified_ns: int,
) -> tuple[RecordDesignSheet, ...]:
    del byte_count, modified_ns
    source_path = Path(path)
    workbook = xlrd.open_workbook(str(source_path), on_demand=True)
    try:
        sheets: list[RecordDesignSheet] = []
        skipped: list[str] = []
        for sheet_name in workbook.sheet_names():
            worksheet = workbook.sheet_by_name(sheet_name)
            try:
                sheets.append(_extract_xls_sheet(worksheet))
            except ValueError as exc:
                if "has no record-design header" not in str(exc):
                    raise
                skipped.append(sheet_name)
        if not sheets:
            skipped_sheets = ", ".join(skipped) if skipped else "none"
            raise RegistryValidationError(
                f"{source_path}: no record-design sheets found; skipped sheets: {skipped_sheets}"
            )
        return tuple(sheets)
    finally:
        workbook.release_resources()


@contextmanager
def _ignore_openpyxl_header_footer_metadata_warnings() -> Iterator[None]:
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message=_OPENPYXL_HEADER_FOOTER_WARNING,
            category=UserWarning,
            module=r"openpyxl\.worksheet\.header_footer",
        )
        warnings.filterwarnings(
            "ignore",
            message=_OPENPYXL_PRINT_AREA_WARNING,
            category=UserWarning,
            module=r"openpyxl\.reader\.workbook",
        )
        yield


def extract_record_design_pdf(path: Path) -> tuple[RecordDesignSheet, ...]:
    """Return fixed-width field rows extracted from an official AEAT PDF."""

    resolved = path.resolve()
    if not resolved.is_file():
        raise FileNotFoundError(f"record-design PDF not found: {path}")
    stat = resolved.stat()
    return _extract_record_design_pdf_cached(str(resolved), stat.st_size, stat.st_mtime_ns)


@lru_cache(maxsize=256)
def _extract_record_design_pdf_cached(
    path: str,
    byte_count: int,
    modified_ns: int,
) -> tuple[RecordDesignSheet, ...]:
    del byte_count, modified_ns
    source_path = Path(path)
    with source_path.open("rb") as pdf_file:
        return _extract_record_design_pdf_stream(pdf_file, source_label=str(source_path))


def extract_record_design_pdf_bytes(
    pdf_bytes: bytes,
    *,
    source_label: str = "in-memory record-design PDF",
) -> tuple[RecordDesignSheet, ...]:
    """Return fixed-width field rows extracted from PDF bytes."""

    return _extract_record_design_pdf_stream(BytesIO(pdf_bytes), source_label=source_label)


def _extract_record_design_pdf_stream(
    stream: BufferedReader | BytesIO,
    *,
    source_label: str,
) -> tuple[RecordDesignSheet, ...]:
    pdf_bytes = stream.read()
    lines = _extract_pdf_text_lines(pdf_bytes, source_label=source_label)
    if _uses_page_record_layout(lines):
        lines = _extract_pdfplumber_text_lines(pdf_bytes, source_label=source_label)
    if not any(line.strip() for line in lines):
        raise RegistryValidationError(f"no text extracted from record-design PDF {source_label}")
    try:
        return _extract_pdf_lines(lines, source_label=source_label)
    except ValueError as pdfium_exc:
        text_fallback_error = pdfium_exc
        try:
            fallback_lines = _extract_pdfplumber_text_lines(pdf_bytes, source_label=source_label)
            return _extract_pdf_lines(fallback_lines, source_label=source_label)
        except ValueError as fallback_exc:
            text_fallback_error = fallback_exc
        if "did not contain parseable field rows" not in str(text_fallback_error):
            raise text_fallback_error from pdfium_exc
        try:
            with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
                pages = tuple(_snapshot_pdf_page(page) for page in pdf.pages)
        except Exception as pdf_exc:  # pragma: no cover - defensive; pdfplumber surface
            raise RegistryValidationError(
                f"pdfplumber could not open record-design PDF {source_label}: {pdf_exc}"
            ) from pdf_exc
        visual_chart = _extract_visual_record_design_chart(pages, source_label=source_label)
        if visual_chart:
            return visual_chart
        raise


def _extract_sheet(worksheet: Worksheet) -> RecordDesignSheet:
    header = _find_header(worksheet)
    return _extract_sheet_rows(
        worksheet.title,
        header,
        enumerate(
            worksheet.iter_rows(min_row=header.row_number + 1, values_only=True),
            start=header.row_number + 1,
        ),
    )


def _extract_xls_sheet(worksheet: XlrdSheet) -> RecordDesignSheet:
    header = _find_xls_header(worksheet)
    return _extract_sheet_rows(
        worksheet.name,
        header,
        ((rowx + 1, tuple(worksheet.row_values(rowx))) for rowx in range(header.row_number, worksheet.nrows)),
    )


def _extract_sheet_rows(
    sheet_name: str,
    header: _WorkbookHeader,
    rows: Iterator[tuple[int, tuple[object, ...]]],
) -> RecordDesignSheet:
    # AEAT Diseño workbooks occasionally carry surrounding whitespace on a
    # sheet tab (e.g. 'DP200026 '). The sheet name is the record-segment
    # identity that segment-qualified casillas and the calculation-
    # completeness derivation match against, so the raw tab whitespace
    # must not leak into that identity.
    sheet_name = sheet_name.strip()
    fields: list[RecordDesignField] = []
    total_positions: int | None = None
    trailing_blank_rows = 0
    for row_number, row in rows:
        values = tuple(row)
        if _is_blank_row(values):
            if fields:
                trailing_blank_rows += 1
                if trailing_blank_rows >= 25:
                    break
            continue
        trailing_blank_rows = 0
        row_total = _total_positions_from_row(values)
        if row_total is not None:
            total_positions = row_total
            continue
        ordinal = _int_or_none(_cell(values, header.ordinal_index))
        offset = _int_or_none(_cell(values, header.offset_index))
        length = _int_or_none(_cell(values, header.length_index))
        if ordinal is None or offset is None or length is None:
            continue
        type_code = _required_text(_cell(values, header.type_index), sheet_name, row_number, "type")
        complementary = _optional_header_text(values, header.complementary_index)
        validation = _optional_header_text(values, header.validation_index)
        content = _optional_header_text(values, header.content_index)
        description = _field_description_text(
            values,
            header=header,
            content=content,
            sheet=sheet_name,
            row=row_number,
        )
        fields.append(
            RecordDesignField(
                sheet=sheet_name,
                row=row_number,
                ordinal=ordinal,
                offset=offset,
                length=length,
                type_code=type_code,
                complementary=complementary,
                description=description,
                validation=validation,
                content=content,
            )
        )
    return RecordDesignSheet(name=sheet_name, fields=tuple(fields), total_positions=total_positions)


def _is_blank_row(values: tuple[object, ...]) -> bool:
    return all(value is None or str(value).strip() == "" for value in values)


def _find_header(worksheet: Worksheet) -> _WorkbookHeader:
    for row_number, row in enumerate(worksheet.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        values = tuple(row)
        if _normalise_header_cell(_cell(values, 0)) not in {"no", "n"}:
            continue
        try:
            offset_index = _required_header_index(values, "posic.")
            length_index = _required_header_index(values, "lon")
            type_index = _required_header_index(values, "tipo")
            description_index = _required_header_index(values, "descripcion")
        except ValueError as header_exc:
            _log.debug(
                "record-design header probe (xlsx %s): row %d missing required columns (%s); trying next",
                worksheet.title,
                row_number,
                header_exc,
            )
            continue
        return _WorkbookHeader(
            row_number=row_number,
            ordinal_index=0,
            offset_index=offset_index,
            length_index=length_index,
            type_index=type_index,
            complementary_index=_optional_header_index(values, "com", "comp"),
            description_index=description_index,
            validation_index=_optional_header_index(values, "validacion", "oblig."),
            content_index=_optional_header_index(values, "contenido"),
        )
    raise RegistryValidationError(f"{worksheet.title!r} has no record-design header")


def _find_xls_header(worksheet: XlrdSheet) -> _WorkbookHeader:
    for rowx in range(min(10, worksheet.nrows)):
        values = tuple(worksheet.row_values(rowx))
        if _normalise_header_cell(_cell(values, 0)) not in {"no", "n"}:
            continue
        try:
            offset_index = _required_header_index(values, "posic.")
            length_index = _required_header_index(values, "lon")
            type_index = _required_header_index(values, "tipo")
            description_index = _required_header_index(values, "descripcion")
        except ValueError as header_exc:
            _log.debug(
                "record-design header probe (xls): row %d missing required columns (%s); trying next",
                rowx + 1,
                header_exc,
            )
            continue
        return _WorkbookHeader(
            row_number=rowx + 1,
            ordinal_index=0,
            offset_index=offset_index,
            length_index=length_index,
            type_index=type_index,
            complementary_index=_optional_header_index(values, "com", "comp"),
            description_index=description_index,
            validation_index=_optional_header_index(values, "validacion", "oblig."),
            content_index=_optional_header_index(values, "contenido"),
        )
    raise RegistryValidationError(f"{worksheet.name!r} has no record-design header")


def _cell(values: tuple[object, ...], index: int) -> object | None:
    return values[index] if index < len(values) else None


def _clean(value: object | None) -> str:
    return "" if value is None else str(value).strip()


def _optional_text(value: object | None) -> str | None:
    cleaned = _clean(value)
    return cleaned or None


def _optional_header_text(values: tuple[object, ...], index: int | None) -> str | None:
    if index is None:
        return None
    return _optional_text(_cell(values, index))


def _required_text(value: object | None, sheet: str, row: int, field: str) -> str:
    cleaned = _clean(value)
    if not cleaned:
        raise RegistryValidationError(f"{sheet!r} row {row} missing {field}")
    return cleaned


def _field_description_text(
    values: tuple[object, ...],
    *,
    header: _WorkbookHeader,
    content: str | None,
    sheet: str,
    row: int,
) -> str:
    description = _optional_text(_cell(values, header.description_index))
    if description is not None:
        return description
    if content is not None:
        return content
    raise RegistryValidationError(f"{sheet!r} row {row} missing description")


def _int_or_none(value: object | None) -> int | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return None


def _normalise_header_cell(value: object | None) -> str:
    return (
        _clean(value)
        .casefold()
        .replace("º", "o")
        .replace("ó", "o")
        .replace("í", "i")
        .replace("á", "a")
        .replace("é", "e")
        .replace("ú", "u")
    )


def _required_header_index(values: tuple[object, ...], header_name: str) -> int:
    index = _optional_header_index(values, header_name)
    if index is None:
        raise RegistryValidationError(f"missing workbook header {header_name!r}")
    return index


def _optional_header_index(values: tuple[object, ...], *header_names: str) -> int | None:
    expected = set(header_names)
    for index, value in enumerate(values):
        if _normalise_header_cell(value) in expected:
            return index
    return None


def _total_positions_from_row(values: tuple[object, ...]) -> int | None:
    for index, value in enumerate(values):
        if _normalise_header_cell(value) != "total":
            continue
        for candidate in values[index + 1 :]:
            total = _int_or_none(candidate)
            if total is not None:
                return total
        return None
    return None


_COMPACT_PDF_ROW_RE = re.compile(
    r"^\s*(?P<ordinal>\d+)\s+(?P<offset>\d+)\s+(?P<length>\d+)\s+(?P<type>An|Num|N|A)\s+(?P<text>.+)$",
    re.IGNORECASE,
)
_COMPACT_PDF_CRLF_ROW_RE = re.compile(
    r"^\s*(?P<ordinal>\d+)\s+(?P<offset>\d+)\s+(?P<type>An|Num|N|A)\s+"
    r"(?P<text>Salto de l[íi]nea\..*CRLF\.?)$",
    re.IGNORECASE,
)
_NARRATIVE_PDF_ROW_RE = re.compile(
    r"^\s*(?P<start>\d+)(?:\s*[-\u2013]\s*(?P<end>\d+))?\s+"
    r"(?P<type>Alfanum[eé]rico|Alfab[eé]tico|Num[eé]rico|[-\u2013]+)\s*"
    r"(?P<text>.*)$",
    re.IGNORECASE,
)
_PDF_PAGE_RECORD_RE = re.compile(r"^P[áa]g\s+(?P<page>\d+)\s+DISE[ÑN]O DE REGISTRO\b", re.IGNORECASE)
_PDF_RECORD_HEADING_RE = re.compile(
    r"^(?:[A-Z]\.?\s*-?\s*)?(?:TIPO DE REGISTRO|Tipo de registro)\s+"
    r"(?P<record>\d+)\s*:\s*(?P<title>.+)$",
    re.IGNORECASE,
)


@dataclass
class _PdfFieldDraft:
    sheet: str
    row: int
    ordinal: int
    offset: int
    length: int
    type_code: str
    description_parts: list[str] = field(default_factory=list)
    content_parts: list[str] = field(default_factory=list)

    def append_continuation(self, line: str) -> None:
        if not self.description_parts or (not self.content_parts and _looks_like_title_continuation(line)):
            self.description_parts.append(line)
            return
        self.content_parts.append(line)

    def finish(self) -> RecordDesignField:
        description = _join_pdf_parts(self.description_parts)
        if not description:
            raise RegistryValidationError(f"{self.sheet!r} PDF row {self.row} missing description")
        return RecordDesignField(
            sheet=self.sheet,
            row=self.row,
            ordinal=self.ordinal,
            offset=self.offset,
            length=self.length,
            type_code=self.type_code,
            complementary=None,
            description=description,
            validation=None,
            content=_join_pdf_parts(self.content_parts) or None,
        )


@dataclass
class _PdfSheetDraft:
    name: str
    fields: list[RecordDesignField] = field(default_factory=list)
    current: _PdfFieldDraft | None = None

    def start_field(self, row: _PdfRow) -> None:
        self.finish_current()
        self.current = _PdfFieldDraft(
            sheet=self.name,
            row=row.source_row,
            ordinal=row.ordinal or len(self.fields) + 1,
            offset=row.offset,
            length=row.length,
            type_code=row.type_code,
            description_parts=[row.description] if row.description else [],
        )

    def finish_current(self) -> None:
        if self.current is None:
            return
        self.fields.append(self.current.finish())
        self.current = None

    def finish(self, *, source_label: str) -> RecordDesignSheet:
        self.finish_current()
        total_positions = max((field.offset + field.length - 1 for field in self.fields), default=None)
        sheet = RecordDesignSheet(name=self.name, fields=tuple(self.fields), total_positions=total_positions)
        _validate_pdf_sheet(sheet, source_label=source_label)
        return sheet


@dataclass(frozen=True)
class _PdfRow:
    source_row: int
    ordinal: int | None
    offset: int
    length: int
    type_code: str
    description: str


@dataclass(frozen=True)
class _PdfWord:
    text: str
    x0: float
    x1: float
    top: float
    bottom: float


@dataclass(frozen=True)
class _PdfRect:
    x0: float
    x1: float
    top: float
    bottom: float
    width: float
    height: float
    fill: object | None


@dataclass(frozen=True)
class _PdfPageSnapshot:
    lines: tuple[str, ...]
    words: tuple[_PdfWord, ...]
    rects: tuple[_PdfRect, ...]


@dataclass(frozen=True)
class _VisualChartFragment:
    start: int
    end: int
    description: str


def _extract_pdf_text_lines(pdf_bytes: bytes, *, source_label: str) -> tuple[str, ...]:
    try:
        document = pdfium.PdfDocument(pdf_bytes)
    except Exception as exc:  # pragma: no cover - pdfium parser surface
        raise RegistryValidationError(f"pypdfium2 could not open record-design PDF {source_label}: {exc}") from exc
    try:
        lines: list[str] = []
        for page in document:
            text_page = page.get_textpage()
            try:
                lines.extend(text_page.get_text_range().splitlines())
            finally:
                text_page.close()
                page.close()
        return tuple(lines)
    finally:
        document.close()


def _extract_pdfplumber_text_lines(pdf_bytes: bytes, *, source_label: str) -> tuple[str, ...]:
    try:
        with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
            return tuple(line for page in pdf.pages for line in _extract_pdf_page_lines(page))
    except Exception as exc:  # pragma: no cover - defensive; pdfplumber surface
        raise RegistryValidationError(f"pdfplumber could not open record-design PDF {source_label}: {exc}") from exc


def _uses_page_record_layout(lines: tuple[str, ...]) -> bool:
    return any(_pdf_page_name(_clean_pdf_line(line)) is not None for line in lines)


def _snapshot_pdf_page(page: Page) -> _PdfPageSnapshot:
    return _PdfPageSnapshot(
        lines=_extract_pdf_page_lines(page),
        words=tuple(
            _PdfWord(
                text=str(word["text"]),
                x0=float(word["x0"]),
                x1=float(word["x1"]),
                top=float(word["top"]),
                bottom=float(word["bottom"]),
            )
            for word in page.extract_words()
        ),
        rects=tuple(
            _PdfRect(
                x0=float(rect["x0"]),
                x1=float(rect["x1"]),
                top=float(rect["top"]),
                bottom=float(rect["bottom"]),
                width=float(rect["width"]),
                height=float(rect["height"]),
                fill=rect.get("non_stroking_color"),
            )
            for rect in page.rects
        ),
    )


def _extract_pdf_page_lines(page: Page) -> tuple[str, ...]:
    text = page.extract_text() or ""
    return tuple(text.splitlines())


class _PdfParseState:
    """Mutable state for the PDF record-design line parser.

    Encapsulates the three locals (``current`` draft sheet,
    ``in_table`` flag, ``pending_name`` carried across page-name
    boundaries) so the per-line dispatch can mutate them without
    threading three out-parameters through every helper.
    """

    __slots__ = ("current", "in_table", "pending_name", "sheets", "source_label")

    def __init__(self, *, source_label: str) -> None:
        self.sheets: list[RecordDesignSheet] = []
        self.current: _PdfSheetDraft | None = None
        self.in_table: bool = False
        self.pending_name: str | None = None
        self.source_label = source_label

    def finalise(self) -> tuple[RecordDesignSheet, ...]:
        if self.current is not None:
            self.sheets.append(self.current.finish(source_label=self.source_label))
        non_empty = tuple(sheet for sheet in self.sheets if sheet.fields)
        if not non_empty:
            raise RegistryValidationError("record-design PDF did not contain parseable field rows")
        return non_empty

    def feed(self, line: str, row_number: int) -> None:
        if not line or _is_pdf_footer(line):
            return
        if self._consume_page_name(line):
            return
        if self._consume_record_heading(line):
            return
        if self._consume_table_header(line):
            return
        if self._consume_title_continuation(line):
            return
        if _is_pdf_page_heading(line):
            return
        if self._consume_field_row(line, row_number):
            return
        self._consume_field_continuation(line)

    def _consume_page_name(self, line: str) -> bool:
        page_name = _pdf_page_name(line)
        if page_name is None:
            return False
        self.pending_name = page_name
        if self.current is not None and self.current.name != page_name:
            self.sheets.append(self.current.finish(source_label=self.source_label))
            self.current = _PdfSheetDraft(page_name)
        return True

    def _consume_record_heading(self, line: str) -> bool:
        heading_name = _pdf_record_heading_name(line)
        if heading_name is None:
            return False
        if self.current is not None:
            self.sheets.append(self.current.finish(source_label=self.source_label))
        self.current = _PdfSheetDraft(heading_name)
        self.in_table = False
        return True

    def _consume_table_header(self, line: str) -> bool:
        if not _is_pdf_header(line):
            return False
        if self.current is None:
            self.current = _PdfSheetDraft(self.pending_name or "PDF record design")
        self.in_table = True
        return True

    def _consume_title_continuation(self, line: str) -> bool:
        if self.in_table or self.current is None or self.current.fields:
            return False
        if not _looks_like_title_continuation(line):
            return False
        self.current.name = _normalise_pdf_sheet_name(_join_pdf_parts([self.current.name, line]))
        return True

    def _consume_field_row(self, line: str, row_number: int) -> bool:
        row = _parse_pdf_row(line, row_number)
        if row is None:
            return False
        if self.current is None:
            self.current = _PdfSheetDraft(self.pending_name or "PDF record design")
        self.current.start_field(row)
        self.in_table = True
        return True

    def _consume_field_continuation(self, line: str) -> None:
        if self.in_table and self.current is not None and self.current.current is not None:
            self.current.current.append_continuation(line)


def _extract_pdf_lines(lines: tuple[str, ...], *, source_label: str) -> tuple[RecordDesignSheet, ...]:
    state = _PdfParseState(source_label=source_label)
    for row_number, raw_line in enumerate(lines, start=1):
        state.feed(_clean_pdf_line(raw_line), row_number)
    return state.finalise()


def _validate_pdf_sheet(sheet: RecordDesignSheet, *, source_label: str) -> None:
    if not sheet.fields:
        return
    first_field = sheet.fields[0]
    if first_field.offset != 1:
        raise RegistryValidationError(
            f"{source_label} {sheet.name!r} first field starts at position {first_field.offset}; expected 1"
        )
    for parsed_field in sheet.fields:
        if parsed_field.offset < 1:
            raise RegistryValidationError(
                f"{source_label} {sheet.name!r} field ordinal {parsed_field.ordinal} has invalid "
                f"position {parsed_field.offset}"
            )
        if parsed_field.length < 1:
            raise RegistryValidationError(
                f"{source_label} {sheet.name!r} field ordinal {parsed_field.ordinal} has invalid "
                f"length {parsed_field.length}"
            )
    terminal_position = max(parsed_field.offset + parsed_field.length - 1 for parsed_field in sheet.fields)
    if sheet.total_positions is not None and terminal_position != sheet.total_positions:
        raise RegistryValidationError(
            f"{source_label} {sheet.name!r} declares {sheet.total_positions} total positions "
            f"but parsed fields fill {terminal_position}"
        )


def _parse_pdf_row(line: str, source_row: int) -> _PdfRow | None:
    compact = _COMPACT_PDF_ROW_RE.match(line)
    if compact is not None:
        return _PdfRow(
            source_row=source_row,
            ordinal=int(compact.group("ordinal")),
            offset=int(compact.group("offset")),
            length=int(compact.group("length")),
            type_code=compact.group("type"),
            description=compact.group("text").strip(),
        )

    crlf = _COMPACT_PDF_CRLF_ROW_RE.match(line)
    if crlf is not None:
        return _PdfRow(
            source_row=source_row,
            ordinal=int(crlf.group("ordinal")),
            offset=int(crlf.group("offset")),
            length=2,
            type_code=crlf.group("type"),
            description=crlf.group("text").strip(),
        )

    narrative = _NARRATIVE_PDF_ROW_RE.match(line)
    if narrative is None:
        return None

    start = int(narrative.group("start"))
    end_group = narrative.group("end")
    end = int(end_group) if end_group is not None else start
    if end < start:
        raise RegistryValidationError(f"PDF row {source_row} has inverted position range {start}-{end}")
    return _PdfRow(
        source_row=source_row,
        ordinal=None,
        offset=start,
        length=end - start + 1,
        type_code=_normalise_pdf_type_code(narrative.group("type")),
        description=narrative.group("text").strip(),
    )


def _normalise_pdf_type_code(value: str) -> str:
    normalised = value.strip(" .").lower()
    if set(normalised) <= {"-", "\u2013"}:
        return "Blancos"
    if normalised.startswith("num"):
        return "Numérico"
    if normalised.startswith("alfanum"):
        return "Alfanumérico"
    if normalised.startswith("alfab"):
        return "Alfabético"
    return value.strip()


def _pdf_page_name(line: str) -> str | None:
    match = _PDF_PAGE_RECORD_RE.match(line)
    if match is None:
        return None
    return f"Pág. {match.group('page')}"


def _pdf_record_heading_name(line: str) -> str | None:
    match = _PDF_RECORD_HEADING_RE.match(line)
    if match is None:
        return None
    title = _normalise_pdf_sheet_name(match.group("title"))
    return f"Tipo {match.group('record')} - {title}"


def _is_pdf_header(line: str) -> bool:
    normalised = line.upper()
    return (
        ("POSICIONES" in normalised or "POSICIÓN" in normalised)
        and "NATURALEZA" in normalised
        and "DESCRIPCI" in normalised
    ) or ("Nº POSIC" in normalised and "LON" in normalised and "TIPO" in normalised and "DESCRIPCI" in normalised)


def _is_pdf_footer(line: str) -> bool:
    return bool(
        re.match(r"^P[áa]gina\s+\d+\s+de\s+\d+$", line, re.IGNORECASE)
        or re.match(r"^Ejercicio\s+\d{4}(?:\s+\d+)?$", line, re.IGNORECASE)
        or re.match(r"^\d+$", line)
    )


def _is_pdf_page_heading(line: str) -> bool:
    return bool(
        line.startswith("Modelo ")
        or line.startswith("Agencia Tributaria")
        or line.startswith("Declaración Informativa")
        or line.startswith("Declaración informativa")
        or line.startswith("determinados ")
        or line.startswith("determinadas ")
        or line == "Resumen anual"
        or line == "MODELO 193"
        or line == "MODELO 190"
        or line == "DISEÑOS DE REGISTRO"
    )


def _looks_like_title_continuation(line: str) -> bool:
    letters = [char for char in line if char.isalpha()]
    if not letters:
        return False
    return not any(char.islower() for char in letters)


def _clean_pdf_line(line: str) -> str:
    return " ".join(line.strip().split())


def _join_pdf_parts(parts: list[str]) -> str:
    return " ".join(part.strip() for part in parts if part.strip())


def _normalise_pdf_sheet_name(value: str) -> str:
    return _join_pdf_parts([value.replace(".", " ").strip()]).strip(". ").title()


def _extract_visual_record_design_chart(
    pages: tuple[_PdfPageSnapshot, ...],
    *,
    source_label: str,
) -> tuple[RecordDesignSheet, ...]:
    pages_by_sheet: dict[str, list[_PdfPageSnapshot]] = {}
    for page in pages:
        sheet_name = _visual_chart_page_sheet_name(page)
        if sheet_name is not None:
            pages_by_sheet.setdefault(sheet_name, []).append(page)
    if not pages_by_sheet:
        return ()

    sheets = tuple(
        _extract_visual_chart_sheet(sheet_name, tuple(sheet_pages), source_label=source_label)
        for sheet_name, sheet_pages in pages_by_sheet.items()
    )
    return sheets if all(sheet.fields for sheet in sheets) else ()


def _visual_chart_page_sheet_name(page: _PdfPageSnapshot) -> str | None:
    for line in page.lines:
        match = _VISUAL_CHART_HEADER_RE.match(_clean_pdf_line(line))
        if match is not None:
            title = _normalise_pdf_sheet_name(match.group("title"))
            return f"Tipo {match.group('record')} - {title}"
    return None


def _extract_visual_chart_sheet(
    name: str,
    pages: tuple[_PdfPageSnapshot, ...],
    *,
    source_label: str,
) -> RecordDesignSheet:
    fragments: list[_VisualChartFragment] = []
    for page in pages:
        fragments.extend(_extract_visual_chart_fragments(page))

    merged = _merge_visual_chart_fragments(sorted(fragments, key=lambda fragment: fragment.start))
    fields = tuple(
        RecordDesignField(
            sheet=name,
            row=ordinal,
            ordinal=ordinal,
            offset=fragment.start,
            length=fragment.end - fragment.start + 1,
            type_code=_VISUAL_CHART_TYPE_CODE,
            description=fragment.description,
            content="Extracted from visual record-design chart geometry.",
        )
        for ordinal, fragment in enumerate(merged, start=1)
    )
    total_positions = max((field.offset + field.length - 1 for field in fields), default=None)
    sheet = RecordDesignSheet(name=name, fields=fields, total_positions=total_positions)
    _validate_pdf_sheet(sheet, source_label=source_label)
    return sheet


def _extract_visual_chart_fragments(page: _PdfPageSnapshot) -> list[_VisualChartFragment]:
    grid = _visual_chart_grid(page)
    if grid is None:
        return []
    left, cell_width, horizontal_rules = grid
    fragments: list[_VisualChartFragment] = []
    number_rows = _visual_chart_number_rows(page)
    for index, (number_top, first_position) in enumerate(number_rows):
        row_rules = _visual_chart_rules_for_number_row(horizontal_rules, number_top)
        if not row_rules:
            continue
        region_top = number_rows[index - 1][0] + 8 if index else 20
        for rule in row_rules:
            start = first_position - 1 + round((rule.x0 - left) / cell_width) + 1
            end = first_position - 1 + round((rule.x1 - left) / cell_width)
            if start > end:
                continue
            fragments.append(
                _VisualChartFragment(
                    start=start,
                    end=end,
                    description=_visual_chart_description(page, rule, region_top=region_top),
                )
            )
    return fragments


def _visual_chart_grid(page: _PdfPageSnapshot) -> tuple[float, float, tuple[_PdfRect, ...]] | None:
    horizontal_rules = tuple(
        rect for rect in page.rects if rect.fill == 0.0 and rect.height <= 2.0 and rect.width >= 8.0
    )
    full_width_rules = tuple(rect for rect in horizontal_rules if rect.width > 700.0)
    if not full_width_rules:
        return None
    left = min(rect.x0 for rect in full_width_rules)
    right = max(rect.x1 for rect in full_width_rules)
    return left, (right - left) / 65, horizontal_rules


def _visual_chart_number_rows(page: _PdfPageSnapshot) -> list[tuple[float, int]]:
    grouped_words: dict[float, list[_PdfWord]] = {}
    for word in page.words:
        if _visual_chart_number_values(word.text):
            grouped_words.setdefault(round(word.top, 1), []).append(word)

    rows: list[tuple[float, int]] = []
    for top, words in grouped_words.items():
        values = [
            value
            for word in sorted(words, key=lambda current: current.x0)
            for value in _visual_chart_number_values(word.text)
        ]
        if len(values) >= 20 and max(values) - min(values) >= 30:
            rows.append((top, min(values)))
    return sorted(rows)


def _visual_chart_number_values(text: str) -> tuple[int, ...]:
    if not text.isdigit():
        return ()
    if len(text) <= 3:
        return (int(text),)
    if len(text) % 3 == 0:
        return tuple(int(text[index : index + 3]) for index in range(0, len(text), 3))
    return ()


def _visual_chart_rules_for_number_row(
    horizontal_rules: tuple[_PdfRect, ...],
    number_top: float,
) -> tuple[_PdfRect, ...]:
    grouped_rules: dict[float, list[_PdfRect]] = {}
    for rule in horizontal_rules:
        if 0 < number_top - rule.top <= 30:
            grouped_rules.setdefault(round(rule.top, 1), []).append(rule)
    if not grouped_rules:
        return ()
    rule_top = max(grouped_rules)
    return tuple(sorted(grouped_rules[rule_top], key=lambda rule: rule.x0))


def _visual_chart_description(
    page: _PdfPageSnapshot,
    rule: _PdfRect,
    *,
    region_top: float,
) -> str:
    words = [
        word
        for word in page.words
        if rule.x0 - 2 <= (word.x0 + word.x1) / 2 <= rule.x1 + 2
        and region_top <= word.top <= rule.top - 1
        and not _is_visual_chart_number_text(word.text)
    ]
    description = _normalise_visual_chart_description(words)
    return description or "BLANCOS."


def _normalise_visual_chart_description(words: list[_PdfWord]) -> str:
    tokens = [word.text for word in sorted(words, key=lambda word: (word.top, word.x0))]
    tokens = [
        _REVERSED_VISUAL_CHART_TOKENS.get(visual_word, visual_word) for visual_word in tokens if visual_word != "D"
    ]
    if not tokens:
        return ""
    if any(token.strip(".").upper() in _REVERSED_VISUAL_CHART_WORDS for token in tokens):
        tokens = [token[::-1] for token in reversed(tokens)]
    return _clean_visual_chart_description(_dedupe_visual_chart_tokens(tokens))


def _dedupe_visual_chart_tokens(tokens: list[str]) -> str:
    deduped: list[str] = []
    for token in tokens:
        if not deduped or deduped[-1] != token:
            deduped.append(token)
    return " ".join(deduped).strip()


def _clean_visual_chart_description(description: str) -> str:
    replacements = {
        "DEL DECLARANTE N.I.F. DEL DECLARANTE": "N.I.F. DEL DECLARANTE",
        "DECLARANTE N.I.F. DECLARANTE": "N.I.F. DECLARANTE",
        "PROVINCIA AICNIVORP OGIDOC": "CODIGO PROVINCIA",
        "CODIGO PAIS SIAP": "CODIGO PAIS",
        "DECIMAL ED": "DECIMAL",
        "I TIPO DE HOJA": "TIPO DE HOJA",
        "REFERENCIA CATASTRAL REFERENCIA CATASTRAL": "REFERENCIA CATASTRAL",
    }
    for before, after in replacements.items():
        description = description.replace(before, after)
    return description


def _is_visual_chart_number_text(text: str) -> bool:
    return bool(re.fullmatch(r"\d+", text) or re.fullmatch(r"\d{3}(?:\d{3})+", text))


def _merge_visual_chart_fragments(
    fragments: list[_VisualChartFragment],
) -> tuple[_VisualChartFragment, ...]:
    merged: list[_VisualChartFragment] = []
    for fragment in fragments:
        if (
            merged
            and fragment.start == merged[-1].end + 1
            and merged[-1].end % 65 == 0
            and _visual_chart_fragments_should_merge(merged[-1], fragment)
        ):
            previous = merged[-1]
            description = _merge_visual_chart_descriptions(previous.description, fragment.description)
            merged[-1] = _VisualChartFragment(start=previous.start, end=fragment.end, description=description)
            continue
        merged.append(fragment)
    return tuple(merged)


def _visual_chart_fragments_should_merge(
    previous: _VisualChartFragment,
    current: _VisualChartFragment,
) -> bool:
    return not (previous.description == "BLANCOS." and current.description != "BLANCOS.")


def _merge_visual_chart_descriptions(previous: str, current: str) -> str:
    parts = [description for description in (previous, current) if description != "BLANCOS."]
    return _clean_visual_chart_description(_join_pdf_parts(parts)) or "BLANCOS."


_VISUAL_CHART_HEADER_RE = re.compile(
    r"^MODELO\s+\d+\s+REGISTRO DE TIPO\s+(?P<record>\d+)\.?\s+(?P<title>REGISTRO DE .+)$",
    re.IGNORECASE,
)
_VISUAL_CHART_TYPE_CODE = "No consta en gráfico"
_REVERSED_VISUAL_CHART_WORDS = {
    "AICNIVORP",
    "AJOH",
    "DNERRA",
    "EVALC",
    "ETROPOS",
    "LACOL",
    "LAMICED",
    "NÓICAREPO",
    "OPIT",
    "ORTSIGER",
}
_REVERSED_VISUAL_CHART_TOKENS = {
    "AIRATNEMELPMOC.CED": "DEC. COMPLEMENTARIA",
    "AVITUTITSUS.CED": "DEC. SUSTITUTIVA",
    ".REPO": "OPER.",
    "ORUGES": "SEGURO",
    "ELBEUMNI": "INMUEBLE",
    ".CAUTIS": "SITUAC.",
    "ARELACSE": "ESCALERA",
    "IMNUEBLE": "INMUEBLE",
}


# ---------------------------------------------------------------------------
# Calculation-completeness manifest derivation and Diseño extraction
# (off-load-path)
# ---------------------------------------------------------------------------
#
# The derivations below run off the snapshot-build hot path; they are
# called only by manifest-authoring scripts, the off-load-path coverage
# report, and the drift re-verification test.
#
# - ``calculation_closure_identities`` enumerates a revision's
#   calculation closure keyed on each closure casilla's own registry
#   ``(segmento, number)`` identity. It is vocabulary-agnostic: it works
#   for Modelo 200's five-digit AEAT Diseño tags and equally for the
#   semantic-slug and short-ordinal casilla numbers the other
#   calculation-bearing modelos use.
#
# - ``derive_calculation_completeness_casillas`` derives the
#   *calculation-completeness manifest* casilla set from that closure:
#   the modelo's calculation surface keyed on the registry identity each
#   closure casilla declares. For a multi-segment modelo it optionally
#   verifies the derived record segments against the AEAT Diseño de
#   Registros. This is the set the load-blocking completeness gate
#   enforces.
#
# - ``derive_diseno_coverage_casillas`` extracts the *full* Diseño
#   casilla set — every five-digit casilla tag AEAT embeds in a field
#   description, accounting-statement data-entry fields included. It
#   parses the multi-megabyte Diseño corpus and is the input to the
#   off-load-path advisory coverage report that inventories form-level
#   data coverage; it is NOT a load-blocking gate.

_CASILLA_TAG_RE = re.compile(r"\[(\d{5})\]")
"""Matches the five-digit casilla tag AEAT embeds in Diseño field text.

The official AEAT Diseño de Registros workbooks annotate every casilla
field with its five-digit casilla number in square brackets within the
field description (e.g. ``Liquidación III - ... - Base imponible
[00552]``). This regex extracts those tags so a derivation can enumerate
the ``(segmento, number)`` casilla set.
"""


@dataclass(frozen=True)
class DerivedDisenoCasilla:
    """One ``(segmento, number)`` casilla derived from an AEAT Diseño workbook.

    ``segmento`` carries the AEAT record-segment code (the workbook sheet
    name) for multi-segment modelos and is ``None`` for single-segment
    modelos. ``number`` is the bare five-digit AEAT casilla number.
    """

    segmento: str | None
    number: str


def _selector_is_cross_modelo(
    selector: Mapping[str, object], modelo_id: str
) -> bool:
    """Return whether a binding / relation selector names a foreign modelo.

    A binding ``selector`` (or a relation's ``source_modelo``) is
    *cross-modelo* when it explicitly names a ``source_modelo`` that is
    not the modelo whose closure is being derived. A selector that omits
    ``source_modelo``, or sets it equal to ``modelo_id``, is a
    *within-modelo* selector: its ``source_casillas`` / ``source_output``
    name casillas on the modelo being derived (a ``previous_filing``
    self-binding or a ``previous_period`` self-relation), and those
    casillas belong in the modelo's own calculation closure.
    """

    source_modelo = selector.get("source_modelo")
    if source_modelo is None:
        return False
    return str(source_modelo) != modelo_id


def calculation_closure_numbers(
    revision: ModeloRevision, modelo_id: str
) -> frozenset[str]:
    """Return the bare casilla numbers in a revision's calculation closure.

    The *calculation closure* is the set of casillas the cross-connecting
    calculation engine traverses **within this modelo revision**:

    - every ``formula.target`` casilla;
    - every casilla referenced inside any ``formula.expression``, walked
      transitively via the runtime-graph ``expression_casilla_refs``
      walker;
    - every casilla that declares a ``formula`` (a computed endpoint) or
      a ``binding`` (a bound endpoint) — the engine-visible casillas;
    - every verification-expectation operand casilla
      (``computed_casillas`` and the ``reconciliation_totals`` targets);
    - every *within-modelo* binding ``source_casillas`` / ``source_output``
      selector casilla, and every *within-modelo*
      ``RelationDefinition.source_output``.

    A binding ``source_casillas`` / ``source_output`` selector — and a
    ``RelationDefinition.source_output`` — is excluded from this closure
    **only when it is genuinely cross-modelo**: when the selector
    explicitly names a ``source_modelo`` that differs from ``modelo_id``.
    A cross-modelo selector's ``source_casillas`` / ``source_output``
    name casillas on that *foreign* modelo, not on the modelo whose
    closure is being derived; the cross-modelo edge enters the current
    modelo through the *bound* casilla — the current-modelo casilla that
    declares the binding (or, for a relation, ``relation.target_binding``)
    — which is already counted above as a binding endpoint. Folding a
    foreign-modelo casilla number into this closure would make the
    completeness gate demand it from the wrong modelo's registry.

    A selector that omits ``source_modelo`` or sets it equal to
    ``modelo_id`` is a *within-modelo* selector: a ``previous_filing``
    self-binding or a ``previous_period`` self-relation names a casilla
    on the modelo being derived, so that casilla is a genuine closure
    member and is kept.

    References are reduced to bare casilla numbers: a reference token may
    be either a casilla ``id`` or a bare ``number``, and a declared
    casilla's ``id`` is mapped back to its ``number`` so the closure is
    expressed in the AEAT bare-number vocabulary the Diseño uses. A
    reference token that matches no declared casilla is kept verbatim so
    a calculation that names a casilla the registry never declared — the
    Modelo 200 defect class — still surfaces in the closure.
    """

    id_to_number = {casilla.id: casilla.number for casilla in revision.casillas}

    def _as_number(token: str) -> str:
        return id_to_number.get(token, token)

    closure: set[str] = set()
    for casilla in revision.casillas:
        if casilla.formula is not None or casilla.binding is not None:
            closure.add(casilla.number)
    for formula in revision.formulas:
        closure.add(_as_number(formula.target))
        for ref in expression_casilla_refs(formula.expression):
            closure.add(_as_number(ref))
    for expectation in revision.verification_expectations:
        for ref in expectation.computed_casillas:
            closure.add(_as_number(ref))
        for ref in expectation.reconciliation_totals.values():
            closure.add(_as_number(ref))
    for binding in revision.bindings:
        if _selector_is_cross_modelo(binding.selector, modelo_id):
            continue
        source_casillas = binding.selector.get("source_casillas")
        if isinstance(source_casillas, tuple):
            for token in source_casillas:
                if isinstance(token, str):
                    closure.add(_as_number(token))
        source_output = binding.selector.get("source_output")
        if isinstance(source_output, str):
            closure.add(_as_number(source_output))
    for relation in revision.relations:
        if relation.source_modelo == modelo_id:
            closure.add(_as_number(relation.source_output))
    return frozenset(closure)


def calculation_closure_identities(
    revision: ModeloRevision, modelo_id: str
) -> frozenset[tuple[str | None, str]]:
    """Return the ``(segmento, number)`` identities in a revision's calculation closure.

    Identity-preserving counterpart of :func:`calculation_closure_numbers`.
    Where the bare-number closure reduces every reference to its casilla
    ``number`` — which discards the record segment a multi-segment modelo
    needs — this function resolves each reference token to the *declared
    casilla* it names and keeps that casilla's full
    ``(segmento, number)`` identity.

    The closure spans the same surface (formula targets, transitive
    formula-expression refs, formula/binding endpoint casillas,
    verification-expectation operands, and within-modelo binding /
    relation source casillas; only genuinely cross-modelo selectors —
    those whose ``source_modelo`` differs from ``modelo_id`` — are
    excluded, see :func:`calculation_closure_numbers`). A reference token
    is resolved against both the casilla ``id`` index and the casilla
    ``number`` index:

    - a token that matches a casilla ``id`` resolves to that exact
      casilla's identity — this is how a multi-segment modelo's formulas,
      which reference casillas by the segment-carrying composite ``id``
      (e.g. ``DP200014:00562``), pin the closure to the correct record
      segment;
    - a token that matches a casilla ``number`` resolves to every
      casilla declared under that number (one for a single-segment
      modelo; possibly several for a multi-segment modelo that reuses the
      number across segments);
    - a token that resolves to no declared casilla is kept as a bare
      ``(None, token)`` identity so a calculation that names a casilla
      the registry never declared still surfaces in the closure.

    This is the identity vocabulary the calculation-completeness manifest
    is keyed on, and it is vocabulary-agnostic: it works for the
    five-digit AEAT Diseño tags of Modelo 200 and equally for the
    semantic-slug and short-ordinal casilla numbers the other
    calculation-bearing modelos use.
    """

    by_id = {casilla.id: casilla for casilla in revision.casillas}
    by_number: dict[str, list[CasillaDefinition]] = {}
    for casilla in revision.casillas:
        by_number.setdefault(casilla.number, []).append(casilla)

    identities: set[tuple[str | None, str]] = set()

    def _resolve(token: str) -> None:
        casilla = by_id.get(token)
        if casilla is not None:
            identities.add((casilla.segmento, casilla.number))
            return
        declared = by_number.get(token)
        if declared:
            for occurrence in declared:
                identities.add((occurrence.segmento, occurrence.number))
            return
        identities.add((None, token))

    for casilla in revision.casillas:
        if casilla.formula is not None or casilla.binding is not None:
            identities.add((casilla.segmento, casilla.number))
    for formula in revision.formulas:
        _resolve(formula.target)
        for ref in expression_casilla_refs(formula.expression):
            _resolve(ref)
    for expectation in revision.verification_expectations:
        for ref in expectation.computed_casillas:
            _resolve(ref)
        for ref in expectation.reconciliation_totals.values():
            _resolve(ref)
    for binding in revision.bindings:
        if _selector_is_cross_modelo(binding.selector, modelo_id):
            continue
        source_casillas = binding.selector.get("source_casillas")
        if isinstance(source_casillas, tuple):
            for token in source_casillas:
                if isinstance(token, str):
                    _resolve(token)
        source_output = binding.selector.get("source_output")
        if isinstance(source_output, str):
            _resolve(source_output)
    for relation in revision.relations:
        if relation.source_modelo == modelo_id:
            _resolve(relation.source_output)
    return frozenset(identities)


def derive_calculation_completeness_casillas(
    revision: ModeloRevision,
    modelo_id: str,
    *,
    multi_segment: bool,
    diseno_path: Path | None = None,
) -> tuple[DerivedDisenoCasilla, ...]:
    """Return the calculation-completeness manifest casilla set for a revision.

    Derives the modelo's *calculation closure*
    (:func:`calculation_closure_numbers`) and keys each closure casilla
    on its **own registry ``(segmento, number)`` identity**. The closure
    bounds the manifest to exactly the casillas the cross-connecting
    calculation engine traverses; the registry's own declared identity —
    not a five-digit AEAT Diseño tag — names each casilla.

    This derivation is *vocabulary-agnostic*. Only Modelo 200's registry
    casilla ``number``\\ s are genuine five-digit AEAT Diseño tags; the
    other calculation-bearing modelos identify casillas by semantic slug
    (``iva.cuota-devengada-total``) or short ordinal (``01``-``19``). The
    manifest is therefore derived from the modelo's calculation surface
    keyed on the registry identity each closure casilla declares, so a
    manifest can be authored for any calculation-bearing modelo
    regardless of its casilla vocabulary.

    For a ``multi_segment`` modelo the result is *segment-aware*. A
    multi-segment modelo reuses the same casilla number across distinct
    record segments and its formulas reference casillas by the
    segment-carrying composite ``id``, so the identity-preserving closure
    (:func:`calculation_closure_identities`) already pins each closure
    casilla to the exact record segment the calculation surface uses.
    When ``diseno_path`` is supplied each segment-scoped identity is
    additionally **verified against the AEAT Diseño de Registros**: the
    Diseño remains authoritative on which record segment carries a
    number, and a pinned ``(segmento, number)`` absent from the Diseño is
    a derivation error.

    For a single-segment modelo ``segmento`` is left unset and the
    closure casilla's registry ``number`` alone identifies it; no Diseño
    is required because a single-segment modelo's identity is unambiguous
    without one.

    A closure reference that resolves to no declared casilla is omitted
    from the derived set — the calculation-completeness gate then fires
    on the missing required casilla when it compares the manifest to the
    declared casillas, which is the gate fulfilling its mission. The
    drift / coverage tests surface such gaps explicitly.

    This is an off-load-path tool. When ``diseno_path`` is supplied it
    parses the multi-megabyte Diseño corpus and must never run on the
    snapshot-build path.
    """

    declared_identities = {
        (casilla.segmento, casilla.number) for casilla in revision.casillas
    }

    diseno_pairs: frozenset[tuple[str, str]] | None = None
    if diseno_path is not None:
        diseno_pairs = frozenset(
            (sheet.name, number)
            for sheet in extract_record_design(diseno_path)
            for number in _sheet_casilla_numbers(sheet)
        )

    ordered: list[DerivedDisenoCasilla] = []
    for segmento, number in sorted(
        calculation_closure_identities(revision, modelo_id),
        key=lambda item: (item[0] or "", item[1]),
    ):
        if (segmento, number) not in declared_identities:
            # The closure references a casilla the registry never
            # declares at this identity. It is omitted here; the
            # completeness gate fires on the omission instead.
            continue
        if not multi_segment:
            ordered.append(DerivedDisenoCasilla(segmento=None, number=number))
            continue
        if (
            diseno_pairs is not None
            and segmento is not None
            and (segmento, number) not in diseno_pairs
        ):
            raise RegistryValidationError(
                f"calculation-completeness derivation: casilla {number!r} is "
                f"declared under segmento {segmento!r} but the AEAT Diseño de "
                "Registros does not carry it under that segment"
            )
        ordered.append(DerivedDisenoCasilla(segmento=segmento, number=number))
    return tuple(ordered)


def derive_diseno_coverage_casillas(
    path: Path,
    *,
    multi_segment: bool,
) -> tuple[DerivedDisenoCasilla, ...]:
    """Return the full ``(segmento, number)`` casilla set declared by a Diseño.

    Runs read-only record-design extraction against the official AEAT
    Diseño de Registros source at ``path`` and collects *every*
    five-digit casilla tag embedded in the field descriptions — including
    the accounting-statement data-entry fields that feed no calculation.

    This is the input to the off-load-path advisory coverage report that
    inventories form-level data coverage. It is intentionally NOT a
    load-blocking gate: a modelo whose registry is not yet exhaustively
    backfilled against the full Diseño is reported as having a coverage
    gap, not failed at load. The load-blocking gate is keyed on the
    bounded calculation closure
    (:func:`derive_calculation_completeness_casillas`) instead.

    For a ``multi_segment`` modelo (e.g. Modelo 200, which reuses the
    same casilla number across distinct record segments) every casilla
    carries the workbook sheet name as its ``segmento``, so the same
    number under two segments yields two distinct identity pairs. For a
    single-segment modelo ``segmento`` is left unset and the bare number
    alone identifies the casilla; a number that recurs across sheets of a
    single-segment Diseño collapses to one identity, matching the
    bare-number registry behaviour.

    This is an off-load-path tool: it parses the multi-megabyte Diseño
    corpus and must never run on the snapshot-build path.
    """

    sheets = extract_record_design(path)
    if multi_segment:
        seen: set[tuple[str | None, str]] = set()
        ordered: list[DerivedDisenoCasilla] = []
        for sheet in sheets:
            for number in _sheet_casilla_numbers(sheet):
                identity = (sheet.name, number)
                if identity in seen:
                    continue
                seen.add(identity)
                ordered.append(DerivedDisenoCasilla(segmento=sheet.name, number=number))
        return tuple(ordered)
    seen_numbers: set[str] = set()
    bare: list[DerivedDisenoCasilla] = []
    for sheet in sheets:
        for number in _sheet_casilla_numbers(sheet):
            if number in seen_numbers:
                continue
            seen_numbers.add(number)
            bare.append(DerivedDisenoCasilla(segmento=None, number=number))
    return tuple(bare)


@dataclass(frozen=True)
class DisenoCoverageReport:
    """An off-load-path advisory inventory of one revision's Diseño coverage.

    Compares a modelo revision's declared casillas against the *full*
    AEAT Diseño de Registros casilla set — every five-digit casilla tag
    AEAT embeds in the form's field descriptions, accounting-statement
    data-entry fields included.

    This report is **advisory**: it is produced off the snapshot-build
    load path and never reds a load. A modelo whose registry is not yet
    exhaustively backfilled against the full Diseño is reported here as
    having a coverage gap, surfaced as information for follow-up
    authoring. The load-blocking gate is the bounded
    calculation-completeness gate, not this full-Diseño inventory — that
    is the ADR-amendment separation: calculation-completeness is enforced
    at load, full-Diseño coverage is inventoried off-load-path.

    Fields:

    - ``modelo_id`` / ``revision_id`` identify the revision inventoried.
    - ``diseno_casillas`` is the full ``(segmento, number)`` set the
      Diseño declares.
    - ``covered_casillas`` is the subset the registry also declares —
      the Diseño casillas the registry has backfilled.
    - ``coverage_gap_casillas`` is the subset the Diseño declares that
      the registry does not — the advisory follow-up inventory.
    """

    modelo_id: str
    revision_id: str
    diseno_casillas: tuple[DerivedDisenoCasilla, ...]
    covered_casillas: tuple[DerivedDisenoCasilla, ...]
    coverage_gap_casillas: tuple[DerivedDisenoCasilla, ...]

    @property
    def diseno_casilla_count(self) -> int:
        """Total ``(segmento, number)`` casillas the Diseño declares."""
        return len(self.diseno_casillas)

    @property
    def covered_count(self) -> int:
        """Diseño casillas the registry also declares."""
        return len(self.covered_casillas)

    @property
    def coverage_gap_count(self) -> int:
        """Diseño casillas the registry does not yet declare."""
        return len(self.coverage_gap_casillas)


def build_diseno_coverage_report(
    path: Path,
    modelo_id: str,
    revision: ModeloRevision,
    *,
    multi_segment: bool,
) -> DisenoCoverageReport:
    """Return the off-load-path full-Diseño coverage advisory report for a revision.

    Extracts the full AEAT Diseño de Registros casilla set
    (:func:`derive_diseno_coverage_casillas`) and compares it against the
    revision's declared casillas, keyed on the ``(segmento, number)``
    identity. The result is a :class:`DisenoCoverageReport` that
    inventories how much of the form's data surface the registry covers
    and which Diseño casillas remain to be authored.

    This is an **advisory** inventory, never a load gate. It is produced
    off the snapshot-build path — it parses the multi-megabyte Diseño
    corpus — and must never run on the load path. A coverage gap reported
    here does not fail any modelo: the load-blocking enforcement is the
    bounded calculation-completeness gate, per the ADR amendment that
    separates calculation-completeness (enforced at load) from
    full-Diseño coverage (inventoried off-load-path).

    For a ``multi_segment`` modelo the comparison is segment-aware: a
    Diseño casilla under segment ``S`` is "covered" only when the
    registry declares a casilla at the same ``(S, number)`` identity. For
    a single-segment modelo ``segmento`` is unset on both sides and the
    bare number alone identifies the casilla.
    """

    diseno = derive_diseno_coverage_casillas(path, multi_segment=multi_segment)
    declared_identities = {(casilla.segmento, casilla.number) for casilla in revision.casillas}
    covered: list[DerivedDisenoCasilla] = []
    gap: list[DerivedDisenoCasilla] = []
    for casilla in diseno:
        if (casilla.segmento, casilla.number) in declared_identities:
            covered.append(casilla)
        else:
            gap.append(casilla)
    return DisenoCoverageReport(
        modelo_id=modelo_id,
        revision_id=revision.id,
        diseno_casillas=diseno,
        covered_casillas=tuple(covered),
        coverage_gap_casillas=tuple(gap),
    )


def _sheet_casilla_numbers(sheet: RecordDesignSheet) -> tuple[str, ...]:
    """Return the casilla tags declared in one record-design sheet, in field order."""

    numbers: list[str] = []
    seen: set[str] = set()
    for design_field in sheet.fields:
        for text in (design_field.description, design_field.validation, design_field.content):
            if not text:
                continue
            for match in _CASILLA_TAG_RE.finditer(text):
                number = match.group(1)
                if number in seen:
                    continue
                seen.add(number)
                numbers.append(number)
    return tuple(numbers)


__all__ = [
    "DerivedDisenoCasilla",
    "DisenoCoverageReport",
    "RecordDesignField",
    "RecordDesignSheet",
    "build_diseno_coverage_report",
    "calculation_closure_identities",
    "calculation_closure_numbers",
    "derive_calculation_completeness_casillas",
    "derive_diseno_coverage_casillas",
    "extract_record_design",
    "extract_record_design_pdf",
    "extract_record_design_pdf_bytes",
    "extract_record_design_workbook",
]
