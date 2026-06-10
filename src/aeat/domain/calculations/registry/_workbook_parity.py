"""Official AEAT workbook parity discovery and verification backend.

Discovers, scans, and executes AEAT official workbooks against a
:class:`RegistrySnapshot` to verify that registry formulas match the
published AEAT calculation workbooks.
"""

from __future__ import annotations

import re
import shutil
import subprocess
import time
from collections.abc import Iterable, Iterator, Mapping
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Any
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.formula import Tokenizer
from openpyxl.formula.tokenizer import TokenizerError
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet

from ....core.config import Settings as _Settings
from ....core.decimal import coerce_decimal
from ....core.errors import CoreError
from ....core.external_constants import XLS_EXTENSION as _XLS_EXTENSION
from ....core.external_constants import XLSX_EXTENSION as _XLSX_EXTENSION
from ....core.hashing import hash_file as _hash_file
from ....core.logging import get_logger
from ._errors import RegistryValidationError
from ._formula_runtime import calculate_registry_snapshot
from ._schema import EvidenceTier, RegistrySnapshot
from ._workbook_parity_models import (
    SyntheticInputSet,
    SyntheticInputValue,
    WorkbookArtefactReport,
    WorkbookBackendVerificationReport,
    WorkbookCellRef,
    WorkbookConversionReport,
    WorkbookModeloCoverage,
    WorkbookParityComparison,
    WorkbookParityModel,
    WorkbookParityRunReport,
    WorkbookRunnerAvailability,
    _WorkbookExtension,
)
from ._workbook_parity_types import (
    ParityStatus,
    WorkbookConversionStatus,
    WorkbookKind,
    WorkbookRunnerEngine,
    WorkbookRunnerStatus,
    WorkbookScanStatus,
)

_PARITY_DEFAULTS = _Settings()

_log = get_logger(__name__)

__all__ = [
    "ParityStatus",
    "SyntheticInputSet",
    "SyntheticInputValue",
    "WorkbookArtefactReport",
    "WorkbookBackendVerificationReport",
    "WorkbookCellRef",
    "WorkbookConversionReport",
    "WorkbookConversionStatus",
    "WorkbookKind",
    "WorkbookModeloCoverage",
    "WorkbookParityComparison",
    "WorkbookParityModel",
    "WorkbookParityRunReport",
    "WorkbookRunnerAvailability",
    "WorkbookRunnerEngine",
    "WorkbookRunnerStatus",
    "WorkbookScanOptions",
    "WorkbookScanStatus",
    "assert_formula_workbook_runner_ready",
    "assert_workbook_scan_clean",
    "compare_registry_to_workbook",
    "convert_binary_xls_with_libreoffice",
    "converted_binary_xls_with_libreoffice",
    "detect_workbook_runner",
    "discover_workbooks",
    "inventory_workbook_coverage",
    "parse_workbook_cell_ref",
    "run_registry_workbook_parity",
    "run_workbook_with_excel_com",
    "run_workbook_with_libreoffice",
    "scan_workbook",
    "verify_workbook_backend",
]


# Runner-engine string constant — single source for every callsite.
_ENGINE_LIBREOFFICE: WorkbookRunnerEngine = "libreoffice-headless"

_WORKBOOK_SUFFIXES = {_XLSX_EXTENSION, _XLS_EXTENSION}
_MODELO_PATTERN = re.compile(r"(?:^|[\\/])modelo[_-](?P<modelo>\d{3})(?:[\\/]|$)", re.IGNORECASE)
_CELL_REF_PATTERN = re.compile(r"(?<![A-Z0-9_])(?:'[^']+'!)?\$?[A-Z]{1,3}\$?\d+(?![A-Z0-9_])")
_CELL_REF_VALUE_PATTERN = re.compile(r"^(?:(?P<sheet>'[^']+'|[^!]+)!)?(?P<coordinate>\$?[A-Z]{1,3}\$?\d+)$")
_LIBREOFFICE_EXECUTABLE_ENV = "AEAT_LIBREOFFICE_EXECUTABLE"
_BINARY_XLS_CONVERSION_BYTES_CACHE: dict[tuple[str, int, str], bytes] = {}


class _BinaryXlsConversionError(CoreError):
    """Failure raised after a valid LibreOffice runner starts XLS conversion."""


@dataclass(frozen=True)
class _BinaryXlsConversionContext:
    resolved_root: Path
    resolved_path: Path
    relative: str
    digest: str
    byte_count: int
    modelo: str | None


@dataclass(frozen=True)
class WorkbookScanOptions:
    """Controls for bounded workbook discovery."""

    per_file_timeout_seconds: float = _PARITY_DEFAULTS.aeat_workbook_parity_per_file_timeout_s
    max_formula_refs: int = 500


def discover_workbooks(root: Path) -> tuple[Path, ...]:
    """Return every official workbook artefact below ``root``."""
    resolved = root.resolve()
    if not resolved.exists():
        raise RegistryValidationError(f"workbook root does not exist: {root}")
    return tuple(sorted(p for p in resolved.rglob("*") if p.suffix.lower() in _WORKBOOK_SUFFIXES and p.is_file()))


def scan_workbook(path: Path, *, root: Path, options: WorkbookScanOptions | None = None) -> WorkbookArtefactReport:
    """Scan one workbook and classify formula coverage.

    Returns:
        A :class:`WorkbookArtefactReport` describing the workbook's formula coverage.
    """
    opts = options or WorkbookScanOptions()
    started = time.monotonic()
    resolved_root = root.resolve()
    resolved_path = path.resolve()
    if resolved_root not in resolved_path.parents and resolved_root != resolved_path:
        raise RegistryValidationError(f"workbook path escapes scan root: {path}")

    relative = resolved_path.relative_to(resolved_root).as_posix()
    digest, byte_count = _hash_file(resolved_path)
    suffix = resolved_path.suffix.lower()
    modelo = _infer_modelo(relative)

    if suffix == _XLS_EXTENSION:
        return _unsupported_binary_xls_report(
            relative=relative,
            modelo=modelo,
            byte_count=byte_count,
            digest=digest,
            started=started,
        )

    try:
        sheets, formulas, references = _scan_xlsx_contents(resolved_path, relative, opts, started)
    except TimeoutError as exc:
        return _failed_report(
            relative=relative,
            modelo=modelo,
            suffix=_XLSX_EXTENSION,
            byte_count=byte_count,
            digest=digest,
            status=WorkbookScanStatus.TIMEOUT,
            error=str(exc),
            started=started,
        )
    except (InvalidFileException, BadZipFile, OSError) as exc:
        _log.warning(
            "workbook parity scan failed for %s: %s",
            relative,
            exc,
            exc_info=True,
        )
        return _failed_report(
            relative=relative,
            modelo=modelo,
            suffix=_XLSX_EXTENSION,
            byte_count=byte_count,
            digest=digest,
            status=WorkbookScanStatus.FAILED,
            error=f"{type(exc).__name__}: {exc}",
            started=started,
        )
    except Exception as exc:
        _log.warning(
            "workbook parity scan unexpected error for %s: %s",
            relative,
            exc,
            exc_info=True,
        )
        raise RegistryValidationError(
            f"Unexpected error scanning workbook {relative}: {type(exc).__name__}: {exc}"
        ) from exc
    kind = _classify_xlsx(relative, formulas)
    evidence_tier, not_evidence_for = _evidence_for_workbook_kind(kind)
    return WorkbookArtefactReport(
        path=relative,
        modelo=modelo,
        extension=_XLSX_EXTENSION,
        bytes=byte_count,
        sha256=digest,
        sheets=tuple(sheets),
        formula_cells=len(formulas),
        input_candidates=tuple(_dedupe_cells(references)),
        output_candidates=tuple(formulas),
        workbook_kind=kind,
        evidence_tier=evidence_tier,
        not_evidence_for=not_evidence_for,
        scan_status=WorkbookScanStatus.SCANNED,
        elapsed_seconds=_elapsed_decimal(started),
    )


def _unsupported_binary_xls_report(
    *,
    relative: str,
    modelo: str | None,
    byte_count: int,
    digest: str,
    started: float,
) -> WorkbookArtefactReport:
    """Stable .xls short-circuit report — binary XLS requires conversion before scanning."""
    evidence_tier, not_evidence_for = _evidence_for_workbook_kind(WorkbookKind.UNSUPPORTED_BINARY_XLS)
    return WorkbookArtefactReport(
        path=relative,
        modelo=modelo,
        extension=_XLS_EXTENSION,
        bytes=byte_count,
        sha256=digest,
        workbook_kind=WorkbookKind.UNSUPPORTED_BINARY_XLS,
        evidence_tier=evidence_tier,
        not_evidence_for=not_evidence_for,
        scan_status=WorkbookScanStatus.UNSUPPORTED,
        formula_cells=0,
        error="binary XLS requires isolated conversion before workbook formula inspection",
        elapsed_seconds=_elapsed_decimal(started),
    )


def _scan_xlsx_contents(
    resolved_path: Path,
    relative: str,
    opts: WorkbookScanOptions,
    started: float,
) -> tuple[list[str], list[WorkbookCellRef], list[WorkbookCellRef]]:
    """Open the workbook in read-only mode and collect (sheet titles, formulas, references)."""
    workbook = load_workbook(resolved_path, data_only=False, read_only=True)
    sheets: list[str] = []
    formulas: list[WorkbookCellRef] = []
    references: list[WorkbookCellRef] = []
    try:
        for worksheet in workbook.worksheets:
            _raise_if_timed_out(started, opts.per_file_timeout_seconds, relative)
            sheets.append(worksheet.title)
            _scan_worksheet_cells(
                worksheet,
                relative=relative,
                opts=opts,
                started=started,
                formulas=formulas,
                references=references,
            )
    finally:
        workbook.close()
    return sheets, formulas, references


# ADAPTER-INTERNAL-ALIAS-RATIONALE-OPENPYXL-WORKSHEET: openpyxl ships incomplete
# stubs for Worksheet; the loose-typed alias keeps the scan helper callable.
def _scan_worksheet_cells(
    worksheet: Any,  # openpyxl Worksheet; loose-typed because openpyxl ships incomplete stubs
    *,
    relative: str,
    opts: WorkbookScanOptions,
    started: float,
    formulas: list[WorkbookCellRef],
    references: list[WorkbookCellRef],
) -> None:
    """Walk one worksheet's cells, appending formulas and bounded references in place."""
    for row in worksheet.iter_rows(values_only=False):
        _raise_if_timed_out(started, opts.per_file_timeout_seconds, relative)
        for cell in row:
            value = cell.value
            if not (isinstance(value, str) and value.startswith("=")):
                continue
            ref = WorkbookCellRef(sheet=worksheet.title, coordinate=cell.coordinate, formula=value)
            formulas.append(ref)
            if len(references) < opts.max_formula_refs:
                references.extend(_formula_references(worksheet.title, value, opts.max_formula_refs - len(references)))


def inventory_workbook_coverage(
    root: Path,
    *,
    options: WorkbookScanOptions | None = None,
    limit: int | None = None,
    previous_reports: Iterable[WorkbookArtefactReport] = (),
) -> tuple[WorkbookArtefactReport, ...]:
    """Scan official workbook artefacts and return :class:`WorkbookArtefactReport` coverage records."""
    paths = discover_workbooks(root)
    if limit is not None:
        paths = paths[:limit]
    previous_by_path = {report.path: report for report in previous_reports}
    reports: list[WorkbookArtefactReport] = []
    resolved_root = root.resolve()
    for path in paths:
        relative = path.resolve().relative_to(resolved_root).as_posix()
        previous = previous_by_path.get(relative)
        if previous is not None and previous.sha256 == _hash_file(path)[0]:
            reports.append(previous)
            continue
        reports.append(scan_workbook(path, root=root, options=options))
    return tuple(reports)


def detect_workbook_runner() -> WorkbookRunnerAvailability:
    """Resolve the local sanctioned spreadsheet recalculation runner.

    LibreOffice headless (or, on Windows, Excel COM) is required infrastructure
    for the workbook parity backend. This resolver does not attempt a graceful
    fallback: if no runner is locatable, it raises so the caller surfaces the
    missing dependency instead of silently downgrading evidence quality.

    Returns:
        A :class:`WorkbookRunnerAvailability` describing the detected runner.
    """
    from ....core.config import load_settings

    settings_configured = load_settings().aeat_libreoffice_executable
    if settings_configured is not None:
        runner = _resolve_libreoffice_runner(str(settings_configured))
        return WorkbookRunnerAvailability(
            status="available",
            engine=_ENGINE_LIBREOFFICE,
            executable=str(runner),
            detail=f"LibreOffice executable configured by {_LIBREOFFICE_EXECUTABLE_ENV}",
        )
    for executable in ("soffice", "libreoffice"):
        found = shutil.which(executable)
        if found:
            return WorkbookRunnerAvailability(
                status="available",
                engine=_ENGINE_LIBREOFFICE,
                executable=found,
                detail="LibreOffice executable found for local workbook recalculation",
            )
    excel_clsid = _detect_excel_com_clsid()
    if excel_clsid is not None:
        return WorkbookRunnerAvailability(
            status="available",
            engine="excel-com",
            executable=excel_clsid,
            detail="Excel COM automation is registered for local read-only workbook recalculation",
        )
    raise RegistryValidationError(
        "No LibreOffice/soffice executable or Excel COM automation found on this host. "
        f"Install LibreOffice or set {_LIBREOFFICE_EXECUTABLE_ENV} to a valid soffice executable."
    )


def run_workbook_with_libreoffice(
    workbook_path: Path,
    *,
    inputs: Mapping[WorkbookCellRef, Decimal | int | str | bool],
    outputs: Mapping[str, WorkbookCellRef],
    executable: str | None = None,
) -> Mapping[str, Decimal | int | str | bool | None]:
    """Run a local XLSX workbook with LibreOffice headless and return outputs."""
    runner = _resolve_libreoffice_runner(executable)
    resolved = workbook_path.resolve()
    if not resolved.is_file():
        raise RegistryValidationError(f"workbook does not exist: {workbook_path}")
    if resolved.suffix.lower() != _XLSX_EXTENSION:
        raise RegistryValidationError("LibreOffice runner currently accepts only XLSX workbooks")

    with TemporaryDirectory(prefix="aeat-workbook-") as tmp:
        tmp_path = Path(tmp)
        output_dir = tmp_path / "output"
        output_dir.mkdir()
        user_installation = (tmp_path / "lo-profile").resolve().as_uri()
        working_copy = tmp_path / resolved.name
        shutil.copy2(resolved, working_copy)
        workbook = load_workbook(working_copy)
        try:
            for cell, value in inputs.items():
                workbook[cell.sheet][cell.coordinate] = _excel_value(value)
            workbook.save(working_copy)
        finally:
            workbook.close()
        try:
            completed = subprocess.run(
                [
                    str(runner),
                    "--headless",
                    "--nologo",
                    "--nodefault",
                    "--nofirststartwizard",
                    f"-env:UserInstallation={user_installation}",
                    "--convert-to",
                    "xlsx",
                    "--outdir",
                    str(output_dir),
                    str(working_copy),
                ],
                check=True,
                capture_output=True,
                text=True,
                timeout=_PARITY_DEFAULTS.aeat_workbook_parity_recalc_timeout_s,
            )
        except subprocess.TimeoutExpired as exc:
            raise RegistryValidationError("LibreOffice workbook recalculation timed out") from exc
        except subprocess.CalledProcessError as exc:
            detail = "\n".join(part for part in (exc.stdout, exc.stderr) if part)
            raise RegistryValidationError(f"LibreOffice workbook recalculation failed: {detail}") from exc
        recalculated_path = output_dir / working_copy.name
        if not recalculated_path.is_file():
            detail = "\n".join(part for part in (completed.stdout, completed.stderr) if part)
            raise RegistryValidationError(f"LibreOffice did not produce recalculated workbook: {detail}")
        recalculated = load_workbook(recalculated_path, data_only=True, read_only=True)
        try:
            return {
                output_id: _coerce_excel_result(recalculated[cell.sheet][cell.coordinate].value)
                for output_id, cell in outputs.items()
            }
        finally:
            recalculated.close()


def convert_binary_xls_with_libreoffice(
    workbook_path: Path,
    *,
    root: Path,
    executable: str | None = None,
) -> WorkbookConversionReport:
    """Convert one official binary XLS in isolated storage and return a :class:`WorkbookConversionReport`."""
    started = time.monotonic()
    context = _binary_xls_conversion_context(workbook_path, root=root)
    runner = _resolve_libreoffice_runner(executable)
    try:
        with _converted_binary_xls_path(context, runner=runner) as converted_path:
            sheets, formulas, references = _inspect_converted_xlsx(
                converted_path,
                original_relative=context.relative,
                started=started,
            )
    except _BinaryXlsConversionError as exc:
        error = str(exc)
        if "timed out" in error:
            return _failed_conversion_report(
                relative=context.relative,
                modelo=context.modelo,
                byte_count=context.byte_count,
                digest=context.digest,
                error=error,
                started=started,
            )
        return _failed_conversion_report(
            relative=context.relative,
            modelo=context.modelo,
            byte_count=context.byte_count,
            digest=context.digest,
            error=error,
            started=started,
        )
    kind = _classify_xlsx(context.relative, formulas)
    evidence_tier, not_evidence_for = _evidence_for_workbook_kind(kind)
    return WorkbookConversionReport(
        path=context.relative,
        modelo=context.modelo,
        bytes=context.byte_count,
        sha256=context.digest,
        converted_extension=_XLSX_EXTENSION,
        sheets=sheets,
        formula_cells=len(formulas),
        input_candidates=tuple(_dedupe_cells(references)),
        output_candidates=formulas,
        workbook_kind=kind,
        evidence_tier=evidence_tier,
        not_evidence_for=not_evidence_for,
        conversion_status="converted",
        elapsed_seconds=_elapsed_decimal(started),
    )


@contextmanager
def converted_binary_xls_with_libreoffice(
    workbook_path: Path,
    *,
    root: Path,
    executable: str | None = None,
) -> Iterator[Path]:
    """Yield a temporary XLSX converted from official binary XLS input."""
    context = _binary_xls_conversion_context(workbook_path, root=root)
    runner = _resolve_libreoffice_runner(executable)
    try:
        with _converted_binary_xls_path(context, runner=runner) as converted_path:
            yield converted_path
    except _BinaryXlsConversionError as exc:
        raise RegistryValidationError(str(exc)) from exc


def _binary_xls_conversion_context(workbook_path: Path, *, root: Path) -> _BinaryXlsConversionContext:
    resolved_root = root.resolve()
    resolved_path = workbook_path.resolve()
    if resolved_root not in resolved_path.parents and resolved_root != resolved_path:
        raise RegistryValidationError(f"workbook path escapes conversion root: {workbook_path}")
    if resolved_path.suffix.lower() != _XLS_EXTENSION:
        raise RegistryValidationError("binary workbook conversion accepts only XLS artefacts")
    relative = resolved_path.relative_to(resolved_root).as_posix()
    digest, byte_count = _hash_file(resolved_path)
    return _BinaryXlsConversionContext(
        resolved_root=resolved_root,
        resolved_path=resolved_path,
        relative=relative,
        digest=digest,
        byte_count=byte_count,
        modelo=_infer_modelo(relative),
    )


@contextmanager
def _converted_binary_xls_path(
    context: _BinaryXlsConversionContext,
    *,
    runner: Path,
) -> Iterator[Path]:
    with TemporaryDirectory(prefix="aeat-xls-conversion-") as tmp:
        tmp_path = Path(tmp)
        output_dir = tmp_path / "output"
        output_dir.mkdir()
        cache_key = (context.digest, context.byte_count, str(runner))
        cached_bytes = _BINARY_XLS_CONVERSION_BYTES_CACHE.get(cache_key)
        if cached_bytes is not None:
            cached_path = output_dir / f"{context.resolved_path.stem}.xlsx"
            cached_path.write_bytes(cached_bytes)
            yield cached_path
            return
        user_installation = (tmp_path / "lo-profile").resolve().as_uri()
        try:
            completed = subprocess.run(
                [
                    str(runner),
                    "--headless",
                    "--nologo",
                    "--nodefault",
                    "--nofirststartwizard",
                    f"-env:UserInstallation={user_installation}",
                    "--convert-to",
                    "xlsx",
                    "--outdir",
                    str(output_dir),
                    str(context.resolved_path),
                ],
                check=True,
                capture_output=True,
                text=True,
                timeout=_PARITY_DEFAULTS.aeat_workbook_parity_libreoffice_timeout_s,
            )
        except subprocess.TimeoutExpired as exc:
            raise _BinaryXlsConversionError("LibreOffice binary XLS conversion timed out") from exc
        except subprocess.CalledProcessError as exc:
            detail = "\n".join(part for part in (exc.stdout, exc.stderr) if part)
            raise _BinaryXlsConversionError(f"LibreOffice binary XLS conversion failed: {detail}") from exc
        outputs = tuple(output_dir.glob("*.xlsx"))
        if len(outputs) != 1:
            detail = "\n".join(part for part in (completed.stdout, completed.stderr) if part)
            raise _BinaryXlsConversionError(f"LibreOffice did not produce exactly one XLSX workbook: {detail}")
        converted_path = outputs[0]
        _BINARY_XLS_CONVERSION_BYTES_CACHE[cache_key] = converted_path.read_bytes()
        yield converted_path


def run_registry_workbook_parity(
    *,
    snapshot: RegistrySnapshot,
    synthetic_input: SyntheticInputSet,
    workbook_path: Path,
    workbook: WorkbookArtefactReport,
    output_cells: Mapping[str, WorkbookCellRef],
    registry_outputs: Mapping[str, str],
    date_context: Mapping[str, date],
    relation_values: Mapping[str, Decimal] | None = None,
    tolerance: Decimal = Decimal("0"),
    executable: str | None = None,
) -> WorkbookParityRunReport:
    """Execute one registry-vs-workbook parity comparison and return a :class:`WorkbookParityRunReport`.

    Args:
        snapshot: The :class:`RegistrySnapshot` providing the registry formulas to compare.
        synthetic_input: :class:`SyntheticInputSet` whose values seed both
            sides of the comparison (operator inputs and registry bindings).
        workbook_path: Path to the AEAT calculation workbook to execute.
        workbook: :class:`WorkbookArtefactReport` describing the workbook
            artefact; must report kind ``FORMULA_FORM``.
        output_cells: Mapping of registry output id to the workbook
            :class:`WorkbookCellRef` carrying its computed value.
        registry_outputs: Mapping of workbook output ids to registry output
            ids, used to align both sides of the comparison.
        date_context: Date-axis context forwarded to the registry formula
            runtime.
        relation_values: Optional resolved registry relation values seeded
            into the registry runtime.
        tolerance: Absolute Decimal tolerance accepted between workbook and
            registry outputs; defaults to ``Decimal("0")``.
        executable: Optional LibreOffice executable override used when
            converting binary XLS workbooks.
    """
    if workbook.workbook_kind != WorkbookKind.FORMULA_FORM:
        raise RegistryValidationError(
            f"workbook {workbook.path!r} is {workbook.workbook_kind!r}, not an executable calculation oracle"
        )
    workbook_inputs: dict[WorkbookCellRef, Decimal | int | str | bool] = {}
    registry_inputs: dict[str, Decimal] = {}
    registry_binding_values: dict[str, Decimal] = {}
    casilla_ids = {casilla.id for casilla in snapshot.revision.casillas}
    binding_ids = {binding.id for binding in snapshot.revision.bindings}
    for value in synthetic_input.values:
        if value.workbook_cell is not None:
            workbook_inputs[value.workbook_cell] = value.value
        if value.registry_binding is None:
            continue
        registry_value = _registry_decimal_value(value.id, value.value)
        if value.registry_binding in casilla_ids:
            registry_inputs[value.registry_binding] = registry_value
        elif value.registry_binding in binding_ids:
            registry_binding_values[value.registry_binding] = registry_value
        else:
            raise RegistryValidationError(
                f"synthetic input {value.id!r} references unknown registry target {value.registry_binding!r}"
            )

    workbook_values = run_workbook_with_libreoffice(
        workbook_path,
        inputs=workbook_inputs,
        outputs=output_cells,
        executable=executable,
    )
    registry_result = calculate_registry_snapshot(
        snapshot,
        inputs=registry_inputs,
        date_context=date_context,
        binding_values=registry_binding_values,
        relation_values=relation_values,
    )
    missing_outputs = sorted(set(registry_outputs.values()).difference(registry_result.values))
    if missing_outputs:
        raise RegistryValidationError(f"registry parity outputs are missing calculated casillas: {missing_outputs!r}")
    registry_values = {
        output_id: registry_result.values[casilla_id] for output_id, casilla_id in registry_outputs.items()
    }
    formulas_by_target = {formula.target: formula for formula in snapshot.revision.formulas}
    legal_refs: dict[str, tuple[str, ...]] = {}
    source_refs: dict[str, tuple[str, ...]] = {}
    for output_id, casilla_id in registry_outputs.items():
        formula = formulas_by_target.get(casilla_id)
        if formula is not None:
            legal_refs[output_id] = tuple(formula.legal_refs)
            source_refs[output_id] = tuple(formula.source_refs)
    runner = _execution_runner_availability(executable)
    return compare_registry_to_workbook(
        synthetic_input=synthetic_input,
        workbook=workbook,
        runner=runner,
        expected_workbook_values=workbook_values,
        actual_registry_values=registry_values,
        output_cells=output_cells,
        registry_snapshot_id=f"{snapshot.modelo.id}:{snapshot.revision.id}",
        legal_refs=legal_refs,
        source_refs=source_refs,
        tolerance=tolerance,
    )


def parse_workbook_cell_ref(value: str, *, default_sheet: str | None = None) -> WorkbookCellRef:
    """Parse a workbook cell reference from registry configuration.

    Returns:
        The parsed :class:`WorkbookCellRef` with sheet and coordinate fields.
    """
    match = _CELL_REF_VALUE_PATTERN.match(value)
    if not match:
        raise RegistryValidationError(f"invalid workbook cell reference {value!r}")
    raw_sheet = match.group("sheet")
    if raw_sheet is None:
        if default_sheet is None:
            raise RegistryValidationError(f"workbook cell reference {value!r} must include a sheet")
        sheet = default_sheet
    else:
        sheet = raw_sheet.strip("'")
    return WorkbookCellRef(sheet=sheet, coordinate=match.group("coordinate").replace("$", ""))


def _resolve_libreoffice_runner(executable: str | None) -> Path:
    """Locate a LibreOffice executable, raising explicitly when none is available."""
    if executable is None:
        from ....core.config import load_settings

        configured = load_settings().aeat_libreoffice_executable
        found = str(configured) if configured is not None else (shutil.which("soffice") or shutil.which("libreoffice"))
        if not found:
            raise RegistryValidationError(
                "LibreOffice or soffice executable is not available on PATH. "
                f"Install LibreOffice or set {_LIBREOFFICE_EXECUTABLE_ENV}."
            )
        return Path(found).resolve()
    candidate = Path(executable).resolve()
    if not candidate.is_file():
        raise RegistryValidationError(f"LibreOffice executable does not exist: {executable}")
    if candidate.name.lower() not in {"soffice", "soffice.exe", "libreoffice", "libreoffice.exe"}:
        raise RegistryValidationError(f"unsupported LibreOffice executable name: {candidate.name}")
    return candidate


def _execution_runner_availability(executable: str | None) -> WorkbookRunnerAvailability:
    if executable is None:
        return detect_workbook_runner()
    runner = _resolve_libreoffice_runner(executable)
    return WorkbookRunnerAvailability(
        status="available",
        engine=_ENGINE_LIBREOFFICE,
        executable=str(runner),
        detail="LibreOffice executable provided for this workbook parity run",
    )


def run_workbook_with_excel_com(
    workbook_path: Path,
    *,
    inputs: Mapping[WorkbookCellRef, Decimal | int | str | bool],
    outputs: Mapping[str, WorkbookCellRef],
) -> Mapping[str, Decimal | int | str | bool | None]:
    """Run a local XLSX workbook with Excel COM and return selected output values.

    The workbook is opened read-only, link updates are disabled, alerts are
    disabled, and it is closed with ``SaveChanges=False``.
    """
    if _detect_excel_com_clsid() is None:
        raise RegistryValidationError("Excel COM automation is not registered")
    resolved = workbook_path.resolve()
    if not resolved.is_file():
        raise RegistryValidationError(f"workbook does not exist: {workbook_path}")
    if resolved.suffix.lower() != _XLSX_EXTENSION:
        raise RegistryValidationError("Excel COM runner currently accepts only XLSX workbooks")

    import pythoncom
    import win32com.client

    pythoncom.CoInitialize()
    excel = win32com.client.DispatchEx("Excel.Application")
    workbook = None
    try:
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.AskToUpdateLinks = False
        workbook = excel.Workbooks.Open(str(resolved), UpdateLinks=0, ReadOnly=True)
        for cell, value in inputs.items():
            workbook.Worksheets(cell.sheet).Range(cell.coordinate).Value = _excel_value(value)
        excel.CalculateFullRebuild()
        result: dict[str, Decimal | int | str | bool | None] = {}
        for output_id, cell in outputs.items():
            result[output_id] = _coerce_excel_result(workbook.Worksheets(cell.sheet).Range(cell.coordinate).Value)
        return result
    finally:
        if workbook is not None:
            workbook.Close(SaveChanges=False)
        excel.Quit()
        pythoncom.CoUninitialize()


def compare_registry_to_workbook(
    *,
    synthetic_input: SyntheticInputSet,
    workbook: WorkbookArtefactReport,
    runner: WorkbookRunnerAvailability,
    expected_workbook_values: Mapping[str, Decimal | int | str | bool | None],
    actual_registry_values: Mapping[str, Decimal | int | str | bool | None],
    output_cells: Mapping[str, WorkbookCellRef],
    registry_snapshot_id: str | None = None,
    legal_refs: Mapping[str, tuple[str, ...]] | None = None,
    source_refs: Mapping[str, tuple[str, ...]] | None = None,
    tolerance: Decimal = Decimal("0"),
) -> WorkbookParityRunReport:
    """Build a deterministic parity comparison report from already-computed values.

    Returns:
        A :class:`WorkbookParityRunReport` comparing registry output to workbook cells.
    """
    comparisons: list[WorkbookParityComparison] = []
    for output_id in sorted(set(expected_workbook_values) | set(actual_registry_values)):
        expected = expected_workbook_values.get(output_id)
        actual = actual_registry_values.get(output_id)
        cell = output_cells.get(output_id)
        if cell is None:
            raise RegistryValidationError(f"missing workbook output cell for {output_id!r}")
        status = _comparison_status(expected, actual, tolerance)
        comparisons.append(
            WorkbookParityComparison(
                output_id=output_id,
                workbook_cell=cell,
                expected_workbook_value=expected,
                actual_registry_value=actual,
                status=status,
                tolerance=tolerance,
                legal_refs=(legal_refs or {}).get(output_id, ()),
                source_refs=(source_refs or {}).get(output_id, ()),
                detail=None if status == "match" else "registry output differs from workbook output",
            )
        )
    run_status: ParityStatus = "match" if all(c.status == "match" for c in comparisons) else "mismatch"
    return WorkbookParityRunReport(
        synthetic_input_id=synthetic_input.id,
        registry_snapshot_id=registry_snapshot_id,
        workbook=workbook,
        runner=runner,
        comparisons=tuple(comparisons),
        status=run_status,
    )


def verify_workbook_backend(
    root: Path,
    *,
    scan_limit: int | None = None,
    per_file_timeout_seconds: float = 10.0,
    previous_report: WorkbookBackendVerificationReport | None = None,
    fail_on_scan_error: bool = True,
    require_formula_runner: bool = False,
) -> WorkbookBackendVerificationReport:
    """Verify the workbook parity backend and return a :class:`WorkbookBackendVerificationReport`."""
    reports = inventory_workbook_coverage(
        root,
        options=WorkbookScanOptions(per_file_timeout_seconds=per_file_timeout_seconds),
        limit=scan_limit,
        previous_reports=previous_report.reports if previous_report is not None else (),
    )
    runner = detect_workbook_runner()
    failed_statuses = {WorkbookScanStatus.FAILED, WorkbookScanStatus.TIMEOUT}
    report = WorkbookBackendVerificationReport(
        root=root.resolve().as_posix(),
        workbook_count=len(discover_workbooks(root)) if root.exists() else 0,
        scanned_count=sum(1 for report in reports if report.scan_status == WorkbookScanStatus.SCANNED),
        formula_workbook_count=sum(1 for report in reports if report.workbook_kind == WorkbookKind.FORMULA_FORM),
        unsupported_xls_count=sum(
            1 for report in reports if report.workbook_kind == WorkbookKind.UNSUPPORTED_BINARY_XLS
        ),
        failed_count=sum(1 for report in reports if report.scan_status in failed_statuses),
        runner=runner,
        reports=reports,
        modelo_coverage=_build_modelo_coverage(reports),
    )
    if fail_on_scan_error:
        assert_workbook_scan_clean(report)
    if require_formula_runner:
        assert_formula_workbook_runner_ready(report)
    return report


def assert_workbook_scan_clean(report: WorkbookBackendVerificationReport) -> None:
    """Raise when discovery could not inspect every workbook artefact."""
    failed_statuses = {WorkbookScanStatus.FAILED, WorkbookScanStatus.TIMEOUT}
    failed = tuple(item for item in report.reports if item.scan_status in failed_statuses)
    if failed:
        details = "\n".join(f" - {item.path}: {item.error}" for item in failed)
        raise RegistryValidationError(f"workbook verification failed to scan {len(failed)} artefact(s):\n{details}")


def assert_formula_workbook_runner_ready(report: WorkbookBackendVerificationReport) -> None:
    """Sanity gate: confirm the verification report carries an available runner.

    `detect_workbook_runner()` raises explicitly when no LibreOffice or Excel COM
    runner can be located, so by the time a `WorkbookBackendVerificationReport`
    exists its `runner.status` must already be `"available"`. This helper is
    retained as a documentation surface and a defensive guard against future
    schema drift; it never fails on a freshly produced report.
    """
    # `WorkbookRunnerStatus` is a single-value literal, so the only way this can
    # fail is if a caller hand-builds a report with a hand-mutated runner; that
    # is intentionally left as an explicit error path rather than silent.
    if report.runner.status != "available":  # pragma: no cover — defensive
        raise RegistryValidationError(
            "formula-bearing official workbooks require a local recalculation runner; "
            f"runner status is {report.runner.status!r}: {report.runner.detail}"
        )


def _build_modelo_coverage(reports: Iterable[WorkbookArtefactReport]) -> tuple[WorkbookModeloCoverage, ...]:
    buckets: dict[str, list[WorkbookArtefactReport]] = {}
    for report in reports:
        modelo = report.modelo or "unknown"
        buckets.setdefault(modelo, []).append(report)
    return tuple(
        WorkbookModeloCoverage(
            modelo=modelo,
            workbook_count=len(modelo_reports),
            formula_workbook_count=sum(
                1 for report in modelo_reports if report.workbook_kind == WorkbookKind.FORMULA_FORM
            ),
            unsupported_xls_count=sum(
                1 for report in modelo_reports if report.workbook_kind == WorkbookKind.UNSUPPORTED_BINARY_XLS
            ),
            failed_count=sum(
                1
                for report in modelo_reports
                if report.scan_status in {WorkbookScanStatus.FAILED, WorkbookScanStatus.TIMEOUT}
            ),
        )
        for modelo, modelo_reports in sorted(buckets.items())
    )


def _detect_excel_com_clsid() -> str | None:
    try:
        import winreg

        with winreg.OpenKey(winreg.HKEY_CLASSES_ROOT, r"Excel.Application\CLSID") as key:
            value, _kind = winreg.QueryValueEx(key, "")
        return str(value)
    except (FileNotFoundError, OSError, ImportError):
        return None


def _infer_modelo(relative_path: str) -> str | None:
    match = _MODELO_PATTERN.search(relative_path)
    return match.group("modelo") if match else None


def _raise_if_timed_out(started: float, timeout_seconds: float, relative: str) -> None:
    if time.monotonic() - started > timeout_seconds:
        raise TimeoutError(f"workbook scan timed out for {relative!r} after {timeout_seconds:.1f}s")


def _elapsed_decimal(started: float) -> Decimal:
    elapsed = coerce_decimal(round(time.monotonic() - started, 6), default=Decimal("0"))
    return elapsed if elapsed is not None else Decimal("0")


def _classify_xlsx(relative: str, formulas: Iterable[WorkbookCellRef]) -> WorkbookKind:
    formula_count = sum(1 for _ in formulas)
    lowered = relative.lower()
    if "valid" in lowered or "valida" in lowered:
        return WorkbookKind.VALIDATION_HINTS
    if _is_record_design_path(lowered):
        return WorkbookKind.RECORD_DESIGN_LAYOUT
    if formula_count > 0:
        return WorkbookKind.FORMULA_FORM
    return WorkbookKind.STATIC_LAYOUT


def _is_record_design_path(lowered_relative_path: str) -> bool:
    return any(marker in lowered_relative_path for marker in ("disenos_registro", "diseños_registro", "registro"))


def _evidence_for_workbook_kind(kind: WorkbookKind) -> tuple[EvidenceTier | None, tuple[EvidenceTier, ...]]:
    if kind == WorkbookKind.FORMULA_FORM:
        return "executable_parity_evidence", ("legal_authority", "layout_authority")
    if kind in {WorkbookKind.RECORD_DESIGN_LAYOUT, WorkbookKind.UNSUPPORTED_BINARY_XLS}:
        return "layout_authority", ("legal_authority", "executable_parity_evidence")
    if kind in {WorkbookKind.VALIDATION_HINTS, WorkbookKind.STATIC_LAYOUT}:
        return "official_source_guidance", ("legal_authority", "executable_parity_evidence")
    return None, (
        "legal_authority",
        "official_source_guidance",
        "executable_parity_evidence",
        "layout_authority",
    )


def _formula_references(sheet: str, formula: str, remaining: int) -> tuple[WorkbookCellRef, ...]:
    if remaining <= 0:
        return ()
    refs: list[WorkbookCellRef] = []
    try:
        tokens = Tokenizer(formula).items
        token_values = (token.value for token in tokens)
    except TokenizerError as exc:
        _log.debug(
            "workbook parity: openpyxl Tokenizer failed on formula %r; falling back to regex (%s)",
            formula[:80],
            exc,
            exc_info=True,
        )
        token_values = (match.group(0) for match in _CELL_REF_PATTERN.finditer(formula))
    for value in token_values:
        for match in _CELL_REF_PATTERN.finditer(value):
            if len(refs) >= remaining:
                return tuple(refs)
            ref_sheet = sheet
            coordinate = match.group(0).replace("$", "")
            if "!" in coordinate:
                raw_sheet, coordinate = coordinate.rsplit("!", 1)
                ref_sheet = raw_sheet.strip("'")
            refs.append(WorkbookCellRef(sheet=ref_sheet, coordinate=coordinate))
    return tuple(refs)


def _dedupe_cells(cells: Iterable[WorkbookCellRef]) -> tuple[WorkbookCellRef, ...]:
    seen: set[tuple[str, str]] = set()
    deduped: list[WorkbookCellRef] = []
    for cell in cells:
        key = (cell.sheet, cell.coordinate)
        if key in seen:
            continue
        seen.add(key)
        deduped.append(cell)
    return tuple(deduped)


def _failed_report(
    *,
    relative: str,
    modelo: str | None,
    suffix: _WorkbookExtension,
    byte_count: int,
    digest: str,
    status: WorkbookScanStatus,
    error: str,
    started: float,
) -> WorkbookArtefactReport:
    return WorkbookArtefactReport(
        path=relative,
        modelo=modelo,
        extension=suffix,
        bytes=byte_count,
        sha256=digest,
        workbook_kind=WorkbookKind.UNREADABLE,
        formula_cells=0,
        evidence_tier=None,
        not_evidence_for=(
            "legal_authority",
            "official_source_guidance",
            "executable_parity_evidence",
            "layout_authority",
        ),
        scan_status=status,
        error=error,
        elapsed_seconds=_elapsed_decimal(started),
    )


def _failed_conversion_report(
    *,
    relative: str,
    modelo: str | None,
    byte_count: int,
    digest: str,
    error: str,
    started: float,
) -> WorkbookConversionReport:
    return WorkbookConversionReport(
        path=relative,
        modelo=modelo,
        bytes=byte_count,
        sha256=digest,
        workbook_kind=WorkbookKind.UNREADABLE,
        formula_cells=0,
        evidence_tier=None,
        not_evidence_for=(
            "legal_authority",
            "official_source_guidance",
            "executable_parity_evidence",
            "layout_authority",
        ),
        conversion_status="failed",
        error=error,
        elapsed_seconds=_elapsed_decimal(started),
    )


_REFERENCE_HARVEST_LIMIT = 500
_INSPECT_CONVERTED_XLSX_TIMEOUT_S = 120


def _inspect_converted_xlsx(
    path: Path,
    *,
    original_relative: str,
    started: float,
) -> tuple[tuple[str, ...], tuple[WorkbookCellRef, ...], tuple[WorkbookCellRef, ...]]:
    workbook = load_workbook(path, data_only=False, read_only=True)
    try:
        sheets: list[str] = []
        formulas: list[WorkbookCellRef] = []
        references: list[WorkbookCellRef] = []
        for worksheet in workbook.worksheets:
            _raise_if_timed_out(started, _INSPECT_CONVERTED_XLSX_TIMEOUT_S, original_relative)
            sheets.append(worksheet.title)
            _collect_sheet_formulas(
                worksheet,
                formulas=formulas,
                references=references,
                original_relative=original_relative,
                started=started,
            )
        return tuple(sheets), tuple(formulas), tuple(references)
    finally:
        workbook.close()


def _collect_sheet_formulas(
    worksheet: Worksheet,
    *,
    formulas: list[WorkbookCellRef],
    references: list[WorkbookCellRef],
    original_relative: str,
    started: float,
) -> None:
    """Walk every row in ``worksheet`` and append formula refs + a bounded set of references."""
    for row in worksheet.iter_rows(values_only=False):
        _raise_if_timed_out(started, _INSPECT_CONVERTED_XLSX_TIMEOUT_S, original_relative)
        for cell in row:
            _record_cell_if_formula(
                cell,
                sheet_title=worksheet.title,
                formulas=formulas,
                references=references,
            )


def _record_cell_if_formula(
    cell: Cell | MergedCell,
    *,
    sheet_title: str,
    formulas: list[WorkbookCellRef],
    references: list[WorkbookCellRef],
) -> None:
    """Append a formula record + (bounded) reference fan-out when ``cell`` carries an ``=…`` value."""
    value = cell.value
    if not (isinstance(value, str) and value.startswith("=")):
        return
    formulas.append(WorkbookCellRef(sheet=sheet_title, coordinate=cell.coordinate, formula=value))
    remaining = _REFERENCE_HARVEST_LIMIT - len(references)
    if remaining > 0:
        references.extend(_formula_references(sheet_title, value, remaining))


def _comparison_status(
    expected: Decimal | int | str | bool | None,
    actual: Decimal | int | str | bool | None,
    tolerance: Decimal,
) -> ParityStatus:
    if expected is None or actual is None:
        return "match" if expected is actual else "mismatch"
    if isinstance(expected, Decimal | int) and isinstance(actual, Decimal | int):
        return "match" if abs(Decimal(expected) - Decimal(actual)) <= tolerance else "mismatch"
    return "match" if expected == actual else "mismatch"


def _excel_value(value: Decimal | int | str | bool) -> str | int | bool:
    if isinstance(value, Decimal):
        return str(value)
    return value


def _coerce_excel_result(value: object) -> Decimal | int | str | bool | None:
    if value is None or isinstance(value, str | bool):
        return value
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        coerced = coerce_decimal(value)
        return coerced if coerced is not None else str(value)
    return str(value)


def _registry_decimal_value(input_id: str, value: Decimal | int | str | bool) -> Decimal:
    if isinstance(value, bool):
        raise RegistryValidationError(f"synthetic input {input_id!r} cannot feed boolean into registry calculation")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, int):
        return Decimal(value)
    try:
        return Decimal(value)
    except (ArithmeticError, ValueError, TypeError) as exc:
        raise RegistryValidationError(
            f"synthetic input {input_id!r} cannot feed non-decimal value into registry calculation"
        ) from exc
