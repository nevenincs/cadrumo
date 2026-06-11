"""Offline XLSX materializer for ``SheetExportPlan`` records."""

from __future__ import annotations

import hashlib
import json
from collections.abc import Iterable, Sequence
from decimal import Decimal
from io import BytesIO
from typing import TYPE_CHECKING, Literal

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties
from pydantic import BaseModel, Field

from ....core import STRICT_FROZEN_CONFIG as _STRICT_FROZEN
from ....core.external_constants import UTF_8_ENCODING
from ._records import (
    SheetAutoFilter,
    SheetColumnWidth,
    SheetEvidenceContributorRow,
    SheetEvidenceFacet,
    SheetEvidenceManualEntry,
    SheetExportMetadata,
    SheetExportPlan,
    SheetFormulaCell,
    SheetFrozenView,
    SheetRowSet,
    SheetStyledRange,
    SheetValueCell,
    TabName,
)
from ._theme import ROLE_STYLES, WORKBOOK_FONT_FAMILY, openpyxl_argb

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet


_EVIDENCE_SIDECAR_SCHEMA_VERSION: Literal["calc-sheets-evidence-sidecar/v1"] = "calc-sheets-evidence-sidecar/v1"
_XLSX_MEDIA_TYPE: Literal["application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"] = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
_JSON_MEDIA_TYPE: Literal["application/json"] = "application/json"

_EVIDENCE_HEADERS: tuple[str, ...] = (
    "Tipo",
    "Casilla",
    "Transaction ID",
    "Amount",
    "Currency",
    "Taxable base",
    "IVA rate",
    "IVA amount",
    "Counterparty",
    "Value",
    "Kind",
    "Note",
    "Attachment IDs",
    "Document link IDs",
    "Legal refs",
    "Source refs",
)


class OfflineWorkbookEvidenceSidecar(BaseModel):
    """Machine-readable evidence sidecar emitted beside an offline workbook."""

    model_config = _STRICT_FROZEN

    schema_version: Literal["calc-sheets-evidence-sidecar/v1"] = _EVIDENCE_SIDECAR_SCHEMA_VERSION
    metadata: SheetExportMetadata
    workbook_sha256: str = Field(min_length=64, max_length=64)
    evidence: SheetEvidenceFacet


class OfflineWorkbookExportResult(BaseModel):
    """Serialized offline workbook plus its machine-readable evidence sidecar."""

    model_config = _STRICT_FROZEN

    workbook_payload: bytes
    workbook_media_type: Literal["application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"] = _XLSX_MEDIA_TYPE
    workbook_filename_extension: Literal["xlsx"] = "xlsx"
    workbook_sha256: str = Field(min_length=64, max_length=64)
    evidence_sidecar_payload: bytes
    evidence_sidecar_media_type: Literal["application/json"] = _JSON_MEDIA_TYPE
    evidence_sidecar_filename_extension: Literal["evidence.json"] = "evidence.json"
    evidence_sidecar_sha256: str = Field(min_length=64, max_length=64)


def evidence_table(plan: SheetExportPlan) -> tuple[str, tuple[str, ...], tuple[tuple[str, ...], ...]]:
    """Return ``(snapshot_fingerprint, header_row, body_rows)`` for the Evidencia surface.

    The single source of truth both transports render: the offline openpyxl
    workbook and the online Google-Sheets apply adapter consume this so the
    Evidencia surface is byte-identical across offline and online exports
    (modelo-export-evidence-parity ADR). Every cell is a string — contributor
    facts are pre-formatted by the same helpers ``_write_evidence`` uses.
    """
    body = tuple(
        tuple(str(value) for value in _contributor_values(row)) for row in plan.evidence.contributor_rows
    ) + tuple(tuple(str(value) for value in _manual_values(row)) for row in plan.evidence.manual_entries)
    return (plan.evidence.snapshot_fingerprint or "", _EVIDENCE_HEADERS, body)


def build_offline_workbook(plan: SheetExportPlan) -> Workbook:
    """Materialise a ``SheetExportPlan`` as an offline openpyxl workbook."""
    workbook = Workbook()
    default = workbook.active
    assert default is not None
    default.title = TabName.ENTRADAS.value
    for tab in TabName:
        if tab.value not in workbook.sheetnames:
            workbook.create_sheet(tab.value)

    _write_value_cells(workbook, plan.value_cells)
    _write_formula_cells(workbook, plan.formula_cells)
    _write_row_set_headers(workbook, plan.row_sets)
    _write_guide(workbook[TabName.GUIDE.value], plan)
    _write_evidence(workbook[TabName.EVIDENCIA.value], plan)
    _apply_styling(workbook, plan)
    return workbook


def _apply_styling(workbook: Workbook, plan: SheetExportPlan) -> None:
    """Apply the design system — font, role fills, widths, freezes, filters.

    The single offline materialisation of the shared ``_theme`` palette: it sets
    the monospace family on every populated cell, tints each styled range by its
    role (header band, section banner, pale-yellow inputs, grey computed, green
    result), wraps the body columns, sizes the columns, freezes the header rows,
    and installs the basic filters — mirroring exactly what the online apply
    adapter emits from the same ``SheetExportPlan`` facets. The phases run in the
    same order as before: base font first, then styled overrides, widths, freezes,
    filters, and finally print setup.
    """
    family = plan.font_family or WORKBOOK_FONT_FAMILY
    _apply_base_font(workbook, family)
    _apply_styled_ranges(workbook, family, plan.styled_ranges)
    _apply_column_widths(workbook, plan.column_widths)
    _apply_frozen_views(workbook, plan.frozen_views)
    _apply_auto_filters(workbook, plan.auto_filters)
    _apply_print_setup(workbook)


def _apply_base_font(workbook: Workbook, family: str) -> None:
    """Set the monospace family on every populated cell across all tabs."""
    base_font = Font(name=family)
    for tab in TabName:
        worksheet = workbook[tab.value]
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.font = base_font


def _apply_styled_ranges(workbook: Workbook, family: str, styled_ranges: Sequence[SheetStyledRange]) -> None:
    """Tint each styled range by its role, in declaration order so later ranges win."""
    for styled in styled_ranges:
        style = ROLE_STYLES[styled.role]
        font = Font(
            name=family,
            bold=style.bold,
            color=openpyxl_argb(style.font_hex) if style.font_hex else None,
        )
        fill = (
            PatternFill(fill_type="solid", fgColor=openpyxl_argb(style.fill_hex))
            if style.fill_hex is not None
            else None
        )
        alignment = Alignment(horizontal=style.align, vertical="top", wrap_text=styled.wrap)
        worksheet = workbook[styled.tab.value]
        for row_index in range(styled.start_row, styled.end_row + 1):
            for column_index in range(styled.start_column, styled.end_column + 1):
                cell = worksheet.cell(row=row_index, column=column_index)
                cell.font = font
                cell.alignment = alignment
                if fill is not None:
                    cell.fill = fill


def _apply_column_widths(workbook: Workbook, column_widths: Sequence[SheetColumnWidth]) -> None:
    """Size each declared column."""
    for width in column_widths:
        worksheet = workbook[width.tab.value]
        worksheet.column_dimensions[get_column_letter(width.column)].width = width.width


def _apply_frozen_views(workbook: Workbook, frozen_views: Sequence[SheetFrozenView]) -> None:
    """Freeze the header rows/columns on each declared tab."""
    for frozen in frozen_views:
        worksheet = workbook[frozen.tab.value]
        worksheet.freeze_panes = f"{get_column_letter(frozen.frozen_columns + 1)}{frozen.frozen_rows + 1}"


def _apply_auto_filters(workbook: Workbook, auto_filters: Sequence[SheetAutoFilter]) -> None:
    """Install the basic filter over each declared range."""
    for filter_range in auto_filters:
        worksheet = workbook[filter_range.tab.value]
        start = f"{get_column_letter(filter_range.start_column)}{filter_range.start_row}"
        end = f"{get_column_letter(filter_range.end_column)}{filter_range.end_row}"
        worksheet.auto_filter.ref = f"{start}:{end}"


def _apply_print_setup(workbook: Workbook) -> None:
    """Landscape, fit all columns to one page width, repeat the header row.

    A printed filing artefact then stays readable across page breaks.
    """
    for tab in TabName:
        worksheet = workbook[tab.value]
        worksheet.page_setup.orientation = "landscape"
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0
        worksheet.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        worksheet.print_title_rows = "1:1"


def build_evidence_sidecar(
    plan: SheetExportPlan,
    *,
    workbook_sha256: str,
) -> OfflineWorkbookEvidenceSidecar:
    """Build the machine-readable evidence sidecar for one workbook export.

    Returns:
        :class:`OfflineWorkbookEvidenceSidecar`: The evidence sidecar.
    """
    return OfflineWorkbookEvidenceSidecar(
        metadata=plan.metadata,
        workbook_sha256=workbook_sha256,
        evidence=plan.evidence,
    )


def serialize_evidence_sidecar(sidecar: OfflineWorkbookEvidenceSidecar) -> bytes:
    """Serialize an evidence sidecar as canonical UTF-8 JSON bytes."""
    payload = sidecar.model_dump(mode="json")
    text = json.dumps(payload, ensure_ascii=True, sort_keys=True, separators=(",", ":"))
    return text.encode(UTF_8_ENCODING)


def serialize_offline_workbook(plan: SheetExportPlan) -> bytes:
    """Serialize an offline workbook plan to XLSX bytes."""
    workbook = build_offline_workbook(plan)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def serialize_offline_export(plan: SheetExportPlan) -> OfflineWorkbookExportResult:
    """Serialize an offline workbook and its adjacent evidence sidecar.

    Returns:
        :class:`OfflineWorkbookExportResult`: The export result.
    """
    workbook_payload = serialize_offline_workbook(plan)
    workbook_sha256 = _sha256(workbook_payload)
    sidecar = build_evidence_sidecar(plan, workbook_sha256=workbook_sha256)
    evidence_sidecar_payload = serialize_evidence_sidecar(sidecar)
    return OfflineWorkbookExportResult(
        workbook_payload=workbook_payload,
        workbook_sha256=workbook_sha256,
        evidence_sidecar_payload=evidence_sidecar_payload,
        evidence_sidecar_sha256=_sha256(evidence_sidecar_payload),
    )


def _write_value_cells(workbook: Workbook, cells: Iterable[SheetValueCell]) -> None:
    for cell in cells:
        worksheet = workbook[cell.address.tab.value]
        target = worksheet.cell(row=cell.address.row, column=cell.address.column, value=_coerce_cell_value(cell.value))
        if cell.note is not None:
            assert isinstance(target, Cell)
            target.comment = Comment(cell.note, "AEAT")


def _write_formula_cells(workbook: Workbook, cells: Iterable[SheetFormulaCell]) -> None:
    for cell in cells:
        worksheet = workbook[cell.address.tab.value]
        worksheet.cell(row=cell.address.row, column=cell.address.column, value=f"={cell.formula}")


def _write_row_set_headers(workbook: Workbook, row_sets: Iterable[SheetRowSet]) -> None:
    for row_set in row_sets:
        worksheet = workbook[row_set.tab.value]
        for column in row_set.columns:
            worksheet.cell(
                row=column.header_address.row,
                column=column.header_address.column,
                value=column.header_label,
            )


def _write_guide(worksheet: Worksheet, plan: SheetExportPlan) -> None:
    worksheet["A1"] = plan.guide.title
    for row, paragraph in enumerate(plan.guide.paragraphs, start=3):
        worksheet.cell(row=row, column=1, value=paragraph)

    metadata = plan.metadata
    base_row = 3 + len(plan.guide.paragraphs) + 2
    stamps = (
        ("Modelo", metadata.modelo_id),
        ("Revision", metadata.revision_id),
        ("Periodo", f"{metadata.period} / {metadata.filing_year}"),
        ("Motor", metadata.engine_version),
        ("Registry SHA", metadata.registry_sha),
        ("Exportado", metadata.exported_at.isoformat()),
    )
    for offset, (label, value) in enumerate(stamps):
        worksheet.cell(row=base_row + offset, column=1, value=label)
        worksheet.cell(row=base_row + offset, column=2, value=value)


def _write_evidence(worksheet: Worksheet, plan: SheetExportPlan) -> None:
    worksheet["A1"] = "Snapshot fingerprint"
    worksheet["B1"] = plan.evidence.snapshot_fingerprint or ""
    for column, header in enumerate(_EVIDENCE_HEADERS, start=1):
        worksheet.cell(row=3, column=column, value=header)

    row_index = 4
    for row in plan.evidence.contributor_rows:
        _write_evidence_row(worksheet, row_index, _contributor_values(row))
        row_index += 1
    for row in plan.evidence.manual_entries:
        _write_evidence_row(worksheet, row_index, _manual_values(row))
        row_index += 1

    worksheet.freeze_panes = "A4"
    worksheet.protection.sheet = True


def _write_evidence_row(worksheet: Worksheet, row_index: int, values: Sequence[str]) -> None:
    for column, value in enumerate(values, start=1):
        worksheet.cell(row=row_index, column=column, value=value)


def _contributor_values(row: SheetEvidenceContributorRow) -> tuple[str, ...]:
    return (
        "ledger",
        row.casilla_id,
        row.transaction_id,
        _format_decimal(row.amount),
        row.currency,
        _format_optional_decimal(row.taxable_base),
        _format_optional_decimal(row.iva_rate),
        _format_optional_decimal(row.iva_amount),
        row.counterparty or "",
        "",
        "",
        "",
        _join(row.attachment_ids),
        _join(row.document_link_ids),
        _join(row.legal_refs),
        _join(row.source_refs),
    )


def _manual_values(row: SheetEvidenceManualEntry) -> tuple[str, ...]:
    return (
        "manual",
        row.casilla_id,
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        row.value,
        row.kind,
        row.note,
        "",
        "",
        _join(row.legal_refs),
        _join(row.source_refs),
    )


def _coerce_cell_value(value: Decimal | str | bool | None) -> str | bool:
    if value is None:
        return ""
    if isinstance(value, Decimal):
        return _format_decimal(value)
    return value


def _format_decimal(value: Decimal) -> str:
    return format(value, "f")


def _format_optional_decimal(value: Decimal | None) -> str:
    if value is None:
        return ""
    return _format_decimal(value)


def _join(values: tuple[str, ...]) -> str:
    return ";".join(values)


def _sha256(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


__all__ = [
    "OfflineWorkbookEvidenceSidecar",
    "OfflineWorkbookExportResult",
    "build_evidence_sidecar",
    "build_offline_workbook",
    "serialize_evidence_sidecar",
    "serialize_offline_export",
    "serialize_offline_workbook",
]
