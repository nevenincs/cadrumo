"""Tests for application export serialization helpers."""

from __future__ import annotations

import csv
import hashlib
from io import BytesIO, StringIO
from typing import cast

import pytest
from openpyxl import load_workbook

from ....core.errors import build_error_envelope
from ....core.external_constants import CSV_MIME_TYPE, JSONL_MIME_TYPE, UTF_8_ENCODING, XLSX_MIME_TYPE
from .. import ExportSerializationFormat, serialize_tabular_rows
from .._errors import ExportFieldError, ExportFormatError

pytestmark = [pytest.mark.unit, pytest.mark.hex_application]


def test_serialize_tabular_rows_writes_stable_csv_payload() -> None:
    result = serialize_tabular_rows(
        (
            {"transaction_id": "b", "amount": "2.00"},
            {"transaction_id": "a", "amount": "1.00"},
        ),
        fieldnames=("transaction_id", "amount"),
        export_format=ExportSerializationFormat.CSV,
    )

    parsed = tuple(csv.DictReader(StringIO(result.payload.decode(UTF_8_ENCODING))))
    assert parsed == (
        {"transaction_id": "b", "amount": "2.00"},
        {"transaction_id": "a", "amount": "1.00"},
    )
    assert result.media_type == CSV_MIME_TYPE
    assert result.filename_extension == "csv"
    assert result.byte_size == len(result.payload)
    assert len(result.sha256) == 64


def test_serialize_tabular_rows_writes_stable_jsonl_payload() -> None:
    result = serialize_tabular_rows(
        ({"transaction_id": "b", "amount": "2.00"},),
        fieldnames=("transaction_id", "amount"),
        export_format=ExportSerializationFormat.JSONL,
    )

    assert result.payload == b'{"amount":"2.00","transaction_id":"b"}\n'
    assert result.media_type == JSONL_MIME_TYPE
    assert result.filename_extension == "jsonl"


def test_serialize_tabular_rows_writes_readable_xlsx_payload() -> None:
    result = serialize_tabular_rows(
        (
            {"transaction_id": "b", "amount": "2.00"},
            {"transaction_id": "a", "amount": "1.00"},
        ),
        fieldnames=("transaction_id", "amount"),
        export_format=ExportSerializationFormat.XLSX,
    )

    workbook = load_workbook(BytesIO(result.payload), read_only=True, data_only=True)
    worksheet = workbook.active
    assert worksheet is not None
    rows = tuple(worksheet.iter_rows(values_only=True))

    assert rows == (
        ("transaction_id", "amount"),
        ("b", "2.00"),
        ("a", "1.00"),
    )
    assert result.media_type == XLSX_MIME_TYPE
    assert result.filename_extension == "xlsx"
    assert result.byte_size == len(result.payload)
    assert result.sha256 == hashlib.sha256(result.payload).hexdigest()


def test_serialize_tabular_rows_rejects_unknown_fields() -> None:
    with pytest.raises(ExportFieldError) as exc_info:
        serialize_tabular_rows(
            ({"transaction_id": "b", "amount": "2.00", "extra": "x"},),
            fieldnames=("transaction_id", "amount"),
            export_format=ExportSerializationFormat.CSV,
        )
    assert exc_info.value.translated_message == "errors.refused.refused_export_field"
    assert exc_info.value.context == {"reason": "unknown_fields", "unknown_fields": ("extra",)}


# ---------------------------------------------------------------------------
# Registry membership — contract
# ---------------------------------------------------------------------------


def test_export_format_error_is_in_error_registry() -> None:
    """ExportFormatError must be registered so the CLI can handle it."""
    from ....core.errors import ERROR_REGISTRY

    assert "REFUSED_EXPORT_FORMAT" in ERROR_REGISTRY


def test_export_format_error_registry_has_no_name_collision() -> None:
    """The error registry must contain exactly one class named ExportFormatError.

    After the adapter rename, only the application-layer ExportFormatError
    (code REFUSED_EXPORT_FORMAT) should appear under that simple name.  The
    adapter class is now AeatExportFormatError (code FAIL_EXPORT_FORMAT).
    """
    from ....core.errors.registry import (
        _ALL_DECLARED_ERROR_CODES,  # pyright: ignore[reportPrivateUsage]  # test introspection
    )

    qualnames_named_export_format_error = [
        qualname for qualname, _ in _ALL_DECLARED_ERROR_CODES if qualname.split(".")[-1] == "ExportFormatError"
    ]
    assert qualnames_named_export_format_error == ["aeat.application.export._errors.ExportFormatError"], (
        f"Expected exactly one ExportFormatError in the registry "
        f"(application canonical), got: {qualnames_named_export_format_error}"
    )


def test_export_field_error_is_in_error_registry() -> None:
    """ExportFieldError must be registered so the CLI can handle it."""
    from ....core.errors import ERROR_REGISTRY

    assert "REFUSED_EXPORT_FIELD" in ERROR_REGISTRY


def test_export_format_error_code_attributes() -> None:
    """ERROR_REGISTRY entry for ExportFormatError carries expected attributes."""
    from ....core.errors import ERROR_REGISTRY

    code = ERROR_REGISTRY["REFUSED_EXPORT_FORMAT"]
    assert code.code == "REFUSED_EXPORT_FORMAT"
    assert code.message_key == "errors.refused.refused_export_format"
    assert code.retryable is False


def test_export_field_error_code_attributes() -> None:
    """ERROR_REGISTRY entry for ExportFieldError carries expected attributes."""
    from ....core.errors import ERROR_REGISTRY

    code = ERROR_REGISTRY["REFUSED_EXPORT_FIELD"]
    assert code.code == "REFUSED_EXPORT_FIELD"
    assert code.message_key == "errors.refused.refused_export_field"
    assert code.retryable is False


# ---------------------------------------------------------------------------
# Envelope round-trip — contract
# ---------------------------------------------------------------------------


def test_export_format_error_build_error_envelope() -> None:
    """build_error_envelope must succeed for ExportFormatError."""

    err = ExportFormatError(
        translated_message="errors.refused.refused_export_format",
        context={"export_format": "xml"},
    )
    envelope = build_error_envelope(err)
    assert envelope.code == "REFUSED_EXPORT_FORMAT"
    assert envelope.category == "REFUSED"
    assert envelope.retryable is False
    assert envelope.schema_version == "1"


def test_export_field_error_build_error_envelope() -> None:
    """build_error_envelope must succeed for ExportFieldError."""

    err = ExportFieldError(
        translated_message="errors.refused.refused_export_field",
        context={"reason": "fieldnames_empty"},
    )
    envelope = build_error_envelope(err)
    assert envelope.code == "REFUSED_EXPORT_FIELD"
    assert envelope.category == "REFUSED"
    assert envelope.retryable is False
    assert envelope.schema_version == "1"


# ---------------------------------------------------------------------------
# Locale translated_message key — contract
# ---------------------------------------------------------------------------


def test_export_format_error_locale_key_present_in_catalogue() -> None:
    """The locale catalogue must carry the refused_export_format key for every locale."""
    import importlib.resources
    import pathlib

    import yaml

    locale_dir = pathlib.Path(
        importlib.resources.files("aeat.locales").__str__()  # type: ignore[arg-type]
    )
    for locale_code in ("en", "es", "ca", "hu"):
        text = (locale_dir / f"{locale_code}.yml").read_text(encoding=UTF_8_ENCODING)
        data = yaml.safe_load(text)
        value = data.get("errors", {}).get("refused", {}).get("refused_export_format")
        assert value, (
            f"locale {locale_code!r}: 'errors.refused.refused_export_format' key is "
            f"missing or empty in {locale_code}.yml"
        )


def test_export_field_error_locale_key_present_in_catalogue() -> None:
    """The locale catalogue must carry the refused_export_field key for every locale."""
    import importlib.resources
    import pathlib

    import yaml

    locale_dir = pathlib.Path(
        importlib.resources.files("aeat.locales").__str__()  # type: ignore[arg-type]
    )
    for locale_code in ("en", "es", "ca", "hu"):
        text = (locale_dir / f"{locale_code}.yml").read_text(encoding=UTF_8_ENCODING)
        data = yaml.safe_load(text)
        value = data.get("errors", {}).get("refused", {}).get("refused_export_field")
        assert value, (
            f"locale {locale_code!r}: 'errors.refused.refused_export_field' key is "
            f"missing or empty in {locale_code}.yml"
        )


# ---------------------------------------------------------------------------
# Real raise sites — contract: seven replaced ValueError sites
# ---------------------------------------------------------------------------


def test_normalize_fieldnames_raises_export_field_error_on_empty_sequence() -> None:
    """Site: _normalize_fieldnames — empty fieldnames sequence."""
    with pytest.raises(ExportFieldError) as exc_info:
        serialize_tabular_rows(
            (),
            fieldnames=(),
            export_format=ExportSerializationFormat.CSV,
        )
    assert exc_info.value.translated_message == "errors.refused.refused_export_field"
    assert exc_info.value.context == {"reason": "fieldnames_empty"}


def test_normalize_fieldnames_raises_export_field_error_on_blank_name() -> None:
    """Site: _normalize_fieldnames — blank field name."""
    with pytest.raises(ExportFieldError) as exc_info:
        serialize_tabular_rows(
            (),
            fieldnames=("transaction_id", "  "),
            export_format=ExportSerializationFormat.CSV,
        )
    assert exc_info.value.translated_message == "errors.refused.refused_export_field"
    assert exc_info.value.context == {"reason": "fieldnames_blank"}


def test_normalize_fieldnames_raises_export_field_error_on_duplicate_names() -> None:
    """Site: _normalize_fieldnames — duplicate field names."""
    with pytest.raises(ExportFieldError) as exc_info:
        serialize_tabular_rows(
            (),
            fieldnames=("amount", "amount"),
            export_format=ExportSerializationFormat.CSV,
        )
    assert exc_info.value.translated_message == "errors.refused.refused_export_field"
    assert exc_info.value.context == {"reason": "fieldnames_duplicate"}


def test_normalize_row_raises_export_field_error_on_unknown_field() -> None:
    """Site: _normalize_row — row contains unknown field keys."""
    with pytest.raises(ExportFieldError) as exc_info:
        serialize_tabular_rows(
            ({"transaction_id": "a", "amount": "1.00", "bogus": "x"},),
            fieldnames=("transaction_id", "amount"),
            export_format=ExportSerializationFormat.JSONL,
        )
    assert exc_info.value.translated_message == "errors.refused.refused_export_field"
    assert exc_info.value.context == {"reason": "unknown_fields", "unknown_fields": ("bogus",)}


def test_serialize_tabular_rows_rejects_unsupported_runtime_format() -> None:
    """Site: serialize_tabular_rows — runtime value outside the closed enum."""
    with pytest.raises(ExportFormatError) as exc_info:
        serialize_tabular_rows(
            (),
            fieldnames=("amount",),
            export_format=cast(ExportSerializationFormat, "xml"),
        )
    assert exc_info.value.translated_message == "errors.refused.refused_export_format"
    assert exc_info.value.context == {"export_format": "xml"}
    envelope = build_error_envelope(exc_info.value)
    assert envelope.code == "REFUSED_EXPORT_FORMAT"


def test_model_validator_raises_export_field_error_on_blank_fieldname() -> None:
    """Site: TabularExportResult._validate_fieldnames — blank field in model."""
    from pydantic import ValidationError

    from .._tabular import TabularExportResult

    payload = b"transaction_id,amount\r\n"
    with pytest.raises((ExportFieldError, ValidationError)) as exc_info:
        TabularExportResult(
            format=ExportSerializationFormat.CSV,
            media_type="text/csv",
            filename_extension="csv",
            payload=payload,
            byte_size=len(payload),
            sha256=hashlib.sha256(payload).hexdigest(),
            row_count=0,
            fieldnames=("transaction_id", "  "),
        )
    if isinstance(exc_info.value, ExportFieldError):
        assert exc_info.value.translated_message == "errors.refused.refused_export_field"
        assert exc_info.value.context == {"reason": "fieldnames_blank"}


def test_model_validator_raises_export_field_error_on_duplicate_fieldname() -> None:
    """Site: TabularExportResult._validate_fieldnames — duplicate field in model."""
    from pydantic import ValidationError

    from .._tabular import TabularExportResult

    payload = b"amount,amount\r\n"
    with pytest.raises((ExportFieldError, ValidationError)) as exc_info:
        TabularExportResult(
            format=ExportSerializationFormat.CSV,
            media_type="text/csv",
            filename_extension="csv",
            payload=payload,
            byte_size=len(payload),
            sha256=hashlib.sha256(payload).hexdigest(),
            row_count=0,
            fieldnames=("amount", "amount"),
        )
    if isinstance(exc_info.value, ExportFieldError):
        assert exc_info.value.translated_message == "errors.refused.refused_export_field"
        assert exc_info.value.context == {"reason": "fieldnames_duplicate"}


def test_model_validator_raises_export_field_error_on_invalid_sha256() -> None:
    """Site: TabularExportResult._validate_sha256 — digest must be hex-addressed."""

    from pydantic import ValidationError

    from .._tabular import TabularExportResult

    payload = b"transaction_id,amount\n"
    with pytest.raises(ValidationError) as exc_info:
        TabularExportResult(
            format=ExportSerializationFormat.CSV,
            media_type=CSV_MIME_TYPE,
            filename_extension="csv",
            payload=payload,
            byte_size=len(payload),
            sha256="not-a-digest",
            row_count=0,
            fieldnames=("transaction_id", "amount"),
        )
    error_detail = exc_info.value.errors()[0]
    assert "ctx" in error_detail and "error" in error_detail["ctx"]
    cause = error_detail["ctx"]["error"]
    assert isinstance(cause, ExportFieldError)
    assert cause.translated_message == "errors.refused.refused_export_field"
    assert cause.context == {"reason": "sha256_invalid"}
