"""XLSX financial provider with header-row detection."""

from __future__ import annotations

from collections.abc import Iterator
from contextlib import suppress
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook import Workbook

from aeat.financial._raw_transaction import RawTransaction, SourceFormat
from aeat.financial.providers._base import (
    FinancialProvider,
    InvalidFinancialSourceError,
    ProviderValidation,
    build_raw_transaction,
    default_currency,
    parse_amount_value,
    parse_date_value,
    synthesize_transaction_id,
)
from aeat.financial.providers._csv import (
    CSV_LAYOUTS,
    CsvBankLayout,
    _find_column,
    _header_lookup,
    _layout_score,
    _required_value,
    _row_is_blank,
    _value_from_aliases,
)


class XlsxProvider(FinancialProvider):
    """Ingest raw transactions from `.xlsx` bank statement exports."""

    name = "XLSX provider"
    supported_extensions = frozenset({".xlsx"})
    source_format = SourceFormat.XLSX

    def __init__(self) -> None:
        """Initialize validation metadata placeholders."""
        self._last_sheet_name = "Sheet1"
        self._last_header_index = 1

    def validate_source(self, path: Path) -> ProviderValidation:
        """Validate workbook accessibility and header detection."""
        try:
            workbook, rows, _, layout, header_row, _, _ = self._locate_sheet(path)
        except InvalidFinancialSourceError as exc:
            return ProviderValidation(is_valid=False, warnings=(str(exc),))
        finally:
            with suppress(Exception):
                workbook.close()  # type: ignore[name-defined]
        warnings: list[str] = []
        if layout is None:
            return ProviderValidation(
                is_valid=False,
                warnings=("Workbook does not contain a supported bank-statement header row",),
            )
        if header_row is None or len(rows) <= self._last_header_index:
            return ProviderValidation(
                is_valid=False,
                warnings=(f"{layout.bank_name} worksheet has no data rows after the header",),
            )
        if not _find_column(_header_lookup(header_row), layout.columns.currency):
            warnings.append(
                f"{layout.bank_name} worksheet has no currency column; falling back to {default_currency()}",
            )
        return ProviderValidation(
            is_valid=True,
            warnings=tuple(warnings),
            detected_encoding=None,
            detected_dialect=f"worksheet={self._last_sheet_name},header_row={self._last_header_index}",
        )

    def ingest(self, path: Path) -> Iterator[RawTransaction]:
        """Yield strict raw transactions from the first matching worksheet."""
        source_bytes = self._read_source_bytes(path)
        source_sha256 = self._compute_sha256(source_bytes)
        workbook, rows, sheet_name, layout, headers, lookup, header_index = self._locate_sheet(path)
        if layout is None or headers is None or lookup is None:
            workbook.close()
            raise InvalidFinancialSourceError("Workbook does not contain a supported bank-statement header row")
        try:
            for source_row_index, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
                raw_fields = {header: row[index] if index < len(row) else "" for index, header in enumerate(headers)}
                if _row_is_blank(raw_fields):
                    continue
                transaction_id = _value_from_aliases(raw_fields, lookup, layout.columns.external_id)
                if not transaction_id:
                    transaction_id = synthesize_transaction_id(
                        provider_name=f"{layout.bank_name}-{sheet_name}",
                        source_sha256=source_sha256,
                        source_row_index=source_row_index,
                    )
                booked_date = parse_date_value(
                    _required_value(raw_fields, lookup, layout.columns.booked_date, "booked_date"),
                    day_first=layout.day_first_dates,
                )
                value_text = _value_from_aliases(raw_fields, lookup, layout.columns.value_date)
                value_date = parse_date_value(value_text, day_first=layout.day_first_dates) if value_text else None
                amount = parse_amount_value(_required_value(raw_fields, lookup, layout.columns.amount, "amount"))
                currency = _value_from_aliases(raw_fields, lookup, layout.columns.currency) or default_currency()
                description = _required_value(raw_fields, lookup, layout.columns.description, "description")
                counterparty = _value_from_aliases(raw_fields, lookup, layout.columns.counterparty)
                yield build_raw_transaction(
                    provider=self,
                    path=path,
                    source_sha256=source_sha256,
                    source_row_index=source_row_index,
                    transaction_id=transaction_id,
                    booked_date=booked_date,
                    value_date=value_date,
                    amount=amount,
                    currency=currency,
                    counterparty=counterparty,
                    description=description,
                    raw_fields=raw_fields,
                )
        finally:
            workbook.close()

    def _locate_sheet(
        self,
        path: Path,
    ) -> tuple[Workbook, list[list[str]], str, CsvBankLayout | None, list[str] | None, dict[str, str] | None, int]:
        """Return the first worksheet that matches a known bank layout."""
        try:
            workbook = load_workbook(filename=path, read_only=True, data_only=True)
        except Exception as exc:  # pragma: no cover - exercised via validation path
            raise InvalidFinancialSourceError(f"could not open workbook: {path}") from exc
        best_rows: list[list[str]] = []
        best_sheet_name = workbook.sheetnames[0] if workbook.sheetnames else "Sheet1"
        best_layout: CsvBankLayout | None = None
        best_headers: list[str] | None = None
        best_lookup: dict[str, str] | None = None
        best_header_index = 0
        best_score = -1
        for worksheet in workbook.worksheets:
            rows = [
                [str(cell).strip() if cell is not None else "" for cell in row]
                for row in worksheet.iter_rows(values_only=True)
            ]
            for index, row in enumerate(rows[:10]):
                if not any(cell.strip() for cell in row):
                    continue
                lookup = _header_lookup(row)
                for layout in CSV_LAYOUTS:
                    score = _layout_score(lookup, layout)
                    if score > best_score:
                        best_rows = rows
                        best_sheet_name = worksheet.title
                        best_layout = layout
                        best_headers = row
                        best_lookup = lookup
                        best_header_index = index
                        best_score = score
        self._last_sheet_name = best_sheet_name
        self._last_header_index = best_header_index + 1
        if best_score < 3:
            return workbook, best_rows, best_sheet_name, None, None, None, best_header_index
        return workbook, best_rows, best_sheet_name, best_layout, best_headers, best_lookup, best_header_index
