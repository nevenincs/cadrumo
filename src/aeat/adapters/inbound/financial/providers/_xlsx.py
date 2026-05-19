"""XLSX financial provider with header-row detection.

Implements :class:`XlsxProvider`, an
:class:`aeat.adapters.inbound.financial.providers._base.FinancialProvider`
backed by ``openpyxl``. Reuses the bank-layout catalogue and scoring
helpers from :mod:`aeat.adapters.inbound.financial.providers._csv` so
the same alias rules apply to spreadsheet exports as to the
matching CSV downloads.
"""

from __future__ import annotations

from collections.abc import Iterator, Mapping, Sequence
from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .....core.logging import get_logger
from .....domain.transactions import RawTransaction, SourceFormat
from ._base import (
    FinancialProvider,
    InvalidFinancialSourceError,
    ProviderValidation,
    build_raw_transaction,
    coerce_cell_text,
    default_currency,
    parse_amount_value,
    parse_date_value,
    synthesize_transaction_id,
)
from ._csv import (
    CSV_LAYOUTS,
    CsvBankLayout,
    _find_column,
    _header_lookup,
    _layout_score,
    _required_value,
    _row_is_blank,
    _value_from_aliases,
)

_logger = get_logger(__name__)


class XlsxProvider(FinancialProvider):
    """Ingest raw transactions from ``.xlsx`` bank statement exports.

    Iterates every worksheet, scoring the first ten rows of each
    against :data:`aeat.adapters.inbound.financial.providers._csv.CSV_LAYOUTS`
    and selecting the worksheet/row pair with the highest match
    score. Numeric and date cell values are read directly from the
    cell types (rather than coerced through their printed strings)
    so locale-formatted ``Decimal`` and ``date`` parsing stays
    accurate.

    Attributes:
        _last_sheet_name: Name of the worksheet selected by the most
            recent :meth:`_locate_sheet` call; surfaced via
            ``ProviderValidation.detected_dialect``.
        _last_header_index: 1-based index of the header row in that
            worksheet.
    """

    name = "XLSX provider"
    supported_extensions = frozenset({".xlsx"})
    source_format = SourceFormat.XLSX

    def __init__(self) -> None:
        """Initialise the validation metadata placeholders."""
        self._last_sheet_name = "Sheet1"
        self._last_header_index = 1

    def validate_source(self, path: Path) -> ProviderValidation:
        """Validate workbook accessibility and header detection."""
        workbook: Workbook | None = None
        try:
            workbook, rows, _, layout, header_row, _, _ = self._locate_sheet(path)
        except InvalidFinancialSourceError as exc:
            return ProviderValidation(is_valid=False, warnings=(str(exc),))
        finally:
            if workbook is not None:
                try:
                    workbook.close()
                except Exception as close_exc:
                    _logger.debug(
                        "xlsx provider: workbook.close() after validate_source failed (%s)",
                        close_exc,
                        exc_info=True,
                    )
        warnings: list[str] = []
        if layout is None:
            return ProviderValidation(
                is_valid=False,
                warnings=("Workbook does not contain a supported bank-statement header row",),
            )
        if header_row is None or len(rows) <= self._last_header_index:
            return ProviderValidation(
                is_valid=False,
                warnings=(f"{layout.bank_name} worksheet has no data rows after the header",),
            )
        if not _find_column(_header_lookup(header_row), layout.columns.currency):
            warnings.append(
                f"{layout.bank_name} worksheet has no currency column; falling back to {default_currency()}",
            )
        return ProviderValidation(
            is_valid=True,
            warnings=tuple(warnings),
            detected_encoding=None,
            detected_dialect=f"worksheet={self._last_sheet_name},header_row={self._last_header_index}",
        )

    def ingest(self, path: Path) -> Iterator[RawTransaction]:
        """Yield strict raw transactions from the first matching worksheet."""
        source_bytes = self._read_source_bytes(path)
        source_sha256 = self._compute_sha256(source_bytes)
        workbook, rows, sheet_name, layout, headers, lookup, header_index = self._locate_sheet(path)
        if layout is None or headers is None or lookup is None:
            workbook.close()
            raise InvalidFinancialSourceError("Workbook does not contain a supported bank-statement header row")
        try:
            for source_row_index, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
                raw_fields = _row_to_mapping(headers, row)
                cell_lookup = _row_to_cells(headers, row)
                if _row_is_blank(raw_fields):
                    continue
                parsed = _parse_xlsx_row(
                    layout=layout,
                    lookup=lookup,
                    raw_fields=raw_fields,
                    cell_lookup=cell_lookup,
                    sheet_name=sheet_name,
                    source_sha256=source_sha256,
                    source_row_index=source_row_index,
                    path=path,
                )
                yield build_raw_transaction(
                    provider=self,
                    path=path,
                    source_sha256=source_sha256,
                    source_row_index=source_row_index,
                    transaction_id=parsed.transaction_id,
                    booked_date=parsed.booked_date,
                    value_date=parsed.value_date,
                    amount=parsed.amount,
                    currency=parsed.currency,
                    counterparty=parsed.counterparty,
                    description=parsed.description,
                    raw_fields=raw_fields,
                )
        finally:
            workbook.close()

    def _locate_sheet(
        self,
        path: Path,
    ) -> tuple[Workbook, list[list[Any]], str, CsvBankLayout | None, list[str] | None, dict[str, str] | None, int]:
        """Return the first worksheet that matches a known bank layout."""
        try:
            workbook = load_workbook(filename=path, read_only=True, data_only=True)
        except Exception as exc:  # pragma: no cover - exercised via validation path
            raise InvalidFinancialSourceError(f"could not open workbook: {path}") from exc
        try:
            best_worksheet: Worksheet | None = workbook.worksheets[0] if workbook.worksheets else None
            best_sheet_name = best_worksheet.title if best_worksheet is not None else "Sheet1"
            best_layout: CsvBankLayout | None = None
            best_headers: list[str] | None = None
            best_lookup: dict[str, str] | None = None
            best_header_index = 0
            best_score = -1
            for worksheet in workbook.worksheets:
                candidate = _best_layout_match_for_worksheet(worksheet)
                if candidate is None or candidate[0] <= best_score:
                    continue
                score, index, row, lookup, layout = candidate
                best_worksheet = worksheet
                best_sheet_name = worksheet.title
                best_layout = layout
                best_headers = row
                best_lookup = lookup
                best_header_index = index
                best_score = score
            best_rows = [list(row) for row in best_worksheet.iter_rows(values_only=True)] if best_worksheet else []
            self._last_sheet_name = best_sheet_name
            self._last_header_index = best_header_index + 1
            if best_score < 3:
                return workbook, best_rows, best_sheet_name, None, None, None, best_header_index
            return workbook, best_rows, best_sheet_name, best_layout, best_headers, best_lookup, best_header_index
        except Exception:
            # Re-raise wrapper that guarantees workbook teardown. Broad
            # catch because the upstream parse can raise openpyxl/xlrd
            # errors, KeyError, ValueError, OSError, IndexError or
            # TypeError depending on file shape; the close() must run
            # uniformly. ``raise`` preserves the original cause.
            try:
                workbook.close()
            except Exception as close_exc:
                _logger.debug(
                    "xlsx provider: workbook.close() during parse-error teardown failed (%s)",
                    close_exc,
                    exc_info=True,
                )
            raise


def _best_layout_match_for_worksheet(
    worksheet: Worksheet,
) -> tuple[int, int, list[str], dict[str, str], CsvBankLayout] | None:
    """Return the best (score, header_index, row, lookup, layout) the worksheet matches, or ``None``."""

    sample_rows = [
        [coerce_cell_text(cell) for cell in row]
        for row in worksheet.iter_rows(min_row=1, max_row=10, values_only=True)
    ]
    best: tuple[int, int, list[str], dict[str, str], CsvBankLayout] | None = None
    for index, row in enumerate(sample_rows):
        if not any(cell.strip() for cell in row):
            continue
        lookup = _header_lookup(row)
        for layout in CSV_LAYOUTS:
            score = _layout_score(lookup, layout)
            if best is None or score > best[0]:
                best = (score, index, row, lookup, layout)
    return best


@dataclass(frozen=True, slots=True)
class _ParsedXlsxRow:
    """Per-row projection used by XlsxProvider.ingest.

    Holds the seven typed fields the build_raw_transaction call
    needs. Constructed inside _parse_xlsx_row so the per-column
    parse + the bank-layout-specific defaulting (synthetic
    transaction ids, default currency, optional value_date) all
    happen in one place.
    """

    transaction_id: str
    booked_date: date
    value_date: date | None
    amount: Decimal
    currency: str
    description: str
    counterparty: str | None


def _parse_xlsx_row(
    *,
    layout: CsvBankLayout,
    lookup: dict[str, str],
    raw_fields: dict[str, str],
    cell_lookup: dict[str, object],
    sheet_name: str,
    source_sha256: str,
    source_row_index: int,
    path: Path,
) -> _ParsedXlsxRow:
    """Project one worksheet row into the typed ``_ParsedXlsxRow`` envelope.

    Re-wraps any ``ValueError`` produced by the per-column parsers
    as :class:`InvalidFinancialSourceError` so the caller sees one
    typed envelope per malformed row. Falls back to
    :func:`synthesize_transaction_id` when the bank does not
    publish an external id; defaults the currency to
    :func:`default_currency` when the bank layout does not declare
    a currency column or the cell is blank.
    """
    try:
        transaction_id = _value_from_aliases(raw_fields, lookup, layout.columns.external_id)
        if not transaction_id:
            transaction_id = synthesize_transaction_id(
                provider_name=f"{layout.bank_name}-{sheet_name}",
                source_sha256=source_sha256,
                source_row_index=source_row_index,
            )
        booked_date = parse_date_value(
            _required_cell_value(cell_lookup, lookup, layout.columns.booked_date, "booked_date"),
            day_first=layout.day_first_dates,
        )
        value_raw = _cell_value_from_aliases(cell_lookup, lookup, layout.columns.value_date)
        value_date = parse_date_value(value_raw, day_first=layout.day_first_dates) if value_raw is not None else None
        amount = parse_amount_value(
            _required_cell_value(cell_lookup, lookup, layout.columns.amount, "amount"),
            decimal_separator=layout.decimal_separator,
        )
        currency = _value_from_aliases(raw_fields, lookup, layout.columns.currency) or default_currency()
        description = _required_value(raw_fields, lookup, layout.columns.description, "description")
        counterparty = _value_from_aliases(raw_fields, lookup, layout.columns.counterparty)
    except ValueError as exc:
        _logger.warning(
            "xlsx_provider: parse error row=%d file=%s",
            source_row_index,
            path.name,
            exc_info=True,
        )
        raise InvalidFinancialSourceError(
            f"worksheet row {source_row_index} could not be parsed: {exc}",
        ) from exc
    return _ParsedXlsxRow(
        transaction_id=transaction_id,
        booked_date=booked_date,
        value_date=value_date,
        amount=amount,
        currency=currency,
        description=description,
        counterparty=counterparty,
    )


def _row_to_mapping(headers: Sequence[str], row: Sequence[object]) -> dict[str, str]:
    """Convert one worksheet row into the stored raw-field mapping."""
    return {header: coerce_cell_text(row[index]) if index < len(row) else "" for index, header in enumerate(headers)}


def _row_to_cells(headers: Sequence[str], row: Sequence[object]) -> dict[str, object]:
    """Map worksheet headers to the original cell values for typed parsing."""
    return {header: row[index] if index < len(row) else "" for index, header in enumerate(headers)}


def _cell_value_from_aliases(
    raw_cells: Mapping[str, object],
    lookup: Mapping[str, str],
    aliases: tuple[str, ...],
) -> object | None:
    """Resolve and read the first non-empty cell value for a logical column."""
    header = _find_column(lookup, aliases)
    if header is None:
        return None
    value = raw_cells.get(header, "")
    return value if coerce_cell_text(value) else None


def _required_cell_value(
    raw_cells: Mapping[str, object],
    lookup: Mapping[str, str],
    aliases: tuple[str, ...],
    field_name: str,
) -> object:
    """Resolve a required logical column from typed worksheet cell values."""
    value = _cell_value_from_aliases(raw_cells, lookup, aliases)
    if value is None:
        raise InvalidFinancialSourceError(f"worksheet row is missing required field {field_name!r}")
    return value
