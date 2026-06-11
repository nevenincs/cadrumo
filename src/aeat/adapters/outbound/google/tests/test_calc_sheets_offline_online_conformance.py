"""Offline/online renderer conformance.

One ``SheetExportPlan`` must render to structurally identical grids whether
materialised offline (openpyxl xls) or online (Google-Sheets apply). This test
compares, cell-for-cell, the offline workbook against the online value/formula/
evidence writes the apply adapter would send — proving the single-builder /
two-transport invariant with no network.
"""

from __future__ import annotations

from datetime import UTC, datetime
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook

from .....core import Period
from .....application.storage.calc_sheets import (
    SheetCellAddress,
    SheetEvidenceContributorRow,
    SheetEvidenceFacet,
    SheetExportMetadata,
    SheetExportPlan,
    SheetFormulaCell,
    SheetGuideContent,
    SheetNumberFormat,
    SheetValueCell,
    TabName,
    serialize_offline_workbook,
)
from .._calc_sheets_apply import (
    _build_emphasis_format_requests,
    _build_evidence_value_data,
    _build_formula_data,
    _build_number_format_requests,
    _build_value_data,
)

pytestmark = [pytest.mark.unit, pytest.mark.hex_outbound_adapter]


def _plan() -> SheetExportPlan:
    return SheetExportPlan(
        metadata=SheetExportMetadata(
            modelo_id="303",
            revision_id="2009-y-siguientes",
            filing_year=2026,
            period=Period.from_year_and_code(2026, "1T"),
            engine_version="test",
            registry_sha="abcd1234",
            exported_at=datetime(2026, 6, 3, 15, 0, tzinfo=UTC),
        ),
        value_cells=(
            SheetValueCell(
                address=SheetCellAddress.at(TabName.ENTRADAS, 2, 4),
                value=Decimal("100.00"),
                role="operator_input",
                casilla="base",
            ),
            SheetValueCell(
                address=SheetCellAddress.at(TabName.ENTRADAS, 1, 1),
                value="Sección",
                role="label",
            ),
        ),
        formula_cells=(
            SheetFormulaCell(
                address=SheetCellAddress.at(TabName.CALCULOS, 2, 4),
                formula="'Entradas'!D2*0.21",
                casilla="cuota",
                rounding_rule="money",
                rounding_scale=2,
            ),
        ),
        guide=SheetGuideContent(title="Modelo 303", paragraphs=("Use Entradas.",)),
        evidence=SheetEvidenceFacet(
            snapshot_fingerprint="f" * 64,
            contributor_rows=(
                SheetEvidenceContributorRow(
                    casilla_id="cuota",
                    transaction_id="tx-001",
                    amount=Decimal("-121.00"),
                    currency="EUR",
                    taxable_base=Decimal("100.00"),
                    iva_rate=Decimal("0.21"),
                    iva_amount=Decimal("21.00"),
                    counterparty="Proveedor SL",
                    legal_refs=("liva-art-99",),
                    source_refs=("boe-a-2026-1",),
                ),
            ),
        ),
    )


def _offline_value(workbook, address: SheetCellAddress) -> object:
    cell = workbook[address.tab.value].cell(row=address.row, column=address.column).value
    return cell if cell is not None else ""


def test_value_and_formula_cells_render_identically_offline_and_online() -> None:
    plan = _plan()
    workbook = load_workbook(BytesIO(serialize_offline_workbook(plan)), data_only=False)

    online_values = {entry["range"]: entry["values"][0][0] for entry in _build_value_data(plan.value_cells)}
    online_formulas = {entry["range"]: entry["values"][0][0] for entry in _build_formula_data(plan.formula_cells)}

    for cell in plan.value_cells:
        addr = cell.address.qualified()
        assert online_values[addr] == _offline_value(workbook, cell.address), addr
    for cell in plan.formula_cells:
        addr = cell.address.qualified()
        assert online_formulas[addr] == _offline_value(workbook, cell.address), addr
        # Both transports emit a live spreadsheet formula (leading '=').
        assert str(online_formulas[addr]).startswith("=")


def test_apply_adapter_emits_number_format_and_emphasis_requests() -> None:
    # contract: the online apply renders number formats + start/final + section
    # headers, not just values. A plan with a money number format yields a
    # NUMBER repeatCell; section headers / anchors yield bold repeatCells.
    from .....application.storage.calc_sheets import SheetAnchor, SheetSectionHeader

    plan = _plan().model_copy(
        update={
            "number_formats": (
                SheetNumberFormat(
                    address=SheetCellAddress.at(TabName.CALCULOS, 2, 4),
                    casilla="cuota",
                    data_type="money",
                    pattern="#,##0.00",
                ),
            ),
            "section_headers": (
                SheetSectionHeader(address=SheetCellAddress.at(TabName.ENTRADAS, 2, 1), text="Sección A"),
            ),
            "anchors": (
                SheetAnchor(address=SheetCellAddress.at(TabName.CALCULOS, 2, 6), kind="final", label="RESULTADO"),
            ),
        },
    )
    sheet_id_by_tab = {tab.value: index for index, tab in enumerate(TabName)}
    number_requests = _build_number_format_requests(plan, sheet_id_by_tab=sheet_id_by_tab)
    assert number_requests, "money casilla must yield a numberFormat request"
    fmt = number_requests[0]["repeatCell"]["cell"]["userEnteredFormat"]["numberFormat"]
    assert fmt == {"type": "NUMBER", "pattern": "#,##0.00"}

    emphasis = _build_emphasis_format_requests(plan, sheet_id_by_tab=sheet_id_by_tab)
    assert emphasis, "section headers / anchors must yield bold requests"
    assert emphasis[0]["repeatCell"]["cell"]["userEnteredFormat"]["textFormat"]["bold"] is True


def test_evidence_surface_conforms_offline_and_online() -> None:
    plan = _plan()
    workbook = load_workbook(BytesIO(serialize_offline_workbook(plan)), data_only=False)
    sheet = workbook[TabName.EVIDENCIA.value]
    online = {entry["range"]: entry["values"][0] for entry in _build_evidence_value_data(plan)}

    tab = TabName.EVIDENCIA.value
    # Fingerprint banner + header + the one contributor row conform cell-for-cell.
    assert online[f"'{tab}'!A1"] == [(sheet.cell(row=1, column=c).value or "") for c in (1, 2)]
    header_width = len(online[f"'{tab}'!A3"])
    assert online[f"'{tab}'!A3"] == [(sheet.cell(row=3, column=c).value or "") for c in range(1, header_width + 1)]
    assert online[f"'{tab}'!A4"] == [(sheet.cell(row=4, column=c).value or "") for c in range(1, header_width + 1)]
