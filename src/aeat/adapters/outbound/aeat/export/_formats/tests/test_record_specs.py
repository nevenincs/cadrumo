"""Structural tests for registry-backed fichero-BOE record specs.

The accepted registry direction is that concrete fixed-width export layouts are
authored as registry TOML, not as per-modelo Python ``modelo_*`` modules under
``_formats``. These tests therefore validate the real registry-backed layouts
that the production serialiser consumes.
"""

from __future__ import annotations

from functools import cache
from pathlib import Path

import pytest
from openpyxl import load_workbook

from .......core.resources import bundled_path, resources
from .......domain.calculations.registry import (
    CasillaFieldKind,
    CasillaId,
    ExportLayoutDefinition,
    ExportRecordDefinition,
    ModeloDefinition,
    ModeloRevision,
    casillas_by_id,
    derive_export_layouts_from_bindings,
)
from .......domain.calculations.registry._export import _verify_record_offsets

pytestmark = [pytest.mark.unit, pytest.mark.hex_outbound_adapter]

_RequiredLayout = tuple[str, str]
_LayoutCase = tuple[ModeloDefinition, ModeloRevision, ExportLayoutDefinition]

_REQUIRED_LAYOUTS: frozenset[_RequiredLayout] = frozenset(
    {
        ("130", "modelo-130-fichero-boe"),
        ("303", "modelo-303-fichero-boe"),
    },
)

_M303_2025_RECORD_DESIGN = (
    bundled_path("corpus", "aeat_official")
    / "disenos_registro"
    / "modelo_303"
    / "files"
    / "06-303-ejercicio-2025-actualizado-04-12-2025-380-kb-xlsx.xlsx"
)


@cache
def _fixed_width_layout_cases() -> tuple[_LayoutCase, ...]:
    """Return all registry-authored fixed-width export layouts."""
    cases: list[_LayoutCase] = []
    for modelo in resources().modelos.authority.modelos:
        for revision in modelo.revisions.values():
            for layout in derive_export_layouts_from_bindings(revision):
                if layout.format == "fixed_width":
                    cases.append((modelo, revision, layout))
    return tuple(cases)


def _layout_case_id(case: _LayoutCase) -> str:
    modelo, revision, layout = case
    return f"M{modelo.id}:{revision.id}:{layout.id}"


def _workbook_record_lengths(workbook_path: Path) -> dict[str, int]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    lengths: dict[str, int] = {}
    try:
        for worksheet in workbook.worksheets:
            last_position: int | None = None
            last_length: int | None = None
            for row in worksheet.iter_rows(min_row=1, max_col=3, values_only=True):
                position, length = row[1], row[2]
                if isinstance(position, int | float) and isinstance(length, int | float):
                    last_position = int(position)
                    last_length = int(length)
            if last_position is not None and last_length is not None:
                lengths[worksheet.title] = last_position + last_length - 1
    finally:
        workbook.close()
    return lengths


def _record_length(record: ExportRecordDefinition) -> int:
    return max((field.offset or 0) + (field.length or 0) - 1 for field in record.fields)


def test_registry_fixed_width_export_layout_surface_is_not_empty() -> None:
    """The fixed-width exporter must be backed by real registry layouts."""
    cases = _fixed_width_layout_cases()

    assert cases, "no registry-backed fixed-width export layouts were discovered"
    present = {(modelo.id, layout.id) for modelo, _revision, layout in cases}
    assert present >= _REQUIRED_LAYOUTS


def test_modelo_303_2025_record_lengths_match_official_workbook() -> None:
    """M303 registry record lengths are grounded in the official 2025 DR workbook."""
    official_lengths = _workbook_record_lengths(_M303_2025_RECORD_DESIGN)
    expected = {
        "modelo-303-page-01": official_lengths["DP30301"],
        "modelo-303-page-02": official_lengths["DP30302"],
        "modelo-303-page-03": official_lengths["DP30303"],
        "modelo-303-page-04": official_lengths["DP30304"],
        "modelo-303-page-05": official_lengths["DP30305"],
        "modelo-303-page-did": official_lengths["DP303DID"],
    }
    modelo = next(item for item in resources().modelos.authority.modelos if item.id == "303")
    revision = modelo.revisions["2023-y-siguientes"]
    layout = next(item for item in derive_export_layouts_from_bindings(revision) if item.id == "modelo-303-fichero-boe")
    actual = {record.id: _record_length(record) for record in layout.records if record.id in expected}

    assert actual == expected


@pytest.mark.parametrize("case", _fixed_width_layout_cases(), ids=_layout_case_id)
def test_registry_fixed_width_record_offsets_are_valid(case: _LayoutCase) -> None:
    """Production record-offset guard accepts every registry fixed-width layout."""
    _modelo, _revision, layout = case

    _verify_record_offsets(layout)


@pytest.mark.parametrize("case", _fixed_width_layout_cases(), ids=_layout_case_id)
def test_registry_fixed_width_casilla_fields_use_declared_canonical_ids(case: _LayoutCase) -> None:
    """Every fixed-width casilla field references a declared canonical ``casilla.id``."""
    modelo, revision, layout = case
    declared_ids: frozenset[CasillaId] = frozenset(casillas_by_id(revision))

    missing: list[str] = []
    for record in layout.records:
        for field in record.fields:
            if field.kind != CasillaFieldKind.CASILLA:
                continue
            if field.casilla_id not in declared_ids:
                missing.append(
                    f"{record.id}.{field.id}: casilla_id={field.casilla_id!r} "
                    f"offset={field.offset} length={field.length}",
                )

    assert not missing, (
        f"modelo {modelo.id} revision {revision.id} layout {layout.id} has export fields "
        "that do not reference declared canonical casilla.id values:\n" + "\n".join(missing)
    )
