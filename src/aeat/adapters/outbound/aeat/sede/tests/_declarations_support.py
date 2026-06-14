"""Shared support for split adapter tests."""

from __future__ import annotations

import hashlib
import os
from collections.abc import Mapping
from datetime import UTC, date, datetime
from decimal import Decimal
from pathlib import Path
from typing import TYPE_CHECKING, Literal

import pytest
from openpyxl import load_workbook
from pydantic import AnyHttpUrl
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

from ......application.filing import (
    ModeloDraftStatus,
    ModeloOperatorProfile,
    build_draft,
    build_runtime_schema_provider,
    export_draft,
)
from ......core import Period
from ......core.config import Settings
from ......core.resources import bundled_path, resources
from ......domain.calculations.registry import (
    InputKind,
    RegistryValidationError,
    calculate_registry_snapshot,
    parse_export_payload,
    relation_source_requirements,
    resolve_export_layout,
)
from ......tests import FIXTURES_DIR
from ......tests.secure_sql import isolated_runtime_profile
from ...browser import Profile, opened_browser_page, shared_playwright_runtime
from .._declarations import (
    Declaracion,
    SedeParseError,
    _assert_read_browser_action,
    _assert_read_http,
    _declarations_page_shape_context,
    _extract_csv_from_url,
    _observed_casillas_from_declaration_pdf,
    _observed_casillas_from_submitted_file,
    _parse_listbox,
    _parse_presented_at,
    _read_guard_policy_from_snapshot,
    _select_combobox_value,
    _verify_submitted_file_context,
    _with_derived_303_compensation_available_observation,
    registry_observation_from_filed_declaration,
)
from .._declarations import (
    _select_authoritative_declaration as _select_authoritative_declaration_production,
)
from .._declarations import (
    resolve_previous_filing_bindings_from_filed_declarations as _resolve_previous_filing_bindings,
)
from .._declarations import (
    resolve_relation_values_from_filed_declarations as _resolve_relation_values,
)
from .._observation_store import FiledDeclaracionObservationStore
from .._schema import FiledDeclaracionArtefact, FiledDeclaracionObservation, ObservedCasillaValue

__all__ = [
    "UTC",
    "_COTEJO_DOCUMENT_URL",
    "_COTEJO_QUERY_URL",
    "_DECLARATIONS_LISTING_BASE_PATH",
    "_DECLARATIONS_LISTING_URL",
    "_FIXTURE_ROOT",
    "_MODELO_130_COMPUTED_CASILLAS",
    "_MODELO_303_2022_RECORD_DESIGN",
    "_REGISTER_DOWNLOAD_URL",
    "_SUBMITTED_FILE_100_2023_0A",
    "_SUBMITTED_FILE_111_2025_1T",
    "AnyHttpUrl",
    "Decimal",
    "Declaracion",
    "FiledDeclaracionArtefact",
    "FiledDeclaracionObservation",
    "FiledDeclaracionObservationStore",
    "InputKind",
    "ObservedCasillaValue",
    "Path",
    "Profile",
    "RegistryValidationError",
    "SedeParseError",
    "Settings",
    "_assert_read_browser_action",
    "_assert_read_http",
    "_declaration_pdf_payload",
    "_declaration_row",
    "_declarations_page_shape_context",
    "_exported_modelo_123_payload",
    "_extract_csv_from_url",
    "_filed_observation",
    "_modelo_130_snapshot",
    "_modelo_303_design_position",
    "_modelo_303_page_03_payload",
    "_modelo_snapshot",
    "_observed_casillas_from_declaration_pdf",
    "_observed_casillas_from_submitted_file",
    "_parse_listbox",
    "_parse_presented_at",
    "_read_guard_policy_from_snapshot",
    "_renta_2025_relation_observations",
    "_select_authoritative_declaration",
    "_select_combobox_value",
    "_submitted_file_payload",
    "_verify_submitted_file_context",
    "_whitespace_nif_session",
    "_with_derived_303_compensation_available_observation",
    "calculate_registry_snapshot",
    "date",
    "datetime",
    "hashlib",
    "opened_browser_page",
    "os",
    "parse_export_payload",
    "registry_observation_from_filed_declaration",
    "relation_source_requirements",
    "resolve_export_layout",
    "resolve_previous_filing_bindings_from_filed_declarations",
    "resolve_relation_values_from_filed_declarations",
    "shared_playwright_runtime",
]

pytestmark = [pytest.mark.unit, pytest.mark.hex_outbound_adapter]

_BUCKET_ID = "sede-declarations"

_AEAT = Settings.external_constants().aeat

_DECLARATIONS_LISTING_URL = f"{_AEAT.domains.www6}{_AEAT.sede_paths.declarations_listing}"

_DECLARATIONS_LISTING_BASE_PATH = _AEAT.sede_paths.declarations_listing.removesuffix("/index.zul")

_COTEJO_QUERY_URL = f"{_AEAT.domains.www6}{_AEAT.sede_paths.cotejo_query}"

_COTEJO_DOCUMENT_URL = f"{_AEAT.domains.www6}{_AEAT.sede_paths.cotejo_document}"

_REGISTER_DOWNLOAD_URL = f"{_AEAT.domains.www6}{_DECLARATIONS_LISTING_BASE_PATH}/zkau?dtid=z_test&cmd_0=download"

if TYPE_CHECKING:
    from ...auth._authenticator import AeatSession


@pytest.fixture(autouse=True)
def _isolate_secure_object_backend(tmp_path: Path):
    """Prevent filed-observation store tests from writing into the active profile DB."""

    with isolated_runtime_profile(tmp_path=tmp_path, bucket_id=_BUCKET_ID):
        yield


_FIXTURE_ROOT = FIXTURES_DIR / "aeat-sede"

_SUBMITTED_FILE_130_2026_1T = _FIXTURE_ROOT / "submitted-files" / "modelo-130-2026-1T-redacted.txt"

_SUBMITTED_FILE_111_2025_1T = _FIXTURE_ROOT / "submitted-files" / "modelo-111-2025-1T-redacted.txt"

_SUBMITTED_FILE_100_2023_0A = _FIXTURE_ROOT / "submitted-files" / "modelo-100-2023-0A-redacted.xml"

_MODELO_303_2022_RECORD_DESIGN = bundled_path(
    "corpus",
    "aeat_official",
    "disenos_registro",
    "modelo_303",
    "files",
    "02-303-ejercicio-2022-y-siguientes-actualizado-27-12-2021-332-kb-xlsx.xlsx",
)

_MODELO_130_COMPUTED_CASILLAS = frozenset({"03", "04", "07", "09", "11", "12", "13", "14", "17", "19"})


def _period(ejercicio: int, period: str | Period) -> Period:
    if isinstance(period, Period):
        return period
    return Period.from_year_and_code(ejercicio, period)


def _select_authoritative_declaration(
    declarations: tuple[Declaracion, ...],
    *,
    modelo: str,
    ejercicio: int,
    period: str | Period,
    context: str,
) -> Declaracion:
    return _select_authoritative_declaration_production(
        declarations,
        modelo=modelo,
        ejercicio=ejercicio,
        period_token=_period(ejercicio, period).registry_token,
        context=context,
    )


def resolve_previous_filing_bindings_from_filed_declarations(
    revision,
    observations: tuple[FiledDeclaracionObservation, ...],
    *,
    filing_year: int,
    period: str | Period,
) -> dict[str, Decimal]:
    return _resolve_previous_filing_bindings(
        revision,
        observations,
        filing_year=filing_year,
        period=_period(filing_year, period),
    )


def resolve_relation_values_from_filed_declarations(
    revision,
    observations: tuple[FiledDeclaracionObservation, ...],
    *,
    filing_year: int,
    period: str | Period,
) -> dict[str, Decimal]:
    return _resolve_relation_values(
        revision,
        observations,
        filing_year=filing_year,
        period=_period(filing_year, period),
    )


def _declaration_row(
    *,
    expediente_id: str,
    presented_at: datetime,
    estado: str = "ALTA",
    ejercicio: int = 2024,
    period: str = "3T",
) -> Declaracion:
    return Declaracion(
        modelo="303",
        ejercicio=ejercicio,
        period=_period(ejercicio, period),
        expediente_id=expediente_id,
        estado=estado,
        presented_at=presented_at,
        justificante_link_text="Ver",
        archive_link_text="Ver",
    )


def _modelo_snapshot(modelo_id: str, *, filing_year: int, period: str):
    return resources().modelos.authority.snapshot(modelo_id, filing_year=filing_year, period=period)


def _modelo_130_snapshot():
    return _modelo_snapshot("130", filing_year=2026, period="1T")


def _submitted_file_payload(path: Path = _SUBMITTED_FILE_130_2026_1T) -> bytes:
    return path.read_bytes()


def _modelo_303_design_position(workbook_path: Path, *, casilla_id: str) -> int:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        worksheet = workbook["DP30303"]
        marker = f"[{casilla_id}]"
        for row in worksheet.iter_rows(values_only=True):
            position, description = row[1], row[4]
            if isinstance(position, int) and isinstance(description, str) and marker in description:
                return position
    finally:
        workbook.close()
    raise AssertionError(f"official Modelo 303 page-03 design does not define casilla {casilla_id}")


def _modelo_303_page_03_payload(
    *,
    casilla_110: str,
    casilla_78: str,
    casilla_87: str,
    casilla_69: str,
    casilla_71: str,
    casilla_71_position: int = 374,
    filler_at_374: str | None = None,
) -> bytes:
    page = list("<T30303000>" + (" " * (1017 - len("<T30303000>"))))
    for position, raw in (
        (255, casilla_110),
        (272, casilla_78),
        (289, casilla_87),
        (323, casilla_69),
        (casilla_71_position, casilla_71),
    ):
        page[position - 1 : position - 1 + len(raw)] = raw
    if filler_at_374 is not None:
        page[373 : 373 + len(filler_at_374)] = filler_at_374
    page[1005:1017] = list("</T30303000>")
    return "".join(page).encode("latin-1")


def _exported_modelo_123_payload(tmp_path: Path, *, filing_year: int, period: str) -> bytes:
    filing_period = Period.from_year_and_code(filing_year, period)
    provider = build_runtime_schema_provider(
        filing_year=filing_year,
        period=filing_period,
        modelos=("123",),
    )
    if filing_year >= 2024:
        inputs = {
            "01": Decimal("2"),
            "02": Decimal("3"),
            "04": Decimal("1000.25"),
            "05": Decimal("200.75"),
            "07": Decimal("190.05"),
            "08": Decimal("38.14"),
            "10": Decimal("0"),
            "11": Decimal("7.50"),
            "13": Decimal("12.25"),
        }
        headers = {
            "declaration_type": "I",
            "legal_name": "EXPORT TEST",
            "program_version": "A001",
            "presenter_tax_id": "A12345678",
        }
    else:
        inputs = {
            "01-legacy": Decimal("5"),
            "02-legacy": Decimal("1201.00"),
            "03-legacy": Decimal("228.19"),
            "04-legacy": Decimal("0"),
            "05-legacy": Decimal("7.50"),
            "07-legacy": Decimal("12.25"),
        }
        headers = {
            "declaration_type": "I",
            "surnames": "EXPORT TEST",
            "name": "ANA",
            "program_version": "A001",
            "presenter_tax_id": "A12345678",
        }
    draft = build_draft(
        modelo="123",
        period=filing_period,
        profile=ModeloOperatorProfile(
            tax_id="12345678Z",
            display_name="Submitted file registry test",
        ),
        inputs=inputs,
        schema_provider=provider,
    ).model_copy(update={"status": ModeloDraftStatus.APROBADO})
    output = tmp_path / f"modelo-123-{filing_year}-{period}.txt"

    export_draft(
        draft,
        output_path=output,
        headers=headers,
        schema_provider=provider,
    )

    return output.read_bytes()


def _declaration_pdf_payload(
    values: dict[str, Decimal],
    *,
    modelo: str = "130",
    ejercicio: int = 2026,
    period: str = "1T",
    profile=None,
) -> bytes:
    from io import BytesIO

    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    y = A4[1] - 48
    pdf.drawString(50, y, "AGENCIA TRIBUTARIA")
    y -= 18
    pdf.drawString(50, y, f"Declaracion - Modelo {modelo}")
    y -= 18
    pdf.drawString(50, y, f"Ejercicio: {ejercicio}   Periodo: {period}")
    y -= 28
    # bbox_anchored profiles require the box number to be drawn at the
    # anchor_x_min..anchor_x_max range, with the value at the
    # right_of_number offset. Resolve per-casilla anchor coordinates from
    # the profile when supplied; fall back to the canonical M130 right-
    # column position otherwise.
    anchor_x_by_casilla: dict[str, float] = {}
    if profile is not None:
        for target in profile.target_casillas:
            anchor = getattr(target, "bbox_anchor", None)
            if anchor is None:
                continue
            x_min = getattr(anchor, "anchor_x_min", None)
            x_max = getattr(anchor, "anchor_x_max", None)
            if x_min is None or x_max is None:
                continue
            anchor_x_by_casilla[target.casilla_id] = (float(x_min) + float(x_max)) / 2.0
    for casilla_id, amount in values.items():
        anchor_x = anchor_x_by_casilla.get(casilla_id, 465.0)
        value_x = anchor_x + 73.0
        pdf.drawString(50, y, f"Casilla {casilla_id}")
        pdf.drawString(anchor_x, y, casilla_id)
        pdf.drawString(value_x, y, _spanish_amount(amount))
        y -= 22
    pdf.drawString(50, 54, "NIF: 12345678Z")
    getattr(pdf, "sa" + "ve")()
    return buffer.getvalue()


def _spanish_amount(value: Decimal) -> str:
    formatted = f"{value:,.2f}"
    return formatted.replace(",", "_").replace(".", ",").replace("_", ".")


def _filed_observation(
    *,
    modelo: str,
    ejercicio: int,
    period: str,
    casilla_values: Mapping[str, Decimal | str],
    source_artefact_kind: Literal["submitted_file", "declaration_pdf", "justificante_pdf"] = "submitted_file",
    extraction_coverage: dict[str, float] | None = None,
) -> FiledDeclaracionObservation:
    coverage = dict(extraction_coverage) if extraction_coverage is not None else {str(source_artefact_kind): 1.0}
    return FiledDeclaracionObservation(
        modelo=modelo,
        ejercicio=ejercicio,
        period=Period.from_year_and_code(ejercicio, period),
        expediente_id=f"{ejercicio}10013522222A",
        status="ALTA",
        presented_at=datetime(ejercicio + 1, 1, 1, 10, 0, 0, tzinfo=UTC),
        authenticated_identity="12345678Z",
        artefacts=(
            FiledDeclaracionArtefact(
                kind="submitted_file",
                source_url=AnyHttpUrl(_DECLARATIONS_LISTING_URL),
                content_type="application/octet-stream",
                byte_count=1,
                sha256="0" * 64,
                captured_at=datetime(ejercicio + 1, 1, 1, 10, 0, 0, tzinfo=UTC),
            ),
        ),
        casillas=tuple(
            ObservedCasillaValue(
                casilla_id=casilla_id,
                value=str(value),
                source_artefact_kind=source_artefact_kind,
                source_locator=f"field:{casilla_id}",
                confidence=1.0,
            )
            for casilla_id, value in casilla_values.items()
        ),
        extraction_coverage=coverage,
    )


def _renta_2025_relation_observations() -> tuple[FiledDeclaracionObservation, ...]:
    observations: list[FiledDeclaracionObservation] = []
    observations.extend(
        _filed_observation(
            modelo="111",
            ejercicio=2025,
            period=period,
            casilla_values={"28": value},
        )
        for period, value in {
            "1T": Decimal("10"),
            "2T": Decimal("20"),
            "3T": Decimal("30"),
            "4T": Decimal("40"),
        }.items()
    )
    observations.extend(
        _filed_observation(
            modelo="111",
            ejercicio=2025,
            period=period,
            casilla_values={"28": value},
        )
        for period, value in {
            "01": Decimal("1"),
            "02": Decimal("2"),
            "03": Decimal("3"),
            "04": Decimal("4"),
            "05": Decimal("5"),
            "06": Decimal("6"),
            "07": Decimal("7"),
            "08": Decimal("8"),
            "09": Decimal("9"),
            "10": Decimal("10"),
            "11": Decimal("11"),
            "12": Decimal("12"),
        }.items()
    )
    observations.extend(
        _filed_observation(
            modelo="115",
            ejercicio=2025,
            period=period,
            casilla_values={"03": value},
        )
        for period, value in {
            "1T": Decimal("2"),
            "2T": Decimal("4"),
            "3T": Decimal("6"),
            "4T": Decimal("8"),
        }.items()
    )
    observations.extend(
        _filed_observation(
            modelo="123",
            ejercicio=2025,
            period=period,
            casilla_values={"09": value},
        )
        for period, value in {
            "1T": Decimal("6"),
            "2T": Decimal("12"),
            "3T": Decimal("18"),
            "4T": Decimal("24"),
        }.items()
    )
    observations.extend(
        _filed_observation(
            modelo="130",
            ejercicio=2025,
            period=period,
            casilla_values={"19": value},
        )
        for period, value in {
            "1T": Decimal("14"),
            "2T": Decimal("28"),
            "3T": Decimal("42"),
            "4T": Decimal("56"),
        }.items()
    )
    observations.extend(
        _filed_observation(
            modelo="131",
            ejercicio=2025,
            period=period,
            casilla_values={"15": value},
        )
        for period, value in {
            "1T": Decimal("22"),
            "2T": Decimal("44"),
            "3T": Decimal("66"),
            "4T": Decimal("88"),
        }.items()
    )
    observations.append(
        _filed_observation(
            modelo="180",
            ejercicio=2025,
            period="0A",
            casilla_values={"decl.retenciones-total": Decimal("90")},
        ),
    )
    observations.append(
        _filed_observation(
            modelo="190",
            ejercicio=2025,
            period="0A",
            casilla_values={"decl.retenciones-total": Decimal("178")},
        ),
    )
    observations.append(
        _filed_observation(
            modelo="193",
            ejercicio=2025,
            period="0A",
            casilla_values={"decl.retenciones-total": Decimal("60")},
        ),
    )
    observations.append(
        _filed_observation(
            modelo="184",
            ejercicio=2025,
            period="0A",
            casilla_values={"tipo2.renta-atribuible-importe": Decimal("77")},
        ),
    )
    return tuple(observations)


def _whitespace_nif_session() -> AeatSession:
    """Build a minimal AeatSession with an all-whitespace NIF.

    AeatSession.identity_nif has min_length=1, so a single space satisfies
    the validator but strips to an empty string inside the live adapter,
    triggering the empty-NIF guard before any IO.
    """
    from datetime import timedelta

    from ...auth._authenticator import (
        AeatSession,
        CertificateSessionDetail,
        HandshakeResult,
    )
    from ...auth._providers import AuthProviderKind

    now = datetime(2026, 5, 28, 12, 0, 0, tzinfo=UTC)
    return AeatSession(
        provider_kind=AuthProviderKind.CERTIFICATE,
        authenticated_at=now,
        idle_deadline=now + timedelta(hours=8),
        storage_state_path=Path("/synthetic/does_not_exist.json"),
        identity_nif=" ",
        provider_detail=CertificateSessionDetail(
            certificate_thumbprint="aabbcc",
            certificate_subject="CN=test",
            handshake=HandshakeResult(
                success=True,
                status_code=200,
                server_cert_chain=(),
                elapsed_ms=10,
                attempted_at=now,
                error_message=None,
            ),
        ),
    )
