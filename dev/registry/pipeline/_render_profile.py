"""Fail-closed authority for absent record-design numeric wire facts.

Render profiles are reviewed inputs, not inference recipes.  A profile may name
only fixed-record numeric fields whose exact official ``Contenido`` cell is
blank.  It must enumerate that eligible set exactly and is rejected before a
renderer can observe it when any authority, anchor, or representation drifts.
"""

from __future__ import annotations

import json
import re
import unicodedata
from collections.abc import Callable, Iterable, Iterator
from contextlib import contextmanager
from pathlib import Path
from typing import Annotated, Final, Literal

import rtoml
from pydantic import BaseModel, BeforeValidator, ConfigDict, Field, model_validator

from cadrumo.core import is_link_like, iter_directory
from cadrumo.core.hashing import content_hash_hex, sha256_file
from cadrumo.domain.calculations.registry import (
    ABSENT_NATURALEZA_TYPE_CODE,
    ExportValuePolicy,
    ModeloId,
    RegistryValidationError,
    RequiredExportValuePolicyValue,
    SourceRefId,
)

from ._record_design_ir import RecordDesignIntermediateField
from ._semantic_map_join import JoinedRecordDesign

__all__ = [
    "RENDER_PROFILE_SCHEMA_VERSION",
    "OfficialSourceEvidence",
    "RenderProfile",
    "RenderProfileAnchor",
    "RenderProfileDesignIdentity",
    "RenderProfileEligibility",
    "RenderProfileFragment",
    "RenderProfileSourceEvidence",
    "RenderProfileSourceEvidenceEntry",
    "ReviewedEvidence",
    "ReviewedPolicyDecision",
    "SingletonNumericRule",
    "Width17MembershipRule",
    "load_and_validate_render_profile",
    "load_render_profile",
    "load_render_profile_source_evidence",
    "project_render_profile_eligibility",
    "render_profile_digest",
    "validate_render_profile",
    "validate_render_profile_authority",
]


RENDER_PROFILE_SCHEMA_VERSION: Final[int] = 1
_RESERVED_DESCRIPTION_MARKER: Final[str] = "reservado"


class _StrictModel(BaseModel):
    """Frozen authored boundary with unknown keys forbidden."""

    model_config = ConfigDict(extra="forbid", frozen=True, strict=True)


class RenderProfileDesignIdentity(_StrictModel):
    """Identity of the exact official design to which rules apply."""

    modelo: ModeloId
    design_epoch: str = Field(min_length=1)
    source_ref: SourceRefId
    source_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")


def _coerce_ordinal(value: object) -> object:
    """Accept a legacy authored int literal alongside the parser's printed str label.

    Committed render-profile authoring data predates the parser's widened
    ``str | None`` ordinal and still writes bare integers (``ordinal = 14``).
    Coercing here lets that authored data hydrate unchanged.
    """
    if value is None or isinstance(value, str):
        return value
    if isinstance(value, int) and not isinstance(value, bool):
        return str(value)
    raise ValueError("ordinal must be a printed str label, a legacy int literal, or None")


type _AnchorOrdinal = Annotated[str | None, BeforeValidator(_coerce_ordinal)]


class RenderProfileAnchor(_StrictModel):
    """Complete parser-owned field identity; no selector or wildcard exists."""

    sheet: str = Field(min_length=1)
    source_row: int = Field(gt=0)
    #: Defaulted to ``None`` so a PDF anchor is authorable at all, mirroring
    #: :class:`SemanticMapAnchor`, which carries the same field for the same
    #: reason: a workbook design has a stable parser-column cell and a PDF design
    #: has none. The type already permitted ``None``; without a default the key
    #: was still required, and TOML has no way to author an explicit null, so
    #: every PDF anchor refused at load.
    source_cell: str | None = Field(default=None, pattern=r"^[A-Z]+[1-9][0-9]*$")
    #: The ordinal AEAT printed, verbatim -- a str because it is a printed LABEL,
    #: never an arithmetic value. Mirrors
    #: :attr:`domain.calculations.registry.RecordDesignField.ordinal`.
    ordinal: _AnchorOrdinal | None = Field(default=None, min_length=1)
    #: Declares that AEAT printed this row with NO ordinal, mirroring
    #: :attr:`SemanticMapAnchor.ordinal_absent` for the same reason: a row whose
    #: naturaleza AEAT omitted is admitted by a gap fill that cannot invent the
    #: ordinal AEAT never printed, and such a row is exactly the one this profile
    #: must be able to anchor. Kept an EXPLICIT opt-in so omitting both keys
    #: still refuses rather than defaulting into an anchor nothing can match.
    ordinal_absent: bool = False
    record_identity: str = Field(min_length=1)

    @model_validator(mode="after")
    def _require_ordinal_or_declared_absence(self) -> RenderProfileAnchor:
        if self.ordinal_absent and self.ordinal is not None:
            raise ValueError("anchor declares ordinal_absent but also names an ordinal")
        if not self.ordinal_absent and self.ordinal is None:
            raise ValueError(
                "anchor names no ordinal; author the ordinal AEAT printed, or declare "
                "ordinal_absent = true when the design printed the row without one",
            )
        return self


class OfficialSourceEvidence(_StrictModel):
    """A profile conclusion grounded in text read from an exact official cell."""

    authority_kind: Literal["official_source"]
    source_sheet: str = Field(min_length=1)
    source_cell: str = Field(pattern=r"^[A-Z]+[1-9][0-9]*$")
    expected_normalized_statement: str = Field(min_length=1)
    justification: str = Field(min_length=1)

    @model_validator(mode="after")
    def _reject_whitespace_only_review_text(self) -> OfficialSourceEvidence:
        if not self.expected_normalized_statement.strip() or not self.justification.strip():
            raise ValueError("official evidence and justification must contain non-whitespace text")
        return self


class ReviewedPolicyDecision(_StrictModel):
    """An exact-anchor policy decision that makes no claim about source text."""

    authority_kind: Literal["reviewed_policy"]
    decision_id: str = Field(min_length=1, pattern=r"^[a-z0-9][a-z0-9-]*$")
    governed_anchor: RenderProfileAnchor
    decision_statement: str = Field(min_length=1)
    justification: str = Field(min_length=1)

    @model_validator(mode="after")
    def _reject_whitespace_only_review_text(self) -> ReviewedPolicyDecision:
        if not self.decision_statement.strip() or not self.justification.strip():
            raise ValueError("reviewed policy and justification must contain non-whitespace text")
        return self


ReviewedEvidence = Annotated[
    OfficialSourceEvidence | ReviewedPolicyDecision,
    Field(discriminator="authority_kind"),
]


class RenderProfileSourceEvidenceEntry(_StrictModel):
    """Actual normalized text read from one exact cell in the verified source."""

    sheet: str = Field(min_length=1)
    cell: str = Field(pattern=r"^[A-Z]+[1-9][0-9]*$")
    normalized_statement: str = Field(min_length=1)

    @model_validator(mode="after")
    def _reject_whitespace_only_statement(self) -> RenderProfileSourceEvidenceEntry:
        if not self.normalized_statement.strip():
            raise ValueError("source evidence statement must contain non-whitespace text")
        return self


class RenderProfileSourceEvidence(_StrictModel):
    """Independent exact-cell evidence extracted from one verified official design."""

    design_identity: RenderProfileDesignIdentity
    entries: tuple[RenderProfileSourceEvidenceEntry, ...]

    @model_validator(mode="after")
    def _require_unique_locators(self) -> RenderProfileSourceEvidence:
        locators = tuple((entry.sheet, entry.cell) for entry in self.entries)
        duplicates = _duplicates(locators)
        if duplicates:
            raise ValueError(f"source evidence contains duplicate exact locators: {duplicates!r}")
        return self


class RenderProfileEligibility(_StrictModel):
    """Production-owned partition of otherwise-unrenderable fixed numeric fields."""

    all_fields: tuple[RecordDesignIntermediateField, ...]
    width_17_fields: tuple[RecordDesignIntermediateField, ...]
    smaller_fields: tuple[RecordDesignIntermediateField, ...]


class Width17MembershipRule(_StrictModel):
    """Reviewed enumeration of width-17 amount anchors of one AEAT type."""

    rule_kind: Literal["width_17_membership"]
    aeat_type: Literal["Num", "N"]
    integer_digits: Literal[15, 14]
    decimal_digits: Literal[2]
    sign_policy: Literal["unsigned", "n-prefix-negative-blank-nonnegative"]
    anchors: tuple[RenderProfileAnchor, ...] = Field(min_length=1)
    evidence: ReviewedEvidence

    @model_validator(mode="after")
    def _require_explicit_type_specific_representation(self) -> Width17MembershipRule:
        if not isinstance(self.evidence, OfficialSourceEvidence):
            raise ValueError("width-17 membership requires verified official-source evidence")
        expected = (15, "unsigned") if self.aeat_type == "Num" else (14, "n-prefix-negative-blank-nonnegative")
        if (self.integer_digits, self.sign_policy) != expected:
            raise ValueError(f"{self.aeat_type} width-17 representation conflicts with its explicit sign policy")
        return self


class SingletonNumericRule(_StrictModel):
    """One reviewed smaller numeric field; grouping is deliberately impossible."""

    rule_kind: Literal["singleton_numeric"]
    aeat_type: Literal["Num"]
    semantic_kind: Literal[
        "integer",
        "decimal",
        "date_yyyymmdd",
        "date_ddmmyyyy",
        "enumeration",
        "percentage_decimal",
        "digit_string",
        "identifier_digits",
        "checkbox",
        "year_yyyy",
        "year_last_two_digits",
        "month_mm",
        "day_dd",
        "mistyped_alphanumeric_text",
    ]
    value_policy: RequiredExportValuePolicyValue
    integer_digits: int = Field(ge=0)
    decimal_digits: int = Field(ge=0)
    sign_policy: Literal["unsigned"]
    allowed_values: tuple[str, ...]
    anchor: RenderProfileAnchor
    evidence: ReviewedEvidence

    @model_validator(mode="after")
    def _require_kind_specific_declaration(self) -> SingletonNumericRule:
        if self.semantic_kind in {"enumeration", "checkbox"}:
            if not self.allowed_values or any(
                not item or not item.isascii() or not item.isdigit() or str(int(item)) != item
                for item in self.allowed_values
            ):
                raise ValueError(
                    f"{self.semantic_kind} rules require explicit canonical ASCII-integer allowed_values",
                )
            if len(set(self.allowed_values)) != len(self.allowed_values):
                raise ValueError("enumeration allowed_values must be unique")
        elif self.allowed_values:
            raise ValueError("allowed_values must be explicitly empty outside an enumeration rule")
        required_policy = {
            "integer": ExportValuePolicy.UNSIGNED_INTEGER,
            "decimal": ExportValuePolicy.IMPLIED_DECIMAL,
            "percentage_decimal": ExportValuePolicy.IMPLIED_DECIMAL,
            "date_yyyymmdd": ExportValuePolicy.YYYYMMDD,
            "date_ddmmyyyy": ExportValuePolicy.DDMMYYYY,
            "enumeration": ExportValuePolicy.ENUMERATED_DIGITS,
            "digit_string": ExportValuePolicy.DIGIT_STRING,
            "identifier_digits": ExportValuePolicy.IDENTIFIER_DIGITS,
            "checkbox": ExportValuePolicy.SELECTED_1_UNSELECTED_0,
            "year_yyyy": ExportValuePolicy.FOUR_DIGIT_YEAR,
            "year_last_two_digits": ExportValuePolicy.FOUR_DIGIT_YEAR_FINAL_TWO_DIGITS,
            "month_mm": ExportValuePolicy.TWO_DIGIT_MONTH,
            "day_dd": ExportValuePolicy.TWO_DIGIT_DAY,
            "mistyped_alphanumeric_text": ExportValuePolicy.MISTYPED_ALPHANUMERIC_TEXT,
        }[self.semantic_kind]
        if self.value_policy != required_policy:
            raise ValueError(f"{self.semantic_kind} requires value_policy {required_policy!r}")
        if self.semantic_kind in {"date_yyyymmdd", "date_ddmmyyyy"} and (
            self.integer_digits,
            self.decimal_digits,
        ) != (8, 0):
            raise ValueError(f"{self.semantic_kind} requires exactly 8 integer digits and 0 decimal digits")
        if self.semantic_kind == "integer" and (self.integer_digits <= 0 or self.decimal_digits != 0):
            raise ValueError("integer requires positive integer digits and 0 decimal digits")
        if self.semantic_kind in {"decimal", "percentage_decimal"} and (
            self.integer_digits <= 0 or self.decimal_digits <= 0
        ):
            raise ValueError(f"{self.semantic_kind} requires positive integer and decimal digits")
        if self.semantic_kind in {"digit_string", "identifier_digits"} and (
            self.integer_digits <= 0 or self.decimal_digits != 0
        ):
            raise ValueError(f"{self.semantic_kind} requires positive integer digits and 0 decimal digits")
        if self.semantic_kind == "enumeration" and self.decimal_digits != 0:
            raise ValueError("enumeration requires 0 decimal digits")
        if self.semantic_kind == "enumeration" and any(len(item) > self.integer_digits for item in self.allowed_values):
            raise ValueError("enumeration allowed_values must fit the declared integer width")
        exact_shapes = {
            # Zero digits in both positions, because the slot has no numeric
            # reading at all. Declaring a width here would re-assert the very
            # naturaleza this kind exists to contradict.
            "mistyped_alphanumeric_text": (0, 0),
            "checkbox": (1, 0),
            "year_yyyy": (4, 0),
            "year_last_two_digits": (2, 0),
            "month_mm": (2, 0),
            "day_dd": (2, 0),
        }
        expected_shape = exact_shapes.get(self.semantic_kind)
        if expected_shape is not None and (self.integer_digits, self.decimal_digits) != expected_shape:
            raise ValueError(f"{self.semantic_kind} requires exactly {expected_shape!r} integer/decimal digits")
        if self.semantic_kind == "checkbox" and self.allowed_values != ("0", "1"):
            raise ValueError("checkbox requires allowed_values ('0', '1')")
        if isinstance(self.evidence, ReviewedPolicyDecision) and self.evidence.governed_anchor != self.anchor:
            raise ValueError("reviewed policy must name the exact governed anchor")
        return self


class RenderProfileFragment(_StrictModel):
    """One deterministic, independently reviewable profile fragment."""

    schema_version: Literal[1]
    fragment_id: str = Field(min_length=1, pattern=r"^[a-z0-9][a-z0-9-]*$")
    design_identity: RenderProfileDesignIdentity
    width_17_rules: tuple[Width17MembershipRule, ...]
    singleton_rules: tuple[SingletonNumericRule, ...]

    @model_validator(mode="after")
    def _require_authored_rules(self) -> RenderProfileFragment:
        if not self.width_17_rules and not self.singleton_rules:
            raise ValueError("render profile fragments must contain at least one authored rule")
        return self


class RenderProfile(_StrictModel):
    """Deterministically compiled complete profile for one official design."""

    schema_version: Literal[1]
    design_identity: RenderProfileDesignIdentity
    fragment_ids: tuple[str, ...]
    width_17_rules: tuple[Width17MembershipRule, ...]
    singleton_rules: tuple[SingletonNumericRule, ...]

    @model_validator(mode="after")
    def _require_unique_fragment_ids(self) -> RenderProfile:
        duplicate_ids = _duplicates(self.fragment_ids)
        if duplicate_ids:
            raise ValueError(f"render profile contains duplicate fragment ids: {duplicate_ids!r}")
        return self


def load_and_validate_render_profile(
    profile_directory: Path,
    joined: JoinedRecordDesign,
    source_evidence: RenderProfileSourceEvidence,
) -> RenderProfile:
    """Load sorted TOML fragments and validate exact coverage against parser IR."""
    profile = load_render_profile(profile_directory)
    validate_render_profile(profile, joined, source_evidence)
    return profile


def load_render_profile(profile_directory: Path) -> RenderProfile:
    """Load sorted TOML fragments without weakening their strict authored schema."""
    if not profile_directory.is_dir() or profile_directory.is_symlink() or profile_directory.is_junction():
        raise RegistryValidationError(f"render profile path must be a real directory: {profile_directory}")
    try:
        paths = tuple(sorted(iter_directory(profile_directory, require_root=True), key=lambda path: path.name))
    except OSError as exc:
        raise RegistryValidationError(f"cannot inspect render profile directory: {profile_directory}") from exc
    if not paths:
        raise RegistryValidationError(f"render profile directory contains no TOML fragments: {profile_directory}")
    non_fragments = tuple(
        path.name
        for path in paths
        if path.suffix.casefold() != ".toml" or path.is_symlink() or path.is_junction() or not path.is_file()
    )
    if non_fragments:
        raise RegistryValidationError(
            "render profile directory accepts only regular TOML fragments; "
            f"refusing non-profile entries: {non_fragments!r}",
        )
    fragments: list[RenderProfileFragment] = []
    for path in paths:
        if path.is_symlink() or path.is_junction() or not path.is_file():
            raise RegistryValidationError(f"render profile fragment must be a regular file: {path}")
        try:
            fragments.append(RenderProfileFragment.model_validate_json(json.dumps(rtoml.load(path))))
        except (OSError, ValueError, TypeError) as exc:
            raise RegistryValidationError(f"invalid render profile fragment {path.name!r}: {exc}") from exc
    profile = _compile_fragments(fragments)
    return profile


#: An `A1`-style reference, which is the only shape `OfficialSourceEvidence`
#: admits. openpyxl resolves it natively; xlrd is index-addressed, so the legacy
#: reader below converts it.
_CELL_REFERENCE_RE: Final[re.Pattern[str]] = re.compile(r"^(?P<column>[A-Z]+)(?P<row>[1-9][0-9]*)$")


def _cell_indices(cell: str) -> tuple[int, int]:
    match = _CELL_REFERENCE_RE.fullmatch(cell)
    if match is None:
        raise RegistryValidationError(f"render profile evidence locator is not an A1 cell reference: {cell!r}")
    column = 0
    for character in match.group("column"):
        column = column * 26 + (ord(character) - ord("A") + 1)
    return int(match.group("row")) - 1, column - 1


@contextmanager
def _ooxml_cell_reader(source_path: Path) -> Iterator[Callable[[str, str], object]]:
    from openpyxl import load_workbook

    workbook = load_workbook(source_path, read_only=True, data_only=True)
    try:

        def read(sheet_name: str, cell: str) -> object:
            if sheet_name not in workbook.sheetnames:
                raise RegistryValidationError(f"render profile evidence sheet does not exist: {sheet_name!r}")
            return workbook[sheet_name][cell].value

        yield read
    finally:
        workbook.close()


@contextmanager
def _legacy_xls_cell_reader(source_path: Path) -> Iterator[Callable[[str, str], object]]:
    """Read evidence cells out of a legacy binary XLS design.

    AEAT still publishes many diseños de registro in the pre-OOXML format, and
    the record-design parser already reads them through xlrd. Refusing them only
    here would leave every such modelo's render profile unauthorable while its
    design parses perfectly, so the evidence path reads exactly the same binaries
    the rest of the pipeline does.
    """
    import xlrd

    workbook = xlrd.open_workbook(str(source_path), on_demand=True)
    try:

        def read(sheet_name: str, cell: str) -> object:
            if sheet_name not in workbook.sheet_names():
                raise RegistryValidationError(f"render profile evidence sheet does not exist: {sheet_name!r}")
            sheet = workbook.sheet_by_name(sheet_name)
            row, column = _cell_indices(cell)
            if row >= sheet.nrows or column >= sheet.ncols:
                raise RegistryValidationError(
                    f"render profile evidence locator does not exist: {(sheet_name, cell)!r}",
                )
            return sheet.cell_value(row, column)

        yield read
    finally:
        workbook.release_resources()


def load_render_profile_source_evidence(
    source_path: Path,
    profile: RenderProfile,
) -> RenderProfileSourceEvidence:
    """Read every claimed official cell from the hash-verified binary itself."""
    if not source_path.is_file() or is_link_like(source_path):
        raise RegistryValidationError(f"render profile source must be a regular file: {source_path}")
    actual_sha256 = sha256_file(source_path)
    if actual_sha256 != profile.design_identity.source_sha256:
        raise RegistryValidationError(
            "render profile source binary SHA-256 does not match the exact design identity",
        )
    suffix = source_path.suffix.lower()
    if suffix not in {".xlsx", ".xlsm", ".xls"}:
        raise RegistryValidationError(
            f"render profile source evidence requires a spreadsheet workbook: {source_path}",
        )

    official_evidence = tuple(
        evidence
        for evidence in (
            *(rule.evidence for rule in profile.width_17_rules),
            *(rule.evidence for rule in profile.singleton_rules),
        )
        if isinstance(evidence, OfficialSourceEvidence)
    )
    if not official_evidence:
        raise RegistryValidationError("render profile contains no official-source evidence to resolve")
    locators = tuple(dict.fromkeys((item.source_sheet, item.source_cell) for item in official_evidence))
    read_cell = _legacy_xls_cell_reader(source_path) if suffix == ".xls" else _ooxml_cell_reader(source_path)
    with read_cell as cell_value:
        entries: list[RenderProfileSourceEvidenceEntry] = []
        for sheet_name, cell in locators:
            value = cell_value(sheet_name, cell)
            normalized = _normalize_source_statement(value)
            if not normalized:
                raise RegistryValidationError(
                    f"render profile evidence locator does not contain source text: {(sheet_name, cell)!r}",
                )
            entries.append(
                RenderProfileSourceEvidenceEntry(
                    sheet=sheet_name,
                    cell=cell,
                    normalized_statement=normalized,
                ),
            )
    return RenderProfileSourceEvidence(
        design_identity=profile.design_identity,
        entries=tuple(entries),
    )


def validate_render_profile(
    profile: RenderProfile,
    joined: JoinedRecordDesign,
    source_evidence: RenderProfileSourceEvidence,
) -> None:
    """Refuse every identity, membership, coverage, and representation conflict."""
    expected_identity = RenderProfileDesignIdentity(
        modelo=joined.modelo,
        design_epoch=joined.source.design_epoch,
        source_ref=joined.source.source_ref,
        source_sha256=joined.source.source_sha256,
    )
    eligibility = project_render_profile_eligibility(joined_field.parser_field for joined_field in joined.fields)
    validate_render_profile_authority(profile, expected_identity, eligibility, source_evidence)


def validate_render_profile_authority(
    profile: RenderProfile,
    expected_identity: RenderProfileDesignIdentity,
    eligibility: RenderProfileEligibility,
    source_evidence: RenderProfileSourceEvidence,
) -> None:
    """Validate a complete profile against source-owned identity and eligibility."""
    if profile.design_identity != expected_identity:
        raise RegistryValidationError(
            f"render profile identity {profile.design_identity!r} does not match exact official design "
            f"{expected_identity!r}",
        )
    if source_evidence.design_identity != expected_identity:
        raise RegistryValidationError(
            "render profile source evidence does not match the exact official design identity",
        )
    _validate_reviewed_evidence(profile, source_evidence)

    eligible = {_field_anchor(field): field for field in eligibility.all_fields}
    governed = tuple(anchor for rule in profile.width_17_rules for anchor in rule.anchors) + tuple(
        rule.anchor for rule in profile.singleton_rules
    )
    duplicates = _duplicates(governed)
    if duplicates:
        raise RegistryValidationError(f"render profile contains duplicate or overlapping exact anchors: {duplicates!r}")
    governed_set = set(governed)
    eligible_set = set(eligible)
    if governed_set != eligible_set:
        missing = sorted(eligible_set - governed_set, key=_anchor_key)
        unknown = sorted(governed_set - eligible_set, key=_anchor_key)
        raise RegistryValidationError(
            "render profile must cover exactly the eligible blank numeric fields; "
            f"missing={missing!r}, unknown={unknown!r}",
        )

    width_17_types = tuple(rule.aeat_type for rule in profile.width_17_rules)
    duplicate_width_17_types = _duplicates(width_17_types)
    if duplicate_width_17_types:
        raise RegistryValidationError(
            "render profile splits one width-17 AEAT type across multiple membership rules: "
            f"{duplicate_width_17_types!r}",
        )
    eligible_width_17_types = {field.aeat_type for field in eligibility.width_17_fields}
    if set(width_17_types) != eligible_width_17_types:
        raise RegistryValidationError(
            "render profile must declare one explicit width-17 membership rule for each eligible AEAT type",
        )

    for rule in profile.width_17_rules:
        for anchor in rule.anchors:
            field = eligible[anchor]
            if field.length != 17 or field.aeat_type != rule.aeat_type:
                raise RegistryValidationError(
                    f"width-17 {rule.aeat_type} membership conflicts with official field at {anchor!r}",
                )
    for rule in profile.singleton_rules:
        field = eligible[rule.anchor]
        # Compared on NUMERIC-NESS, not on the spelling. A singleton rule pins
        # ``aeat_type = "Num"`` and ``sign_policy = "unsigned"`` by construction,
        # so its type token carries no information beyond "numeric, unsigned" --
        # the signed ``N`` form only ever appears in a width-17 membership rule,
        # which is matched exactly above and keeps its Num/N distinction. String
        # equality here would refuse every PDF design, whose naturaleza the
        # parser canonicalises to ``Numerico``, leaving those fields eligible for
        # a rule that could never be written.
        # An absent naturaleza is admitted for the same reason it is eligible:
        # AEAT printed the type cell EMPTY, so there is no naturaleza to agree
        # with, and the reviewed rule is the only thing that can state the wire
        # representation. Refusing here would make that field eligible for a
        # rule the validator then rejects, which is the contradiction the
        # comment above warns about, one case further along.
        if field.length == 17 or not (
            _is_numeric_aeat_type(field.aeat_type) or _has_absent_naturaleza(field)
        ):
            raise RegistryValidationError(f"smaller singleton rule conflicts with official field at {rule.anchor!r}")
        # A digit budget is checked against the slot for every kind that HAS
        # one. The mistyped-alphanumeric kind has none by construction -- it
        # exists to say the slot carries no number -- so requiring its zeros to
        # equal a 100-position width would refuse the only rule that can be
        # written for it, the same contradiction the eligibility comments above
        # guard against.
        if rule.semantic_kind != "mistyped_alphanumeric_text" and (
            rule.integer_digits + rule.decimal_digits != field.length
        ):
            raise RegistryValidationError(
                f"smaller singleton representation width conflicts with official length at {rule.anchor!r}",
            )
        if any(len(item) > field.length for item in rule.allowed_values):
            raise RegistryValidationError(
                f"enumeration value width conflicts with official length at {rule.anchor!r}",
            )


def render_profile_digest(
    profile: RenderProfile,
    source_evidence: RenderProfileSourceEvidence,
) -> str:
    """Digest every reviewed profile fact and resolved source-evidence fact.

    Authored fragment, rule, membership, allowed-value, and evidence ordering is
    deliberately irrelevant. Exact anchors and every policy/evidence payload
    remain part of the digest.
    """
    if source_evidence.design_identity != profile.design_identity:
        raise RegistryValidationError(
            "render profile source evidence does not match the profile design identity",
        )
    width_rules: list[dict[str, object]] = [
        {
            **rule.model_dump(mode="json", exclude={"anchors"}),
            "anchors": [anchor.model_dump(mode="json") for anchor in sorted(rule.anchors, key=_anchor_key)],
        }
        for rule in sorted(profile.width_17_rules, key=lambda item: item.aeat_type)
    ]
    singleton_rules: list[dict[str, object]] = [
        {
            **rule.model_dump(mode="json", exclude={"allowed_values"}),
            "allowed_values": sorted(rule.allowed_values),
        }
        for rule in sorted(profile.singleton_rules, key=lambda item: _anchor_key(item.anchor))
    ]
    evidence_entries = [
        entry.model_dump(mode="json")
        for entry in sorted(source_evidence.entries, key=lambda item: (item.sheet, item.cell))
    ]
    return content_hash_hex(
        {
            "schema_version": profile.schema_version,
            "design_identity": profile.design_identity.model_dump(mode="json"),
            "fragment_ids": sorted(profile.fragment_ids),
            "width_17_rules": width_rules,
            "singleton_rules": singleton_rules,
            "source_evidence": {
                "design_identity": source_evidence.design_identity.model_dump(mode="json"),
                "entries": evidence_entries,
            },
        },
    )


def _is_source_reserved_field(field: RecordDesignIntermediateField) -> bool:
    """Report whether the official description marks this slot reserved.

    Reservation is read from the DESCRIPTION and never from the type column,
    because the type column is unreliable for exactly these slots.  When AEAT
    retires a numeric field it rewrites the description to mark the slot
    reserved and leaves the original numeric type in place, so a retired
    quantity keeps announcing itself as ``Num`` forever.  Modelo 303 shows both
    states of that edit within one modelo: its 2023 design types every reserved
    slot ``An``, while its 2025 design types five of thirteen ``Num``, and four
    of those five sit exactly where live employee-count quantities stood before
    they were retired.  The surviving ``Num`` is therefore a residue of the
    slot's former life, not a claim that a reserved run carries a number.

    The match is case-insensitive because official designs disagree on casing
    for the same marker.
    """
    return _RESERVED_DESCRIPTION_MARKER in field.normalized_description.casefold()


def _has_absent_naturaleza(field: RecordDesignIntermediateField) -> bool:
    """Whether AEAT printed this row's naturaleza cell EMPTY.

    The strongest case for a reviewed rule there is, and it was excluded. A
    numeric field at least tells the renderer it is numeric; a row AEAT printed
    with no naturaleza at all states nothing about its wire representation, and
    the shipped parser records exactly that by stamping the absent-naturaleza
    marker rather than guessing a type. Modelo 184's ``151-155 PORCENTAJE DE
    RENTA ATRIBUIBLE A MIEMBROS RESIDENTES`` is the worked case: its own prose
    says "campo numerico" and subdivides it into ENTERO and DECIMAL, so the wire
    fact is knowable and reviewable -- but it lives in the description, which is
    exactly the evidence a render profile exists to carry.

    Excluding these did not make the renderer safe: it made the field invisible
    to the profile's exhaustive-coverage check and then crashed the renderer on
    an unsupported type, so the gap surfaced as a late refusal instead of as the
    reviewable rule it should have demanded.
    """
    return field.aeat_type.strip() == ABSENT_NATURALEZA_TYPE_CODE


def _is_numeric_aeat_type(aeat_type: str) -> bool:
    """Whether ``aeat_type`` names a numeric naturaleza, however AEAT spelled it.

    A workbook prints the abbreviation (``Num``, and ``N`` for the signed form);
    a PDF design prints the word, which the shipped parser canonicalises to
    ``Numerico``. Selecting on the abbreviations alone made every PDF design's
    numeric fields ineligible, so no reviewed rule was ever demanded for them and
    an empty profile satisfied exhaustive coverage completely.

    Matched on an accent-stripped stem for the same reason
    ``_naturaleza_or_none`` is: AEAT does not spell consistently, and every
    unmatched spelling is a field that silently escapes review.
    """
    normalised = unicodedata.normalize("NFKD", aeat_type.strip(" .")).encode("ascii", "ignore").decode("ascii").lower()
    return normalised in {"num", "n"} or normalised.startswith("numeric")


def _states_no_wire_fact(field: RecordDesignIntermediateField) -> bool:
    """Whether the design left this field's wire fact unstated at its anchor.

    A WORKBOOK field has a Contenido cell, so a non-blank one is the design
    stating the fact and the field needs no reviewed rule. A PDF design has no
    such column: the parser fills ``content`` with the field's DESCRIPTIVE PROSE,
    which states the fact sometimes, partially, or not at all. Modelo 347 carries
    all of those side by side -- one field giving sign and decimals in full, one
    giving only a width, a bare cross-reference, and a purely semantic
    description -- so no rule over the text can separate them and every numeric
    PDF anchor is put to a reviewed rule instead.

    The source shape is read off the field rather than threaded in: a workbook
    anchor carries a ``source_cell`` and a PDF anchor does not, which is the
    distinction :class:`RecordDesignIntermediateField` already documents.

    Where the prose DOES state a fact the reviewed rule must agree with it, which
    keeps the official design's veto intact; that agreement is checked where the
    rule's own evidence is validated, not here.
    """
    if field.source_cell is None:
        return True
    return field.content is None or not field.content.strip()


def project_render_profile_eligibility(
    fixed_fields: Iterable[RecordDesignIntermediateField],
) -> RenderProfileEligibility:
    """Partition fixed joined fields eligible for reviewed absent-wire authority.

    A source-reserved slot is never eligible.  A render profile exists to state
    a wire fact the official design left unstated, and a reserved run has no
    wire fact beyond being filler, so admitting one would force an author to
    model numeric meaning onto a slot that carries none.
    """
    eligible = tuple(
        field
        for field in fixed_fields
        if (_is_numeric_aeat_type(field.aeat_type) or _has_absent_naturaleza(field))
        and _states_no_wire_fact(field)
        and not _is_source_reserved_field(field)
    )
    return RenderProfileEligibility(
        all_fields=eligible,
        width_17_fields=tuple(field for field in eligible if field.length == 17),
        smaller_fields=tuple(field for field in eligible if field.length != 17),
    )


def _validate_reviewed_evidence(
    profile: RenderProfile,
    source_evidence: RenderProfileSourceEvidence,
) -> None:
    actual_by_locator = {(entry.sheet, entry.cell): entry.normalized_statement for entry in source_evidence.entries}
    reviewed = tuple(rule.evidence for rule in profile.width_17_rules) + tuple(
        rule.evidence for rule in profile.singleton_rules
    )
    for evidence in reviewed:
        if isinstance(evidence, ReviewedPolicyDecision):
            continue
        locator = evidence.source_sheet, evidence.source_cell
        actual_statement = actual_by_locator.get(locator)
        if actual_statement is None:
            raise RegistryValidationError(f"render profile evidence locator does not exist: {locator!r}")
        if actual_statement != evidence.expected_normalized_statement:
            raise RegistryValidationError(
                f"render profile evidence statement does not match verified source at {locator!r}",
            )


def _compile_fragments(fragments: Iterable[RenderProfileFragment]) -> RenderProfile:
    ordered = tuple(fragments)
    if not ordered:
        raise RegistryValidationError("render profile requires at least one fragment")
    ids = tuple(fragment.fragment_id for fragment in ordered)
    duplicate_ids = _duplicates(ids)
    if duplicate_ids:
        raise RegistryValidationError(f"render profile contains duplicate fragment ids: {duplicate_ids!r}")
    design_identity = ordered[0].design_identity
    mismatched = tuple(fragment.fragment_id for fragment in ordered if fragment.design_identity != design_identity)
    if mismatched:
        raise RegistryValidationError(f"render profile fragments have inapplicable design identities: {mismatched!r}")
    width_rules = _compile_width_17_rules(rule for fragment in ordered for rule in fragment.width_17_rules)
    return RenderProfile(
        schema_version=RENDER_PROFILE_SCHEMA_VERSION,
        design_identity=design_identity,
        fragment_ids=ids,
        width_17_rules=width_rules,
        singleton_rules=tuple(rule for fragment in ordered for rule in fragment.singleton_rules),
    )


def _compile_width_17_rules(rules: Iterable[Width17MembershipRule]) -> tuple[Width17MembershipRule, ...]:
    by_type: dict[str, list[Width17MembershipRule]] = {}
    for rule in rules:
        by_type.setdefault(rule.aeat_type, []).append(rule)
    compiled: list[Width17MembershipRule] = []
    for aeat_type in ("Num", "N"):
        type_rules = by_type.get(aeat_type, [])
        if not type_rules:
            continue
        authority = type_rules[0]
        if any(
            (
                rule.integer_digits,
                rule.decimal_digits,
                rule.sign_policy,
                rule.evidence,
            )
            != (
                authority.integer_digits,
                authority.decimal_digits,
                authority.sign_policy,
                authority.evidence,
            )
            for rule in type_rules[1:]
        ):
            raise RegistryValidationError(
                f"render profile fragments conflict on width-17 {aeat_type} authority",
            )
        compiled.append(
            authority.model_copy(
                update={"anchors": tuple(anchor for rule in type_rules for anchor in rule.anchors)},
            ),
        )
    return tuple(compiled)


def _field_anchor(field: RecordDesignIntermediateField) -> RenderProfileAnchor:
    # ``ordinal_absent`` is DERIVED here rather than authored: this anchor is
    # built from the parser field itself, so a missing ordinal is an observed
    # fact about the design and not an authoring claim. The authored side keeps
    # the explicit declaration, which is what the comparison below is against.
    return RenderProfileAnchor(
        sheet=field.sheet,
        source_row=field.source_row,
        source_cell=field.source_cell,
        ordinal=field.ordinal,
        ordinal_absent=field.ordinal is None,
        record_identity=field.record_identity,
    )


def _duplicates[T](values: Iterable[T]) -> tuple[T, ...]:
    seen: set[T] = set()
    duplicates: set[T] = set()
    for value in values:
        if value in seen:
            duplicates.add(value)
        seen.add(value)
    return tuple(sorted(duplicates, key=repr))


def _anchor_key(anchor: RenderProfileAnchor) -> tuple[str, int, str, str, str]:
    """Return a deterministic, total sort key -- presentation order, not AEAT order.

    Plain string ordering on ``ordinal`` is fine here: every use is a stable,
    reproducible listing (an error message, a persisted artefact), never an
    AEAT-numeric position a downstream index depends on.
    """
    return (
        anchor.sheet,
        anchor.source_row,
        anchor.ordinal or "",
        anchor.source_cell or "",
        anchor.record_identity,
    )


def _normalize_source_statement(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).split())
