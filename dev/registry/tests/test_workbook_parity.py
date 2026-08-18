"""Tests for official AEAT workbook parity backend infrastructure.

These tests drive LibreOffice directly to convert binary ``.xls`` workbooks and
to recalculate formulas, so they need an external tool the project's dependency
set does not install. They therefore carry ``external_tool`` alongside ``unit``,
and the default lane selects ``unit and not external_tool`` -- so the marker,
not the path ``--ignore``, is what holds them out. The ``--ignore`` beside it is
belt-and-braces. ``unit`` is retained because the marker-integrity contract
requires exactly one of ``{unit, integration, aeat_live}`` on every item.

Run them through ``just test-workbook-parity``.
"""

from __future__ import annotations

from datetime import date
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from pydantic import ValidationError

from cadrumo.core import CasillaId, validated_casilla_id
from cadrumo.core.resources import bundled_path
from cadrumo.domain.calculations.registry import (
    RegistrySnapshot,
    RegistryValidationError,
    build_snapshot,
)
from cadrumo.domain.calculations.registry.tests._registry_schema_support import _committed_modelo

from ..parity._parity_tapes import ParityScenario
from ..parity._workbook_parity import (
    SyntheticInputSet,
    SyntheticInputValue,
    WorkbookCellRef,
    WorkbookScanOptions,
    _BinaryXlsConversionError,
    assert_workbook_scan_clean,
    compare_registry_to_workbook,
    convert_binary_xls_with_libreoffice,
    converted_binary_xls_with_libreoffice,
    discover_workbooks,
    inventory_workbook_coverage,
    run_registry_workbook_parity,
    run_workbook_with_libreoffice,
    scan_workbook,
    verify_workbook_backend,
)

pytestmark = [pytest.mark.unit, pytest.mark.external_tool, pytest.mark.hex_domain]

_M130_CASILLA_19: CasillaId = validated_casilla_id("19", surface="_M130_CASILLA_19")


def _write_formula_workbook(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.title = "Modelo"
    worksheet["A1"] = Decimal("10")
    worksheet["A2"] = Decimal("21")
    worksheet["B1"] = "=A1+A2"
    worksheet["B2"] = "=SUM(A1:A2)"
    workbook.save(path)


def _write_static_workbook(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.title = "Modelo"
    worksheet["A1"] = "Casilla"
    worksheet["B1"] = "Descripcion"
    worksheet["A2"] = "001"
    worksheet["B2"] = "Importe"
    workbook.save(path)


def _committed_modelo_130_snapshot() -> RegistrySnapshot:
    modelo, catalogues = _committed_modelo("130")
    return build_snapshot(
        modelo,
        catalogues,
        source_root=bundled_path(),
        filing_year=2026,
        period="1T",
    )


def test_scan_workbook_discovers_xlsx_formula_cells(tmp_path: Path) -> None:
    workbook_path = tmp_path / "modelo_303" / "files" / "303-test.xlsx"
    _write_formula_workbook(workbook_path)

    report = scan_workbook(workbook_path, root=tmp_path, options=WorkbookScanOptions(per_file_timeout_seconds=5))

    assert report.path == "modelo_303/files/303-test.xlsx"
    assert report.modelo == "303"
    assert report.extension == ".xlsx"
    assert report.scan_status == "scanned"
    assert report.workbook_kind == "formula_form"
    assert report.evidence_tier == "executable_parity_evidence"
    assert "layout_authority" in report.not_evidence_for
    assert report.formula_cells == 2
    assert {cell.coordinate for cell in report.output_candidates} == {"B1", "B2"}
    assert {cell.coordinate for cell in report.input_candidates} >= {"A1", "A2"}


def test_scan_workbook_classifies_static_layout_as_layout_authority(tmp_path: Path) -> None:
    workbook_path = tmp_path / "modelo_151" / "files" / "151-static-layout.xlsx"
    _write_static_workbook(workbook_path)

    report = scan_workbook(workbook_path, root=tmp_path, options=WorkbookScanOptions(per_file_timeout_seconds=5))

    assert report.path == "modelo_151/files/151-static-layout.xlsx"
    assert report.modelo == "151"
    assert report.workbook_kind == "static_layout"
    assert report.evidence_tier == "layout_authority"
    assert report.formula_cells == 0
    assert "executable_parity_evidence" in report.not_evidence_for


def test_committed_record_design_xlsx_is_not_tax_formula_parity_oracle() -> None:
    root = bundled_path("corpus", "aeat_official")
    workbook_path = (
        root
        / "disenos_registro"
        / "modelo_303"
        / "files"
        / "01-303-ejercicio-2026-y-siguientes-actualizado-28-01-26-378-kb-xlsx.xlsx"
    )

    report = scan_workbook(workbook_path, root=root, options=WorkbookScanOptions(per_file_timeout_seconds=5))

    assert report.formula_cells > 0
    assert report.workbook_kind == "record_design_layout"
    assert report.evidence_tier == "layout_authority"
    assert "executable_parity_evidence" in report.not_evidence_for


def test_committed_modelo_131_record_designs_cover_current_and_historical_layouts() -> None:
    root = bundled_path("corpus", "aeat_official")
    workbook_paths = (
        root / "disenos_registro" / "modelo_131" / "files" / "05-131-ejercicios-2019-a-2023-116-kb-xlsx.xlsx",
        root
        / "disenos_registro"
        / "modelo_131"
        / "files"
        / "06-131-ejercicios-2024-actualizado-13-12-24-180-kb-xlsx.xlsx",
        root
        / "disenos_registro"
        / "modelo_131"
        / "files"
        / "07-131-ejercicios-2025-actualizado-11-12-25-179-kb-xlsx.xlsx",
        root
        / "disenos_registro"
        / "modelo_131"
        / "files"
        / "01-131-ejercicios-2026-actualizado-04-03-26-180-kb-xlsx.xlsx",
    )

    reports = tuple(
        scan_workbook(path, root=root, options=WorkbookScanOptions(per_file_timeout_seconds=5))
        for path in workbook_paths
    )

    assert {report.modelo for report in reports} == {"131"}
    assert all(report.workbook_kind == "record_design_layout" for report in reports)
    assert all(report.evidence_tier == "layout_authority" for report in reports)
    assert all("executable_parity_evidence" in report.not_evidence_for for report in reports)
    assert all(report.formula_cells > 0 for report in reports)


def test_committed_binary_xls_converts_to_layout_evidence_only() -> None:
    root = bundled_path("corpus", "aeat_official")
    workbook_path = (
        root
        / "disenos_registro"
        / "modelo_130"
        / "files"
        / "01-130-orden-hap-258-2015-ejercicios-2019-y-siguientes-actualizado-marzo-2019-176-kb-xls.xls"
    )

    report = convert_binary_xls_with_libreoffice(workbook_path, root=root)

    assert report.conversion_status == "converted"
    assert report.converted_extension == ".xlsx"
    assert report.formula_cells > 0
    assert report.workbook_kind == "record_design_layout"
    assert report.evidence_tier == "layout_authority"
    assert "executable_parity_evidence" in report.not_evidence_for


def test_committed_binary_xls_context_yields_temporary_xlsx_workbook() -> None:
    root = bundled_path("corpus", "aeat_official")
    workbook_path = (
        root
        / "disenos_registro"
        / "modelo_130"
        / "files"
        / "01-130-orden-hap-258-2015-ejercicios-2019-y-siguientes-actualizado-marzo-2019-176-kb-xls.xls"
    )

    with converted_binary_xls_with_libreoffice(workbook_path, root=root) as converted_path:
        assert converted_path.suffix == ".xlsx"
        workbook = load_workbook(converted_path, read_only=True, data_only=True)
        try:
            assert workbook.sheetnames
        finally:
            workbook.close()

    assert not converted_path.exists()


def test_libreoffice_runner_recalculates_xlsx_formula_workbook(tmp_path: Path) -> None:
    workbook_path = tmp_path / "modelo_303" / "files" / "303-live.xlsx"
    _write_formula_workbook(workbook_path)

    result = run_workbook_with_libreoffice(
        workbook_path,
        inputs={
            WorkbookCellRef(sheet="Modelo", coordinate="A1"): Decimal("12"),
            WorkbookCellRef(sheet="Modelo", coordinate="A2"): Decimal("30"),
        },
        outputs={"total": WorkbookCellRef(sheet="Modelo", coordinate="B1")},
    )

    assert result == {"total": 42}


def test_registry_workbook_parity_rejects_record_design_as_calculation_oracle(tmp_path: Path) -> None:
    snapshot = _committed_modelo_130_snapshot()
    workbook_path = tmp_path / "disenos_registro" / "modelo_130" / "files" / "130-layout.xlsx"
    _write_formula_workbook(workbook_path)
    workbook = scan_workbook(workbook_path, root=tmp_path)
    synthetic = SyntheticInputSet(
        id="modelo-130-record-design-rejected",
        modelo="130",
        revision="2019-y-siguientes",
        values=(
            SyntheticInputValue(
                id="casilla-01",
                value=Decimal("10000"),
                workbook_cell=WorkbookCellRef(sheet="Modelo", coordinate="A1"),
                registry_binding="01",
            ),
        ),
    )

    with pytest.raises(RegistryValidationError, match="not an executable calculation oracle"):
        run_registry_workbook_parity(
            snapshot=snapshot,
            synthetic_input=synthetic,
            workbook_path=workbook_path,
            workbook=workbook,
            output_cells={"casilla-19": WorkbookCellRef(sheet="Modelo", coordinate="B1")},
            registry_outputs={"casilla-19": _M130_CASILLA_19},
            date_context={"filing_period": date(2026, 3, 31)},
        )


def test_scan_workbook_records_binary_xls_as_unsupported(tmp_path: Path) -> None:
    workbook_path = tmp_path / "modelo_111" / "files" / "111-test.xls"
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    workbook_path.write_bytes(b"\xd0\xcf\x11\xe0")

    report = scan_workbook(workbook_path, root=tmp_path)

    assert report.modelo == "111"
    assert report.extension == ".xls"
    assert report.scan_status == "unsupported"
    assert report.workbook_kind == "unsupported_binary_xls"
    assert report.evidence_tier == "layout_authority"
    assert report.formula_cells == 0
    assert "isolated conversion" in (report.error or "")


def test_inventory_workbook_coverage_is_deterministic(tmp_path: Path) -> None:
    _write_formula_workbook(tmp_path / "modelo_303" / "files" / "b.xlsx")
    _write_formula_workbook(tmp_path / "modelo_303" / "files" / "a.xlsx")

    reports = inventory_workbook_coverage(tmp_path, options=WorkbookScanOptions(per_file_timeout_seconds=5))

    assert [report.path for report in reports] == [
        "modelo_303/files/a.xlsx",
        "modelo_303/files/b.xlsx",
    ]


def test_inventory_workbook_coverage_reuses_unchanged_previous_report(tmp_path: Path) -> None:
    workbook_path = tmp_path / "modelo_303" / "files" / "a.xlsx"
    _write_formula_workbook(workbook_path)
    first = inventory_workbook_coverage(tmp_path, options=WorkbookScanOptions(per_file_timeout_seconds=5))

    reports = inventory_workbook_coverage(
        tmp_path,
        options=WorkbookScanOptions(per_file_timeout_seconds=5),
        previous_reports=first,
    )

    assert reports == first


def test_discover_workbooks_requires_existing_root(tmp_path: Path) -> None:
    missing = tmp_path / "missing"

    with pytest.raises(RegistryValidationError, match="workbook root does not exist"):
        discover_workbooks(missing)


def test_compare_registry_to_workbook_reports_mismatch(tmp_path: Path) -> None:
    workbook_path = tmp_path / "modelo_303" / "files" / "303-test.xlsx"
    _write_formula_workbook(workbook_path)
    workbook = scan_workbook(workbook_path, root=tmp_path)
    synthetic = SyntheticInputSet(
        id="synthetic-303-basic",
        modelo="303",
        revision="2026",
        values=(
            SyntheticInputValue(
                id="base",
                value=Decimal("10"),
                workbook_cell=WorkbookCellRef(sheet="Modelo", coordinate="A1"),
                registry_binding="iva.base",
            ),
        ),
    )

    report = compare_registry_to_workbook(
        synthetic_input=synthetic,
        workbook=workbook,
        runner=verify_workbook_backend(tmp_path, scan_limit=1).runner,
        expected_workbook_values={"result": Decimal("31")},
        actual_registry_values={"result": Decimal("30")},
        output_cells={"result": WorkbookCellRef(sheet="Modelo", coordinate="B1", formula="=A1+A2")},
        registry_snapshot_id="303:2026:1T",
        legal_refs={"result": ("ley-37-1992:art-90",)},
        source_refs={"result": ("aeat-dr-303-2026",)},
    )

    assert report.status == "mismatch"
    assert report.registry_snapshot_id == "303:2026:1T"
    assert report.comparisons[0].status == "mismatch"
    assert report.comparisons[0].expected_workbook_value == Decimal("31")
    assert report.comparisons[0].legal_refs == ("ley-37-1992:art-90",)
    assert report.comparisons[0].source_refs == ("aeat-dr-303-2026",)


def test_compare_registry_to_workbook_rejects_missing_grounding(tmp_path: Path) -> None:
    workbook_path = tmp_path / "modelo_303" / "files" / "303-test.xlsx"
    _write_formula_workbook(workbook_path)
    workbook = scan_workbook(workbook_path, root=tmp_path)
    synthetic = SyntheticInputSet(
        id="synthetic-303-basic",
        modelo="303",
        revision="2026",
        values=(
            SyntheticInputValue(
                id="base",
                value=Decimal("10"),
                workbook_cell=WorkbookCellRef(sheet="Modelo", coordinate="A1"),
                registry_binding="iva.base",
            ),
        ),
    )

    with pytest.raises(RegistryValidationError, match="missing legal_refs"):
        compare_registry_to_workbook(
            synthetic_input=synthetic,
            workbook=workbook,
            runner=verify_workbook_backend(tmp_path, scan_limit=1).runner,
            expected_workbook_values={"result": Decimal("31")},
            actual_registry_values={"result": Decimal("31")},
            output_cells={"result": WorkbookCellRef(sheet="Modelo", coordinate="B1", formula="=A1+A2")},
            registry_snapshot_id="303:2026:1T",
        )

    with pytest.raises(ValidationError, match="legal_refs"):
        compare_registry_to_workbook(
            synthetic_input=synthetic,
            workbook=workbook,
            runner=verify_workbook_backend(tmp_path, scan_limit=1).runner,
            expected_workbook_values={"result": Decimal("31")},
            actual_registry_values={"result": Decimal("31")},
            output_cells={"result": WorkbookCellRef(sheet="Modelo", coordinate="B1", formula="=A1+A2")},
            registry_snapshot_id="303:2026:1T",
            legal_refs={"result": ("",)},
            source_refs={"result": ("aeat-dr-303-2026",)},
        )

    with pytest.raises(ValidationError, match="source_refs"):
        compare_registry_to_workbook(
            synthetic_input=synthetic,
            workbook=workbook,
            runner=verify_workbook_backend(tmp_path, scan_limit=1).runner,
            expected_workbook_values={"result": Decimal("31")},
            actual_registry_values={"result": Decimal("31")},
            output_cells={"result": WorkbookCellRef(sheet="Modelo", coordinate="B1", formula="=A1+A2")},
            registry_snapshot_id="303:2026:1T",
            legal_refs={"result": ("ley-37-1992:art-90",)},
            source_refs={"result": (" ",)},
        )


def test_compare_registry_to_workbook_rejects_missing_registry_output(tmp_path: Path) -> None:
    workbook_path = tmp_path / "modelo_303" / "files" / "303-test.xlsx"
    _write_formula_workbook(workbook_path)
    workbook = scan_workbook(workbook_path, root=tmp_path)
    synthetic = SyntheticInputSet(
        id="synthetic-303-basic",
        modelo="303",
        revision="2026",
        values=(
            SyntheticInputValue(
                id="base",
                value=Decimal("10"),
                workbook_cell=WorkbookCellRef(sheet="Modelo", coordinate="A1"),
                registry_binding="iva.base",
            ),
        ),
    )

    with pytest.raises(RegistryValidationError, match="output ids must match exactly"):
        compare_registry_to_workbook(
            synthetic_input=synthetic,
            workbook=workbook,
            runner=verify_workbook_backend(tmp_path, scan_limit=1).runner,
            expected_workbook_values={"result": None},
            actual_registry_values={},
            output_cells={"result": WorkbookCellRef(sheet="Modelo", coordinate="B1", formula="=A1+A2")},
            registry_snapshot_id="303:2026:1T",
        )


def test_parity_scenario_rejects_malformed_output_identifier(tmp_path: Path) -> None:
    synthetic = SyntheticInputSet(
        id="modelo-130-basic",
        modelo="130",
        revision="2019-y-siguientes",
        values=(
            SyntheticInputValue(
                id="casilla-01",
                value=Decimal("10000"),
                workbook_cell=WorkbookCellRef(sheet="Modelo", coordinate="A1"),
                registry_binding="01",
            ),
        ),
    )

    with pytest.raises(ValidationError):
        ParityScenario(
            id="modelo-130-basic",
            modelo="130",
            revision="2019-y-siguientes",
            filing_year=2026,
            period="1T",
            workbook_path=tmp_path / "modelo-130.xlsx",
            synthetic_input=synthetic,
            output_cells={"bad output": WorkbookCellRef(sheet="Modelo", coordinate="B1")},
            registry_outputs={"bad output": _M130_CASILLA_19},
        )


def test_parity_scenario_rejects_negative_tolerance(tmp_path: Path) -> None:
    synthetic = SyntheticInputSet(
        id="modelo-130-negative-tolerance",
        modelo="130",
        revision="2019-y-siguientes",
        values=(
            SyntheticInputValue(
                id="casilla-01",
                value=Decimal("10000"),
                workbook_cell=WorkbookCellRef(sheet="Modelo", coordinate="A1"),
                registry_binding="01",
            ),
        ),
    )

    with pytest.raises(ValidationError, match="tolerance"):
        ParityScenario(
            id="modelo-130-negative-tolerance",
            modelo="130",
            revision="2019-y-siguientes",
            filing_year=2026,
            period="1T",
            workbook_path=tmp_path / "modelo-130.xlsx",
            synthetic_input=synthetic,
            output_cells={"casilla-19": WorkbookCellRef(sheet="Modelo", coordinate="B1")},
            registry_outputs={"casilla-19": _M130_CASILLA_19},
            tolerance=Decimal("-0.01"),
        )


def test_parity_scenario_zero_tolerance_accepts_exact_comparison(tmp_path: Path) -> None:
    workbook_path = tmp_path / "modelo_303" / "files" / "303-test.xlsx"
    _write_formula_workbook(workbook_path)
    workbook = scan_workbook(workbook_path, root=tmp_path)
    synthetic = SyntheticInputSet(
        id="modelo-303-zero-tolerance",
        modelo="303",
        revision="2026",
        values=(
            SyntheticInputValue(
                id="base",
                value=Decimal("10"),
                workbook_cell=WorkbookCellRef(sheet="Modelo", coordinate="A1"),
                registry_binding="iva.base",
            ),
        ),
    )
    scenario = ParityScenario(
        id="modelo-303-zero-tolerance",
        modelo="303",
        revision="2026",
        filing_year=2026,
        period="1T",
        workbook_path=workbook_path,
        synthetic_input=synthetic,
        output_cells={"result": WorkbookCellRef(sheet="Modelo", coordinate="B1", formula="=A1+A2")},
        registry_outputs={"result": _M130_CASILLA_19},
        tolerance=Decimal("0"),
    )

    report = compare_registry_to_workbook(
        synthetic_input=scenario.synthetic_input,
        workbook=workbook,
        runner=verify_workbook_backend(tmp_path, scan_limit=1).runner,
        expected_workbook_values={"result": Decimal("31")},
        actual_registry_values={"result": Decimal("31")},
        output_cells=scenario.output_cells,
        registry_snapshot_id="303:2026:1T",
        legal_refs={"result": ("ley-37-1992:art-90",)},
        source_refs={"result": ("aeat-dr-303-2026",)},
        tolerance=scenario.tolerance,
    )

    assert report.status == "match"
    assert report.comparisons[0].tolerance == Decimal("0")


def test_verify_workbook_backend_reports_existing_backend(tmp_path: Path) -> None:
    _write_formula_workbook(tmp_path / "modelo_390" / "files" / "390-test.xlsx")

    report = verify_workbook_backend(tmp_path, scan_limit=10)

    assert report.backend_exists
    assert report.workbook_count == 1
    assert report.scanned_count == 1
    assert report.formula_workbook_count == 1
    assert report.failed_count == 0
    assert report.modelo_coverage[0].modelo == "390"
    assert report.modelo_coverage[0].formula_workbook_count == 1


def test_verify_workbook_backend_scans_all_workbooks_by_default(tmp_path: Path) -> None:
    _write_formula_workbook(tmp_path / "modelo_202" / "files" / "202-test.xlsx")
    _write_formula_workbook(tmp_path / "modelo_390" / "files" / "390-test.xlsx")

    report = verify_workbook_backend(tmp_path)

    assert report.workbook_count == 2
    assert report.scanned_count == 2
    assert {coverage.modelo for coverage in report.modelo_coverage} == {"202", "390"}


def test_verify_workbook_backend_fails_on_scan_errors_by_default(tmp_path: Path) -> None:
    workbook_path = tmp_path / "modelo_303" / "files" / "broken.xlsx"
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    workbook_path.write_text("not a workbook", encoding="utf-8")

    with pytest.raises(RegistryValidationError, match="failed to scan"):
        verify_workbook_backend(tmp_path)


def test_inventory_workbook_coverage_can_report_scan_errors_for_audit(tmp_path: Path) -> None:
    workbook_path = tmp_path / "modelo_303" / "files" / "broken.xlsx"
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    workbook_path.write_text("not a workbook", encoding="utf-8")

    reports = inventory_workbook_coverage(tmp_path)

    assert reports[0].scan_status == "failed"
    with pytest.raises(RegistryValidationError, match="failed to scan"):
        assert_workbook_scan_clean(
            verify_workbook_backend(
                tmp_path,
                fail_on_scan_error=False,
            ),
        )


def test_libreoffice_runner_rejects_explicit_missing_executable(tmp_path: Path) -> None:
    workbook_path = tmp_path / "formula.xlsx"
    _write_formula_workbook(workbook_path)

    with pytest.raises(RegistryValidationError, match="LibreOffice executable does not exist"):
        run_workbook_with_libreoffice(
            workbook_path,
            inputs={},
            outputs={},
            executable=str(tmp_path / "missing-soffice"),
        )


def test_binary_xls_conversion_error_code_is_registered() -> None:
    assert issubclass(_BinaryXlsConversionError, RuntimeError)
