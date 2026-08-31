"""Real-binary tests for the record-design generator intermediate representation."""

from __future__ import annotations

import ast
import inspect
from pathlib import Path

import pytest
from openpyxl import load_workbook

from cadrumo.core.resources import bundled_path
from cadrumo.domain.calculations.registry.corpus_catalogue import resolve_record_design_binary
from cadrumo.domain.calculations.registry.errors import RegistryValidationError
from cadrumo.domain.calculations.registry.loader import load_catalogue_file
from cadrumo.domain.calculations.registry.record_design import extract_record_design
from cadrumo.domain.calculations.registry.record_design_schema import (
    RecordDesignCompositeRelativeClosing,
    RecordDesignRelativeSuffixMarker,
)

from ..pipeline import _record_design_ir
from ..pipeline._record_design_ir import (
    RECORD_DESIGN_INTERMEDIATE_SCHEMA_VERSION,
    RecordDesignIntermediateCompositeRelativeClosing,
    RecordDesignIntermediateRelativeSuffixMarker,
    RecordDesignWorkbookFormat,
    load_record_design_intermediate,
)

pytestmark = [pytest.mark.unit, pytest.mark.hex_core]

_MODELO_303_DESIGNS = (
    ("aeat-dr-303-2023", 2023, "2023"),
    ("aeat-dr-303-2024-early", 2024, "2024-early"),
    ("aeat-dr-303-2024-late", 2024, "2024-late"),
    ("aeat-dr-303-2025", 2025, "2025"),
    ("aeat-dr-303-2026", 2026, "2026"),
)
_MODELO_220_DESIGNS = (
    ("aeat-dr-220-2023", 2023, "2023"),
    ("aeat-dr-220-2024", 2024, "2024"),
    ("aeat-dr-220-2025", 2025, "2025"),
)
_MODELO_390_DESIGNS = (
    ("aeat-dr-390-2022", 2022, "2022"),
    ("aeat-dr-390-2023", 2023, "2023"),
    ("aeat-dr-390-2024", 2024, "2024"),
    ("aeat-dr-390-2025", 2025, "2025"),
)


def test_intermediate_is_a_complete_total_preserving_projection_of_the_verified_workbook() -> None:
    """Every parsed record and official total reaches the generator IR unchanged."""
    source_root = bundled_path()
    catalogues = load_catalogue_file(bundled_path("registry", "aeat", "legal", "is.toml"))
    resolved = resolve_record_design_binary(
        source_root,
        catalogues.sources,
        source_ref="aeat-dr-200-2025",
        filing_year=2025,
        design_epoch="2025",
    )
    parsed_sheets = extract_record_design(resolved.path).accept_partial()

    intermediate = load_record_design_intermediate(
        source_root,
        catalogues.sources,
        source_ref="aeat-dr-200-2025",
        filing_year=2025,
        design_epoch="2025",
    )

    assert intermediate.source.source_ref == resolved.source.id
    assert intermediate.source.source_sha256 == resolved.source.sha256
    assert intermediate.source.design_epoch == resolved.source.record_design_epoch
    assert intermediate.source.workbook_format is RecordDesignWorkbookFormat.XLS
    fixed_parser_sheets = tuple(sheet for sheet in parsed_sheets if sheet.variable_envelope is None)
    assert tuple(sheet.sheet for sheet in intermediate.sheets) == tuple(sheet.name for sheet in fixed_parser_sheets)
    assert tuple(envelope.sheet for envelope in intermediate.variable_envelopes) == tuple(
        sheet.name for sheet in parsed_sheets if sheet.variable_envelope is not None
    )

    for parser_sheet, intermediate_sheet in zip(fixed_parser_sheets, intermediate.sheets, strict=True):
        assert intermediate_sheet.sheet == parser_sheet.name
        assert intermediate_sheet.record_identity == parser_sheet.name
        assert intermediate_sheet.declared_total == parser_sheet.total_positions
        assert intermediate_sheet.declared_total is not None
        assert (
            max(field.offset + field.length - 1 for field in parser_sheet.fields) == intermediate_sheet.declared_total
        )
        assert tuple(
            (
                field.sheet,
                field.record_identity,
                field.source_row,
                field.source_cell,
                field.ordinal,
                field.offset,
                field.length,
                field.aeat_type,
                field.normalized_description,
                field.validation,
                field.content,
            )
            for field in intermediate_sheet.fields
        ) == tuple(
            (
                field.sheet,
                parser_sheet.name,
                field.row,
                f"A{field.row}",
                field.ordinal,
                field.offset,
                field.length,
                field.type_code,
                field.description,
                field.validation,
                field.content,
            )
            for field in parser_sheet.fields
        )

    envelope = next(envelope for envelope in intermediate.variable_envelopes if envelope.sheet == "DP200000")
    parser_envelope = next(sheet.variable_envelope for sheet in parsed_sheets if sheet.name == envelope.sheet)
    assert parser_envelope is not None
    assert isinstance(envelope.closing, RecordDesignIntermediateRelativeSuffixMarker)
    assert isinstance(parser_envelope.closing, RecordDesignRelativeSuffixMarker)
    assert envelope.sheet == "DP200000"
    assert envelope.record_identity == "DP200000"
    assert tuple(
        (
            field.sheet,
            field.record_identity,
            field.source_row,
            field.source_cell,
            field.ordinal,
            field.offset,
            field.length,
            field.aeat_type,
            field.normalized_description,
            field.validation,
            field.content,
        )
        for field in envelope.prefix_fields
    ) == tuple(
        (
            field.sheet,
            parser_envelope.name,
            field.row,
            f"A{field.row}",
            field.ordinal,
            field.offset,
            field.length,
            field.type_code,
            field.description,
            field.validation,
            field.content,
        )
        for field in parser_envelope.prefix_fields
    )
    assert envelope.prefix_extent == 328
    assert max(field.offset + field.length - 1 for field in envelope.prefix_fields) == 328
    assert (
        envelope.body_source_row,
        envelope.body_source_cell,
        envelope.body_ordinal,
        envelope.body_offset,
        envelope.body_length,
        envelope.body_aeat_type,
        envelope.body_normalized_description,
        envelope.body_validation,
        envelope.body_content,
    ) == (
        parser_envelope.body.row,
        f"A{parser_envelope.body.row}",
        parser_envelope.body.ordinal,
        parser_envelope.body.offset,
        parser_envelope.body.length,
        parser_envelope.body.type_code,
        parser_envelope.body.description,
        parser_envelope.body.validation,
        parser_envelope.body.content,
    )
    assert (envelope.body_source_cell, envelope.body_offset, envelope.body_length) == (
        "A14",
        329,
        "Variable",
    )
    assert envelope.body_aeat_type == "An"
    assert envelope.body_normalized_description.startswith("Contenido del fichero")
    assert envelope.body_validation is None
    assert envelope.body_content is None
    assert (
        envelope.closing.source_row,
        envelope.closing.source_cell,
        envelope.closing.ordinal,
        envelope.closing.offset,
        envelope.closing.length,
        envelope.closing.aeat_type,
        envelope.closing.normalized_description,
        envelope.closing.validation,
        envelope.closing.content,
    ) == (
        parser_envelope.closing.row,
        f"A{parser_envelope.closing.row}",
        parser_envelope.closing.ordinal,
        parser_envelope.closing.offset,
        parser_envelope.closing.length,
        parser_envelope.closing.type_code,
        parser_envelope.closing.description,
        parser_envelope.closing.validation,
        parser_envelope.closing.content,
    )
    assert (
        envelope.closing.source_cell,
        envelope.closing.offset,
        envelope.closing.length,
    ) == ("A15", "***", 18)
    assert envelope.closing.aeat_type == "An"
    assert envelope.closing.normalized_description.startswith("Constante. </T")
    assert envelope.closing.validation is None
    assert envelope.closing.content == '"</T200020250A0000>"'
    assert (
        envelope.total_source_row,
        envelope.total_source_cell,
        envelope.total_label,
        envelope.total_length,
    ) == (
        parser_envelope.variable_total.row,
        f"A{parser_envelope.variable_total.row}",
        parser_envelope.variable_total.label,
        parser_envelope.variable_total.length,
    )
    assert (
        envelope.total_source_cell,
        envelope.total_label,
        envelope.total_length,
    ) == ("A16", "total", "Variable")


def _official_totals(path: Path) -> dict[str, int]:
    """Read every sheet's official ``Total:`` cell straight from the real binary.

    The reader is chosen by the CATALOGUE path's suffix. AEAT publishes many
    disenos in the pre-OOXML format, and taking the format from the AEAT
    filename would be wrong: this design's name embeds "xls" mid-string
    (``01-200-ejercicio-2025-10-9-mb-xls.xls``) and other designs carry the same
    token while being genuine .xlsx. openpyxl refuses a legacy .xls outright, so
    that branch reads through xlrd, which the record-design parser already uses
    for these binaries.
    """
    if path.suffix.casefold() == ".xls":
        import xlrd

        book = xlrd.open_workbook(str(path), on_demand=True)
        try:
            totals: dict[str, int] = {}
            for name in book.sheet_names():
                sheet = book.sheet_by_name(name)
                for row in range(sheet.nrows):
                    if sheet.ncols < 3:
                        continue
                    label = sheet.cell_value(row, 0)
                    value = sheet.cell_value(row, 2)
                    if not isinstance(label, str) or label.strip().casefold() != "total:":
                        continue
                    # xlrd surfaces every number as float; only a whole number is
                    # a declared total, matching the openpyxl branch's int check.
                    if isinstance(value, float) and value.is_integer():
                        totals[name.strip()] = int(value)
            return totals
        finally:
            book.release_resources()

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return {
            worksheet.title.strip(): value
            for worksheet in workbook.worksheets
            for label, _unused, value in worksheet.iter_rows(min_col=1, max_col=3, values_only=True)
            if isinstance(label, str) and label.strip().casefold() == "total:" and isinstance(value, int)
        }
    finally:
        workbook.close()


def test_intermediate_recovers_every_official_total_colon_without_fixing_the_variable_envelope() -> None:
    """The real workbook's ``Total:`` cells govern fixed totals; DP200000 stays variable."""
    source_root = bundled_path()
    catalogues = load_catalogue_file(bundled_path("registry", "aeat", "legal", "is.toml"))
    resolved = resolve_record_design_binary(
        source_root,
        catalogues.sources,
        source_ref="aeat-dr-200-2025",
        filing_year=2025,
        design_epoch="2025",
    )
    intermediate = load_record_design_intermediate(
        source_root,
        catalogues.sources,
        source_ref="aeat-dr-200-2025",
        filing_year=2025,
        design_epoch="2025",
    )
    official_totals = _official_totals(resolved.path)

    assert official_totals
    assert {sheet.sheet for sheet in intermediate.sheets} == set(official_totals)
    assert {sheet.sheet: sheet.declared_total for sheet in intermediate.sheets} == official_totals
    envelope = next(envelope for envelope in intermediate.variable_envelopes if envelope.sheet == "DP200000")
    assert envelope.prefix_extent == 328
    assert envelope.body_offset == 329
    assert envelope.body_length == "Variable"
    assert envelope.total_length == "Variable"


@pytest.mark.parametrize(("source_ref", "filing_year", "design_epoch"), _MODELO_303_DESIGNS)
def test_intermediate_preserves_each_modelo_303_variable_envelope(
    source_ref: str,
    filing_year: int,
    design_epoch: str,
) -> None:
    """Every real M303 parser envelope reaches the typed IR without becoming fixed."""
    source_root = bundled_path()
    catalogues = load_catalogue_file(bundled_path("registry", "aeat", "legal", "iva.toml"))
    resolved = resolve_record_design_binary(
        source_root,
        catalogues.sources,
        source_ref=source_ref,
        filing_year=filing_year,
        design_epoch=design_epoch,
    )
    parsed = extract_record_design(resolved.path).accept_partial()

    intermediate = load_record_design_intermediate(
        source_root,
        catalogues.sources,
        source_ref=source_ref,
        filing_year=filing_year,
        design_epoch=design_epoch,
    )

    parser_envelopes = tuple(sheet.variable_envelope for sheet in parsed if sheet.variable_envelope is not None)
    assert len(intermediate.sheets) == 6
    assert all(sheet.record_identity != "DP30300" for sheet in intermediate.sheets)
    assert len(parser_envelopes) == len(intermediate.variable_envelopes) == 1
    parser_envelope = parser_envelopes[0]
    assert parser_envelope is not None
    envelope = intermediate.variable_envelopes[0]
    assert isinstance(envelope.closing, RecordDesignIntermediateRelativeSuffixMarker)
    assert isinstance(parser_envelope.closing, RecordDesignRelativeSuffixMarker)
    assert envelope.sheet == envelope.record_identity == parser_envelope.name == "DP30300"
    assert envelope.prefix_extent == parser_envelope.prefix_extent == 328
    assert (envelope.body_source_row, envelope.body_ordinal, envelope.body_offset, envelope.body_length) == (
        parser_envelope.body.row,
        parser_envelope.body.ordinal,
        parser_envelope.body.offset,
        parser_envelope.body.length,
    )
    assert (
        envelope.closing.source_row,
        envelope.closing.ordinal,
        envelope.closing.offset,
        envelope.closing.length,
    ) == (
        parser_envelope.closing.row,
        parser_envelope.closing.ordinal,
        parser_envelope.closing.offset,
        parser_envelope.closing.length,
    )
    assert (envelope.total_source_row, envelope.total_label, envelope.total_length) == (
        parser_envelope.variable_total.row,
        parser_envelope.variable_total.label,
        parser_envelope.variable_total.length,
    )


@pytest.mark.parametrize(("source_ref", "filing_year", "design_epoch"), _MODELO_220_DESIGNS)
def test_intermediate_preserves_each_modelo_220_composite_relative_closing(
    source_ref: str,
    filing_year: int,
    design_epoch: str,
) -> None:
    """The public SHA-bound loader retains all six M220 closing rows in typed IR."""
    source_root = bundled_path()
    catalogues = load_catalogue_file(bundled_path("registry", "aeat", "legal", "is.toml"))
    resolved = resolve_record_design_binary(
        source_root,
        catalogues.sources,
        source_ref=source_ref,
        filing_year=filing_year,
        design_epoch=design_epoch,
    )
    parser_envelope = next(
        sheet.variable_envelope
        for sheet in extract_record_design(resolved.path).accept_partial()
        if sheet.variable_envelope is not None
    )
    intermediate = load_record_design_intermediate(
        source_root,
        catalogues.sources,
        source_ref=source_ref,
        filing_year=filing_year,
        design_epoch=design_epoch,
    )

    assert resolved.source.sha256 == intermediate.source.source_sha256
    assert RECORD_DESIGN_INTERMEDIATE_SCHEMA_VERSION == 4
    assert len(intermediate.variable_envelopes) == 1
    envelope = intermediate.variable_envelopes[0]
    assert isinstance(parser_envelope.closing, RecordDesignCompositeRelativeClosing)
    assert isinstance(envelope.closing, RecordDesignIntermediateCompositeRelativeClosing)
    assert tuple(
        (
            part.source_row,
            part.source_cell,
            part.ordinal,
            part.offset,
            part.length,
            part.aeat_type,
            part.normalized_description,
            part.validation,
            part.content,
        )
        for part in envelope.closing.parts
    ) == tuple(
        (
            part.row,
            f"A{part.row}",
            part.ordinal,
            part.offset,
            part.length,
            part.type_code,
            part.description,
            part.validation,
            part.content,
        )
        for part in parser_envelope.closing.parts
    )
    assert tuple(part.content for part in envelope.closing.parts) == (
        "</T",
        "220",
        "(*)[A|E|I|0]",
        None,
        "0A",
        "0000>",
    )


@pytest.mark.parametrize(("source_ref", "filing_year", "design_epoch"), _MODELO_390_DESIGNS)
def test_intermediate_classifies_each_modelo_390_page_zero_as_a_total_less_auxiliary_header(
    source_ref: str,
    filing_year: int,
    design_epoch: str,
) -> None:
    """Real M390 page zero is retained once as its 13-anchor non-fixed composition header."""
    source_root = bundled_path()
    catalogues = load_catalogue_file(bundled_path("registry", "aeat", "legal", "iva.toml"))
    parsed = extract_record_design(
        resolve_record_design_binary(
            source_root,
            catalogues.sources,
            source_ref=source_ref,
            filing_year=filing_year,
            design_epoch=design_epoch,
        ).path,
    ).accept_partial()
    intermediate = load_record_design_intermediate(
        source_root,
        catalogues.sources,
        source_ref=source_ref,
        filing_year=filing_year,
        design_epoch=design_epoch,
    )

    page_zero = next(sheet for sheet in parsed if sheet.name == "Pág. 0")
    assert page_zero.total_positions is None
    assert len(intermediate.auxiliary_envelope_headers) == 1
    (header,) = intermediate.auxiliary_envelope_headers
    assert header.sheet == header.record_identity == page_zero.name
    assert header.emitted_extent == 328
    assert len(header.fields) == len(page_zero.fields) == 13
    assert tuple(field.parser_field.offset for field in header.fields) == (
        1,
        3,
        6,
        7,
        11,
        13,
        18,
        23,
        93,
        97,
        101,
        110,
        323,
    )
    assert tuple(field.parser_field.length for field in header.fields) == (2, 3, 1, 4, 2, 5, 5, 70, 4, 4, 9, 213, 6)
    assert tuple(field.parser_field.content for field in header.fields)[-1] == '"</AUX>"'
    assert all(sheet.record_identity != page_zero.name for sheet in intermediate.sheets)


@pytest.mark.parametrize(
    ("source_ref", "filing_year", "design_epoch", "message"),
    (
        ("aeat-dr-200-2024", 2025, "2024", "does not apply to filing year 2025"),
        ("aeat-dr-200-2025", 2025, "2024", "declares design epoch '2025', not requested '2024'"),
    ),
)
def test_intermediate_refuses_an_inapplicable_or_wrong_epoch_source_before_parsing(
    source_ref: str,
    filing_year: int,
    design_epoch: str,
    message: str,
) -> None:
    """The generator cannot parse a source which registry authority did not select."""
    source_root = bundled_path()
    catalogues = load_catalogue_file(bundled_path("registry", "aeat", "legal", "is.toml"))

    with pytest.raises(RegistryValidationError, match=message):
        load_record_design_intermediate(
            source_root,
            catalogues.sources,
            source_ref=source_ref,
            filing_year=filing_year,
            design_epoch=design_epoch,
        )


def test_intermediate_refuses_a_hash_drifting_source_before_parser_projection() -> None:
    """A changed catalogue digest blocks the real workbook before parser output exists."""
    source_root = bundled_path()
    catalogues = load_catalogue_file(bundled_path("registry", "aeat", "legal", "is.toml"))
    source = catalogues.sources["aeat-dr-200-2025"]
    drifting_source = source.model_copy(update={"sha256": "0" * 64})

    with pytest.raises(RegistryValidationError, match="sha256 mismatch"):
        load_record_design_intermediate(
            source_root,
            {str(drifting_source.id): drifting_source},
            source_ref="aeat-dr-200-2025",
            filing_year=2025,
            design_epoch="2025",
        )


def test_intermediate_loader_has_no_derivative_or_legacy_fallback_access_path() -> None:
    """The loader stays a two-authority handoff: catalogue selection then shipped parser."""
    module = ast.parse(inspect.getsource(_record_design_ir))
    loader = next(
        node
        for node in module.body
        if isinstance(node, ast.FunctionDef) and node.name == "load_record_design_intermediate"
    )
    call_names = {
        node.func.id for node in ast.walk(loader) if isinstance(node, ast.Call) and isinstance(node.func, ast.Name)
    }

    assert call_names == {"_build_record_design_intermediate", "extract_record_design", "resolve_record_design_binary"}
