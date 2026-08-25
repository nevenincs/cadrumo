"""Diseno de Registro workbook extractor for the dev-side RAG index.

Materialises the official AEAT Diseno de Registro workbooks (the
``casilla-number to field-position`` tables that the registry parity gates
consume) as schema-conformant extraction sidecars so the resident
``vaultspec-rag`` walker indexes the highest-value grounding surface. A
"what is field/casilla X" query needs exactly this table: the field number,
its byte position in the fichero, its length, type, description, validation,
and constant content.

The workbooks ship in two formats. The modern ``.xlsx`` is read with
``openpyxl``; the 28 legacy binary ``.xls`` are read with ``xlrd`` (which
opens the old BIFF format ``openpyxl`` cannot). Both are locked project
dependencies (the offline xls export surface already uses ``openpyxl``).

Each worksheet becomes one :class:`PreprocessUnit`: a readable rendering of
the field table headed by the sheet name. The unit text is plain rows so the
walker's text chunker keeps each field row's tokens (the field number, the
position, the Spanish description) together. Attribution is pulled from the
modelo's ``manifest.json`` (the AEAT Sede source page and the per-artefact
download URL), honouring the BOE/AEAT reuse-with-attribution obligation.

This is the production Diseno de Registro preprocessor. When the upstream
``vaultspec-rag`` preprocess-hook lands, this module re-targets the upstream
sink and the committed sidecars are retired; ``PreprocessOutput`` is the
forward-compatible precursor that makes that migration mechanical.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Final

from ..._paths import UTF_8
from ._parts import (
    split_units_by_budget,
    stamp_part_anchors,
    write_part_sidecars,
)
from ._schema import (
    ExtractionStatus,
    PreprocessOutput,
    PreprocessUnit,
    SourceDocumentKind,
)
from ._sidecar import (
    PreprocessSidecarError,
    sha256_of,
)

#: Stable id of this extractor, recorded in every sidecar's provenance.
WORKBOOK_EXTRACTOR_ID = "diseno-registro-workbook"

#: Version of this extractor; part of the cache identity. Bump when the
#: rendering changes so a regeneration is distinguishable from a no-op.
WORKBOOK_EXTRACTOR_VERSION = "1.0"
_UTF_8: Final[str] = UTF_8

#: Column separator in the rendered field table. A pipe keeps each row on
#: one line and survives the text chunker without being mistaken for prose.
_CELL_SEP = " | "


def _norm_cell(value: object) -> str:
    """Render one cell value as a trimmed string.

    ``xlrd`` surfaces numeric cells as floats (``1.0``); whole numbers are
    rendered without the trailing ``.0`` so a field number reads as ``1``
    rather than ``1.0``. Non-numeric values stringify as-is. ``None`` (an
    empty cell) becomes the empty string.
    """
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _render_sheet(name: str, rows: list[list[object]]) -> str:
    """Render one worksheet's field table as readable plain text.

    Args:
        name: The worksheet name (e.g. ``DP30300``) used as the unit title.
        rows: The sheet's rows as lists of raw cell values.

    Returns:
        A plain-text block: the sheet name heading followed by one line per
        non-empty row, cells joined by the pipe separator. Fully empty rows
        are dropped so blank spacer rows do not bloat the index.
    """
    lines: list[str] = []
    for row in rows:
        cells = [_norm_cell(c) for c in row]
        if not any(cells):
            continue
        # Drop trailing empty cells so a 7-column sheet with a short row
        # does not render a run of dangling separators.
        while cells and cells[-1] == "":
            cells.pop()
        if cells:
            lines.append(_CELL_SEP.join(cells))
    return "\n".join(lines)


def _read_xlsx_sheets(source: Path) -> list[tuple[str, list[list[object]]]]:
    """Read every worksheet of an ``.xlsx`` workbook via openpyxl.

    Returns a list of ``(sheet_name, rows)`` pairs; ``rows`` is a list of
    lists of raw cell values (``data_only`` resolves formulas to their last
    cached value, which is what a Diseno table carries).
    """
    import openpyxl

    sheets: list[tuple[str, list[list[object]]]] = []
    workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
    try:
        for name in workbook.sheetnames:
            worksheet = workbook[name]
            rows: list[list[object]] = [[cell for cell in row] for row in worksheet.iter_rows(values_only=True)]
            sheets.append((name, rows))
    finally:
        workbook.close()
    return sheets


def _read_xls_sheets(source: Path) -> list[tuple[str, list[list[object]]]]:
    """Read every worksheet of a legacy ``.xls`` workbook via xlrd.

    ``xlrd`` 2.x opens the binary BIFF ``.xls`` format that openpyxl cannot.
    Cell values come through typed (numbers as floats); :func:`_norm_cell`
    normalises whole-number floats so positions and lengths read as ints.
    """
    import xlrd

    sheets: list[tuple[str, list[list[object]]]] = []
    # ``xlrd`` (loose-typed: it ships no precise type stubs for the workbook
    # and sheet objects) returns ``Book`` / ``Sheet`` handles.
    workbook: Any = xlrd.open_workbook(str(source))
    for worksheet in workbook.sheets():
        rows: list[list[object]] = [
            [worksheet.cell_value(r, c) for c in range(worksheet.ncols)] for r in range(worksheet.nrows)
        ]
        sheets.append((worksheet.name, rows))
    return sheets


def _attribution_for(source: Path) -> str:
    """Build the attribution string for a workbook from its modelo manifest.

    Reads the sibling ``manifest.json`` (one directory up from the
    ``files/`` folder), matching the artefact whose ``stored_path`` ends with
    the source filename, and composes the AEAT Sede source plus the official
    download URL. Falls back to the standing AEAT Sede attribution when the
    manifest or the matching artefact is absent (an extraction must never
    ship corpus text with no attribution).
    """
    base = (
        "Source: Agencia Estatal de Administracion Tributaria (AEAT) - "
        "Disenos de registro (Sede Electronica). Reused with attribution."
    )
    manifest_path = source.parent.parent / "manifest.json"
    if not manifest_path.is_file():
        return base
    try:
        manifest = json.loads(manifest_path.read_text(encoding=_UTF_8))
    except (OSError, json.JSONDecodeError):
        return base
    artefacts = manifest.get("artefacts", [])
    if not isinstance(artefacts, list):
        return base
    for artefact in artefacts:
        if not isinstance(artefact, dict):
            continue
        stored = str(artefact.get("stored_path", ""))
        if stored.endswith(source.name):
            url = str(artefact.get("url", "")).strip()
            if url:
                return f"{base} Official source: {url}"
            return base
    return base


def build_outputs(source: Path, *, repo_root: Path) -> list[PreprocessOutput]:
    """Extract a Diseno de Registro workbook into one or more records.

    One :class:`PreprocessUnit` is produced per worksheet. When the combined
    rendered text would exceed the per-sidecar byte budget the units are
    grouped so each resulting :class:`PreprocessOutput` renders under the
    walker's file cap; a single workbook then yields one sidecar pair per
    group, each carrying a ``part`` anchor so the groups are distinguishable.

    Args:
        source: Path to the ``.xls`` or ``.xlsx`` workbook.
        repo_root: Repository root for the POSIX-relative source locator.

    Returns:
        A list of validated records (one per sidecar pair to write).

    Raises:
        PreprocessSidecarError: If the workbook cannot be read at all (a
            hard failure - never a silently empty ``ok`` sidecar).
    """
    suffix = source.suffix.lower()
    try:
        if suffix in {".xlsx", ".xlsm"}:
            sheets = _read_xlsx_sheets(source)
        elif suffix == ".xls":
            sheets = _read_xls_sheets(source)
        else:
            raise PreprocessSidecarError(f"unsupported workbook extension {suffix!r} for {source}")
    except PreprocessSidecarError:
        raise
    except Exception as exc:  # reader-specific failure: hard-fail, do not skip
        raise PreprocessSidecarError(f"cannot read workbook {source}: {exc}") from exc

    relpath = source.resolve().relative_to(repo_root.resolve()).as_posix()
    digest = sha256_of(source)
    attribution = _attribution_for(source)

    units: list[PreprocessUnit] = []
    for name, rows in sheets:
        rendered = _render_sheet(name, rows)
        if rendered:
            units.append(PreprocessUnit(text=rendered, title=name, section=name))

    if not units:
        return [
            PreprocessOutput(
                source_kind=SourceDocumentKind.DISENO_REGISTRO_WORKBOOK,
                status=ExtractionStatus.EMPTY,
                source_relpath=relpath,
                source_sha256=digest,
                preprocessor_id=WORKBOOK_EXTRACTOR_ID,
                preprocessor_version=WORKBOOK_EXTRACTOR_VERSION,
                attribution=attribution,
                units=(),
            ),
        ]

    groups = stamp_part_anchors(split_units_by_budget(units))
    return [
        PreprocessOutput(
            source_kind=SourceDocumentKind.DISENO_REGISTRO_WORKBOOK,
            status=ExtractionStatus.OK,
            source_relpath=relpath,
            source_sha256=digest,
            preprocessor_id=WORKBOOK_EXTRACTOR_ID,
            preprocessor_version=WORKBOOK_EXTRACTOR_VERSION,
            attribution=attribution,
            units=tuple(group),
        )
        for group in groups
    ]


def extract_workbook(source: Path, *, repo_root: Path) -> list[Path]:
    """Extract a workbook and write its committed sidecar pair(s).

    Args:
        source: Path to the ``.xls`` or ``.xlsx`` workbook.
        repo_root: Repository root for the POSIX-relative source locator.

    Returns:
        The list of ``*.extracted.md`` text-sidecar paths written (one per
        part), in part order.

    Raises:
        PreprocessSidecarError: If the workbook cannot be read or a sidecar
            cannot be written.
    """
    outputs = build_outputs(source, repo_root=repo_root)
    return write_part_sidecars(source, outputs)
